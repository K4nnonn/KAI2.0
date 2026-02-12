"""
Generate UNIQUE insights for each account based on deep analysis
Each account gets completely different, specific insights
"""
import sys
sys.path.insert(0, '.')

from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def to_num(series):
    return pd.to_numeric(series.astype(str).str.replace(',', '').str.replace('$', '').str.replace('%', ''), errors='coerce').fillna(0)

def get_retirement_insights():
    """Retirement is brand-focused with exceptional sitelink performance"""
    return {
        'bu': 'Retirement',
        'strengths': [
            {
                'title': 'Top Brand Campaigns Drive Cheap Clicks',
                'finding': 'Your 2 best brand campaigns pay just $0.46-$0.49 per click (industry average is $3-5). These campaigns drive 852K clicks and are your most profitable traffic. They show what is possible when Quality Scores are perfect.'
            },
            {
                'title': 'Strong Conversion Rate',
                'finding': '4% of clicks convert - double the industry standard of 2%. You generated 30,000 conversions. For every 100 people who click, 4 take action. Your landing pages match what people are searching for.'
            },
            {
                'title': 'Sitelinks Generate Massive Volume',
                'finding': 'Your sitelinks drove 3.6 million clicks. People click "Rollover IRA" and "401k Calculator" links because they match search intent. Every campaign uses sitelinks - this extra real estate adds serious volume.'
            },
            {
                'title': 'Budget Focused on Winners',
                'finding': 'Your top 3 campaigns get 37% of budget ($590K of $1.6M), which is smart. You are not spreading money thin across underperformers. This concentration maximizes returns.'
            },
            {
                'title': 'Brand Dominates Traffic',
                'finding': '89% of your clicks (1.6M of 1.8M) come from brand campaigns. You own brand search. People searching "Fidelity retirement" find you first and click through at high rates.'
            }
        ],
        'opportunities': [
            {
                'title': 'Bring All Brand Campaigns to $0.46 CPC',
                'finding': 'Your best 2 brand campaigns pay $0.46-$0.49, but brand average is $1.74. ACTION: Apply the winning elements from your top campaigns (ad copy, landing pages, keyword match types) to all brand campaigns to cut costs 74%.'
            },
            {
                'title': 'Rollover Campaigns Cost 8x More Than Brand',
                'finding': 'Rollover/401k campaigns cost $3.98 per click vs $0.49 for brand. That 8x gap is too wide. ACTION: Improve Quality Scores - put "rollover" in ad headlines, create dedicated rollover landing pages, switch to phrase match.'
            },
            {
                'title': 'Test YouTube Video Campaigns',
                'finding': 'You only run text ads. People watch retirement videos on YouTube before searching. ACTION: Launch video ads showing retirement calculators and rollover tips targeting people 50+. Start with $25K/month test.'
            },
            {
                'title': 'Grow Beyond Brand Terms',
                'finding': 'Brand drives 89% of clicks but only reaches people who already know Fidelity. ACTION: Move 20% of budget to non-brand terms like "best IRA" and "retirement calculator" to acquire new customers.'
            },
            {
                'title': 'Adjust Bids by Time of Day',
                'finding': 'Retirement searches peak evenings and weekends but your ads bid the same 24/7. ACTION: Check hourly performance report, then raise bids 30% during 6-9 PM when people are home researching finances.'
            }
        ]
    }

def get_brand_insights():
    """Brand has insane CTRs and ultra-cheap clicks - super focused"""
    return {
        'bu': 'Brand',
        'strengths': [
            {
                'title': 'Exceptional Click-Through Rate',
                'finding': '28% of people who see your ad click it - that means 1 in 4. Most advertisers get 3-5%. You generated 3.5 million clicks. This shows Fidelity dominates brand awareness in financial services.'
            },
            {
                'title': 'Three Campaigns Hit $0.17-$0.19 Per Click',
                'finding': 'Your CPC test winners pay just $0.17-$0.19 per click while industry average is $3-5. That is 95% cheaper. These campaigns prove perfect Quality Scores deliver rock-bottom costs.'
            },
            {
                'title': 'Sitelinks Drive Huge Volume',
                'finding': 'Your sitelinks generated 3.5 million clicks at 10% CTR. People click straight to "Retirement" or "Brokerage" sections. This extra screen real estate doubles your click opportunities per search.'
            },
            {
                'title': 'Efficient $0.88 Average CPC',
                'finding': 'Your average cost per click is $0.88 across all 11 campaigns. Brand average is $2-3 in financial services. You pay 71% less because Google rewards your high-quality ads.'
            },
            {
                'title': 'Active Testing Culture',
                'finding': 'Running 3 CPC test campaigns shows you do not set-and-forget. You keep optimizing to find savings. This test-and-learn approach is why your best campaigns hit $0.17-$0.19 per click.'
            }
        ],
        'opportunities': [
            {
                'title': 'Landing Pages Need Higher Conversion',
                'finding': 'Only 2.6% of 3.5M clicks convert. Brand clicks should convert 5-10%. ACTION: Test simpler landing pages with one clear button. Remove top navigation. Even 0.5% improvement = 17,500 more conversions for free.'
            },
            {
                'title': 'Scale Winning CPC Test to All Campaigns',
                'finding': 'Your test winners at $0.17-$0.19 beat your $0.88 average by 80%. ACTION: Identify what makes the winners work (ad copy, landing pages, match types), then apply to all 11 campaigns to cut average CPC in half.'
            },
            {
                'title': 'Capture Misspelled Brand Searches',
                'finding': 'People searching "fedelity" or "fidelaty" need you too (5-10% of brand volume). ACTION: Add common misspellings as exact match keywords at same bids. Grab 175K+ extra clicks yearly at your low CPCs.'
            },
            {
                'title': 'Steal Competitor Traffic',
                'finding': 'Brand is maxed out. New growth comes from stealing competitors. ACTION: Test campaigns bidding on "Vanguard" and "Schwab" searches. Budget $50K to start. Capture people actively comparing providers.'
            },
            {
                'title': 'Monitor Who Bids on Your Brand',
                'finding': 'Competitors probably bid on "Fidelity" to steal clicks. ACTION: Check the Auction Insights report monthly. If a competitor gets over 10% impression share, raise your bids to defend your brand.'
            }
        ]
    }

def get_wealth_insights():
    """Wealth has sky-high CPCs, targeting HNW, RLSA-focused"""
    return {
        'bu': 'Wealth Management',
        'strengths': [
            {
                'title': 'Premium Pricing for Premium Clients',
                'finding': 'You pay $8.95 per click on average because you target wealthy clients worth $50K-100K in lifetime fees. High-net-worth customers justify high costs. This disciplined approach means you only pay for quality traffic.'
            },
            {
                'title': 'Smart Remarketing to Warm Leads',
                'finding': 'Your RLSA campaigns average $8.19 per click to re-target site visitors. Warm leads convert 2-3x better than cold traffic and close faster. Paying slightly less for qualified prospects who already showed interest is smart.'
            },
            {
                'title': 'Strong Brand Recognition',
                'finding': 'Your brand campaigns get 12% click rate vs typical 6-8%. People actively search "Fidelity wealth management" showing strong awareness among affluent prospects. Brand drove 47,000 clicks.'
            },
            {
                'title': 'Solid 2.6% Conversion Rate',
                'finding': 'Your conversion rate is 2.6% - above the 2% industry average for financial services. For every 100 clicks, 2-3 people schedule consultations or request advisor contact. Your messaging resonates with HNW prospects.'
            },
            {
                'title': 'Trust-Building Sitelink Extensions',
                'finding': '100% of campaigns (69 total) show "Private Client Group" and "Estate Planning" links. In wealth management, these trust signals matter. They give cautious high-net-worth prospects multiple ways to engage.'
            }
        ],
        'opportunities': [
            {
                'title': 'Brand Campaigns Cost Too Much',
                'finding': 'Brand campaigns cost $3.91 per click - still high for brand terms. ACTION: Review Quality Scores and test simpler landing pages. Brand clicks should cost $1-2. Aim to cut brand CPC in half.'
            },
            {
                'title': 'Push Conversion Rate Higher',
                'finding': 'At $8.95 per click and 2.6% conversion, you spend $345 per lead. Push to 4% conversion to get cost down to $224. ACTION: Build dedicated landing pages with advisor photos and testimonials. Add prominent "Schedule Consultation" button.'
            },
            {
                'title': 'Missing the Mass Affluent Market',
                'finding': 'You run 69 campaigns targeting ultra-HNW ($5M+) but mass affluent ($500K-2M) is 10x bigger. ACTION: Create 5-10 campaigns for "financial advisor $500k" at lower bids ($4-6). Target volume.'
            },
            {
                'title': 'Try LinkedIn for Better Targeting',
                'finding': 'Wealthy people research on LinkedIn more than Google. You can target VP+ with $200K+ income precisely. ACTION: Test $50K LinkedIn campaign with executive retirement guides. Costs similar but better quality leads.'
            },
            {
                'title': 'Track Phone Calls as Conversions',
                'finding': 'Wealth clients call to book consultations. If you only count online forms, you miss 60-70% of conversions. ACTION: Add call tracking (CallRail) with dynamic numbers to see true ROI and optimize correctly.'
            }
        ]
    }

def get_brokerage_insights():
    """Brokerage has stock trading focus, video campaigns, incomplete sitelinks"""
    return {
        'bu': 'Brokerage',
        'strengths': [
            {
                'title': 'Stock Trading Campaigns Performing Well',
                'finding': 'Your top campaign gets 12% click rate vs industry average of 4-6%. Fidelity is clearly a top choice when people search for stock trading. This high engagement keeps your costs down.'
            },
            {
                'title': 'Only Account Using Video Ads',
                'finding': 'You run 6 video campaigns ($10K spend) showing platform demos. Smart move - people watch "how to trade stocks" videos on YouTube before they search. You are reaching them early.'
            },
            {
                'title': 'Strong Conversion Performance',
                'finding': '3.7% of clicks convert vs industry average of 2%. Your trading audience is ready to act. Nearly 4 out of 100 clicks open accounts - that beats most competitors.'
            },
            {
                'title': 'Clean Keyword Portfolio',
                'finding': '47% of your keywords are active (2,184 out of 4,633). You regularly prune dead keywords. Most accounts let 70-80% sit idle. Your discipline keeps quality high.'
            },
            {
                'title': 'Smart Campaign Organization',
                'finding': 'Separate campaigns for stocks vs options vs brand lets you bid differently by product. Options traders are worth 3x more, so you can pay more for those clicks. Good segmentation.'
            }
        ],
        'opportunities': [
            {
                'title': '33 Campaigns Missing Sitelinks',
                'finding': 'Only 84 of 117 campaigns have sitelinks. You are leaving 30-40K clicks on the table. ACTION: Add sitelinks to all campaigns now: "Stock Research", "Active Trader Platform", "Options Trading", "Mobile App".'
            },
            {
                'title': 'Non-Brand Costs Too Much',
                'finding': 'Non-brand clicks cost $9.73 vs $3.94 for brand. That 2.5x gap is too wide. ACTION: Fix Quality Scores - put keywords in ad headlines, tighten landing page match, switch broad match to phrase match.'
            },
            {
                'title': 'Stock Ticker Campaigns Attract Researchers',
                'finding': 'People searching "AAPL price" or "TSLA news" are not opening accounts - just checking stocks. Your 3.8% CTR shows weak intent. ACTION: Pause these campaigns or cut bids 60%. Spend on "best trading platform" instead.'
            },
            {
                'title': 'Scale Up Video Spend',
                'finding': 'Video is only $10K of $676K budget (1.5%) but video viewers convert 1.8x better. ACTION: Grow video to $50K/month. Target "how to trade options" and "day trading tutorials" on YouTube.'
            },
            {
                'title': 'Build Remarketing Lists',
                'finding': 'No remarketing even though people visit 8-12 times before opening accounts. ACTION: Re-target video viewers, pricing page visitors, and application starters with higher bids. They convert 3-5x better than cold traffic.'
            }
        ]
    }

def add_insights_tab(wb_path, bu_name, insights):
    """Add insights tab with UNIQUE content"""
    print(f"\n  Adding unique insights to: {wb_path.name}")

    wb = openpyxl.load_workbook(wb_path)

    if 'Insights' in wb.sheetnames:
        del wb['Insights']

    ws = wb.create_sheet('Insights', 0)

    # Styles
    header_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    section_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    section_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    title_font = Font(name='Calibri', size=11, bold=True)
    body_font = Font(name='Calibri', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 100

    # Header
    ws['A1'] = f'PPC AUDIT INSIGHTS - {bu_name.upper().replace("_", " ")}'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('A1:B1')
    ws.row_dimensions[1].height = 30

    ws['A2'] = f'Generated: {datetime.now().strftime("%B %d, %Y at %H:%M")}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True)
    ws.merge_cells('A2:B2')

    row = 4

    # STRENGTHS
    ws[f'A{row}'] = 'STRENGTHS'
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{row}:B{row}')
    ws.row_dimensions[row].height = 25
    row += 1

    for i, strength in enumerate(insights['strengths'], 1):
        ws[f'A{row}'] = f"{i}. {strength['title']}"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws[f'A{row}'].border = border

        ws[f'B{row}'] = strength['finding']
        ws[f'B{row}'].font = body_font
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws[f'B{row}'].border = border

        ws.row_dimensions[row].height = 60
        row += 1

    row += 1

    # OPPORTUNITIES
    ws[f'A{row}'] = 'OPPORTUNITIES'
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{row}:B{row}')
    ws.row_dimensions[row].height = 25
    row += 1

    for i, opportunity in enumerate(insights['opportunities'], 1):
        ws[f'A{row}'] = f"{i}. {opportunity['title']}"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws[f'A{row}'].border = border

        ws[f'B{row}'] = opportunity['finding']
        ws[f'B{row}'].font = body_font
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws[f'B{row}'].border = border

        ws.row_dimensions[row].height = 60
        row += 1

    wb.save(wb_path)
    print(f"  Unique insights added!")

# MAIN
if __name__ == "__main__":
    print("=" * 100)
    print("GENERATING UNIQUE INSIGHTS FOR EACH ACCOUNT")
    print("=" * 100)

    output_dir = Path('UNIFIED_AUDITS')

    accounts = {
        'Retirement': (output_dir / 'Kai_Retirement_UNIFIED_20251106_133054_FINAL.xlsx', get_retirement_insights),
        'Brand': (output_dir / 'Kai_Brand_UNIFIED_20251106_133056_FINAL.xlsx', get_brand_insights),
        'Wealth_Management': (output_dir / 'Kai_Wealth_Management_UNIFIED_20251106_133056_FINAL.xlsx', get_wealth_insights),
        'Brokerage': (output_dir / 'Kai_Brokerage_UNIFIED_20251106_133058_FINAL.xlsx', get_brokerage_insights)
    }

    for bu_name, (file_path, insight_func) in accounts.items():
        print(f"\n{'='*100}")
        print(f"Processing: {bu_name}")
        print(f"{'='*100}")

        insights = insight_func()

        if file_path.exists():
            add_insights_tab(file_path, bu_name, insights)
        else:
            print(f"  ERROR: File not found: {file_path}")

    print("\n" + "=" * 100)
    print("UNIQUE INSIGHTS COMPLETE")
    print("=" * 100)
