"""
Azure Functions HTTP Trigger for Kai Audit Generation
"""
from __future__ import annotations

import logging
import os
import re
import sys
import tempfile
import textwrap
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import azure.functions as func
from azure.storage.blob import BlobServiceClient, ContainerClient
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from kai_core.account_utils import normalize_account_name, slugify_account_name
from kai_core.shared.vector_index import index_audit_workbook
from config import get_deployment_mode, get_custom_data_path
try:
    from kai_core.core_logic import UnifiedAuditEngine  # type: ignore
    ENGINE_SOURCE = "source"
except Exception:
    from kai_core_engine import UnifiedAuditEngine  # type: ignore
    ENGINE_SOURCE = "compiled"

try:
    from azure.search.documents import SearchClient
    from azure.core.credentials import AzureKeyCredential
except ImportError:
    SearchClient = None  # type: ignore[assignment]
    AzureKeyCredential = None  # type: ignore[assignment]

# Add parent directory to path for imports
function_dir = Path(__file__).resolve().parent
sys.path.insert(0, str(function_dir))
root_dir = function_dir.parent.parent
if str(root_dir) not in sys.path:
    sys.path.insert(0, str(root_dir))

from kai_core.shared.business_physics import (  # noqa: E402
    AccountContext,
    GuardrailFinding,
    evaluate_guardrails,
)

try:
    from generate_unique_insights import (  # noqa: E402
        get_brand_insights,
        get_brokerage_insights,
        get_retirement_insights,
        get_wealth_insights,
    )

    INSIGHT_FUNCTIONS = {
        "Retirement": get_retirement_insights,
        "Brand": get_brand_insights,
        "Wealth Management": get_wealth_insights,
        "Wealth_Management": get_wealth_insights,
        "Brokerage": get_brokerage_insights,
    }
except ImportError:
    INSIGHT_FUNCTIONS = {}

INSIGHT_TABLE_COLUMNS = [
    "Section",
    "Line of Business / Product",
    "Dimension",
    "Type",
    "Level",
    "Theme Group",
    "Insight Summary",
    "Supporting Data / Evidence",
    "Business Impact",
    "Recommended Action",
    "Executive Narrative",
    "Priority",
    "Phase",
    "Slide Usage / Notes",
]

# ============================================================
# Weighted scoring scaffolding (additive, flag-guarded)
# ============================================================
def _get_scoring_mode() -> str:
    """
    Returns the requested scoring mode.
    Allowed: AVERAGE (default), WEIGHTED, STRICT.
    """
    mode = os.environ.get("AUDIT_SCORING_MODE", "AVERAGE").strip().upper()
    if mode not in {"AVERAGE", "WEIGHTED", "STRICT"}:
        return "AVERAGE"
    return mode


_WEIGHTED_CONFIG_CACHE: Optional[Dict[str, object]] = None


def _load_weighted_config() -> Optional[Dict[str, object]]:
    """
    Loads weighted scoring config if present.
    Order: env AUDIT_SCORING_CONFIG_PATH, azure/weighted_scoring_config.json,
    azure/weighted_scoring_config.template.json.
    """
    global _WEIGHTED_CONFIG_CACHE
    if _WEIGHTED_CONFIG_CACHE is not None:
        return _WEIGHTED_CONFIG_CACHE

    candidates: List[Path] = []
    env_path = os.environ.get("AUDIT_SCORING_CONFIG_PATH")
    if env_path:
        candidates.append(Path(env_path))
    repo_root = Path(__file__).resolve().parent
    candidates.append(repo_root / "weighted_scoring_config.json")
    candidates.append(repo_root / "weighted_scoring_config.template.json")

    for path in candidates:
        try:
            if path.exists():
                with path.open("r", encoding="utf-8") as f:
                    _WEIGHTED_CONFIG_CACHE = json.load(f)
                    logging.info("[WeightedScoring] Loaded config from %s", path)
                    return _WEIGHTED_CONFIG_CACHE
        except Exception as exc:  # pragma: no cover - defensive
            logging.warning("[WeightedScoring] Failed to load config %s: %s", path, exc)
            continue

    _WEIGHTED_CONFIG_CACHE = None
    return None


def _resolve_vertical() -> str:
    """
    Returns the active vertical context for weighting (default ALL).
    """
    return os.environ.get("AUDIT_VERTICAL", "ALL").upper()

SCORECARD_CATEGORY_COL = 2  # Column B
SCORECARD_ITEM_COL = 3      # Column C
SCORECARD_RATING_COL = 4    # Column D
SCORECARD_DETAILS_COL = 5   # Column E
SCORECARD_ACTIONS_COL = SCORECARD_DETAILS_COL + 1  # Column F
SCORECARD_RATIONALE_COL = SCORECARD_DETAILS_COL + 2  # Column G

DIMENSION_KEYWORDS = {
    "Structure & Hygiene": ["structure", "match", "naming", "campaign", "split", "consolidated", "coverage"],
    "Keywords & Queries": ["keyword", "query", "search term", "match type", "broad"],
    "Audiences & Remarketing": ["audience", "remarketing", "customer match", "lookalike", "rlsa"],
    "Ads & Creative": ["ad", "creative", "copy", "rsa", "headline", "asset"],
    "Landing Pages & CX": ["landing", "page", "experience", "site", "cx"],
    "Measurement & Tracking": ["measurement", "tracking", "conversion", "attribution", "signal"],
    "Bid Strategy & Budgets": ["bid", "budget", "smart bidding", "tcpa", "troas", "spend"],
    "Cross-Channel": ["pmax", "cross-channel", "youtube", "display", "microsoft"],
    "Strategies": ["strategy", "roadmap", "testing", "expansion", "guardrail"],
}

DIMENSION_PHASE_MAP = {
    "Structure & Hygiene": "1",
    "Measurement & Tracking": "1",
    "Keywords & Queries": "1",
    "Audiences & Remarketing": "2",
    "Ads & Creative": "2",
    "Landing Pages & CX": "2",
    "Bid Strategy & Budgets": "2",
    "Cross-Channel": "3",
    "Strategies": "3",
}

BUSINESS_IMPACT_TEMPLATES = {
    "Structure & Hygiene": "Impacts scalability and controllable growth because structure dictates how budgets and optimizations flow.",
    "Keywords & Queries": "Dictates how efficiently we capture qualified demand and avoid waste on irrelevant terms.",
    "Audiences & Remarketing": "Determines our ability to influence high-value users and protect conversion efficiency.",
    "Ads & Creative": "Directly influences CTR and conversion rate, shaping the cost to acquire qualified traffic.",
    "Landing Pages & CX": "Controls whether paid traffic converts into qualified leads or abandons the experience.",
    "Measurement & Tracking": "Without reliable signals we can't optimize bids, budgets, or evaluate performance credibly.",
    "Bid Strategy & Budgets": "Alignment between budgets and bid strategy controls volume ramp and CPA stability.",
    "Cross-Channel": "Channel mix determines incremental reach and protects brand demand from competitors.",
    "Strategies": "Strategic clarity sets the tone for future testing, innovation, and stakeholder trust.",
}

ACTION_BLUEPRINTS = {
    "Structure & Hygiene": {
        "strength": [
            "Document the exact structure behind {summary_lower} and publish it as the template for future launches.",
            "Move {summary_lower} into the operating manual so every net-new campaign inherits the same guardrails.",
            "Use {summary_lower} as the control sample in automation workshops to keep quality high at scale.",
            "Highlight {summary_lower} in the next governance review so budget owners keep backing this approach.",
        ],
        "opportunity": [
            "Re-map campaigns so {summary_lower} no longer bleeds spend across intent tiers.",
            "Assign an owner to rebuild the structure until {summary_lower} stabilizes.",
            "Stage the restructure in a controlled pilot so {summary_lower} improves without disrupting budgets.",
            "Pair the rebuild with fresh QA checkpoints so {summary_lower} does not regress.",
        ],
    },
    "Keywords & Queries": {
        "strength": [
            "Package the keyword and negative logic behind {summary_lower} into a playbook for every squad.",
            "Share the query insights that produced {summary_lower} during the next ops sync so the coverage doesn’t drift.",
            "Turn {summary_lower} into a benchmark slide so new hires understand what good looks like.",
            "Use {summary_lower} as the reference set when evaluating new match-type tests.",
        ],
        "opportunity": [
            "Regroup keywords and negatives so {summary_lower} stops leaking budget.",
            "Pair the diagnostic checklist with fresh negatives until {summary_lower} improves.",
            "Host a query-triage session to rebuild the structure behind {summary_lower}.",
            "Audit long-tail intent and reassign match types so {summary_lower} moves closer to the benchmark.",
        ],
    },
    "Audiences & Remarketing": {
        "strength": [
            "Sync this audience recipe so {summary_lower} keeps Smart Bidding calibrated.",
            "Expand lookalikes using the segments that drove {summary_lower}.",
            "Document audience refresh cadences so {summary_lower} remains reliable even as CRM data shifts.",
            "Use {summary_lower} to brief analytics/CRM partners on what signals deliver the best ROI.",
        ],
        "opportunity": [
            "Stand up remarketing/customer lists so {summary_lower} feeds richer signals.",
            "Partner with CRM to backfill audiences until {summary_lower} is fixed.",
            "Create a data-refresh calendar so {summary_lower} does not stall again after fixes.",
            "Add modeled audiences or seed lists to jump-start improvements to {summary_lower}.",
        ],
    },
    "Ads & Creative": {
        "strength": [
            "Archive the RSA variants that produced {summary_lower} and reuse them in new LOBs.",
            "Keep creative QA tight so {summary_lower} stays on brand and high-performing.",
            "Turn the best-performing copy behind {summary_lower} into a swipe file for the next brief.",
            "Celebrate {summary_lower} in the creative stand-up so future drafts follow the same formula.",
        ],
        "opportunity": [
            "Refresh RSA assets and landing hooks so {summary_lower} is no longer a blocker.",
            "Launch a creative sprint focused on reversing {summary_lower}.",
            "Run A/B tests on headlines/CTAs tied to {summary_lower} until engagement lifts.",
            "Coordinate with brand/legal so {summary_lower} improvements can move quickly without rework.",
        ],
    },
    "Landing Pages & CX": {
        "strength": [
            "Share the CX checklist that delivered {summary_lower} before any new launches.",
            "Reuse these layouts in the next cohort so {summary_lower} stays consistent.",
            "Publish the UX learnings from {summary_lower} in the growth wiki so tests build off a strong base.",
            "Record a quick loom walkthrough of the experience behind {summary_lower} for partner teams.",
        ],
        "opportunity": [
            "Pair media with new CX tests so {summary_lower} stops wasting paid traffic.",
            "Assign a CRO squad to resolve {summary_lower} before scaling budgets.",
            "Map the drop-off points that create {summary_lower} and prioritize fixes with UX/analytics.",
            "Prototype a simplified flow so {summary_lower} starts trending in the right direction.",
        ],
    },
    "Measurement & Tracking": {
        "strength": [
            "Lock in the tagging recipe that gives us {summary_lower} quality signals.",
            "Keep analytics and media in sync so {summary_lower} stays trustworthy.",
            "Snapshots of {summary_lower} should be captured in every monthly business review to keep compliance aligned.",
            "Use {summary_lower} as the QA checklist whenever a new conversion or signal is added.",
        ],
        "opportunity": [
            "Close the measurement gaps so {summary_lower} no longer undermines Smart Bidding.",
            "Run the SST/Consent Mode checklist until {summary_lower} is resolved.",
            "Escalate {summary_lower} as a risk in the next steering committee so resources land on the fix.",
            "Backfill missing identifiers or hashed data so {summary_lower} can stabilize quickly.",
        ],
    },
    "Bid Strategy & Budgets": {
        "strength": [
            "Preserve this bid/budget mix so {summary_lower} keeps pacing smoothly.",
            "Share the playbook that keeps {summary_lower} aligned to profit bands.",
            "Highlight {summary_lower} in finance ops so future rebalances follow the same guardrails.",
            "Use {summary_lower} as proof when requesting incremental headroom.",
        ],
        "opportunity": [
            "Rebalance bids and budgets so {summary_lower} stops causing volatility.",
            "Rebuild the forecast plan until {summary_lower} meets finance guardrails.",
            "Pair the new bid targets with pacing alerts so {summary_lower} stays on track after the fix.",
            "Document the resilience plan for {summary_lower} once the remediation lands.",
        ],
    },
    "Cross-Channel": {
        "strength": [
            "Keep the channel mix that achieved {summary_lower} visible in the roadmap.",
            "Extend this tactic to adjacent networks so {summary_lower} scales.",
            "Codify how channels share signals when {summary_lower} is working and reuse it in partners’ briefs.",
            "Host a mini post-mortem on why {summary_lower} succeeded so other funnels can copy it.",
        ],
        "opportunity": [
            "Launch the PMAX/video plan so {summary_lower} evolves beyond search-only coverage.",
            "Add incremental pilots to close the gaps highlighted by {summary_lower}.",
            "Bring creative, media, and analytics together to relaunch the mix until {summary_lower} rebounds.",
            "Stage the new channel plan in crawl/walk/run phases so {summary_lower} improves with measured risk.",
        ],
    },
    "Strategies": {
        "strength": [
            "Codify this roadmap clarity so {summary_lower} remains sponsor-ready.",
            "Use this win as the template for next quarter's governance forums.",
            "Brief leadership on {summary_lower} so upcoming investments reference the same logic.",
            "Archive the story behind {summary_lower} in the strategy wiki for future pitch decks.",
        ],
        "opportunity": [
            "Rebuild the roadmap so {summary_lower} has an owner, timeline, and success metric.",
            "Sequence fixes using the diagnostic framework until {summary_lower} is no longer a risk.",
            "Host a planning workshop so stakeholders align on how to reverse {summary_lower}.",
            "Tie {summary_lower} to measurable OKRs so progress is tracked every QBR.",
        ],
    },
    "default": {
        "strength": [
            "Keep this proof point visible so {summary_lower} continues to scale.",
            "Share the approach behind {summary_lower} across pods so everyone benefits.",
            "Capture the lessons from {summary_lower} in the enablement library before the details fade.",
        ],
        "opportunity": [
            "Assign a remediation plan so {summary_lower} is resolved before the next QBR.",
            "Flag {summary_lower} in the weekly stand-up until the fix is complete.",
            "Pair {summary_lower} with specific milestones so owners stay accountable.",
        ],
    },
}

DIMENSION_STRENGTH_STORIES = {
    "Structure & Hygiene": "It shows structure and pacing stay predictable, so budgets deploy cleanly.",
    "Keywords & Queries": "It proves query coverage is pulling qualified demand without excess waste.",
    "Audiences & Remarketing": "It confirms audience signals steer Smart Bidding toward high-value cohorts.",
    "Ads & Creative": "It demonstrates creative depth keeps CTR and CVR competitive.",
    "Landing Pages & CX": "It reassures stakeholders that landing flows convert without extra friction.",
    "Measurement & Tracking": "It means the measurement pipe is trustworthy enough to steer bids and reporting.",
    "Bid Strategy & Budgets": "It shows bids and budgets are aligned to volume and margin goals.",
    "Cross-Channel": "It proves the channel mix defends brand demand while expanding incremental reach.",
    "Strategies": "It signals the roadmap is sequenced so leaders see exactly where investment goes next.",
}

DIMENSION_GAP_STORIES = {
    "Structure & Hygiene": "Structure still feels messy, so budgets and bid strategies bleed across lanes.",
    "Keywords & Queries": "Query coverage is leaking into low-quality demand and missing intent pockets.",
    "Audiences & Remarketing": "Audience signals stay thin, forcing Smart Bidding to guess who matters most.",
    "Ads & Creative": "Creative depth is shallow, which drags on CTR, CVR, and ultimately CPC efficiency.",
    "Landing Pages & CX": "Landing flows are stalling conversions and wasting paid traffic.",
    "Measurement & Tracking": "Measurement gaps make every bid, budget, and KPI harder to trust.",
    "Bid Strategy & Budgets": "Bid strategy and budgets are misaligned, creating volume swings and margin risk.",
    "Cross-Channel": "Channel mix leaves reach gaps or cannibalizes brand demand instead of expanding it.",
    "Strategies": "The roadmap feels hazy, so stakeholders cannot prioritize or fund the right fixes.",
}

SCOPE_BLUEPRINTS = {
    "campaigns": {
        "dimension": "Structure & Hygiene",
        "type": "Strength",
        "impact": "Confirms {value} active campaigns were audited, covering growth, brand defense, and nurture lanes.",
        "action": "Map those {value} campaigns to strategy lanes before reorganizing budgets or bid policies.",
    },
    "ad_groups": {
        "dimension": "Structure & Hygiene",
        "type": "Strength",
        "impact": "{value} ad groups show how granular the execution currently is and where controls can tighten.",
        "action": "Review the {value} ad groups for intent overlap and merge thin groupings to boost data density.",
    },
    "keywords": {
        "dimension": "Keywords & Queries",
        "type": "Strength",
        "impact": "{value} keywords illustrate the reach of our query coverage and the workload on negative lists.",
        "action": "Classify the top cohorts of these {value} keywords into high-impact, experimental, and defense swim lanes.",
    },
    "ads": {
        "dimension": "Ads & Creative",
        "type": "Strength",
        "impact": "{value} ads indicates how much creative variety is in market and how RSAs rotate messaging.",
        "action": "Audit the {value} ads for asset diversity, mobile readiness, and coverage of key value props.",
        "impact_zero": "No active ads were detected, which means the creative library needs to be rebuilt before scaling media.",
        "action_zero": "Stand up refreshed RSAs with the mandated CTA, offer, and compliance language before re-launching campaigns.",
    },
    "audiences": {
        "dimension": "Audiences & Remarketing",
        "type": "Strength",
        "impact": "{value} audiences highlight how much first-party or remarketing signal is available.",
        "action": "Ensure these {value} audiences stay synced and build lookalikes or exclusions from the highest value segments.",
        "impact_zero": "No audience segments were found, so Smart Bidding lacks first-party signals to prioritize.",
        "action_zero": "Build baseline remarketing and customer-match lists, then feed them into every search and PMAX lane.",
        "type_zero": "Opportunity",
    },
    "conversions_tracked": {
        "dimension": "Measurement & Tracking",
        "type": "Strength",
        "impact": "{value} conversion actions are feeding reporting and Smart Bidding.",
        "action": "Keep these {value} conversions aligned to the 2025 measurement framework and verify values are accurate.",
        "impact_low": "{value} conversion actions are being tracked, which is thin for a program of this size.",
        "action_low": "Expand conversion coverage (lead qualifiers, revenue proxies, or offline uploads) so Smart Bidding sees the full business impact.",
        "type_low": "Opportunity",
        "low_threshold": 3,
    },
    "default": {
        "dimension": "Structure & Hygiene",
        "type": "Strength",
        "impact": "{value} data points were audited for {label}, showing how deep the diagnostic went.",
        "action": "Reference these counts when sizing effort for the remediation roadmap.",
    },
}

PRIOR_INSIGHT_SECTIONS = [
    "Insights",
    "Key Themes & Takeaways",
    "Executive Summary",
    "Havas Accelerate SEM",
    "Theme Summary",
]

ACCOUNT_TO_ENGINE = {
    "Asset Management": "Asset_Management",
    "Brand": "Brand",
    "Brokerage": "Brokerage",
    "Crypto": "Crypto",
    "FILI": "FILI",
    "FILI Term Life": "FILI_Term_Life",
    "Medicare": "Medicare",
    "Retirement": "Retirement",
    "Wealth Management": "Wealth_Management",
    "WI": "WI",
}

DEFAULT_LOCAL_DATA = Path(r"Z:\Kel\PPC_Audit_Project\Total Files")
CACHED_DATA_DIR = Path(tempfile.gettempdir()) / "generateaudit_ppc_data"


def _download_data_from_storage(account_name: str) -> Optional[Path]:
    connection_string = os.environ.get("AZURE_STORAGE_CONNECTION_STRING")
    container_name = os.environ.get("PPC_DATA_BLOB_CONTAINER")
    if not connection_string or not container_name:
        logging.warning(
            "[GenerateAudit] Storage download skipped (missing AZURE_STORAGE_CONNECTION_STRING or PPC_DATA_BLOB_CONTAINER)"
        )
        return None

    prefix_template = os.environ.get("PPC_DATA_BLOB_PREFIX", "")
    slug = slugify_account_name(account_name)
    prefix = (prefix_template or "").format(account=account_name, account_slug=slug).strip("/")

    logging.info(
        "[GenerateAudit] Downloading PPC data from container '%s' with prefix '%s'",
        container_name,
        prefix or "<root>",
    )

    cache_dir = CACHED_DATA_DIR / (slug or "default")
    if cache_dir.exists() and any(cache_dir.glob("*.csv")):
        logging.info("[GenerateAudit] Using cached PPC data at %s", cache_dir)
        return cache_dir

    try:
        blob_service = BlobServiceClient.from_connection_string(connection_string)
        container_client: ContainerClient = blob_service.get_container_client(container_name)
    except Exception as exc:
        logging.error("[GenerateAudit] Failed to connect to storage: %s", exc)
        return None

    cache_dir.mkdir(parents=True, exist_ok=True)

    downloaded = 0
    try:
        for blob in container_client.list_blobs(name_starts_with=prefix or None):
            if not blob.name.lower().endswith(".csv"):
                continue
            destination = cache_dir / Path(blob.name).name
            with open(destination, "wb") as handle:
                container_client.download_blob(blob).readinto(handle)
            downloaded += 1
    except Exception as exc:
        logging.error("[GenerateAudit] Failed to download PPC data: %s", exc, exc_info=True)
        return None

    if downloaded == 0:
        logging.warning("[GenerateAudit] No CSV blobs found for prefix '%s'", prefix or "<root>")
        return None

    logging.info("[GenerateAudit] Downloaded %s CSV files to %s", downloaded, cache_dir)
    return cache_dir


def _resolve_data_directory(account_name: str) -> Path:
    mode = get_deployment_mode()
    candidates = []
    custom_path = get_custom_data_path()
    if custom_path:
        candidates.append(custom_path)
    configured_path = os.environ.get("PPC_DATA_FOLDER")
    if configured_path:
        candidates.append(Path(configured_path))

    local_data_dir = Path(__file__).parent / "data"
    slug = slugify_account_name(account_name)
    if slug:
        account_specific = (Path(__file__).parent / "data_accounts" / slug)
        candidates.append(account_specific)

    candidates.append(local_data_dir)

    # In LOCAL mode, prefer the longstanding Z: default early to preserve existing behavior.
    if mode == "LOCAL":
        candidates.append(DEFAULT_LOCAL_DATA)

    for candidate in candidates:
        if candidate and candidate.exists() and any(candidate.glob("*.csv")):
            logging.info("[GenerateAudit] Using PPC data directory: %s (mode=%s)", candidate, mode)
            return candidate

    # Enterprise-first download before falling back to LOCAL defaults, keeping LOCAL unchanged.
    download_dir = _download_data_from_storage(account_name)
    if download_dir and download_dir.exists() and any(download_dir.glob("*.csv")):
        logging.info("[GenerateAudit] Using downloaded PPC data directory: %s (mode=%s)", download_dir, mode)
        return download_dir

    # In ENTERPRISE mode, avoid failing early on the missing Z: path; use it only as a final fallback.
    if mode == "ENTERPRISE" and DEFAULT_LOCAL_DATA.exists() and any(DEFAULT_LOCAL_DATA.glob("*.csv")):
        logging.info("[GenerateAudit] Falling back to LOCAL data directory: %s", DEFAULT_LOCAL_DATA)
        return DEFAULT_LOCAL_DATA

    raise RuntimeError(
        "No PPC data available. Set PPC_DATA_FOLDER to a directory with CSV files or configure "
        "PPC_DATA_BLOB_CONTAINER (+ optional PPC_DATA_BLOB_PREFIX) so the function can pull data from storage."
    )


def _load_request_body(req: func.HttpRequest) -> Dict[str, object]:
    try:
        payload = req.get_json()
        if isinstance(payload, dict):
            return payload
    except ValueError:
        pass
    return {}


def _extract_account(req: func.HttpRequest, body: Optional[Dict[str, object]] = None) -> Optional[str]:
    account_name = req.params.get("account")
    if not account_name:
        body = body or {}
        account_name = body.get("account") if isinstance(body, dict) else None
    return account_name


def _extract_context_overrides(
    req: func.HttpRequest,
    body: Optional[Dict[str, object]] = None,
) -> Dict[str, object]:
    """Return optional context hints provided by the caller."""

    def _first_value(keys: List[str]) -> Optional[object]:
        for source in filter(None, [body, req.params]):
            for key in keys:
                if key in source:
                    return source.get(key)
        return None

    def _as_float(value: Optional[object]) -> Optional[float]:
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _as_int(value: Optional[object]) -> Optional[int]:
        if value is None or value == "":
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def _as_bool(value: Optional[object]) -> Optional[bool]:
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value != 0
        if isinstance(value, str):
            lowered = value.strip().lower()
            if lowered in {"true", "1", "yes", "y"}:
                return True
            if lowered in {"false", "0", "no", "n"}:
                return False
        return None

    overrides: Dict[str, object] = {}
    body = body or {}
    overrides["account_id"] = _first_value(["account_id", "accountId"])
    overrides["vertical"] = _first_value(["vertical", "businessModel"])
    overrides["margin"] = _as_float(_first_value(["margin", "grossMargin", "gross_margin"]))
    overrides["avg_lead_close_rate"] = _as_float(
        _first_value(["avg_lead_close_rate", "avgLeadCloseRate"])
    )
    overrides["verification_status"] = _as_bool(
        _first_value(["verification_status", "verificationStatus"])
    ) or False
    country_value = _first_value(["target_country", "targetCountry"])
    overrides["target_country"] = str(country_value) if country_value else None
    overrides["target_roas"] = _as_float(_first_value(["target_roas", "targetROAS"]))
    overrides["bidding_strategy"] = _first_value(["bidding_strategy", "biddingStrategy"])
    overrides["conversions_30d"] = _as_int(_first_value(["conversions_30d", "conversions30d"]))
    overrides["current_roas"] = _as_float(_first_value(["current_roas", "currentROAS"]))
    return overrides


def _normalize_vertical_label(label: Optional[object]) -> str:
    if not label:
        return "LEAD_GEN"
    text = str(label).strip().upper()
    mapping = {
        "ECOMMERCE": "ECOMMERCE",
        "E-COMMERCE": "ECOMMERCE",
        "D2C": "ECOMMERCE",
        "RETAIL": "ECOMMERCE",
        "SAAS": "SAAS",
        "SOFTWARE": "SAAS",
        "FINANCE": "FINANCE",
        "FINANCIAL": "FINANCE",
        "BANKING": "FINANCE",
        "INSURANCE": "FINANCE",
        "HEALTH": "LEAD_GEN",
        "HEALTHCARE": "LEAD_GEN",
        "LOCAL": "LOCAL",
        "LEAD_GEN": "LEAD_GEN",
        "LEADGEN": "LEAD_GEN",
    }
    return mapping.get(text, "LEAD_GEN")


def _infer_vertical_from_account(account_name: str, overrides: Dict[str, object]) -> str:
    if overrides.get("vertical"):
        return _normalize_vertical_label(overrides["vertical"])

    lowered = (account_name or "").lower()
    finance_keywords = ["retire", "wealth", "broker", "asset", "fili", "crypto", "invest", "medicare"]
    if any(keyword in lowered for keyword in finance_keywords):
        return "FINANCE"
    if "brand" in lowered:
        return "LEAD_GEN"
    return "LEAD_GEN"


def _build_account_context(account_name: str, overrides: Dict[str, object]) -> AccountContext:
    detected_vertical = _infer_vertical_from_account(account_name, overrides)
    target_country = overrides.get("target_country")
    target_country = target_country.lower() if isinstance(target_country, str) else None
    if not target_country:
        target_country = os.environ.get("DEFAULT_TARGET_COUNTRY", "us").lower()
    bidding_strategy = overrides.get("bidding_strategy")
    bidding_strategy = str(bidding_strategy).upper() if bidding_strategy else None

    context = AccountContext(
        account_id=str(overrides.get("account_id")) if overrides.get("account_id") else None,
        account_name=account_name,
        detected_vertical=detected_vertical,
        margin_input=overrides.get("margin"),
        avg_lead_close_rate=overrides.get("avg_lead_close_rate"),
        verification_status=bool(overrides.get("verification_status", False)),
        target_country=target_country,
        target_roas=overrides.get("target_roas"),
        bidding_strategy=bidding_strategy,
        conversions_30d=overrides.get("conversions_30d"),
    )
    return context


def _summarize_guardrails(findings: List[GuardrailFinding]) -> str:
    if not findings:
        return "All guardrails passed."
    parts = []
    for finding in findings:
        parts.append(f"{finding.level}: {finding.section} — {finding.detail}")
    return " | ".join(parts)


def _write_insights_tab(
    workbook_path: Path,
    account_name: str,
    highlights: Optional[Dict[str, object]] = None,
) -> None:
    func = INSIGHT_FUNCTIONS.get(account_name) or INSIGHT_FUNCTIONS.get(account_name.replace('_', ' '))
    insights: Dict[str, object] = {}
    if func:
        try:
            raw = func()
            if isinstance(raw, dict):
                insights = raw
            else:
                logging.warning("[GenerateAudit] Insight hook for %s returned non-dict payload; falling back to highlights.", account_name)
        except Exception as exc:
            logging.warning("[GenerateAudit] Failed to load insights for %s (%s); using highlights-only narrative.", account_name, exc)
            insights = {}
    else:
        logging.info("[GenerateAudit] No bespoke insights configured for %s; deriving narratives from scorecard context.", account_name)

    insights.setdefault("bu", account_name)

    try:
        workbook = load_workbook(workbook_path)
    except Exception as exc:
        logging.error("[GenerateAudit] Unable to open workbook for insights injection: %s", exc)
        return

    sheet_name = "Insights"
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
    sheet.delete_rows(1, sheet.max_row)

    sheet.cell(1, 1, f"{account_name.replace('_', ' ')} - Executive Insights")

    row = 3
    precomputed = highlights or {}
    theme_rows = _build_key_theme_sentences(precomputed, insights)
    if theme_rows:
        sheet.cell(row, 1, "Key Themes & Takeaways")
        row += 1
        for line in theme_rows:
            sheet.cell(row, 1, line)
            row += 1
        row += 1

    forensic_rows = (highlights or {}).get("theme_forensics") or []
    if forensic_rows:
        sheet.cell(row, 1, "Forensic Theme Assessment")
        row += 1
        headers = ["Theme", "Score", "Status", "Forensic Defense"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row, col, header)
            cell.font = Font(bold=True)
        row += 1
        for payload in forensic_rows:
            sheet.cell(row, 1, payload.get("theme", ""))
            score_value = payload.get("score")
            sheet.cell(row, 2, f"{score_value:.2f}" if isinstance(score_value, (int, float)) else score_value)
            sheet.cell(row, 3, payload.get("status", ""))
            sheet.cell(row, 4, payload.get("defense", ""))
            row += 1
        row += 1

    scope = precomputed.get('scope') or insights.get('scope') or {}
    scope_rows = [
        ("Campaigns audited", scope.get('campaigns')),
        ("Ad groups evaluated", scope.get('ad_groups')),
        ("Keywords reviewed", scope.get('keywords')),
        ("Ads reviewed", scope.get('ads')),
        ("Audiences analyzed", scope.get('audiences')),
        ("Conversion actions tracked", scope.get('conversions_tracked')),
    ]
    if any(value for _, value in scope_rows):
        sheet.cell(row, 1, "Scope of Audit")
        row += 1
        for label, value in scope_rows:
            sheet.cell(row, 1, label)
            sheet.cell(row, 2, value if value not in (None, '') else '--')
            row += 1
        row += 1

    context_block = precomputed.get("business_context") or {}
    if context_block:
        sheet.cell(row, 1, "Business Physics Context")
        row += 1
        constraint_block = context_block.get("constraints") or {}
        context_lines = [
            ("Detected Vertical", context_block.get("detected_vertical") or "Unknown"),
            ("Primary KPI", constraint_block.get("primary_kpi")),
            ("Min. conversions for Smart Bidding", constraint_block.get("min_conversion_volume")),
            ("Target Country", (context_block.get("target_country") or "unknown").upper()),
            ("Break-even ROAS", context_block.get("break_even_roas")),
        ]
        for label, value in context_lines:
            if value in (None, ""):
                continue
            sheet.cell(row, 1, label)
            sheet.cell(row, 2, value)
            row += 1
        row += 1

    guardrails = precomputed.get("guardrails") or []
    if guardrails:
        sheet.cell(row, 1, "Guardrail Monitor")
        row += 1
        sheet.cell(row, 1, "Level / Section").font = Font(bold=True)
        sheet.cell(row, 2, "Detail").font = Font(bold=True)
        sheet.cell(row, 3, "Recommended Action").font = Font(bold=True)
        row += 1
        for finding in guardrails:
            level = finding.get("level", "INFO")
            section = finding.get("section", "General")
            sheet.cell(row, 1, f"{level} — {section}")
            sheet.cell(row, 2, finding.get("detail", ""))
            sheet.cell(row, 3, finding.get("action", ""))
            row += 1
        row += 1

    prior_context = _load_prior_insight_snippets(account_name)
    rows = _build_insight_rows(account_name, insights, highlights, prior_context)
    if rows:
        row += 1
        header_row = row
        for idx, column in enumerate(INSIGHT_TABLE_COLUMNS, start=1):
            cell = sheet.cell(header_row, idx, column)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EFEFEF")
            sheet.column_dimensions[get_column_letter(idx)].width = max(len(column) + 4, 18)
        row += 1
        for record in rows:
            for idx, column in enumerate(INSIGHT_TABLE_COLUMNS, start=1):
                sheet.cell(row, idx, record.get(column, ""))
            row += 1

    try:
        workbook.save(workbook_path)
    except Exception as exc:
        logging.error("[GenerateAudit] Failed to save workbook after writing insights: %s", exc)

THEME_ORDER = [
    "Transparency & Campaign Settings",
    "Strategic Alignment",
    "Media & Channel Planning",
    "Data, Targeting, and Technology",
    "Optimization & Automation",
    "Creative, Messaging, and Formats",
    "Measurement",
]

THEME_INDEX = {theme: idx for idx, theme in enumerate(THEME_ORDER)}

THEME_KEYWORDS = {
    "Transparency & Campaign Settings": [
        "workflow",
        "training",
        "ownership",
        "process",
        "team",
        "enablement",
        "governance",
        "people",
        "operations",
        "documentation",
        "naming convention",
        "playbook",
        "hand-off",
        "aks/approval",
        "roles",
        "escalation",
        "setting",
        "schedule",
        "ad schedule",
        "location setting",
        "geo setting",
        "brand safety",
        "exclusion",
        "negative list",
        "transparency",
    ],
    "Strategic Alignment": [
        "strategy",
        "strategic",
        "alignment",
        "objective",
        "goal",
        "roadmap",
        "north star",
        "business priority",
        "vision",
        "kpi ladder",
        "pillar",
        "charter",
        "governance goal",
    ],
    "Media & Channel Planning": [
        "channel",
        "campaign mix",
        "planning",
        "allocation",
        "media",
        "coverage",
        "expansion",
        "portfolio",
        "network",
        "match type",
        "negative keyword",
        "keyword structure",
        "geo",
        "location",
        "inventory",
        "upper funnel",
        "mid funnel",
    ],
    "Data, Targeting, and Technology": [
        "data",
        "target",
        "audience",
        "technology",
        "signal",
        "pmax",
        "automation",
        "ai readiness",
        "first-party",
        "tagging",
        "integration",
        "api",
        "segment",
        "tech stack",
        "feed",
        "ga4",
        "measurement tag",
    ],
    "Optimization & Automation": [
        "bid",
        "bidding",
        "buying strategy",
        "auction",
        "cpc",
        "budget pacing",
        "manual",
        "tcpa",
        "troas",
        "day-part",
        "daypart",
        "smart bidding",
        "budget",
        "guardrail",
        "pacing",
        "portfolio",
    ],
    "Creative, Messaging, and Formats": [
        "creative",
        "messaging",
        "ad copy",
        "format",
        "asset",
        "landing page",
        "headline",
        "cta",
        "extension",
        "video",
        "responsive search ad",
        "story",
        "creative testing",
    ],
    "Measurement": [
        "measurement",
        "attribution",
        "tracking",
        "report",
        "kpi",
        "experiment",
        "test",
        "incrementality",
        "data manipulation",
        "qa",
        "benchmark",
        "lift study",
        "conversion tracking",
    ],
}

# Targeted hints let us connect higher-level intent (section headings, prompts) to themes.
SECTION_THEME_HINTS = {
    "goal-consistent": "Strategic Alignment",
    "business goal": "Strategic Alignment",
    "guardrail": "Optimization & Automation",
    "smart bidding": "Optimization & Automation",
    "signal density": "Data, Targeting, and Technology",
    "audience targeting": "Data, Targeting, and Technology",
    "creative excellence": "Creative, Messaging, and Formats",
    "brilliant basics": "Transparency & Campaign Settings",
    "workflow": "Transparency & Campaign Settings",
    "measurement framework": "Measurement",
}

ROW_INTENT_HINTS = {
    "roadmap": "Strategic Alignment",
    "org objective": "Strategic Alignment",
    "portfolio mix": "Media & Channel Planning",
    "match type": "Media & Channel Planning",
    "audience segment": "Data, Targeting, and Technology",
    "first-party": "Data, Targeting, and Technology",
    "automation": "Buying Strategy",
    "manual cpc": "Buying Strategy",
    "creative testing": "Creative, Messaging, and Formats",
    "messaging": "Creative, Messaging, and Formats",
    "lift study": "Measurement",
    "tracking": "Measurement",
}

CATEGORY_THEME_DEFAULTS: List[Tuple[str, str]] = [
    ("smart bidding", "Optimization & Automation"),
    ("keywords", "Media & Channel Planning"),
    ("negative keywords", "Media & Channel Planning"),
    ("ad copy", "Creative, Messaging, and Formats"),
    ("landing pages", "Creative, Messaging, and Formats"),
    ("bids", "Optimization & Automation"),
    ("budgets", "Optimization & Automation"),
    ("settings", "Transparency & Campaign Settings"),
    ("audience targeting", "Media & Channel Planning"),
    ("audience segments", "Data, Targeting, and Technology"),
    ("technology", "Data, Targeting, and Technology"),
    ("data integration", "Data, Targeting, and Technology"),
    ("data collection", "Data, Targeting, and Technology"),
    ("data manipulation", "Measurement"),
    ("measurement", "Measurement"),
    ("location", "Media & Channel Planning"),
]


def _clean_text(value: Optional[str]) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _determine_theme(context: Dict[str, str]) -> Optional[str]:
    scores: Dict[str, float] = {theme: 0.0 for theme in THEME_ORDER}

    def _bump(theme: str, weight: float = 1.0) -> None:
        if theme in scores:
            scores[theme] += weight

    section_text = context.get("section", "").lower()
    category_text = context.get("category", "").lower()
    item_text = context.get("item", "").lower()
    question_text = context.get("question", "").lower()
    example_text = context.get("example", "").lower()
    detail_text = context.get("detail", "").lower()
    all_text = " ".join(
        filter(None, [section_text, category_text, item_text, question_text, example_text, detail_text])
    )

    for theme, keywords in THEME_KEYWORDS.items():
        if not keywords:
            continue
        if any(keyword in all_text for keyword in keywords):
            _bump(theme, 1.0)

    for hint, theme in SECTION_THEME_HINTS.items():
        if hint in section_text:
            _bump(theme, 2.0)

    for hint, theme in ROW_INTENT_HINTS.items():
        if hint in all_text:
            _bump(theme, 1.5)

    if category_text.startswith("naming") or "workflow" in category_text or "setting" in category_text:
        _bump("Transparency & Campaign Settings", 1.25)

    if any(term in category_text for term in ("budget", "bidding", "pacing", "cpa", "roas")):
        _bump("Optimization & Automation", 1.25)

    if "channel" in category_text or "match" in category_text:
        _bump("Media & Channel Planning", 1.0)

    if not any(score > 0 for score in scores.values()):
        fallback_text = " ".join(filter(None, [category_text, section_text, item_text]))
        for pattern, theme in CATEGORY_THEME_DEFAULTS:
            if pattern in fallback_text:
                return theme
        return THEME_ORDER[0]

    best_theme, _ = max(
        scores.items(),
        key=lambda item: (item[1], -THEME_INDEX[item[0]]),
    )
    return best_theme


# ------------------------------------------------------------
# Weighted scoring helpers (optional, flag-controlled)
# ------------------------------------------------------------
def _match_rule(entry: Dict[str, object], rules: List[Dict[str, object]]) -> Optional[Dict[str, object]]:
    """
    Heuristic match of a rule to a scorecard entry using item/category text.
    If no match, returns None and defaults apply.
    """
    item = str(entry.get("item") or "").strip().lower()
    category = str(entry.get("category") or "").strip().lower()
    if not item and not category:
        return None
    for rule in rules:
        name = str(rule.get("rule_name") or "").strip().lower()
        control_code = str(rule.get("control_code") or "").strip().lower()
        if name and (name == item or name == category):
            return rule
        if control_code and (control_code == item or control_code == category):
            return rule
    return None


def _apply_vertical_override(rule: Dict[str, object], vertical: str) -> Tuple[float, bool, Optional[float]]:
    """
    Returns (weight, must_pass, failure_threshold) for a rule given a vertical.
    Defaults to base values if no override matches.
    """
    weight = float(rule.get("base_weight", 1.0) or 1.0)
    must_pass = bool(rule.get("base_must_pass", False))
    threshold = rule.get("failure_threshold")

    for override in rule.get("vertical_overrides", []) or []:
        if str(override.get("vertical", "")).upper() == vertical:
            weight = float(override.get("weight", weight) or weight)
            must_pass = bool(override.get("must_pass", must_pass))
            threshold = override.get("failure_threshold", threshold)
            break
    return weight, must_pass, threshold


def _compute_weighted_scores(entries: List[Dict[str, object]]) -> Optional[Dict[str, object]]:
    """
    Computes weighted theme/overall scores from scorecard entries (0-5 scale).
    Returns None when mode is AVERAGE or config is unavailable to preserve current behavior.
    """
    mode = _get_scoring_mode()
    if mode == "AVERAGE":
        return None

    config = _load_weighted_config()
    if not config:
        logging.info("[WeightedScoring] Config missing; falling back to average.")
        return None

    vertical = _resolve_vertical()
    rules = config.get("rules") or []

    theme_weighted_points: Dict[str, float] = {t: 0.0 for t in THEME_ORDER}
    theme_possible_points: Dict[str, float] = {t: 0.0 for t in THEME_ORDER}
    critical_errors: List[str] = []

    for entry in entries:
        score = entry.get("score")
        try:
            rating = float(score)
        except (TypeError, ValueError):
            continue

        theme = _determine_theme(entry)
        if not theme:
            continue

        rule = _match_rule(entry, rules)
        if rule:
            weight, must_pass, threshold = _apply_vertical_override(rule, vertical)
        else:
            weight, must_pass, threshold = 1.0, False, None

        # Skip zero-weight rules to avoid diluting the denominator.
        if weight <= 0:
            continue

        # Must-pass check on 0-5 scale. If no threshold is provided, use <1.0 as the default fail gate.
        if must_pass:
            fail_threshold = 1.0 if threshold is None else threshold
            if rating < fail_threshold:
                critical_errors.append(
                    f"Must-pass failed: {entry.get('item') or entry.get('category') or 'Unnamed'} "
                    f"(score {rating:.2f} < {fail_threshold})"
                )
                if mode == "STRICT":
                    return {
                        "final_score": 0.0,
                        "theme_scores": {t: 0.0 for t in THEME_ORDER},
                        "critical_errors": critical_errors,
                    }

        earned = rating * weight
        possible = 5.0 * weight  # ratings are 0-5
        theme_weighted_points[theme] += earned
        theme_possible_points[theme] += possible

    if critical_errors and mode in {"WEIGHTED", "STRICT"}:
        return {
            "final_score": 0.0,
            "theme_scores": {t: 0.0 for t in THEME_ORDER},
            "critical_errors": critical_errors,
        }

    theme_scores: Dict[str, float] = {}
    total_earned = 0.0
    total_possible = 0.0
    for theme in THEME_ORDER:
        possible = theme_possible_points.get(theme, 0.0)
        earned = theme_weighted_points.get(theme, 0.0)
        total_earned += earned
        total_possible += possible
        theme_scores[theme] = round((earned / possible) * 5.0, 2) if possible > 0 else 0.0

    final_score = round((total_earned / total_possible) * 5.0, 2) if total_possible > 0 else 0.0
    return {
        "final_score": final_score,
        "theme_scores": theme_scores,
        "critical_errors": critical_errors,
    }


def _write_theme_information(workbook_path: Path) -> None:
    try:
        wb = load_workbook(workbook_path)
    except Exception as exc:
        logging.error("[GenerateAudit] Unable to open workbook for theme summary: %s", exc)
        return

    scorecard = wb.active
    theme_values: Dict[str, List[float]] = {theme: [] for theme in THEME_ORDER}
    theme_cell_refs: Dict[str, List[str]] = {theme: [] for theme in THEME_ORDER}
    theme_scores: Dict[str, float] = {}
    rating_col = SCORECARD_RATING_COL

    def _context_for_row(row_idx: int) -> Dict[str, Optional[str]]:
        section = _clean_text(scorecard.cell(row_idx, 1).value)
        category = _clean_text(scorecard.cell(row_idx, SCORECARD_CATEGORY_COL).value)
        item = _clean_text(scorecard.cell(row_idx, SCORECARD_ITEM_COL).value)
        detail = _clean_text(scorecard.cell(row_idx, SCORECARD_DETAILS_COL).value)
        return {
            "section": section,
            "question": category or section,
            "example": item,
            "category": category,
            "item": item,
            "detail": detail,
        }

    for row in range(12, scorecard.max_row + 1):
        context = _context_for_row(row)
        rating_cell = scorecard.cell(row, rating_col).value
        try:
            rating = float(rating_cell)
        except (TypeError, ValueError):
            continue
        theme = _determine_theme(context)
        if theme:
            theme_values[theme].append(rating)
            cell_ref = f"{get_column_letter(SCORECARD_RATING_COL)}{row}"
            theme_cell_refs.setdefault(theme, []).append(cell_ref)

    theme_examples: Dict[str, List[str]] = {theme: [] for theme in THEME_ORDER}

    for row in range(12, scorecard.max_row + 1):
        context = _context_for_row(row)
        rating_cell = scorecard.cell(row, rating_col).value
        try:
            rating = float(rating_cell)
        except (TypeError, ValueError):
            continue
        theme = _determine_theme(context)
        if not theme:
            continue
        descriptor = context.get("item") or context.get("detail") or context.get("question")
        if descriptor:
            theme_examples[theme].append(descriptor)

    for theme in THEME_ORDER:
        values = theme_values.get(theme, [])
        if values:
            avg_score = round(sum(values) / len(values), 2)
        else:
            avg_score = 0.0
        theme_scores[theme] = avg_score

    try:
        _write_theme_dashboard(scorecard, theme_scores, theme_examples, theme_cell_refs)
    except Exception as exc:
        logging.warning("[GenerateAudit] Failed to update dashboard theme table: %s", exc)

    try:
        wb.save(workbook_path)
    except Exception as exc:
        logging.error("[GenerateAudit] Failed to save workbook after writing theme context: %s", exc)


def _score_bucket(score: float) -> str:
    if score >= 4.2:
        return "elite"
    if score >= 3.8:
        return "steady"
    if score >= 3.3:
        return "mixed"
    if score >= 2.8:
        return "fragile"
    return "critical"


def _select_variant(options: List[str], seed: str) -> str:
    if not options:
        return ""
    idx = abs(hash(seed)) % len(options)
    return options[idx]


def _summarize_theme_drivers(examples: List[str]) -> Tuple[str, int]:
    cleaned: List[str] = []
    for descriptor in examples:
        text = (descriptor or "").strip()
        if not text or text in cleaned:
            continue
        cleaned.append(text)
        if len(cleaned) == 3:
            break
    if not cleaned:
        return "the scorecard signals reviewed", 0
    if len(cleaned) == 1:
        return cleaned[0], 1
    if len(cleaned) == 2:
        return f"{cleaned[0]} and {cleaned[1]}", 2
    return f"{cleaned[0]}, {cleaned[1]}, and {cleaned[2]}", 3


def _build_theme_rationale_text(theme: str, score: float, examples: List[str]) -> str:
    persona = THEME_PERSONA_BLUEPRINTS.get(theme, {})
    drivers_text, driver_sample = _summarize_theme_drivers(examples)
    total_controls = max(sum(1 for item in examples if (item or "").strip()), 1)
    bucket = _score_bucket(score)
    voice = SCORE_BUCKET_VOICE.get(bucket, SCORE_BUCKET_VOICE["mixed"])

    client_voice = _select_variant(voice["client"], f"{theme}:{drivers_text}") or ""
    client_voice = client_voice.format(theme=theme, theme_lower=theme.lower())
    action_voice = _select_variant(voice["action"], f"{drivers_text}:{theme}") or ""
    action_voice = action_voice.format(theme=theme, theme_lower=theme.lower())

    client_summary = persona.get("client_summary", "")
    if client_summary:
        client_summary = client_summary.format(theme=theme, theme_lower=theme.lower())

    ops_template = persona.get("ops_focus_template", "Key drivers: {drivers}.")
    ops_line = ops_template.format(drivers=drivers_text, theme=theme, theme_lower=theme.lower())

    detail_template = persona.get("detail_focus_template", "")
    detail_line = ""
    if detail_template:
        detail_line = detail_template.format(drivers=drivers_text, theme=theme, theme_lower=theme.lower())

    definition = THEME_DEFINITIONS.get(theme, "").strip()
    definition_line = f"{definition}" if definition else ""
    score_line = f"{theme} averages {score:.2f}/5 across {max(driver_sample, total_controls)} evaluated controls."

    diagnostic_note = persona.get("diagnostic_summary") or THEME_DIAGNOSTIC_HOOKS.get(theme, "")
    diagnostic_line = f"Diagnostic focus: {diagnostic_note}" if diagnostic_note else ""

    segments = [
        client_summary,
        client_voice,
        ops_line,
        detail_line,
        definition_line,
        score_line,
        action_voice,
        diagnostic_line,
    ]
    return " ".join(part.strip() for part in segments if part)


def _write_theme_dashboard(
    scorecard,
    theme_scores: Dict[str, float],
    theme_examples: Dict[str, List[str]],
    theme_cell_refs: Dict[str, List[str]],
) -> None:
    if not theme_scores:
        return

    start_row = 95
    theme_col = 2
    definition_col = 3
    score_col = 4
    rationale_col = 5
    header_row = start_row - 1

    scorecard.cell(header_row, theme_col, "Theme")
    scorecard.cell(header_row, definition_col, "Definition")
    scorecard.cell(header_row, score_col, "Score")
    scorecard.cell(header_row, rationale_col, "Rationale")

    for idx, theme in enumerate(THEME_ORDER):
        row = start_row + idx
        score = round(theme_scores.get(theme, 0.0), 2)
        scorecard.cell(row, theme_col, theme)
        scorecard.cell(row, definition_col, THEME_DEFINITIONS.get(theme, ""))
        formula_refs = theme_cell_refs.get(theme, [])
        if formula_refs:
            avg_formula = "AVERAGE({})".format(",".join(formula_refs))
            scorecard.cell(row, score_col).value = f"=ROUND({avg_formula},2)"
        else:
            scorecard.cell(row, score_col, score)
        scorecard.cell(
            row,
            rationale_col,
            _build_theme_rationale_text(theme, score, theme_examples.get(theme, [])),
        )


def _load_prior_insight_snippets(account_name: str, limit: int = 3) -> List[Dict[str, str]]:
    """
    Pull prior insights for this account from the Cognitive Search index so we can
    enrich the Insights tab with historical context.
    """

    endpoint = os.environ.get("CONCIERGE_SEARCH_ENDPOINT")
    key = os.environ.get("CONCIERGE_SEARCH_KEY")
    index = os.environ.get("CONCIERGE_SEARCH_INDEX", "kai-audits")

    if not endpoint or not key or SearchClient is None or AzureKeyCredential is None:
        return []

    safe_account = (account_name or "").replace("'", "''")
    if not safe_account:
        return []

    section_filters = " or ".join([f"section eq '{sec}'" for sec in PRIOR_INSIGHT_SECTIONS])
    filter_expr = f"accountName eq '{safe_account}' and ({section_filters})"

    try:
        client = SearchClient(endpoint=endpoint, index_name=index, credential=AzureKeyCredential(key))
        results = client.search(
            search_text="*",
            filter=filter_expr,
            order_by=["auditDate desc"],
            top=limit * 2,
        )
    except Exception as exc:  # pragma: no cover - depends on env
        logging.warning("[GenerateAudit] Unable to load prior insights for %s: %s", account_name, exc)
        return []

    snippets: List[Dict[str, str]] = []
    for doc in results:
        content = (doc.get("content") or "").strip()
        if not content:
            continue
        snippets.append(
            {
                "content": content,
                "title": doc.get("title") or doc.get("section") or "",
                "auditDate": doc.get("auditDate", ""),
            }
        )
        if len(snippets) >= limit:
            break
    return snippets


def _collect_scorecard_entries(sheet) -> List[Dict[str, Optional[str]]]:
    entries: List[Dict[str, Optional[str]]] = []
    for row in range(12, sheet.max_row + 1):
        category = sheet.cell(row, SCORECARD_CATEGORY_COL).value
        item = sheet.cell(row, SCORECARD_ITEM_COL).value
        rating_cell = sheet.cell(row, SCORECARD_RATING_COL).value
        details = sheet.cell(row, SCORECARD_DETAILS_COL).value
        action = sheet.cell(row, SCORECARD_ACTIONS_COL).value
        rationale = sheet.cell(row, SCORECARD_RATIONALE_COL).value
        try:
            rating = float(rating_cell)
        except (TypeError, ValueError):
            continue
        entries.append(
            {
                "category": str(category or "").strip(),
                "item": str(item or "").strip(),
                "score": rating,
                "details": str(details or "").strip(),
                "action": _normalize_action_text(str(action or "")),
                "rationale": str(rationale or "").strip(),
            }
        )
    return entries

def _build_audit_highlights(workbook_path: Path) -> Optional[Dict[str, object]]:
    """Return precomputed strengths/risks/averages for downstream sections."""
    try:
        wb = load_workbook(workbook_path, data_only=True)
    except Exception as exc:
        logging.error("[GenerateAudit] Unable to open workbook for highlight extraction: %s", exc)
        return None

    sheet = wb.active
    entries = _collect_scorecard_entries(sheet)
    if not entries:
        wb.close()
        return None

    strengths = sorted((e for e in entries if e["score"] >= 4), key=lambda e: e["score"], reverse=True)[:5]
    risks = sorted((e for e in entries if e["score"] <= 2), key=lambda e: e["score"])[:5]
    avg_score = round(sum(e["score"] for e in entries) / len(entries), 2)
    theme_forensics = _build_theme_forensics(entries)
    weighted_result = _compute_weighted_scores(entries)
    wb.close()

    result: Dict[str, object] = {
        "entries": entries,
        "strengths": strengths,
        "risks": risks,
        "average_score": avg_score,
        "theme_forensics": theme_forensics,
    }
    if weighted_result:
        result["weighted"] = weighted_result
    return result


def _score_descriptor(score: float) -> str:
    if score >= 4.2:
        return "best-in-class maturity"
    if score >= 3.8:
        return "strong but still evolving performance"
    if score >= 3.3:
        return "steady yet inconsistent execution"
    if score >= 2.8:
        return "fragile performance that needs reinforcement"
    return "critical remediation requirements"


def _theme_status(score: float) -> str:
    if score >= 4.5:
        return "Dominance"
    if score >= 4.0:
        return "Strength"
    if score >= 3.0:
        return "Adequate"
    if score >= 2.0:
        return "Fragile"
    return "Absence"


def _format_forensic_driver(row: Optional[Dict[str, object]]) -> str:
    if not row:
        return "No validated metric recorded."
    item = (row.get("item") or row.get("category") or "control").strip()
    try:
        rating = float(row.get("score", 0.0))
    except (TypeError, ValueError):
        rating = 0.0
    detail_source = row.get("details") or row.get("rationale") or row.get("action") or ""
    detail = textwrap.shorten(str(detail_source).strip(), width=170, placeholder="�?�")
    if detail:
        return f"{item} scored {rating:.2f}/5 — {detail}"
    return f"{item} scored {rating:.2f}/5."


def _select_forensic_push(rows: List[Dict[str, object]]) -> Optional[Dict[str, object]]:
    if not rows:
        return None
    sorted_rows = sorted(rows, key=lambda r: float(r.get("score", 0.0)), reverse=True)
    for row in sorted_rows:
        try:
            if float(row.get("score", 0.0)) >= 4.5:
                return row
        except (TypeError, ValueError):
            continue
    return sorted_rows[0]


def _select_forensic_pull(rows: List[Dict[str, object]]) -> Optional[Dict[str, object]]:
    if not rows:
        return None
    sorted_rows = sorted(rows, key=lambda r: float(r.get("score", 0.0)))
    for row in sorted_rows:
        try:
            if float(row.get("score", 5.0)) <= 3.0:
                return row
        except (TypeError, ValueError):
            continue
    return sorted_rows[0]


def _build_theme_forensics(entries: List[Dict[str, object]]) -> List[Dict[str, object]]:
    theme_entries: Dict[str, List[Dict[str, object]]] = {theme: [] for theme in THEME_ORDER}
    for entry in entries:
        context = {
            "section": entry.get("category", ""),
            "question": entry.get("category", ""),
            "example": entry.get("item", ""),
            "category": entry.get("category", ""),
            "item": entry.get("item", ""),
            "detail": entry.get("details", ""),
        }
        theme = _determine_theme(context)
        if not theme:
            continue
        theme_entries.setdefault(theme, []).append(entry)

    forensic_rows: List[Dict[str, object]] = []
    for theme in THEME_ORDER:
        rows = [r for r in theme_entries.get(theme, []) if isinstance(r.get("score"), (int, float))]
        if not rows:
            continue
        scores = [float(r["score"]) for r in rows]
        score = round(sum(scores) / len(scores), 2)
        status = _theme_status(score)
        push_row = _select_forensic_push(rows)
        pull_row = _select_forensic_pull(rows)
        defense = f"Push — {_format_forensic_driver(push_row)}. Pull — {_format_forensic_driver(pull_row)}."
        forensic_rows.append(
            {
                "theme": theme,
                "score": score,
                "status": status,
                "defense": defense,
            }
        )
    return forensic_rows


def _build_key_theme_sentences(
    highlights: Optional[Dict[str, object]],
    insights: Dict[str, object],
) -> List[str]:
    sentences: List[str] = []
    highlights = highlights or {}

    entries = highlights.get("entries") or []
    avg_score = highlights.get("average_score")
    if avg_score and entries:
        descriptor = _score_descriptor(avg_score)
        sentences.append(
            f"Overall score: {avg_score:.2f}/5 across {len(entries)} checks, indicating {descriptor}."
        )

    strengths = highlights.get("strengths") or []
    if strengths:
        top_strength = strengths[0]
        detail = textwrap.shorten(
            (top_strength.get("details") or top_strength.get("category") or "").strip(),
            width=130,
            placeholder="…",
        )
        sentences.append(f"Momentum driver: {top_strength.get('item')} — {detail}")

    risks = highlights.get("risks") or []
    if risks:
        top_risk = risks[0]
        detail = textwrap.shorten(
            (top_risk.get("details") or top_risk.get("category") or "").strip(),
            width=130,
            placeholder="…",
        )
        sentences.append(f"Urgent gap: {top_risk.get('item')} — {detail}")

    if not sentences:
        base = insights.get("key_themes") or []
        sentences = [
            textwrap.shorten(line.strip(), width=140, placeholder="…")
            for line in base
            if line.strip()
        ]

    return sentences[:3]


def _derive_theme_group(text: str, dimension: str) -> str:
    lowered = text.lower()
    mapping = [
        ("Brand Efficiency", ["brand", "exact", "defend", "branded"]),
        ("Non-Brand Scale & Efficiency", ["non-brand", "prospecting", "rollover", "generic"]),
        ("Upper Funnel & Cross-Channel", ["pmax", "youtube", "display", "upper funnel", "cross-channel"]),
        ("Measurement & Tracking", ["conversion", "tracking", "measurement", "attribution", "signal"]),
        ("Audience & Remarketing", ["audience", "remarketing", "customer match", "lists"]),
        ("Bid Strategy & Budgets", ["budget", "bid", "smart bidding", "tcpa", "troas"]),
        ("Structure & Hygiene", ["structure", "match type", "naming", "campaign split", "redundant"]),
    ]
    for label, keywords in mapping:
        if any(keyword in lowered for keyword in keywords):
            return label
    return dimension


def _infer_dimension_from_text(text: str) -> str:
    text_lower = text.lower()
    for dimension, keywords in DIMENSION_KEYWORDS.items():
        if any(keyword in text_lower for keyword in keywords):
            return dimension
    return "Strategies"


def _section_for_row(row_type: str, lob_name: str) -> str:
    if row_type == "Opportunity":
        return "Roadmap & Phasing"
    if lob_name == "All":
        return "Cross-Program Insights"
    return "LOB Deep Dive"


def _business_impact_for_dimension(dimension: str, row_type: str, summary: str) -> str:
    base = BUSINESS_IMPACT_TEMPLATES.get(
        dimension,
        "Clarifies how this area impacts paid search efficiency and growth velocity.",
    )
    headline = summary.split(".")[0].strip()
    if row_type == "Strength":
        extra = DIMENSION_STRENGTH_STORIES.get(dimension, "")
        parts = [headline]
        if extra:
            parts.append(extra)
        parts.append(base)
        return " ".join(part.strip() for part in parts if part)
    extra = DIMENSION_GAP_STORIES.get(dimension, "This gap slows down our ability to scale cleanly.")
    return f"{headline}. {extra} {base}"


def _phase_for_dimension(dimension: str) -> str:
    return DIMENSION_PHASE_MAP.get(dimension, "2")


def _priority_for_phase(phase: str, row_type: str) -> str:
    if row_type == "Strength":
        return "M"
    return {"1": "H", "2": "M", "3": "L"}.get(phase, "M")


def _recommended_action_for_dimension(dimension: str, text: str, row_type: str) -> str:
    lower = text.lower()
    if "action:" in lower:
        idx = lower.index("action:")
        extracted = text[idx + len("action:") :].strip()
        return _normalize_action_text(extracted)
    template = _choose_action_template(dimension, row_type, text)
    if template:
        return _normalize_action_text(template)
    if row_type == "Strength":
        return "Keep this win documented so it scales across teams."
    return "Assign an owner and timeline to close this gap."


def _split_sentences(text: str) -> List[str]:
    return [
        sentence.strip()
        for sentence in re.split(r"(?<=\.)\s+", text.strip())
        if sentence.strip()
    ]


def _compose_insight_summary(title: str, evidence: str, row_type: str) -> str:
    sentences = _split_sentences(evidence)
    fact = sentences[0].rstrip(".") if sentences else title.rstrip(".")
    interpretation = title.rstrip(".")
    if row_type == "Strength":
        return f"{fact}. It proves {interpretation.lower()} and reinforces momentum."
    return f"{fact}. It signals {interpretation.lower()} needs to tighten before we scale."


def _compose_supporting_data(evidence: str) -> str:
    sentences = _split_sentences(evidence)
    if not sentences:
        return evidence
    if len(sentences) == 1:
        return sentences[0]
    return " ".join(sentences[:2])


def _clean_summary_for_template(text: str) -> str:
    if not text:
        return "this area"
    cleaned = text.replace("ACTION:", "").strip()
    cleaned = cleaned.split(".")[0].strip()
    return cleaned or "this area"


def _choose_action_template(dimension: str, row_type: str, summary: str) -> Optional[str]:
    bucket = ACTION_BLUEPRINTS.get(dimension) or ACTION_BLUEPRINTS["default"]
    key = "strength" if row_type == "Strength" else "opportunity"
    options = bucket.get(key) or ACTION_BLUEPRINTS["default"].get(key, [])
    if not options:
        return None
    seed = f"{dimension}:{row_type}:{summary}"
    idx = abs(hash(seed)) % len(options)
    summary_clean = _clean_summary_for_template(summary)
    template = options[idx]
    return template.format(summary=summary_clean, summary_lower=summary_clean.lower())


def _normalize_action_text(text: str) -> str:
    cleaned = (text or "").strip()
    if not cleaned:
        return ""
    cleaned = re.sub(r"^\s*action:\s*", "", cleaned, flags=re.IGNORECASE).strip()
    return cleaned.rstrip(".")


def _compose_executive_narrative(summary: str, data: str, impact: str, action: str) -> str:
    parts = []
    if summary:
        parts.append(summary.rstrip("."))
    if data:
        parts.append(f"Data: {data.rstrip('.')}")
    if impact:
        parts.append(f"This means {impact.rstrip('.')}.")
    if action:
        action_clean = _normalize_action_text(action)
        if action_clean:
            parts.append(action_clean.rstrip(".") + ".")
    return " ".join(parts).strip()


def _format_scope_label(label: str) -> str:
    return label.replace("_", " ").title()


def _scope_blueprint(label: str, value: int) -> Tuple[str, str, str, str]:
    normalized = label.lower()
    blueprint = dict(SCOPE_BLUEPRINTS.get(normalized, SCOPE_BLUEPRINTS["default"]))
    dimension = blueprint.get("dimension", "Structure & Hygiene")
    row_type = blueprint.get("type", "Strength")

    if value == 0 and blueprint.get("impact_zero"):
        impact_template = blueprint["impact_zero"]
        action_template = blueprint.get("action_zero", blueprint["impact_zero"])
        row_type = blueprint.get("type_zero", row_type)
    elif blueprint.get("low_threshold") and value < blueprint["low_threshold"]:
        impact_template = blueprint.get("impact_low", blueprint.get("impact", ""))
        action_template = blueprint.get("action_low", blueprint.get("action", ""))
        row_type = blueprint.get("type_low", row_type)
    else:
        impact_template = blueprint.get("impact", "")
        action_template = blueprint.get("action", "")

    formatted_label = _format_scope_label(label)
    impact = impact_template.format(value=value, label=formatted_label)
    action = action_template.format(value=value, label=formatted_label)
    return dimension, row_type, impact, action


def _build_insight_rows(
    account_name: str,
    insights: Dict[str, object],
    highlights: Optional[Dict[str, object]] = None,
    prior_context: Optional[List[Dict[str, str]]] = None,
) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    detail_rows: List[Dict[str, str]] = []
    lob_name = (insights.get("bu") or account_name).replace("_", " ")

    scope = (highlights or {}).get("scope")
    if scope:
        for label, value in scope.items():
            summary = f"{_format_scope_label(label)}: {value}"
            dimension, row_type, impact_text, action_text = _scope_blueprint(label, value)
            phase = _phase_for_dimension(dimension)
            priority = _priority_for_phase(phase, row_type)
            rows.append(
                {
                    "Section": "Scope of Audit",
                    "Line of Business / Product": lob_name,
                    "Dimension": dimension,
                    "Type": row_type,
                    "Level": "Detail",
                    "Theme Group": "Scope of Audit",
                    "Insight Summary": summary,
                    "Supporting Data / Evidence": "",
                    "Business Impact": impact_text,
                    "Recommended Action": action_text,
                    "Executive Narrative": _compose_executive_narrative(
                        summary, "", impact_text, action_text
                    ),
                    "Priority": priority,
                    "Phase": phase,
                    "Slide Usage / Notes": "Scope of Audit metrics",
                }
            )

    def _prepare_entry(payload: Dict[str, object]) -> Dict[str, str]:
        return {
            "title": str(
                payload.get("title")
                or payload.get("item")
                or payload.get("name")
                or payload.get("criterion")
                or "Insight"
            ).strip(),
            "finding": str(
                payload.get("finding")
                or payload.get("details")
                or payload.get("summary")
                or ""
            ).strip(),
            "action": _normalize_action_text(str(payload.get("action") or "")),
            "rationale": str(
                payload.get("rationale")
                or payload.get("supporting")
                or payload.get("why")
                or ""
            ).strip(),
        }

    entries: List[Tuple[str, Dict[str, str]]] = []
    for entry in insights.get("strengths", []):
        entries.append(("Strength", _prepare_entry(entry)))
    for entry in insights.get("opportunities", []):
        entries.append(("Opportunity", _prepare_entry(entry)))

    if highlights:
        for entry in (highlights.get("strengths") or [])[:2]:
            entries.append(
                (
                    "Strength",
                    _prepare_entry(
                        {
                            "title": entry.get("item"),
                            "finding": entry.get("details"),
                            "action": entry.get("action"),
                            "rationale": entry.get("rationale"),
                        }
                    ),
                )
            )
        for entry in (highlights.get("risks") or [])[:2]:
            entries.append(
                (
                    "Opportunity",
                    _prepare_entry(
                        {
                            "title": entry.get("item"),
                            "finding": entry.get("details"),
                            "action": entry.get("action"),
                            "rationale": entry.get("rationale"),
                        }
                    ),
                )
            )

        avg_score = highlights.get("average_score")
        total = len(highlights.get("entries") or [])
        if avg_score and total:
            entries.insert(
                0,
                (
                    "Strength",
                    _prepare_entry(
                        {
                            "title": "Overall maturity snapshot",
                            "finding": f"Overall weighted score is {avg_score:.2f}/5 across {total} scored criteria, indicating the current automation baseline.",
                        }
                    ),
                ),
            )

    if not entries:
        entries = [
            (
                "Opportunity",
                {
                    "title": "Capture qualitative insights",
                    "finding": "Current dataset does not include qualitative insights; schedule stakeholder interviews before finalizing the deck.",
                },
            )
        ]

    for row_type, entry in entries:
        summary_raw = entry.get("title", "Insight").strip()
        evidence_raw = entry.get("finding", "").strip()
        evidence_text = evidence_raw
        action_override = entry.get("action")
        if not action_override:
            match = re.search(r"action:\s*(.+)", evidence_raw, flags=re.IGNORECASE)
            if match:
                action_override = _normalize_action_text(match.group(1).strip())
                evidence_text = evidence_raw[: match.start()].strip()

        dimension = _infer_dimension_from_text(f"{summary_raw} {evidence_text}")
        theme_group = _derive_theme_group(f"{summary_raw} {evidence_text}".lower(), dimension)

        phase = _phase_for_dimension(dimension)
        priority = _priority_for_phase(phase, row_type)
        insight_summary = textwrap.shorten(
            _compose_insight_summary(summary_raw, evidence_text or summary_raw, row_type),
            width=220,
            placeholder="…",
        )
        supporting_seed = entry.get("rationale") or evidence_text
        supporting_data = _compose_supporting_data(supporting_seed) or "Additional qualitative evidence required."
        business_impact = _business_impact_for_dimension(dimension, row_type, insight_summary)
        recommended_action = (
            action_override
            if action_override
            else _recommended_action_for_dimension(dimension, supporting_data, row_type)
        )
        detail_rows.append(
            {
                "Section": _section_for_row(row_type, lob_name),
                "Line of Business / Product": lob_name,
                "Dimension": dimension,
                "Type": row_type,
                "Level": "Detail",
                "Theme Group": theme_group,
                "Insight Summary": insight_summary,
                "Supporting Data / Evidence": supporting_data,
                "Business Impact": business_impact,
                "Recommended Action": recommended_action,
                "Executive Narrative": _compose_executive_narrative(
                    insight_summary, supporting_data, business_impact, recommended_action
                ),
                "Priority": priority,
                "Phase": phase,
                "Slide Usage / Notes": f"{lob_name} – {row_type} ({dimension})",
            }
        )

    themes: Dict[str, List[Dict[str, str]]] = {}
    for row in detail_rows:
        themes.setdefault(row["Theme Group"], []).append(row)

    for theme_name, theme_rows in themes.items():
        if not theme_rows:
            continue
        type_counter = {"Strength": 0, "Opportunity": 0}
        for row in theme_rows:
            type_counter[row["Type"]] += 1
        theme_type = "Opportunity" if type_counter["Opportunity"] >= type_counter["Strength"] else "Strength"
        dimension = theme_rows[0]["Dimension"]
        phase = str(
            min(
                [
                    int(r["Phase"]) if str(r["Phase"]).isdigit() else 3
                    for r in theme_rows
                ]
            )
        )

    if prior_context:
        for snippet in prior_context:
            snippet_text = (snippet.get("content") or "").strip()
            if not snippet_text:
                continue
            snippet_title = (snippet.get("title") or "").strip()
            summary_raw = snippet_title or textwrap.shorten(snippet_text, width=140, placeholder="…")
            date_text = snippet.get("auditDate") or "previous audit"
            entry_finding = f"{snippet_text} (Source: {date_text})"
            sentiment = snippet_text.lower()
            row_type = "Opportunity" if any(term in sentiment for term in ("gap", "risk", "opportunity", "issue", "challenge")) else "Strength"
            entries.append(
                (
                    row_type,
                    {
                        "title": summary_raw,
                        "finding": entry_finding,
                    },
                )
            )
        if any(r["Priority"] == "H" for r in theme_rows):
            priority = "H"
        elif any(r["Priority"] == "M" for r in theme_rows):
            priority = "M"
        else:
            priority = "L"
        summary_sentences = [r["Insight Summary"] for r in theme_rows[:2]]
        supporting = "; ".join(r["Supporting Data / Evidence"] for r in theme_rows[:2])
        summary_text = _compose_insight_summary(
            f"{theme_name} {('continues to perform' if theme_type == 'Strength' else 'needs reinforcement')}",
            supporting or "Multiple findings roll into this theme.",
            theme_type,
        )
        business_impact = _business_impact_for_dimension(dimension, theme_type, summary_text)
        recommended_action = _recommended_action_for_dimension(
            dimension, supporting or summary_text, theme_type
        )
        rows.append(
            {
                "Section": "Executive Insights" if lob_name != "All" else "Cross-Program Insights",
                "Line of Business / Product": lob_name,
                "Dimension": dimension,
                "Type": theme_type,
                "Level": "Theme",
                "Theme Group": theme_name,
                "Insight Summary": summary_text,
                "Supporting Data / Evidence": supporting,
                "Business Impact": business_impact,
                "Recommended Action": recommended_action,
                "Executive Narrative": _compose_executive_narrative(
                    summary_text, supporting, business_impact, recommended_action
                ),
                "Priority": priority,
                "Phase": phase,
                "Slide Usage / Notes": f"Key Theme – {theme_name}",
            }
        )

    # Append diagnostic framework guidance so Insights tab mirrors the updated system logic.
    diagnostic_rows: List[Dict[str, str]] = []
    for theme_name, hook in THEME_DIAGNOSTIC_HOOKS.items():
        if not hook:
            continue
        clean_hook = hook.strip()
        if clean_hook.lower().startswith("2025 diagnostic:"):
            clean_hook = clean_hook.split(":", 1)[1].strip()
        dimension = THEME_DIAGNOSTIC_DIMENSIONS.get(theme_name, "Strategies")
        row_type = "Opportunity"
        phase = _phase_for_dimension(dimension)
        priority = _priority_for_phase(phase, row_type)
        persona = THEME_PERSONA_BLUEPRINTS.get(theme_name, {})
        summary_stub = persona.get(
            "diagnostic_summary",
            f"Apply the {theme_name} safeguards from the 2025 diagnostic framework.",
        )
        action_stub = persona.get(
            "diagnostic_action",
            f"Follow the {theme_name} checklist from the 2025 diagnostic framework.",
        )
        insight_summary = textwrap.shorten(
            f"{theme_name} ({lob_name}): {summary_stub}",
            width=220,
            placeholder="...",
        )
        supporting = textwrap.shorten(clean_hook, width=280, placeholder="...")
        business_impact = _business_impact_for_dimension(dimension, row_type, insight_summary)
        recommended_action = textwrap.shorten(action_stub, width=260, placeholder="...")
        diagnostic_rows.append(
            {
                "Section": "Executive Insights",
                "Line of Business / Product": lob_name,
                "Dimension": dimension,
                "Type": row_type,
                "Level": "Theme",
                "Theme Group": theme_name,
                "Insight Summary": insight_summary,
                "Supporting Data / Evidence": supporting,
                "Business Impact": business_impact,
                "Recommended Action": recommended_action,
                "Executive Narrative": _compose_executive_narrative(
                    insight_summary, supporting, business_impact, recommended_action
                ),
                "Priority": priority,
                "Phase": phase,
                "Slide Usage / Notes": f"Diagnostic Framework – {theme_name}",
            }
        )

    final_rows: List[Dict[str, str]] = []
    unique_keys = set()
    for entry in rows + detail_rows + diagnostic_rows:
        key = (entry["Section"], entry["Theme Group"], entry["Insight Summary"])
        if key in unique_keys:
            continue
        unique_keys.add(key)
        final_rows.append(entry)

    return final_rows


def _post_process_workbook(
    workbook_path: Path,
    canonical_account: str,
    metadata: Optional[Dict[str, object]] = None,
) -> None:
    highlights = _build_audit_highlights(workbook_path) or {}
    scope_summary = (metadata or {}).get("scope_summary")
    if scope_summary:
        highlights["scope"] = scope_summary
    business_context = (metadata or {}).get("business_context")
    if business_context:
        highlights["business_context"] = business_context
    guardrails = (metadata or {}).get("guardrail_findings")
    if guardrails:
        highlights["guardrails"] = guardrails

    _write_insights_tab(workbook_path, canonical_account, highlights)
    _write_theme_information(workbook_path)


def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("[GenerateAudit] HTTP trigger function received a request")

    try:
        body = _load_request_body(req)
        requested_account = _extract_account(req, body)
        if not requested_account:
            return func.HttpResponse(
                "Please provide an account name in the 'account' query parameter.",
                status_code=400,
            )

        canonical_account = normalize_account_name(requested_account)
        if canonical_account not in ACCOUNT_TO_ENGINE:
            allowed = ", ".join(sorted(ACCOUNT_TO_ENGINE.keys()))
            return func.HttpResponse(
                f"Invalid account name '{requested_account}'. Must be one of: {allowed}",
                status_code=400,
            )
        engine_account = ACCOUNT_TO_ENGINE[canonical_account]
        context_overrides = _extract_context_overrides(req, body)
        business_context = _build_account_context(canonical_account, context_overrides)
        guardrail_findings = evaluate_guardrails(
            business_context, context_overrides.get("current_roas")
        )

        logging.info(
            "[GenerateAudit] Generating audit for account: %s (canonical: %s)",
            engine_account,
            canonical_account,
        )
        logging.info(
            "[GenerateAudit] Context detected: vertical=%s, KPI=%s, break-even=%s",
            business_context.detected_vertical,
            business_context.constraints.primary_kpi,
            business_context.break_even_roas,
        )
        for finding in guardrail_findings:
            logging.warning(
                "[GenerateAudit] Guardrail %s/%s — %s",
                finding.level,
                finding.section,
                finding.detail,
            )

        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        output_file = Path(tempfile.gettempdir()) / f"Kai_{engine_account}_{timestamp}.xlsx"

        template_path = Path(__file__).parent / "template.xlsx"
        data_dir = _resolve_data_directory(engine_account)

        logging.info("[GenerateAudit] Template: %s", template_path)
        logging.info("[GenerateAudit] Data dir: %s", data_dir)

        logging.info("[GenerateAudit] Engine source: %s", ENGINE_SOURCE)
        engine = UnifiedAuditEngine(
            template_path=template_path,
            data_directory=data_dir,
            business_unit=engine_account,
            business_context=business_context.model_dump(),
        )

        result = engine.generate_audit(account_name=engine_account, output_path=output_file)
        result["business_context"] = business_context.model_dump()
        result["guardrail_findings"] = [finding.model_dump() for finding in guardrail_findings]
        _post_process_workbook(output_file, canonical_account, result)
        try:
            result["vector_index"] = index_audit_workbook(output_file, canonical_account)
        except Exception as exc:
            logging.warning("[GenerateAudit] Vector indexing failed: %s", exc)

        logging.info("[GenerateAudit] Audit generated successfully")
        overall_score = result.get("overall_score")
        if overall_score is not None:
            logging.info("[GenerateAudit] Overall score: %.2f/5.0", overall_score)

        with open(output_file, "rb") as handle:
            file_content = handle.read()

        output_file.unlink(missing_ok=True)

        download_name = f"Kai_{engine_account}_{timestamp}.xlsx"
        return func.HttpResponse(
            file_content,
            status_code=200,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{download_name}"',
                "Access-Control-Allow-Origin": "*",
            },
        )

    except Exception as exc:
        logging.error("[GenerateAudit] Error: %s", exc, exc_info=True)
        return func.HttpResponse(
            f"An error occurred: {str(exc)}",
            status_code=500,
            headers={"Access-Control-Allow-Origin": "*"},
        )
THEME_DEFINITIONS = {
    "Transparency & Campaign Settings": "Evaluates how clearly campaigns are named, configured, and governed so settings, guardrails, and coverage are transparent and repeatable.",
    "Strategic Alignment": "Checks whether campaigns and objectives ladder up to business goals.",
    "Media & Channel Planning": "Looks at how coverage and campaign mix balance reach and focus.",
    "Data, Targeting, and Technology": "Measures the quality of signals, audiences, and tech adoption.",
    "Optimization & Automation": "Rates how well bidding, budgets, and automation are used to manage performance and scale.",
    "Creative, Messaging, and Formats": "Assesses how ad copy and formats tell the brand story.",
    "Measurement": "Validates tracking, attribution, and reporting rigor.",
}

# Theme voice profiles ensure each narrative feels human-written and grounded
# in the Paid Search Diagnostic Framework.
THEME_PERSONA_BLUEPRINTS = {
    "Transparency & Campaign Settings": {
        "client_summary": "Transparency & Campaign Settings reassure stakeholders that every campaign is auditable and privacy-safe.",
        "ops_focus_template": "This review centers on {drivers}; pressure-test those workflows against the SST plus Consent Mode checklist before scaling.",
        "detail_focus_template": "Start with the top 20% of spend-driving campaigns—if they lack SST endpoints or use \"Basic\" consent, Smart Bidding is optimizing on fragile data.",
        "diagnostic_summary": "Reconfirm SST endpoints, Consent Mode v2 (Advanced), and account-level brand exclusions so measurement loss does not undermine bidding.",
        "diagnostic_action": "Document SST coverage, enforce Consent Mode (Advanced), and refresh suitability controls before releasing additional budget.",
    },
    "Strategic Alignment": {
        "client_summary": "Strategic Alignment keeps leadership anchored to profit instead of vanity ROAS.",
        "ops_focus_template": "The drivers called out—{drivers}—need to ladder to profit-centric KPIs and margin bands, not just revenue volume.",
        "detail_focus_template": "Quantify break-even ROAS per margin tier and make sure SA360/portfolio bidding can flex budgets between Google and Microsoft accordingly.",
        "diagnostic_summary": "Tie objectives to profit KPIs and SA360 portfolio strategies so automation chases the right economics across engines.",
        "diagnostic_action": "Publish the margin-based guardrails, confirm portfolios manage budgets across engines, and sunset engine-specific bidding policies.",
    },
    "Media & Channel Planning": {
        "client_summary": "Media & Channel Planning shows whether structure gives automation enough data density.",
        "ops_focus_template": "The constraints surfaced around {drivers}; consolidate or isolate inventory so Hagakure/STAG heuristics are respected.",
        "detail_focus_template": "Use <3,000 impressions/ad group and <30 conversions/campaign as signals to consolidate, and move zero-impression SKUs into a dedicated PMax test bed.",
        "diagnostic_summary": "Apply Hagakure/STAG consolidation and isolate zombie SKUs so every campaign clears the data thresholds automation expects.",
        "diagnostic_action": "Re-map campaigns to Hagakure/STAG, spin up dedicated PMax for dormant SKUs, and add Search Themes aligned to each asset concept.",
    },
    "Data, Targeting, and Technology": {
        "client_summary": "Data, Targeting, and Technology determines how rich the signals are for Smart Bidding.",
        "ops_focus_template": "{drivers} highlight where the data pipe thins; stopgap fixes must happen before adding more automation.",
        "detail_focus_template": "Enhanced Conversions, OCI cadence, and SST coverage must be daily, hashed, and within the 63/90-day upload windows to avoid signal loss.",
        "diagnostic_summary": "Verify Enhanced Conversions plus daily OCI uploads so Smart Bidding sees true profit and not just click-to-lead noise.",
        "diagnostic_action": "Audit OCI automation, hashed identifiers, and server-side endpoints; escalate any manual uploads or missing identifiers immediately.",
    },
    "Optimization & Automation": {
        "client_summary": "Optimization & Automation ensures bids and budgets chase profitable growth instead of empty calories.",
        "ops_focus_template": "{drivers} reveal where automation still needs guardrails; lead with margin math before touching bid targets.",
        "detail_focus_template": "Import COGS into Merchant Center, enforce profit-based tROAS, and use Auction Insights (budget vs rank) before raising spend.",
        "diagnostic_summary": "Run value-based bidding with profit gates and interpret Auction Insights to separate budget caps from rank constraints.",
        "diagnostic_action": "Import COGS, reset bid targets by margin tier, and document how the team responds to Lost IS (Budget) versus Lost IS (Rank).",
    },
    "Creative, Messaging, and Formats": {
        "client_summary": "Creative, Messaging, and Formats is now targeting—the asset depth determines reach.",
        "ops_focus_template": "The findings around {drivers} show where missing asset types or HITL reviews are capping delivery.",
        "detail_focus_template": "Fill the PMax asset matrix (15 headlines, 5 long headlines, 5 descriptions, multi-aspect imagery, video, logos) and keep a HITL gate for AI-produced copy.",
        "diagnostic_summary": "Ensure every asset group hits the matrix and route all AI outputs through HITL so Ad Strength stays high across placements.",
        "diagnostic_action": "Backfill missing formats, add vertical video, and formalize HITL approvals before turning on Automatically Created Assets.",
    },
    "Measurement": {
        "client_summary": "Measurement proves the business story and reconciles platform discrepancies.",
        "ops_focus_template": "{drivers} signal where attribution or consent gaps distort KPIs—resolve those before debating performance.",
        "detail_focus_template": "Consent Mode v2, SA360 vs Google Ads timing, and Quality Score finance triage need to be interpreted together to explain deltas.",
        "diagnostic_summary": "Enforce Consent Mode (Advanced) and align SA360/Google Ads attribution windows so modeled conversions are trusted.",
        "diagnostic_action": "Document modeled conversion uplift, reconcile reporting windows, and only chase QS gains where the savings justify the lift.",
    },
}

# Score-bucket voice options to avoid repetitive phrasing.
SCORE_BUCKET_VOICE = {
    "elite": {
        "client": [
            "{theme} operates with disciplined guardrails leadership can hold up as proof of maturity.",
            "Stakeholders can point to {theme_lower} as a model for scaled execution—it already behaves like a best-in-class program.",
        ],
        "action": [
            "Protect the current routines by documenting handoffs so the next campaign inheres the same rigor.",
            "Keep the proof points visible: publish the playbooks and scorecard wins so future budget requests land faster.",
        ],
    },
    "steady": {
        "client": [
            "{theme} is steady but still has friction points to sand down.",
            "Leaders see {theme_lower} tracking well, yet the review shows a few seams worth tightening.",
        ],
        "action": [
            "Tighten the noted gaps before layering new pilots so quality does not drift.",
            "Stabilize the weak spots now so the next wave of testing rides on a clean foundation.",
        ],
    },
    "mixed": {
        "client": [
            "{theme} delivers mixed execution—some routines land, others feel improvised.",
            "Expect uneven outcomes from {theme_lower} until the highlighted workflows stop swinging between owners.",
        ],
        "action": [
            "Prioritize the listed blockers and set a short diagnostic sprint to rebuild confidence.",
            "Assign owners to the cited gaps so this pillar becomes consistent before QBR season.",
        ],
    },
    "fragile": {
        "client": [
            "{theme} remains fragile, and leadership can feel that inconsistency.",
            "Signals show {theme_lower} slipping into reactive mode; core disciplines need restoration.",
        ],
        "action": [
            "Stabilize the fundamentals—close the gaps before expanding budgets or tests.",
            "Treat the cited controls as a remediation list and track them in the operating review.",
        ],
    },
    "critical": {
        "client": [
            "{theme} requires critical remediation before any scale discussion continues.",
            "{theme_lower} is signaling emergency territory—automation is guessing without the basics in place.",
        ],
        "action": [
            "Escalate the fixes at the program level and pause net-new investments until the baseline is rebuilt.",
            "Treat this as a stop-work list: remediate, re-measure, then reopen the roadmap.",
        ],
    },
}

# Theme-specific diagnostic hooks sourced from the Paid Search Diagnostic Framework.
# These append reasoning guidance directly into workbook narratives so the audit output
# always references the system knowledge base (SST/Consent Mode/OCI/etc.).
THEME_DIAGNOSTIC_HOOKS = {
    "Transparency & Campaign Settings": (
        "2025 Diagnostic: Cross-check server-side tagging (SST) via first-party endpoints (e.g., metrics.brand.com) "
        "and Consent Mode v2 (Advanced)—Basic mode blocks tags entirely, losing ~18–19% data recovery from modeled conversions. "
        "Verify account-level brand safety exclusions and suitability settings. Missing SST or legacy consent = compliance risk + "
        "measurement attrition that invalidates Smart Bidding outcomes. These are non-negotiable foundations for 2025."
    ),
    "Strategic Alignment": (
        "2025 Diagnostic: Campaign objectives must ladder to profit-centric KPIs—break-even ROAS = 1 / margin (e.g., 20% margin → 5.0 tROAS). "
        "Single tROAS for mixed margins guarantees misallocation. SA360 portfolio bid strategies optimize marginal cost across Google/Microsoft; "
        "engine-specific strategies leave efficiency on the table. Audit for documented roadmaps that sequence foundational fixes before advanced expansion."
    ),
    "Media & Channel Planning": (
        "2025 Diagnostic: Consolidation for data density—hyper-segmentation/SKAGs are obsolete. Use Hagakure (group by landing page) or STAG (strict semantic themes). "
        "Heuristics: <3,000 impressions/week per ad group → too thin; <30 conversions/month per campaign → consolidate. "
        "PMax 'zombie SKU' risk: isolate zero-impression SKUs into separate PMax with lower ROAS targets to force testing. Use Search Themes (up to 25) aligned to asset concepts."
    ),
    "Data, Targeting, and Technology": (
        "2025 Diagnostic: Validate Enhanced Conversions for Leads (hashed email/phone) + Offline Conversion Imports (OCI) uploaded daily, within 63d (EC) / 90d (OCI). "
        "If lead-to-sale >63d, shift to proxy measurement. Signal density = match rates + SST endpoints + OCI freshness. "
        "Privacy attrition (cookie deprecation, ITP) makes server-side tracking mandatory. Failing to audit the data pipe invalidates automation outcomes. "
        "Check for first-party audience signals in feeds—quality of financial data (COGS/margin) + first-party signals = automation advantage."
    ),
    "Optimization & Automation": (
        "2025 Diagnostic: Value/profit-based bidding—import COGS; optimize to Gross Profit (Revenue – COGS). "
        "Economic pressure: Google CPCs up ~45% YoY; ROAS alone inadequate—couple with margin discipline to avoid 'empty calorie' revenue. "
        "Auction Insights triage: Lost IS (Budget) → increase budget or lower target; Lost IS (Rank) → check QS vs bid. Use Outranking Share + Overlap to detect missed segments. "
        "Performance Max must have brand exclusions to chase incremental demand, not cannibalize brand. Automation coverage should be high, but guardrails mandatory."
    ),
    "Creative, Messaging, and Formats": (
        "2025 Diagnostic: PMax asset completeness—Headlines (15), Long Headlines (5), Descriptions (5 incl. 1×60), Images (landscape/square/portrait), Video (5; include vertical to avoid auto-slideshows), Logos (5). "
        "Missing video triggers low-quality auto-generated video. Ad Strength affects eligibility; 'Poor' can disqualify auctions. "
        "Generative AI is fine for scale but Human-in-the-Loop (HITL) mandatory to prevent hallucination/brand violations. In regulated verticals, avoid Automatically Created Assets where non-compliant claims risk is high. "
        "Ensure keyword presence + semantic diversity in RSA assets for auction eligibility."
    ),
    "Measurement": (
        "2025 Diagnostic: Consent Mode v2 (Advanced) enables cookieless pings for modeled conversions; recovers ~18–19% lost data vs Basic. "
        "SA360 vs Google Ads discrepancies = click-date (GA) vs conversion-date (Floodlight) attribution; cross-channel logic can credit Display. Lookback windows differ (Floodlight up to 90d). "
        "Discrepancy ≠ error—interpret by model/window. Audit modeled conversions and attribution windows before trusting KPIs. "
        "Framework flags signal loss as #1 reason Smart Bidding underperforms. Quality Score as finance lever: improving QS 5→10 can ~halve CPC—focus where potential savings × volume justify effort."
    ),
}

THEME_DIAGNOSTIC_DIMENSIONS = {
    "Transparency & Campaign Settings": "Structure & Hygiene",
    "Strategic Alignment": "Strategies",
    "Media & Channel Planning": "Cross-Channel",
    "Data, Targeting, and Technology": "Measurement & Tracking",
    "Optimization & Automation": "Bid Strategy & Budgets",
    "Creative, Messaging, and Formats": "Ads & Creative",
    "Measurement": "Measurement & Tracking",
}
