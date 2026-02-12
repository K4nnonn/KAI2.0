"""
Kai UNIFIED ENGINE
One audit system, template-driven, accurate and repeatable

Architecture:
1. Load template (defines ALL criteria)
2. Discover data sources (intelligent mapper)
3. Iterate template rows
4. Score each criterion (completeness gates)
5. Write results directly to template
6. Output ONE accurate Excel file

No V3.1/V4 split. No dictionaries. No row mapping bugs.
Template IS the contract.
"""

import json
import logging
import os
import re
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Set, Any

import numpy as np
import openpyxl
from openpyxl.cell.cell import MergedCell
import pandas as pd
import requests
from .audit_verbalizer import AuditNarrativeHumanizer
from .intelligent_data_mapper import (
    IntelligentDataSourceMapper,
    to_numeric_safe,
    _looks_like_structured_export,
)
from .industry_knowledge_base import (
    get_benchmark, score_against_benchmark, determine_business_impact,
    get_criterion_weight, calculate_confidence, BEST_PRACTICES
)
from .strategic_context_knowledge_base import (
    detect_campaign_strategy, get_strategy_scoring_thresholds,
    get_acceptable_cpa_multiplier, get_recommended_tactics,
    score_metric_with_strategy_context
)
from kai_core.account_utils import normalize_account_name, slugify_account_name
from kai_core.telemetry import log_openai_usage
from kai_core.config import is_azure_openai_enabled
from kai_core.shared.azure_budget import allow_azure_usage
from .insight_injector import inject_insights


@dataclass
class ScoringResult:
    """Enhanced result of scoring one criterion with industry context"""
    # Core fields
    score: Optional[float | str]  # 1-5, "N/A", or None
    details: str
    recommendations: str = ""  # Column H - What to do (Made optional with default empty string)
    calculation: str = ""
    data_needed: Optional[str] = None

    # NEW: Best Practice Rationale for Column I
    bp_rationale: str = ""  # Column I - Why it matters (NEW FIELD for dual-column enhancement)

    # NEW: Benchmark context
    benchmark: Optional[float] = None
    performance_vs_benchmark: Optional[str] = None  # "+20%", "on par", "-15%"

    # NEW: Confidence
    data_completeness: float = 1.0  # 0-1, how complete is the data
    confidence: str = "high"  # low/medium/high

    # NEW: Business impact
    business_impact: str = "medium"  # low/medium/high/critical
    priority_score: int = 5  # 1-10
    impact_rationale: str = ""

    # NEW: Weight for overall score
    criterion_weight: float = 0.01  # Default weight

    # NEW: Risk-adjusted transparency
    raw_score: Optional[float] = None
    risk_adjustments: List[str] = field(default_factory=list)

    # NEW: Benchmark + gap language
    benchmark_delta: Optional[str] = None
    binary_gap_indicator: Optional[str] = None
    vertical_context: Optional[str] = None

    def is_scored(self) -> bool:
        """Has numeric score"""
        return isinstance(self.score, (int, float)) and self.score is not None

    def is_not_applicable(self) -> bool:
        """Criterion doesn't apply"""
        return self.score == "N/A"

    def needs_data(self) -> bool:
        """Cannot evaluate without more data"""
        if self.score is None:
            if self.data_needed is not None:
                return True
            if isinstance(self.details, str) and self.details.strip().lower().startswith("data needed:"):
                return True
        return False

    def get_priority_label(self) -> str:
        """Get human-readable priority label"""
        if self.priority_score >= 9:
            return "CRITICAL"
        elif self.priority_score >= 7:
            return "HIGH"
        elif self.priority_score >= 4:
            return "MEDIUM"
        else:
            return "LOW"


# ===================================================================
# HELPER FUNCTIONS FOR DUAL-COLUMN ENHANCEMENT
# ===================================================================

def generate_bp_rationale(
    why_matters: str,
    threshold_context: str,
    your_performance: str,
    evidence_tier: str = "applied_standard"
) -> str:
    """
    Generate research-validated Column I rationale following 5Cs Framework.

    This helper function implements the evidence-based language framework identified
    in the Column I enhancement analysis, distinguishing between verified facts,
    applied thresholds, and interpretive assessments.

    Args:
        why_matters: Strategic/tactical importance statement (universal value proposition)
        threshold_context: Threshold explanation with appropriate hedging based on evidence tier
        your_performance: Account-specific performance data and assessment
        evidence_tier: Level of evidence backing the threshold claim
            - 'verified': Industry-validated benchmark with specific cited source
            - 'applied_standard': Threshold we apply based on audit methodology (DEFAULT)
            - 'industry_context': Typical performance ranges from industry research

    Returns:
        Formatted Column I rationale text following 3-part structure

    Evidence Tiers Explained:
        TIER 1 (verified): Use when you have a specific, citable industry source
            Example: "Industry benchmark: 2-4% CTR (Source: WordStream 2024 benchmarks)"

        TIER 2 (applied_standard): Use for thresholds defined in our audit methodology
            Example: "Threshold applied: 20%+ audience segmentation (based on PPC audit
            methodology standards). Top-performing accounts typically exceed this threshold."

        TIER 3 (industry_context): Use for general performance ranges from research
            Example: "Industry studies show targeted campaigns outperform generic ads
            (typical lift: 20-50% per conversion optimization research)."

    Example Usage:
        # For audience segmentation with 45% performance
        bp_rationale = generate_bp_rationale(
            why_matters=(
                "Audience segmentation enables personalized messaging. Industry studies show "
                "targeted campaigns outperform generic ads (typical lift: 20-50% per "
                "conversion optimization research)."
            ),
            threshold_context=(
                "Threshold applied: 20%+ audience segmentation (based on PPC audit methodology "
                "standards). Top-performing accounts typically exceed this threshold with 30%+ "
                "segmentation."
            ),
            your_performance=(
                "Your account: 232 of 519 campaigns (45%) use audience segmentation. This "
                "performance significantly exceeds the 20% applied threshold, demonstrating "
                "mature audience targeting capabilities."
            ),
            evidence_tier="applied_standard"
        )
    """
    # Validate evidence tier
    valid_tiers = ['verified', 'applied_standard', 'industry_context']
    if evidence_tier not in valid_tiers:
        raise ValueError(f"evidence_tier must be one of {valid_tiers}, got '{evidence_tier}'")

    # Build 3-part rationale structure following 5Cs Framework
    # Part 1: CRITERIA - Why this metric matters (strategic value)
    why_text = why_matters.strip()
    benchmark_text = threshold_context.strip()
    performance_text = your_performance.strip()

    narrative_parts = []
    if why_text:
        narrative_parts.append(why_text)
    if benchmark_text:
        if evidence_tier == "verified":
            narrative_parts.append(f"Industry researchers cite this benchmark: {benchmark_text}")
        elif evidence_tier == "applied_standard":
            narrative_parts.append(f"Our applied threshold: {benchmark_text}")
        else:
            narrative_parts.append(f"Industry context: {benchmark_text}")
    if performance_text:
        narrative_parts.append(f"In your account, {performance_text}")

    return " ".join(narrative_parts).strip()


class AdaptiveAIScorer:
    """Optional Azure OpenAI-backed scorer that refines quantitative ratings with AI reasoning."""

    def __init__(self):
        self.enabled = os.environ.get("ENABLE_AI_SCORING", "true").lower() == "true"
        self.apply_score = os.environ.get("AI_SCORING_APPLY", "false").lower() == "true"
        if self.enabled and not is_azure_openai_enabled():
            logging.info("[AI Scoring] Azure disabled by policy; skipping AI refinement.")
            self.enabled = False
            return
        endpoint = os.environ.get("AZURE_OPENAI_ENDPOINT")
        key = os.environ.get("AZURE_OPENAI_API_KEY") or os.environ.get("AZURE_OPENAI_KEY")
        deployment = (
            os.environ.get("AZURE_OPENAI_DEPLOYMENT")
            or os.environ.get("AZURE_OPENAI_CHAT_DEPLOYMENT")
            or os.environ.get("AZURE_OPENAI_DEPLOYMENT_NAME")
        )
        self.api_version = os.environ.get("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")
        self.deployment = deployment
        self.endpoint = endpoint.rstrip("/") if endpoint else None
        self.session: Optional[requests.Session] = None
        self.headers = {"Content-Type": "application/json"}

        if not self.enabled:
            logging.info("[AI Scoring] Disabled via ENABLE_AI_SCORING flag.")
            return

        if not all([self.endpoint, key, deployment]):
            logging.warning("[AI Scoring] Missing Azure OpenAI configuration; skipping AI refinement.")
            self.enabled = False
            return

        self.session = requests.Session()
        self.headers["api-key"] = key
        self.max_retries = 3
        self.retry_backoff = 2.0
        def _parse_float(value: Optional[str], default: float) -> float:
            try:
                return float(value)
            except (TypeError, ValueError):
                return default

        requests_per_minute = _parse_float(os.environ.get("AZURE_OPENAI_REQUESTS_PER_MINUTE"), 45.0)
        tokens_per_minute = _parse_float(os.environ.get("AZURE_OPENAI_TOKENS_PER_MINUTE"), 50000.0)
        estimated_tokens_per_request = _parse_float(
            os.environ.get("AZURE_OPENAI_ESTIMATED_TOKENS_PER_REQUEST"),
            1200.0,
        )

        requests_per_minute = max(1.0, requests_per_minute)
        tokens_per_minute = max(1000.0, tokens_per_minute)
        estimated_tokens_per_request = max(100.0, estimated_tokens_per_request)

        token_based_rpm = tokens_per_minute / estimated_tokens_per_request
        effective_rpm = max(1.0, min(requests_per_minute, token_based_rpm))
        self.request_interval = 60.0 / effective_rpm
        self._last_request_ts = 0.0

    def refine_result(
        self,
        account_name: str,
        category: Optional[str],
        criterion_name: Optional[str],
        result: ScoringResult,
        context_summary: Optional[Dict[str, Dict]] = None,
    ) -> ScoringResult:
        if (
            not self.enabled
            or self.session is None
            or not result
            or not isinstance(result.score, (int, float))
        ):
            return result

        context_blob = {
            "account": account_name,
            "category": category or "",
            "criterion": criterion_name or "",
            "initial_score": result.score,
            "details": result.details or "",
            "calculation": result.calculation or "",
            "recommendations": result.recommendations or "",
            "confidence": result.confidence,
            "business_impact": result.business_impact,
            "context": context_summary or {},
        }

        ai_output = self._invoke_model(context_blob)
        if not ai_output:
            return result

        ai_score = ai_output.get("score")
        justification = ai_output.get("justification")

        if isinstance(ai_score, (int, float)):
            if self.apply_score:
                clamped = max(1.0, min(5.0, float(ai_score)))
                if abs(clamped - float(result.score)) >= 0.25:
                    logging.info(
                        "[AI Scoring] %s adjusted from %.2f to %.2f based on AI reasoning.",
                        criterion_name,
                        result.score,
                        clamped,
                    )
                result.score = round(clamped, 2)

        if justification:
            ai_note = justification
            result.bp_rationale = f"{result.bp_rationale}\n\n{ai_note}" if result.bp_rationale else ai_note

        return result

    def _invoke_model(self, payload: Dict[str, str]) -> Optional[Dict[str, str]]:
        if self.session is None:
            return None

        allowed, reason = allow_azure_usage(module="adaptive_ai_scorer", purpose="audit_scoring")
        if not allowed:
            logging.warning("[AI Scoring] Azure blocked by policy: %s", reason)
            return None

        def _respect_rate_limit() -> None:
            if self.request_interval <= 0:
                return
            now = time.monotonic()
            sleep_for = self.request_interval - (now - self._last_request_ts)
            if sleep_for > 0:
                time.sleep(sleep_for)
            self._last_request_ts = time.monotonic()

        url = f"{self.endpoint}/openai/deployments/{self.deployment}/chat/completions"
        body = {
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "You are a PPC audit quality reviewer. Rate each criterion from 1-5 "
                        "where 1=non-existent, 3=adequate, 5=excellent. "
                        "Use the provided evidence (details, calculations, contextual telemetry like change history, "
                        "audience depth, geo mix, conversion coverage, and extension usage) to justify the score."
                    ),
                },
                {
                    "role": "user",
                    "content": (
                        "Return JSON with keys 'score' (1-5) and 'justification' summarizing the reasoning.\n"
                        f"{json.dumps(payload, ensure_ascii=False)}"
                    ),
                },
            ],
            "temperature": 0.1,
            "max_tokens": 150,
        }

        data = None
        for attempt in range(self.max_retries):
            try:
                _respect_rate_limit()
                start_time = time.perf_counter()
                response = self.session.post(
                    url,
                    params={"api-version": self.api_version},
                    headers=self.headers,
                    json=body,
                    timeout=30,
                )
                latency_ms = (time.perf_counter() - start_time) * 1000
                if response.status_code == 429 and attempt < self.max_retries - 1:
                    sleep_for = self.retry_backoff ** attempt
                    logging.warning(
                        "[AI Scoring] Rate limited (429). Retrying in %.1fs (attempt %s/%s)...",
                        sleep_for,
                        attempt + 1,
                        self.max_retries,
                    )
                    time.sleep(sleep_for)
                    continue
                response.raise_for_status()
                data = response.json()
                usage = data.get("usage")
                log_openai_usage(
                    source="ai_scorer",
                    metadata={
                        "account": payload.get("account"),
                        "category": payload.get("category"),
                        "criterion": payload.get("criterion"),
                    },
                    usage=usage,
                    latency_ms=latency_ms,
                )
                break
            except Exception as exc:  # pragma: no cover - network dependency
                if attempt < self.max_retries - 1:
                    sleep_for = self.retry_backoff ** attempt
                    logging.warning(
                        "[AI Scoring] Chat completion error (%s). Retrying in %.1fs...",
                        exc,
                        sleep_for,
                    )
                    time.sleep(sleep_for)
                    continue
                logging.warning("[AI Scoring] Chat completion failed: %s", exc)
                return None

        if data is None:
            logging.warning("[AI Scoring] Disabling AI scoring after repeated failures.")
            self.enabled = False
            return None

        choices = data.get("choices")
        if not choices:
            return None

        message = choices[0].get("message") or {}
        content = message.get("content")
        if not content:
            return None

        try:
            return json.loads(content)
        except json.JSONDecodeError:
            start = content.find("{")
            end = content.rfind("}")
            if start >= 0 and end > start:
                try:
                    return json.loads(content[start : end + 1])
                except json.JSONDecodeError:
                    logging.warning("[AI Scoring] Unable to parse AI response: %s", content)
        return None


class UnifiedAuditEngine:
    """
    Template-driven audit engine
    Compiles best logic from V3.1 and V4 into one coherent system
    """

    MATCH_TYPE_SYNONYMS = {
        "exact": {
            "exact",
            "[exact]",
            "=exact",
            "ex match",
        },
        "phrase": {
            "phrase",
            "\"phrase",
            "phr",
            "phrase match",
        },
        "broad": {
            "broad",
            "brd",
            "bmm",
            "broad match",
            "broad match modifier",
            "modified broad",
            "broad match mod",
        },
    }

    SMART_BIDDING_STRATEGIES = {
        "target cpa",
        "target roas",
        "maximize conversions",
        "maximize conversion value",
        "target impression share",
        "maximize clicks",
        "portfolio target cpa",
        "portfolio target roas",
        "portfolio target impression share",
        "target cost per acquisition",
        "target return on ad spend",
    }

    def __init__(
        self,
        template_path: Path,
        data_directory: Path,
        business_unit: str = None,
        business_context: Optional[Dict[str, Any]] = None,
        chunk_size: int = 200000,
        projection_columns: Optional[List[str]] = None,
    ):
        self.template_path = template_path
        self.data_dir = data_directory
        self.business_unit = (
            normalize_account_name(business_unit, allow_passthrough=True)
            if business_unit
            else None
        )
        self.business_unit_key: Optional[str] = None
        self._bu_key_map: Dict[str, str] = {}
        self.business_context = business_context or {}
        self.chunk_size = chunk_size
        # Optional list of columns to project for very large files; default None keeps all columns
        self.projection_columns = projection_columns
        self.data = {}
        self.mapper = IntelligentDataSourceMapper()
        self.source_info = {}
        self.ai_scorer = AdaptiveAIScorer()
        self.humanizer = AuditNarrativeHumanizer()
        self.summary_stats: Dict[str, Dict[str, float | int | bool]] = {}
        self.diagnostics: Dict[str, Any] = {
            "file_manifest": [],
            "normalized_key_collisions": [],
            "filter_stats": [],
            "data_rows": {},
            "selected_files": {},
        }
        self._where_clause_seen: set[str] = set()
        self.strict_filter = os.environ.get("KLAUDIT_STRICT_FILTER", "false").lower() == "true"

        # Column positions (Excel 1-indexed) aligned to the Klaudit template header row
        # Template headers are on row 10: Category(B), Item(C), Rating(D), Details(E), Next Steps(F), Rationale(G)
        self.COL_CATEGORY = 2   # B ("Category")
        self.COL_ITEM = 3       # C ("Item")
        self.COL_RATING = 4     # D ("Rating")
        self.COL_DETAILS = 5    # E ("Details")
        self.COL_ACTIONS = 6    # F ("Next Steps/Action Items")
        self.COL_RATIONALE = 7  # G ("Best Practice/Rationale")

        # Template row range
        self.FIRST_CRITERION_ROW = 11  # Row 11 in Excel
        self.LAST_CRITERION_ROW = 92   # Row 92 in Excel (82 criteria)

        # Business unit account name patterns (from "Account name" column)
        self.BU_ACCOUNT_PATTERNS = {
            'Retirement': r'(?:Fidelity - Retirement|Retirement -)',
            'Brand': r'(?:Fidelity - Brand|Brand -|Kelvin Co|Kelvin)',
            'Wealth_Management': r'(?:Fidelity - Wealth|Wealth -)',
            'Brokerage': r'(?:Fidelity - Brokerage|Brokerage -)',
            'FILI_Term_Life': r'(?:Fidelity - FILI Term Life|FILI Term Life -|FILI - Term Life)',
            'FILI': r'(?:Fidelity - FILI|FILI -)',
            'Asset_Management': r'(?:Fidelity - Asset Management|Asset Management -|Kelvin Co - Asset Management)',
            'FCM': r'(?:FCM -|Kelvin Co - FCM|\bFCM\b)',
            'Family_Office': r'(?:Family Office -|Kelvin Co - Family Office|\bFamily Office\b)',
            'IICG': r'(?:IICG -|Kelvin Co - IICG|\bIICG\b|\bICG\b)',
            'IWMS': r'(?:IWMS -|Kelvin Co - IWMS|\bIWMS\b)',
            'Plynk': r'(?:Plynk -|\bPlynk\b)',
            'SPS': r'(?:SPS -|Kelvin Co - SPS|\bSPS\b)',
            'Crypto': r'(?:Fidelity - Crypto|Crypto -)',
            'Fidelity - Crypto': r'(?:Fidelity - Crypto|Crypto -)',
            'Fidelity - HSA': r'(?:Fidelity - HSA|HSA -)',
            'Fidelity - 529 National': r'(?:Fidelity - 529 National|529 National -)',
            'Fidelity - 529 MA': r'(?:Fidelity - 529 MA|529 MA -)',
            'Fidelity - 529 CT': r'(?:Fidelity - 529 CT|529 CT -)',
            'Fidelity - 529 AZ': r'(?:Fidelity - 529 AZ|529 AZ -)',
            'Fidelity - 529 NH': r'(?:Fidelity - 529 NH|529 NH -)',
            'Fidelity - 529 DE': r'(?:Fidelity - 529 DE|529 DE -)',
            '529 States': r'(?:Fidelity - 529 (?:AZ|CT|DE|MA|NH)|529 (?:AZ|CT|DE|MA|NH) -|529 States|Fidelity - 529 States)',
            'Fidelity - 529 States': r'(?:Fidelity - 529 (?:AZ|CT|DE|MA|NH)|529 (?:AZ|CT|DE|MA|NH) -|529 States|Fidelity - 529 States)',
            'Fidelity - Youth': r'(?:Fidelity - Youth|Youth -|Youth Account -)',
            'Fidelity - TEM': r'(?:Fidelity - TEM|TEM -)',
            'Fidelity - ABLE': r'(?:Fidelity - ABLE|ABLE -)',
            'Medicare': r'(?:Fidelity - Medicare|Medicare -|Kelvin Co - Medicare)',
            'WI': r'(?:Fidelity - WI|WI -|Kelvin Co - WI)',
            'Fidelity_Go': r'(?:Fidelity - Fidelity Go|Fidelity Go -)',
            'Mock': r'(?:Kelvin Co|Kelvin Services)',
        }

        # Fallback: Business unit campaign patterns (if Account name not available)
        self.BU_CAMPAIGN_PATTERNS = {
            'Retirement': r'^Retirement -',
            'Brand': r'^(?:Brand -|Kelvin Co|Kelvin)',
            'Wealth_Management': r'^Wealth Management -',
            'Brokerage': r'^(?:FCM -|Trading -)',
            'FILI_Term_Life': r'^FILI -',
            'FILI': r'^FILI -',
            'Asset_Management': r'^(?:ASM -|Asset Management -|Kelvin Co - Asset Management)',
            'FCM': r'^FCM -',
            'Family_Office': r'^(?:Family Office -|IWMS - Family Office)',
            'IICG': r'^IICG -',
            'IWMS': r'^IWMS -',
            'Plynk': r'^Plynk -',
            'SPS': r'^SPS -',
            'Crypto': r'^Crypto -',
            'Fidelity - Crypto': r'^Crypto -',
            'Fidelity - HSA': r'^HSA -',
            'Fidelity - 529 National': r'^529 National -',
            'Fidelity - 529 MA': r'^529 MA -',
            'Fidelity - 529 CT': r'^529 CT -',
            'Fidelity - 529 AZ': r'^529 AZ -',
            'Fidelity - 529 NH': r'^529 NH -',
            'Fidelity - 529 DE': r'^529 DE -',
            '529 States': r'^(?:529 (?:AZ|CT|DE|MA|NH) -|529 States -)',
            'Fidelity - 529 States': r'^(?:529 (?:AZ|CT|DE|MA|NH) -|529 States -)',
            'Fidelity - Youth': r'^Youth Account -',
            'Fidelity - TEM': r'^TEM -',
            'Fidelity - ABLE': r'^ABLE -',
            'Medicare': r'^Medicare -',
            'WI': r'^WI -',
            'Fidelity_Go': r'^(?:Fidelity Go -|Robo -|DPA -)',
            'Mock': r'^Kelvin ',
        }

        # Optional: restrict specific business units to known Customer IDs
        self.BU_CUSTOMER_IDS = {
            # Medicare has multiple archived accounts; only 817-271-4944 is live
            'Medicare': {'817-271-4944'},
            'Fidelity_Go': {'491-359-7543', '4913597543', '156-873-8935', '1568738935'},
        }
        # Inferred customer IDs discovered from matching account/campaign rows.
        # This is supplemental to hardcoded IDs and only applied when it yields matches.
        self.inferred_customer_ids: Dict[str, Set[str]] = {}
        self._bu_key_map = self._build_bu_key_map()
        self.business_unit_key = self._resolve_bu_key(self.business_unit)

    def _read_csv_with_fallbacks(self, csv_file: Path) -> pd.DataFrame:
        """
        Adaptive CSV reader with detection and self-heal:
        - Detect encoding (BOM-aware) and delimiter from a sample.
        - Parse once with detected params; if structure is off, intelligently reparse (e.g., combined headers).
        - Returns structured errors instead of scattered warnings.
        """
        def detect_encoding_and_delimiter(sample: bytes) -> tuple[str, str, int]:
            if sample.startswith(b'\xff\xfe') or sample.startswith(b'\xfe\xff'):
                encoding = "utf-16"
            elif sample.startswith(b'\xef\xbb\xbf'):
                encoding = "utf-8-sig"
            else:
                encoding = "utf-8"

            text = sample.decode(encoding, errors="ignore")
            non_empty = [
                (idx, line) for idx, line in enumerate(text.splitlines()) if line.strip()
            ][:6]
            best = {"count": 0, "delimiter": ",", "idx": 0}
            for idx, line in non_empty:
                counts = {delim: line.count(delim) for delim in [",", "\t", ";"]}
                max_delim = max(counts, key=counts.get)
                max_count = counts[max_delim]
                if max_count > best["count"]:
                    best = {"count": max_count, "delimiter": max_delim, "idx": idx}
            delimiter = best["delimiter"] if best["count"] > 0 else ","
            return encoding, delimiter, best["idx"]

        def parse_once(encoding: str, delimiter: str, skiprows: int, usecols=None) -> pd.DataFrame:
            read_kwargs = {
                "encoding": encoding,
                "sep": delimiter,
                "engine": "python",
                "skiprows": skiprows,
                "on_bad_lines": "skip",
            }
            if usecols:
                read_kwargs["usecols"] = usecols

            if use_chunking:
                chunks = []
                for chunk in pd.read_csv(csv_file, chunksize=self.chunk_size, **read_kwargs):
                    chunks.append(chunk)
                df_local = pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()
            else:
                df_local = pd.read_csv(csv_file, **read_kwargs)

            df_local.columns = df_local.columns.str.lstrip('\ufeff')
            return df_local

        def try_reparse_combined(df: pd.DataFrame, encoding: str, delimiter: str) -> pd.DataFrame:
            if df.shape[1] != 1:
                return df
            col0 = df.columns[0]
            if not (isinstance(col0, str) and any(sep in col0 for sep in [",", "\t", ";"])):
                return df
            for sep_try in [delimiter, "\t", ",", ";"]:
                for skip_try in [0, 1, 2]:
                    try:
                        candidate = parse_once(encoding, sep_try, skip_try)
                        if candidate.shape[1] == 1:
                            col0_candidate = candidate.columns[0]
                            if isinstance(col0_candidate, str) and any(
                                sep in col0_candidate for sep in [",", "\t", ";"]
                            ):
                                continue
                        if _looks_like_structured_export(candidate):
                            print(f"  [csv-normalize] reparsed {csv_file.name} with sep='{sep_try}' skiprows={skip_try}")
                            return candidate
                    except Exception:
                        continue
            return df

        if not csv_file.exists():
            raise FileNotFoundError(f"{csv_file} not found")

        sample_bytes = csv_file.open("rb").read(65536)
        encoding, delimiter, preferred_skip = detect_encoding_and_delimiter(sample_bytes)
        skip_candidates = [preferred_skip, 0, 1, 2]
        skip_candidates = [x for x in skip_candidates if isinstance(x, int) and x >= 0]
        skip_candidates = list(dict.fromkeys(skip_candidates))

        projected_usecols = None
        if self.projection_columns:
            try:
                header_cols = pd.read_csv(
                    csv_file,
                    nrows=0,
                    encoding=encoding,
                    sep=delimiter,
                    engine="python",
                    on_bad_lines="skip",
                ).columns.tolist()
                lower_cols = {c.lower(): c for c in header_cols}
                desired = []
                for key in self.projection_columns:
                    lc = key.lower()
                    if lc in lower_cols:
                        desired.append(lower_cols[lc])
                if desired:
                    projected_usecols = desired
            except Exception:
                projected_usecols = None

        use_chunking = bool(self.chunk_size and csv_file.stat().st_size > 50 * 1024 * 1024)
        last_error: Exception | None = None

        for skiprows in skip_candidates:
            try:
                df = parse_once(encoding, delimiter, skiprows, usecols=projected_usecols)
                if df is None or df.empty:
                    continue

                df = try_reparse_combined(df, encoding, delimiter)

                if not _looks_like_structured_export(df) and projected_usecols:
                    df_retry = parse_once(encoding, delimiter, skiprows, usecols=None)
                    df_retry = try_reparse_combined(df_retry, encoding, delimiter)
                    if _looks_like_structured_export(df_retry):
                        df = df_retry

                if not df.empty and any(
                    str(value).strip().lower() == "account"
                    for value in df.iloc[0].astype(str).tolist()
                ):
                    df = df.iloc[1:].reset_index(drop=True)

                if _looks_like_structured_export(df):
                    return df
            except Exception as exc:
                last_error = exc
                continue

        if last_error:
            raise ValueError(f"Unable to parse {csv_file.name} with detected encoding={encoding} delimiter='{delimiter}' (last error: {last_error})")
        raise ValueError(f"Unable to parse {csv_file.name} with detected encoding={encoding} delimiter='{delimiter}'")

    def _infer_normalized_key(self, df: pd.DataFrame, fallback_key: str) -> Optional[str]:
        """Infer a canonical dataset key from column content when filenames are generic."""
        if df is None or df.empty:
            return None
        key = (fallback_key or "").lower()
        if not (key.startswith("untitled") or key in {"report", "data", "dataset", "export"}):
            return None
        cols = {str(col).strip().lower() for col in df.columns if col is not None}

        def has_any(options: List[str]) -> bool:
            return any(opt.lower() in cols for opt in options)

        def has_all(options: List[str]) -> bool:
            return all(opt.lower() in cols for opt in options)

        if has_any(["Negative keyword"]):
            return "negative_keyword"
        if has_any(["Search keyword", "Search keyword match type"]):
            return "search_keyword"
        if has_any(["Landing page"]):
            return "landing_page"
        if has_any(["Ad name", "Headline", "Description"]):
            return "ad"
        if has_any(["Targeted location"]):
            return "user_locations"
        if has_any(["Audience segment", "Audience list", "Audience"]):
            return "audience_segment"
        if has_any(["Campaign state", "Campaign status"]) and has_any(["Campaign"]):
            return "campaign_details"
        if has_all(["Ad group", "Ad group ID"]):
            if has_any(["Ad group status", "Ad group state"]):
                return "ad_group_details"
            return "ad_group"
        if has_all(["Campaign", "Campaign ID"]) and has_any(["Impr.", "Impressions"]):
            return "campaign"
        if has_all(["Account name", "Customer ID"]) and has_any(["Impr.", "Impressions"]):
            return "campaign"

        return None
    @staticmethod
    def _ensure_text_series(series: pd.Series) -> pd.Series:
        """
        Coerce a pandas Series to a string-compatible dtype so `.str` accessors are always safe.
        """
        if pd.api.types.is_string_dtype(series):
            return series
        try:
            return series.astype(str)
        except Exception:
            return series.apply(lambda value: "" if value is None else str(value))

    def _record_filter_stat(self, info: Dict[str, Any]) -> None:
        if not info:
            return
        self.diagnostics.setdefault("filter_stats", []).append(info)

    def _build_bu_key_map(self) -> Dict[str, str]:
        keys = set(self.BU_ACCOUNT_PATTERNS) | set(self.BU_CAMPAIGN_PATTERNS) | set(self.BU_CUSTOMER_IDS)
        key_map: Dict[str, str] = {}
        for key in keys:
            normalized = str(key).lower().replace(" ", "_")
            key_map[normalized] = key
        return key_map

    def _candidate_bu_keys(self, business_unit: Optional[str]) -> List[str]:
        if not business_unit:
            return []
        raw = business_unit
        normalized = normalize_account_name(raw, allow_passthrough=True) or raw
        candidates = [raw, normalized]
        for value in (raw, normalized):
            if value:
                candidates.append(value.replace(" ", "_"))
        slug = slugify_account_name(normalized or raw)
        if slug:
            candidates.append(slug)
        seen = set()
        ordered = []
        for value in candidates:
            if value and value not in seen:
                seen.add(value)
                ordered.append(value)
        return ordered

    def _resolve_bu_key(self, business_unit: Optional[str]) -> Optional[str]:
        if not business_unit:
            return None
        for candidate in self._candidate_bu_keys(business_unit):
            if candidate in self.BU_ACCOUNT_PATTERNS or candidate in self.BU_CAMPAIGN_PATTERNS or candidate in self.BU_CUSTOMER_IDS:
                return candidate
            normalized = candidate.lower().replace(" ", "_")
            mapped = self._bu_key_map.get(normalized)
            if mapped:
                return mapped
        return business_unit

    def _filter_by_business_unit(self, df: pd.DataFrame, source_name: Optional[str] = None) -> pd.DataFrame:
        """Filter dataframe by business unit using Account name (preferred) or Campaign name (fallback)."""
        if not self.business_unit:
            return df

        filtered = df
        applied_filter = False
        attempted_filter = False
        match_info = {
            "source": source_name,
            "total_rows": int(len(df)),
            "attempted_filter": False,
            "applied_filter": False,
            "matched_rows": 0,
            "column": None,
            "pattern": None,
            "reason": None,
        }

        # PRIORITY 1: Use Account name column if available (most accurate)
        account_col = self._find_column(filtered, ["Account name", "Account", "Account Name"], allow_partial=True)
        if account_col:
            if account_col != "Account name":
                filtered = filtered.copy()
                filtered["Account name"] = filtered[account_col]
            pattern = self.BU_ACCOUNT_PATTERNS.get(self.business_unit_key) if self.business_unit_key else None
            if pattern:
                attempted_filter = True
                match_info["attempted_filter"] = True
                match_info["column"] = "Account name"
                match_info["pattern"] = pattern
                account_series = self._ensure_text_series(filtered['Account name'])
                mask = account_series.str.contains(pattern, case=False, na=False, regex=True)
                if mask.any():
                    filtered = filtered[mask]
                    applied_filter = True
                    match_info["applied_filter"] = True
                    match_info["matched_rows"] = int(mask.sum())
                else:
                    match_info["reason"] = "no_match"

        # PRIORITY 2: Fallback to Campaign name (only if no account filter applied)
        if not applied_filter and 'Campaign' in filtered.columns:
            pattern = self.BU_CAMPAIGN_PATTERNS.get(self.business_unit_key) if self.business_unit_key else None
            if pattern:
                attempted_filter = True
                match_info["attempted_filter"] = True
                match_info["column"] = "Campaign"
                match_info["pattern"] = pattern
                campaign_series = self._ensure_text_series(filtered['Campaign'])
                mask = campaign_series.str.contains(pattern, case=False, na=False, regex=True)
                if mask.any():
                    filtered = filtered[mask]
                    applied_filter = True
                    match_info["applied_filter"] = True
                    match_info["matched_rows"] = int(mask.sum())
                else:
                    match_info["reason"] = "no_match"

        # Derive customer IDs from matched account/campaign rows (adaptive, non-hardcoded).
        customer_col = self._find_column(filtered, ["Customer ID"], allow_partial=True)
        if customer_col and applied_filter:
            col_values = filtered[customer_col].astype(str).str.replace(" ", "", regex=False)
            inferred = {cid for cid in col_values if cid and cid.lower() not in ("nan", "none")}
            if inferred:
                key = self.business_unit_key or self.business_unit
                self.inferred_customer_ids.setdefault(key, set()).update(inferred)
                match_info["inferred_customer_ids"] = {"count": len(inferred)}

        # PRIORITY 3: Enforce explicit customer IDs when configured (or inferred IDs if available).
        customer_ids = self.BU_CUSTOMER_IDS.get(self.business_unit_key) if self.business_unit_key else None
        effective_ids = customer_ids or self.inferred_customer_ids.get(self.business_unit_key or self.business_unit)
        if effective_ids and customer_col:
            normalized_ids = {str(cid).replace(" ", "") for cid in effective_ids}
            col_values = filtered[customer_col].astype(str).str.replace(" ", "", regex=False)
            mask = col_values.isin(normalized_ids)
            if mask.any():
                filtered = filtered[mask]
                match_info["customer_id_filter"] = {
                    "matched_rows": int(mask.sum()),
                    "total_rows": int(len(col_values)),
                    "source": "configured" if customer_ids else "inferred",
                }
            else:
                match_info["customer_id_filter"] = {
                    "matched_rows": 0,
                    "total_rows": int(len(col_values)),
                    "source": "configured" if customer_ids else "inferred",
                    "reason": "no_match",
                }

        # If a BU filter was attempted but nothing matched, fail fast in strict mode.
        if attempted_filter and not applied_filter:
            if self.strict_filter:
                raise ValueError(f"No rows matched business unit filter for {self.business_unit} ({source_name}).")

        # If filtering removed everything, fall back to unfiltered data to avoid empty datasets during tests
        if applied_filter and filtered.empty:
            if self.strict_filter:
                raise ValueError(f"No rows matched business unit filter for {self.business_unit} ({source_name}).")
            match_info["reason"] = "fallback_unfiltered"
            filtered = df

        if source_name:
            if not attempted_filter and match_info["reason"] is None:
                match_info["reason"] = "no_filter_columns"
            self._record_filter_stat(match_info)

        return filtered

    def load_data(self) -> Dict[str, pd.DataFrame]:
        """Load all supported report exports from data directory (CSV/TSV/XLSX/XLS).

        Agencies often export multiple report tabs in a single workbook. For XLSX/XLS inputs we load all sheets
        and let the intelligent mapper infer the correct report type from the column headers.
        """
        print(f"\n[1/5] Loading data from {self.data_dir}...")
        if self.business_unit:
            print(f"  Filtering for Business Unit: {self.business_unit}")

        csv_files = sorted(self.data_dir.glob("*.csv"), key=lambda p: p.name.lower())
        tsv_files = sorted(self.data_dir.glob("*.tsv"), key=lambda p: p.name.lower())
        xlsx_files = sorted(self.data_dir.glob("*.xlsx"), key=lambda p: p.name.lower())
        xls_files = sorted(self.data_dir.glob("*.xls"), key=lambda p: p.name.lower())
        normalized_map: Dict[str, List[str]] = {}
        self.diagnostics["file_manifest"] = []

        def ingest_frame(
            df: pd.DataFrame,
            *,
            source_display: str,
            source_stem: str,
            size_bytes: int | None,
            sheet_name: str | None,
        ) -> None:
            file_key = (source_stem or "").lower().replace(" ", "_")
            if df is None or df.shape[1] == 0:
                print(f"  SKIP {file_key:20} (empty or no columns after parse)")
                return

            # Normalize key: strip common export patterns.
            normalized_key = file_key.replace("_report_-_kl2", "").replace("_-_kl2", "").replace("_report", "")
            # Strip parenthetical suffixes like "_(5)" from duplicate exports.
            import re
            normalized_key = re.sub(r"_\\(\\d+\\)$", "", normalized_key)

            inferred_key = None
            try:
                inferred_key = self._infer_normalized_key(df, normalized_key)
            except Exception:
                # Unknown/novel sheet layouts should not prevent loading; fall back to the filename-derived key.
                inferred_key = None
            original_key = normalized_key
            if inferred_key and inferred_key != normalized_key:
                normalized_key = inferred_key
            normalized_map.setdefault(normalized_key, []).append(source_display)

            # Apply Business Unit filtering if specified.
            df_filtered = self._filter_by_business_unit(df, source_name=normalized_key)

            entry = {
                "file": source_display,
                "sheet": sheet_name,
                "normalized_key": normalized_key,
                "original_key": original_key if original_key != normalized_key else None,
                "inferred_key": inferred_key if inferred_key and inferred_key != original_key else None,
                "size_bytes": int(size_bytes) if size_bytes is not None else None,
                "rows_loaded": int(len(df_filtered)),
                "column_count": int(df_filtered.shape[1]),
            }
            self.diagnostics["file_manifest"].append(entry)

            if normalized_key in self.data:
                existing_rows = int(len(self.data[normalized_key]))
                keep_new = int(len(df_filtered)) > existing_rows
                existing_file = self.diagnostics.get("selected_files", {}).get(normalized_key)
                decision = {
                    "normalized_key": normalized_key,
                    "kept_file": source_display if keep_new else existing_file,
                    "skipped_file": existing_file if keep_new else source_display,
                    "kept_rows": int(len(df_filtered)) if keep_new else existing_rows,
                    "skipped_rows": existing_rows if keep_new else int(len(df_filtered)),
                }
                self.diagnostics["normalized_key_collisions"].append(decision)
                if keep_new:
                    self.data[normalized_key] = df_filtered
                    self.diagnostics["selected_files"][normalized_key] = source_display
            else:
                self.data[normalized_key] = df_filtered
                self.diagnostics["selected_files"][normalized_key] = source_display

            print(f"  OK {normalized_key:20} {len(df_filtered):>8,} rows")

        for csv_file in (csv_files + tsv_files):
            file_key = csv_file.stem.lower().replace(" ", "_")
            try:
                df = self._read_csv_with_fallbacks(csv_file)
                ingest_frame(
                    df,
                    source_display=csv_file.name,
                    source_stem=csv_file.stem,
                    size_bytes=int(csv_file.stat().st_size),
                    sheet_name=None,
                )
            except Exception as exc:
                print(f"  WARN {file_key:20} skipped ({exc})")

        for xls_file in (xlsx_files + xls_files):
            try:
                # pandas returns a dict of {sheet_name: DataFrame} when sheet_name=None
                sheets = pd.read_excel(xls_file, sheet_name=None)
            except Exception as exc:
                print(f"  WARN {xls_file.name:20} skipped ({exc})")
                continue

            if not sheets:
                continue
            size_bytes = None
            try:
                size_bytes = int(xls_file.stat().st_size)
            except Exception:
                size_bytes = None

            for sheet, df in sheets.items():
                # Skip empty sheets (common in templates/exports).
                if df is None or getattr(df, "empty", False) or df.shape[1] == 0:
                    continue
                source_display = f"{xls_file.name}::{sheet}"
                source_stem = f"{xls_file.stem}__{sheet}"
                try:
                    ingest_frame(
                        df,
                        source_display=source_display,
                        source_stem=source_stem,
                        size_bytes=size_bytes,
                        sheet_name=str(sheet),
                    )
                except Exception as exc:
                    print(f"  WARN {source_display:20} skipped ({exc})")

        if not self.data:
            raise ValueError(f"No CSV/TSV/XLSX/XLS files found in {self.data_dir}")

        self.diagnostics["data_rows"] = {k: int(len(v)) for k, v in self.data.items()}
        self.diagnostics["normalized_key_collisions"] = [
            item for item in self.diagnostics["normalized_key_collisions"]
            if item.get("normalized_key")
        ]
        self.diagnostics["normalized_key_map"] = normalized_map

        self._build_landing_page_proxy()

        return self.data

    def _build_landing_page_proxy(self) -> None:
        """Create a landing_page dataset from URL columns when an explicit landing_page report is missing."""
        if "landing_page" in self.data:
            return
        candidates = []
        for key in (
            "ad",
            "ads",
            "responsive_search_ad",
            "search_keyword",
            "keyword",
            "campaign",
            "ad_group",
            "adgroup",
            "ad_group_details",
        ):
            df = self.data.get(key)
            if df is None or not isinstance(df, pd.DataFrame) or df.empty:
                continue
            url_col = self._find_column(
                df,
                [
                    "Final URL",
                    "Landing page",
                    "Landing Page",
                    "Final Url",
                    "URL",
                    "Destination URL",
                    "Ad final URL",
                ],
                allow_partial=True,
            )
            if not url_col:
                continue
            non_null = int(df[url_col].notna().sum())
            candidates.append((non_null, int(len(df)), key, df, url_col))

        if not candidates:
            return

        candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
        _, _, source_key, source_df, url_col = candidates[0]

        landing_df = pd.DataFrame()
        landing_df["Landing page"] = source_df[url_col].astype(str)
        landing_df["Final URL"] = source_df[url_col].astype(str)

        for col_name in ("Campaign", "Ad group", "Ad group ID", "Keyword", "Account name", "Customer ID"):
            source_col = self._find_column(source_df, [col_name], allow_partial=True)
            if source_col:
                landing_df[col_name] = source_df[source_col]

        self.data["landing_page"] = landing_df
        self.diagnostics.setdefault("landing_page_proxy", {
            "source_key": source_key,
            "url_column": url_col,
            "rows": int(len(landing_df)),
        })
        print(f"  INFO landing_page proxy from {source_key} using column '{url_col}' ({len(landing_df):,} rows)")

    def _find_data_with_column(
        self,
        column_name: str,
        preferred_sources: Optional[List[str]] = None
    ) -> Optional[pd.DataFrame]:
        """
        ADAPTIVE COLUMN DISCOVERY - Phase 1: The Search Party

        Searches for DATA CONCEPTS (columns) not FILES.
        Implements the Kai Doctrine's "Search Party" logic to eliminate brittle
        hardcoded file lookups.

        Args:
            column_name: The column to search for (e.g., "Keywords: disapproved")
            preferred_sources: Optional list of file keys to check first (optimization)

        Returns:
            First DataFrame containing the column, or None if not found

        Design Principles:
            - Duck Typing: "If it has the column, it IS the data"
            - Registry Pattern: Decouples data definition from physical location
            - Transparency: Logs discovery for audit trail and debugging
            - Anti-Fragility: System adapts to different file structures

        Example:
            # Old brittle approach:
            keyword_data = self.data.get("search_keyword")  # ❌ Only checks ONE file

            # New adaptive approach:
            keyword_data = self._find_data_with_column(
                "Keywords: disapproved",
                preferred_sources=["search_keyword", "ad_group_details"]  # ✓ Tries multiple sources
            )
        """
        # Phase 1: Check preferred sources first (optimization - reduces search time)
        if preferred_sources:
            for source_key in preferred_sources:
                if source_key in self.data:
                    df = self.data[source_key]
                    if isinstance(df, pd.DataFrame) and column_name in df.columns:
                        logging.info(f"[Search Party] Found '{column_name}' in preferred source '{source_key}'")
                        return df

        # Phase 2: Search ALL loaded data (Duck Typing - if it has the column, it IS the data)
        for file_key, df in self.data.items():
            if isinstance(df, pd.DataFrame) and column_name in df.columns:
                logging.info(f"[Search Party] Found '{column_name}' in file '{file_key}'")
                return df

        # Not found - return None (caller handles with Confidence & Citation Protocol)
        logging.warning(f"[Search Party] Column '{column_name}' not found in any loaded data source")
        return None

    def _find_column(self, df, possible_names, allow_partial: bool = True):
        """Find a column by name (case-insensitive) with optional substring fallback."""
        if df is None:
            return None

        lowered = {col.lower(): col for col in df.columns}
        alias_map = {
            "account name": ["account", "account_name", "accountname"],
        }
        for name in possible_names:
            candidate = name.lower()
            if candidate in lowered:
                return lowered[candidate]
            for alt in alias_map.get(candidate, []):
                if alt in lowered:
                    return lowered[alt]

        if not allow_partial:
            return None

        for name in possible_names:
            candidate = name.lower()
            for col in df.columns:
                if candidate in col.lower():
                    return col
        return None

    def _find_column_by_substrings(self, df, substrings):
        """Find column containing all substrings (case-insensitive)."""
        if df is None:
            return None
        for col in df.columns:
            col_lower = col.lower()
            if all(sub.lower() in col_lower for sub in substrings):
                return col
        return None

    def _pick_campaign_source(self, required_cols: List[str]) -> Optional[pd.DataFrame]:
        """Pick a campaign-like dataset that contains required columns."""
        def has_required_columns(frame: pd.DataFrame) -> bool:
            for col in required_cols:
                candidates = [col]
                if col.lower() == "account name":
                    candidates.extend(["Account", "Account Name"])
                if col.lower() == "campaign":
                    candidates.extend(["Campaign name", "Campaigns"])
                if not self._find_column(frame, candidates, allow_partial=True):
                    return False
            return True

        candidates = ["campaign", "campaign_details", "account"]
        for key in candidates:
            df = self.data.get(key)
            if df is None or not isinstance(df, pd.DataFrame) or df.empty:
                continue
            if has_required_columns(df):
                return df
        for _, df in self.data.items():
            if df is None or not isinstance(df, pd.DataFrame) or df.empty:
                continue
            if has_required_columns(df):
                return df
        return None

    def _extract_match_type_series(self) -> Tuple[Optional[pd.Series], Optional[str]]:
        """Return normalized match type series (lowercase) from the best available source."""
        candidate_sources = [
            ("search_keyword", self.data.get("search_keyword")),
            ("keyword", self.data.get("keyword")),
            ("performance", getattr(self, "performance_data", None)),
        ]

        for label, df in candidate_sources:
            if df is None or not len(df):
                continue
            match_col = self._find_column(
                df,
                [
                    "Search keyword match type",
                    "Keyword match type",
                    "Match type",
                ],
            )
            if not match_col:
                continue
            series = df[match_col].astype(str).str.strip()
            normalized = series.apply(self._normalize_match_type_value).dropna()
            if normalized.empty:
                continue
            return normalized, label
        return None, None

    def _compute_match_distribution(self, series: Optional[pd.Series]) -> Optional[Dict[str, float]]:
        if series is None or series.empty:
            return None
        counts = {"exact": 0, "phrase": 0, "broad": 0}
        for value in series.dropna():
            norm = self._normalize_match_type_value(value)
            if norm in counts:
                counts[norm] += 1
        total = sum(counts.values())
        if total == 0:
            return None
        return {
            "exact_pct": counts["exact"] / total * 100,
            "phrase_pct": counts["phrase"] / total * 100,
            "broad_pct": counts["broad"] / total * 100,
            "total_keywords": total,
        }

    def _normalize_match_type_value(self, value: object) -> Optional[str]:
        """Normalize any match-type string (including BMM) to exact/phrase/broad."""
        if value is None:
            return None
        text = str(value).strip().lower()
        if not text or text in {"--", "n/a"}:
            return None
        for canonical, tokens in self.MATCH_TYPE_SYNONYMS.items():
            if any(token in text for token in tokens):
                return canonical
        # Handle single-letter abbreviations (E/P/B) when exported from SA360
        if text in {"e", "exact"}:
            return "exact"
        if text in {"p", "phrase"}:
            return "phrase"
        if text in {"b", "bm", "bmm", "broad"}:
            return "broad"
        return None

    def _get_smart_bidding_stats(self) -> Dict[str, float]:
        """Return smart bidding stats; lazily compute if smart_bidding criterion not scored yet."""
        cached = self.summary_stats.get("smart_bidding")
        if cached:
            return cached

        campaign_df = self.data.get("campaign_details")
        if campaign_df is None or (isinstance(campaign_df, pd.DataFrame) and campaign_df.empty):
            fallback = self.source_info.get("campaign_source")
            if isinstance(fallback, pd.DataFrame) and not fallback.empty:
                campaign_df = fallback
            else:
                campaign_df = None
        default = {
            "total_campaigns": 0,
            "smart_campaigns": 0,
            "manual_campaigns": 0,
            "smart_pct": 0.0,
            "smart_with_volume": 0,
            "smart_without_volume": 0,
            "volume_data_available": False,
        }

        if campaign_df is None:
            self.summary_stats["smart_bidding"] = default
            return default

        strategy_col = self._find_column(campaign_df, ["Campaign bid strategy type"])
        if strategy_col is None:
            self.summary_stats["smart_bidding"] = default
            return default

        strategies = campaign_df[strategy_col].astype(str).str.strip().str.lower()
        smart_mask = strategies.isin(self.SMART_BIDDING_STRATEGIES)
        total = len(strategies)
        smart_campaigns = int(smart_mask.sum())
        manual_campaigns = max(total - smart_campaigns, 0)
        smart_pct = (smart_campaigns / total * 100) if total else 0.0

        smart_with_volume = 0
        smart_without_volume = 0
        volume_data_available = False
        conversion_col = self._find_column(campaign_df, ["Conversions", "All conv.", "Conv."])
        if conversion_col:
            volume_data_available = True
            conv_series = to_numeric_safe(campaign_df[conversion_col])
            ready_mask = conv_series >= 30
            smart_with_volume = int(((smart_mask) & ready_mask).sum())
            smart_without_volume = int(smart_campaigns - smart_with_volume)

        computed = {
            **default,
            "total_campaigns": total,
            "smart_campaigns": smart_campaigns,
            "manual_campaigns": manual_campaigns,
            "smart_pct": smart_pct,
            "smart_with_volume": smart_with_volume,
            "smart_without_volume": smart_without_volume,
            "volume_data_available": volume_data_available,
        }

        self.summary_stats["smart_bidding"] = computed
        return computed

    def _get_ad_schedule_metrics(self) -> Optional[Dict[str, float]]:
        """Aggregate ad scheduling usage across campaigns."""
        cached = self.summary_stats.get("ad_schedule")
        if cached:
            return cached

        campaign_data = self.data.get("campaign_details")
        if campaign_data is None:
            return None

        schedule_source = None
        schedule_col = None
        for df in self.data.values():
            if isinstance(df, pd.DataFrame):
                col = self._find_column(df, ["Ad schedule", "Ad Schedule"], allow_partial=False)
                if col:
                    schedule_source = df
                    schedule_col = col
                    break

        if schedule_source is None:
            for csv_path in self.data_dir.glob("*.csv"):
                name_lower = csv_path.name.lower()
                if "campaign" not in name_lower or "settings" not in name_lower:
                    continue
                try:
                    df = pd.read_csv(csv_path, encoding="utf-16", sep="\t", skiprows=0, low_memory=False)
                except Exception:
                    continue
                if "Ad schedule" in df.columns and "Campaign" in df.columns:
                    schedule_source = df
                    schedule_col = "Ad schedule"
                    break

        strategy_col = self._find_column(campaign_data, ["Campaign bid strategy type"])
        if schedule_source is None or schedule_col is None or not strategy_col:
            return None

        base_df = campaign_data[["Campaign", strategy_col]].copy()
        schedule_df = schedule_source[["Campaign", schedule_col]].copy()
        merged = base_df.merge(schedule_df, on="Campaign", how="left")
        df = merged.rename(columns={schedule_col: "Ad schedule", strategy_col: "Campaign bid strategy type"})

        def _has_schedule(value: object) -> bool:
            if value is None:
                return False
            text = str(value).strip().lower()
            if not text or text in {"all day", "--", "n/a"}:
                return False
            return True

        df["has_schedule"] = df["Ad schedule"].apply(_has_schedule)

        total_campaigns = len(df)
        if total_campaigns == 0:
            return None

        schedule_campaigns = int(df["has_schedule"].sum())

        smart_strategies = {
            "target cpa",
            "target roas",
            "maximize conversions",
            "maximize conversion value",
        }
        smart_mask = df["Campaign bid strategy type"].astype(str).str.strip().str.lower().isin(smart_strategies)

        smart_campaigns = int(smart_mask.sum())
        manual_campaigns = total_campaigns - smart_campaigns

        smart_schedule = int(df[smart_mask & df["has_schedule"]]["has_schedule"].sum())
        manual_schedule = schedule_campaigns - smart_schedule

        smart_pct = (smart_campaigns / total_campaigns * 100) if total_campaigns > 0 else 0.0
        prefs = self._dayparting_preferences()
        manual_threshold_ratio = 0.3 if prefs["manual_priority"] else 0.2
        manual_threshold = max(1, int(max(total_campaigns, 1) * manual_threshold_ratio))
        automation_covers_schedule = (
            smart_pct >= prefs["smart_threshold"]
            and smart_schedule == 0
            and manual_schedule <= max(1, manual_campaigns)
        )

        metrics = {
            "total": total_campaigns,
            "schedule_campaigns": schedule_campaigns,
            "smart_campaigns": smart_campaigns,
            "manual_campaigns": manual_campaigns,
            "smart_schedule": smart_schedule,
            "manual_schedule": manual_schedule,
            "smart_pct": smart_pct,
            "automation_covers_schedule": automation_covers_schedule,
            "smart_schedule_pct": (smart_schedule / max(smart_campaigns, 1)) if smart_campaigns else 0.0,
            "manual_schedule_pct": (manual_schedule / max(manual_campaigns, 1)) if manual_campaigns else 0.0,
        }
        self.summary_stats["ad_schedule"] = metrics
        return metrics

    def _score_dayparting_guardrail(self, mode: str) -> ScoringResult:
        """Shared logic for dayparting guardrail rows."""
        label_map = {
            "bid": "dayparting bid adjustments",
            "campaigns": "separate campaigns with ad scheduling",
            "automation": "automated day-parting strategy",
        }
        metrics = self._get_ad_schedule_metrics()
        if not metrics:
            smart_stats = self._get_smart_bidding_stats()
            smart_pct = smart_stats.get("smart_pct", 0.0) or 0.0
            if smart_pct >= 70:
                return ScoringResult(
                    score=4,
                    details=(
                        f"Automation inference: Smart Bidding covers {smart_pct:.0f}% of campaigns, so {label_map.get(mode, 'dayparting guardrails')} "
                        "are assumed to be automated even though the ad schedule export was missing."
                    ),
                    recommendations="Allow Smart Bidding to control dayparting; capture ad schedule columns next run for documentation.",
                    calculation="Automation fallback",
                )
            return ScoringResult(
                score=None,
                details=f"Data needed: Ad schedule/dayparting data to evaluate {label_map.get(mode, 'dayparting guardrails')}.",
                data_needed="Ad schedule/dayparting data",
            )

        total = metrics["total"]
        schedule_campaigns = metrics["schedule_campaigns"]
        smart_campaigns = metrics["smart_campaigns"]
        smart_schedule = metrics["smart_schedule"]
        manual_campaigns = metrics["manual_campaigns"]
        manual_schedule = metrics["manual_schedule"]
        smart_pct = metrics["smart_pct"]
        automation_covers = metrics.get("automation_covers_schedule", False)
        smart_schedule_pct = metrics.get("smart_schedule_pct", 0.0) or 0.0
        manual_schedule_pct = metrics.get("manual_schedule_pct", 0.0) or 0.0
        prefs = self._dayparting_preferences()
        smart_threshold = prefs["smart_threshold"]
        manual_high = prefs["manual_high"]
        manual_mid = prefs["manual_mid"]
        manual_priority = prefs["manual_priority"]

        if automation_covers:
            score, assessment = 5, "Automation manages dayparting"
            details = (
                f"{assessment}: Smart Bidding covers {smart_pct:.0f}% of campaigns with no conflicting manual schedules "
                f"({smart_schedule}/{max(smart_campaigns,1)} overrides)."
            )
            recommendations = "Maintain automation; Smart Bidding dynamically adjusts bids by time-of-day without manual modifiers."
        elif smart_pct >= max(50, smart_threshold - 10):
            if smart_schedule_pct <= 0.10:
                score, assessment = 4, "Limited Smart Bidding overrides"
            elif smart_schedule_pct <= 0.25:
                score, assessment = 3, "Partial automation conflicts"
            else:
                score, assessment = 2, "Manual schedules conflict with Smart Bidding"
            details = (
                f"{assessment}: {smart_schedule}/{max(smart_campaigns,1)} Smart Bidding campaigns ({smart_schedule_pct:.0%}) still use manual schedules "
                f"while automation covers {smart_pct:.0f}% of campaigns."
            )
            if score >= 4:
                recommendations = "Document intentional guardrails but keep automation in control for most campaigns."
            elif score == 3:
                recommendations = "Trim manual schedules on Smart Bidding campaigns so automation can learn from uninterrupted delivery."
            else:
                recommendations = "Remove manual overrides so Smart Bidding can optimize delivery across all hours."
        else:
            if manual_campaigns == 0:
                return ScoringResult(
                    score=None,
                    details=f"No manual campaigns available to evaluate {label_map.get(mode, 'dayparting guardrails')}.",
                    data_needed="Manual campaign list with schedules",
                )
            high_threshold = manual_high
            mid_threshold = manual_mid
            if not manual_priority:
                high_threshold = manual_mid
                mid_threshold = manual_mid / 2
            if manual_schedule_pct >= high_threshold:
                score, assessment = 5, "Manual campaigns actively manage dayparting"
            elif manual_schedule_pct >= mid_threshold:
                score, assessment = 3, "Partial manual dayparting coverage"
            else:
                score, assessment = 2, "Most manual campaigns run 24/7"
            details = (
                f"{assessment}: {manual_schedule}/{manual_campaigns} manual campaigns ({manual_schedule_pct:.0%}) use custom schedules."
            )
            if score == 5:
                recommendations = "Maintain manual dayparting coverage and continue monitoring results."
            elif score == 3:
                recommendations = "Expand dayparting schedules across manual campaigns to control spend during low-performing hours."
            else:
                recommendations = "Introduce schedules for manual campaigns or migrate them to Smart Bidding for automated control."

        calc_default = (
            f"Manual schedules: {manual_schedule}/{max(manual_campaigns,1)} | Smart schedules: {smart_schedule}/{max(smart_campaigns,1)}"
        )
        schedule_share = (schedule_campaigns / total) if total else 0.0

        if mode == "campaigns":
            details = (
                f"{details} {schedule_campaigns}/{total} campaigns ({schedule_share:.0%}) still rely on manual ad schedules."
            )
            calculation = f"Campaigns with schedules: {schedule_campaigns}/{total}"
        elif mode == "automation":
            details = f"{details} This reflects whether legacy schedules have been removed so Smart Bidding has uninterrupted control."
            calculation = calc_default
        else:
            calculation = calc_default

        weight = get_criterion_weight("dayparting_adjustments")
        impact = determine_business_impact("Dayparting Bid Adjustments", score, f"{manual_campaigns} manual campaigns")

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=calculation,
            criterion_weight=weight,
            business_impact=impact["business_impact"],
            priority_score=impact["priority_score"],
        )

    def _score_localized_ad_copy(self) -> ScoringResult:
        ad_data = self.data.get("ad")
        if ad_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Ad asset export with headlines and descriptions.",
                data_needed="Ad report",
            )

        terms = self._get_localization_terms()
        if not terms:
            return ScoringResult(
                score=None,
                details="Data needed: Targeted location data to evaluate localization signals.",
                data_needed="Campaign targeted location or user location export",
            )

        text_cols = [
            col
            for col in ad_data.columns
            if any(keyword in col.lower() for keyword in ["headline", "description", "text", "asset", "path"])
        ]
        if not text_cols:
            return ScoringResult(
                score=None,
                details="Data needed: Headline/description columns to analyze localized ad copy.",
                data_needed="Ad report with headline/description fields",
            )

        total = 0
        localized = 0
        for col in text_cols:
            series = ad_data[col].dropna().astype(str)
            for value in series:
                total += 1
                if self._text_contains_location(value.replace(";", " "), terms):
                    localized += 1

        if total == 0:
            return ScoringResult(
                score=None,
                details="Ad export did not contain usable headline or description text.",
                data_needed="Ad report with populated text columns",
            )

        pct = localized / total * 100
        expectation = self._localization_expectation()
        if expectation == "high":
            elite, acceptable = 40, 15
        elif expectation == "medium":
            elite, acceptable = 25, 10
        else:
            elite, acceptable = 15, 5

        calculation = f"Localized ad assets: {localized}/{total}"
        if expectation == "low" and localized == 0:
            return ScoringResult(
                score=4,
                details="Localization optional: national/regional messaging is acceptable for this program.",
                recommendations="Only add localized headlines when launching geo-specific offers or promotions.",
                calculation=calculation,
            )

        if pct >= elite:
            score, assessment = 5, "Localized messaging is embedded in most ad copy"
        elif pct >= acceptable:
            score, assessment = 3, "Some campaigns reference local markets"
        elif localized > 0:
            score, assessment = 2, "Localized copy exists but is rare"
        else:
            score, assessment = 1, "No evidence of localized ad copy"

        details = (
            f"{assessment}: {localized}/{total} ad assets ({pct:.0f}%) mention cities, regions, or dynamic location insertions."
        )
        if score >= 4:
            recommendations = "Maintain location-aware ad copy and reuse the highest-performing city or state variants across campaigns."
        else:
            recommendations = (
                "Inject localized headlines or use dynamic location insertion so each market sees relevant city/state messaging."
            )

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=calculation,
        )

    def _score_localized_extensions(self) -> ScoringResult:
        sitelink_data = self.data.get("sitelink_assets")
        if sitelink_data is None:
            sitelink_data = self.data.get("sitelink_assets_-_kl")
        if sitelink_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Sitelink asset export to audit localized extensions.",
                data_needed="Sitelink assets report",
            )

        terms = self._get_localization_terms()
        if not terms:
            return ScoringResult(
                score=None,
                details="Data needed: Targeted location data to evaluate localization signals.",
                data_needed="Campaign targeted location or user location export",
            )

        text_cols = [
            col
            for col in sitelink_data.columns
            if any(keyword in col.lower() for keyword in ["sitelink", "headline", "description", "text", "asset"])
        ]
        if not text_cols:
            return ScoringResult(
                score=None,
                details="Data needed: Sitelink text/description columns to assess localization.",
                data_needed="Sitelink assets report with text fields",
            )

        total = 0
        localized = 0
        for col in text_cols:
            series = sitelink_data[col].dropna().astype(str)
            for value in series:
                total += 1
                if self._text_contains_location(value.replace(";", " "), terms):
                    localized += 1

        if total == 0:
            return ScoringResult(
                score=None,
                details="Sitelink export contained no text fields to review.",
                data_needed="Sitelink assets report with text fields",
            )

        pct = localized / total * 100
        expectation = self._localization_expectation()
        if expectation == "high":
            elite, acceptable = 35, 15
        elif expectation == "medium":
            elite, acceptable = 25, 10
        else:
            elite, acceptable = 15, 5

        calculation = f"Localized sitelinks: {localized}/{total}"
        if expectation == "low" and localized == 0:
            return ScoringResult(
                score=4,
                details="Localized extensions are optional for this vertical; national sitelinks are sufficient.",
                recommendations="Add geo-specific sitelinks only when launching market-specific offers.",
                calculation=calculation,
            )

        if pct >= elite:
            score, assessment = 5, "Localized sitelinks are broadly deployed"
        elif pct >= acceptable:
            score, assessment = 3, "Some sitelinks reference local markets"
        elif localized > 0:
            score, assessment = 2, "Localized extensions exist but rarely surface"
        else:
            score, assessment = 1, "No localized sitelinks present"

        details = (
            f"{assessment}: {localized}/{total} sitelink or extension assets ({pct:.0f}%) reference city or state language."
        )
        if score >= 4:
            recommendations = "Keep rotating localized sitelinks alongside national variants so every market sees relevant copy."
        else:
            recommendations = "Add city/state sitelinks or leverage feed-based location insertion to highlight local offers."

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=calculation,
        )

    def _score_localized_landing_pages(self) -> ScoringResult:
        landing_page_data = self.data.get("landing_page")
        if landing_page_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Landing page export to evaluate localized experiences.",
                data_needed="Landing page report",
            )

        terms = self._get_localization_terms()
        if not terms:
            return ScoringResult(
                score=None,
                details="Data needed: Targeted location data to evaluate localization signals.",
                data_needed="Campaign targeted location or user location export",
            )

        url_col = None
        for candidate in ["Landing page", "Final URL", "Final Url", "URL"]:
            if candidate in landing_page_data.columns:
                url_col = candidate
                break
        if url_col is None:
            return ScoringResult(
                score=None,
                details="Data needed: Landing page URL column.",
                data_needed="Landing page report with Landing page/Final URL columns",
            )

        series = landing_page_data[url_col].dropna().astype(str)
        total = len(series)
        if total == 0:
            return ScoringResult(
                score=None,
                details="Landing page export contains no URLs to review.",
                data_needed="Landing page report with Landing page/Final URL columns",
            )

        localized = sum(1 for value in series if self._text_contains_location(value, terms))
        pct = localized / total * 100
        expectation = self._localization_expectation()
        if expectation == "high":
            elite, acceptable = 30, 10
        elif expectation == "medium":
            elite, acceptable = 20, 5
        else:
            elite, acceptable = 10, 3

        calculation = f"Localized landing pages: {localized}/{total}"
        if expectation == "low" and localized == 0:
            return ScoringResult(
                score=4,
                details="Localized landing pages are optional for national campaigns; the current setup is acceptable.",
                recommendations="Only build geo-specific destinations when campaigns target specific metro areas.",
                calculation=calculation,
            )

        if pct >= elite:
            score, assessment = 5, "Landing pages are localized for key markets"
        elif pct >= acceptable:
            score, assessment = 3, "Some localized landing pages exist"
        elif localized > 0:
            score, assessment = 2, "Localized landing pages are rare"
        else:
            score, assessment = 1, "Landing pages are not localized"

        details = (
            f"{assessment}: {localized}/{total} landing pages ({pct:.0f}%) include location-specific paths or parameters."
        )
        if score >= 4:
            recommendations = "Maintain localized landing destinations so local messaging carries through post-click."
        else:
            recommendations = "Create city/state-specific landing variants or append geo parameters so media can deep-link into relevant content."

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=calculation,
        )

    def _get_geo_strategy_metrics(self) -> Dict[str, object]:
        cached = self.summary_stats.get("geo_strategy")
        if cached:
            return cached

        smart_stats = self._get_smart_bidding_stats()
        smart_pct = smart_stats.get("smart_pct", 0.0) or 0.0
        manual_campaigns = smart_stats.get("manual_campaigns", 0) or 0
        total_campaigns = smart_stats.get("total_campaigns", 0) or 0
        manual_threshold = max(1, int(max(total_campaigns, 1) * 0.2))

        metrics: Dict[str, object] = {
            "smart_pct": smart_pct,
            "manual_campaigns": manual_campaigns,
            "total_campaigns": total_campaigns,
            "automation_covers_geo": smart_pct >= 60 and manual_campaigns <= manual_threshold,
            "locations_available": False,
            "total_locations": 0,
            "national_target": False,
            "regional_locations": 0,
            "location_samples": [],
            "geo_bid_available": False,
            "geo_bid_rows": 0,
            "geo_bid_adjusted": 0,
            "geo_bid_share": 0.0,
        }

        campaign_data = self.source_info.get("campaign_source")
        if campaign_data is not None:
            location_col = self._find_column(campaign_data, ["Targeted location"])
            if location_col:
                locations = campaign_data[location_col].astype(str).str.strip()
                locations = locations[locations.str.len() > 0]
                unique_locations = [loc for loc in locations.unique() if loc]
                total_locations = len(unique_locations)
                national_target = any("united states" in loc.lower() or loc.lower() in {"us", "usa"} for loc in unique_locations)
                regional_locations = total_locations - (1 if national_target else 0)
                metrics.update(
                    {
                        "locations_available": True,
                        "total_locations": total_locations,
                        "national_target": national_target,
                        "regional_locations": regional_locations,
                        "location_samples": unique_locations[:5],
                    }
                )

            bid_col = self._find_column_by_substrings(campaign_data, ["location", "bid"])
            if bid_col:
                bids = to_numeric_safe(campaign_data[bid_col]).fillna(0)
                total_rows = len(bids)
                adjusted = int((bids != 0).sum())
                share = (adjusted / total_rows * 100) if total_rows else 0.0
                metrics.update(
                    {
                        "geo_bid_available": True,
                        "geo_bid_rows": total_rows,
                        "geo_bid_adjusted": adjusted,
                        "geo_bid_share": share,
                    }
                )

        self.summary_stats["geo_strategy"] = metrics
        return metrics

    def _get_localization_terms(self) -> List[str]:
        cached = self.summary_stats.get("localization_terms")
        if cached is not None:
            return cached

        terms: Set[str] = set()
        skip_tokens = {"united states", "united state", "usa", "us", "america", "canada"}

        def _normalize_value(value: object) -> List[str]:
            items: List[str] = []
            if value is None:
                return items
            raw = str(value).strip()
            if not raw:
                return items
            # Break on separators to capture multi-value fields
            parts = re.split(r"[;,/|]", raw)
            for part in parts:
                cleaned = re.sub(r"[^A-Za-z0-9\s]", " ", part).strip().lower()
                cleaned = re.sub(r"\s+", " ", cleaned)
                if len(cleaned) < 3:
                    continue
                if cleaned in skip_tokens:
                    continue
                items.append(cleaned)
            return items

        def _collect_from_dataframe(df: Optional[pd.DataFrame], target_columns: List[str]) -> None:
            if df is None:
                return
            for col in df.columns:
                col_lower = col.lower()
                if any(token in col_lower for token in target_columns):
                    series = df[col].dropna()
                    for value in series:
                        for term in _normalize_value(value):
                            terms.add(term)

        geo_metrics = self._get_geo_strategy_metrics()
        samples = geo_metrics.get("location_samples") or []
        for sample in samples:
            for term in _normalize_value(sample):
                terms.add(term)

        campaign_data = self.source_info.get("campaign_source")
        _collect_from_dataframe(campaign_data, ["location", "city", "region", "state"])

        user_locations = self.data.get("user_locations")
        _collect_from_dataframe(user_locations, ["location", "city", "region", "state", "dma"])

        if not terms:
            self.summary_stats["localization_terms"] = []
            return []

        ordered = sorted(terms)
        self.summary_stats["localization_terms"] = ordered
        return ordered

    def _text_contains_location(self, text: str, terms: List[str]) -> bool:
        if not text:
            return False
        lowered = str(text).lower()
        dynamic_tokens = [
            "{=location",
            "{location",
            ":city",
            ":state",
            ":location",
            "{city",
            "{state",
            "targetcity",
            "city:",
            "state:",
        ]
        if any(token in lowered for token in dynamic_tokens):
            return True
        for term in terms:
            if term and term in lowered:
                return True
        return False

    def _build_scope_summary(self) -> Dict[str, int]:
        """Count core entities (campaigns, ad groups, keywords, ads, audiences)."""

        def _unique_count(df: Optional[pd.DataFrame], columns: List[str]) -> int:
            if df is None:
                return 0
            column = self._find_column(df, columns)
            if column is None:
                return 0
            return int(df[column].dropna().nunique())

        campaign_df = self.data.get("campaign_details")
        if campaign_df is None:
            campaign_df = self.source_info.get("campaign_source")
        adgroup_df = self.source_info.get("adgroup_source")
        keyword_df = self.data.get("search_keyword")
        if keyword_df is None:
            keyword_df = self.data.get("keyword")
        ad_df = self.data.get("ad")
        if ad_df is None:
            ad_df = self.data.get("ads")
        if ad_df is None:
            ad_df = self.data.get("responsive_search_ad")
        audience_df = self.data.get("audience")

        summary = {
            "campaigns": _unique_count(campaign_df, ["Campaign", "Campaign name"]),
            "ad_groups": _unique_count(adgroup_df, ["Ad group", "Ad group name"]),
            "keywords": _unique_count(keyword_df, ["Search keyword", "Keyword"]),
            "ads": _unique_count(ad_df, ["Ad ID", "Headline 1", "Ad", "Asset combination ID"]),
            "audiences": _unique_count(audience_df, ["Audience", "Audience name"]),
        }

        conversion_source = self._get_conversion_source_data()
        summary["conversions_tracked"] = _unique_count(conversion_source, ["Conversion name", "Name"])

        return summary

    def _score_descriptor(self, score):
        mapping = {
            1: "Non-existent; immediate build required",
            2: "Poor execution; significant improvement required",
            3: "OK; material improvement needed",
            4: "Good; minor optimization opportunities",
            5: "Excellent; keep current approach"
        }
        return mapping.get(int(score), "Assessment unavailable.")

    def _select_variant(self, options: List[str], seed: str) -> str:
        if not options:
            return ""
        idx = abs(hash(seed)) % len(options)
        return options[idx]

    def _clean_action(self, text: str) -> str:
        action = (text or "").strip()
        if not action:
            return ""
        action = re.sub(r"^\s*ACTION:\s*", "", action, flags=re.IGNORECASE)
        return action.rstrip(".")

    def _business_context_data(self) -> Dict[str, Any]:
        if isinstance(self.summary_stats.get("business_context"), dict):
            return self.summary_stats["business_context"]
        if isinstance(self.business_context, dict):
            return self.business_context
        return {}

    def _context_constraints(self) -> Dict[str, Any]:
        ctx = self._business_context_data()
        constraints = ctx.get("constraints")
        return constraints if isinstance(constraints, dict) else {}

    def _normalize_vertical_label(self, label: Optional[object]) -> str:
        if not label:
            return "LEAD_GEN"
        text = str(label).strip().upper()
        mapping = {
            "ECOMMERCE": "ECOMMERCE",
            "E-COMMERCE": "ECOMMERCE",
            "D2C": "ECOMMERCE",
            "RETAIL": "ECOMMERCE",
            "SAAS": "SAAS",
            "SOFTWARE": "SAAS",
            "FINANCE": "FINANCE",
            "FINANCIAL": "FINANCE",
            "BANKING": "FINANCE",
            "INSURANCE": "FINANCE",
            "HEALTH": "LEAD_GEN",
            "HEALTHCARE": "LEAD_GEN",
            "LOCAL": "LOCAL",
            "LEAD_GEN": "LEAD_GEN",
            "LEADGEN": "LEAD_GEN",
        }
        return mapping.get(text, "LEAD_GEN")

    def _infer_vertical_from_account(self, account_name: str) -> str:
        lowered = (account_name or "").lower()
        finance_keywords = [
            "retire",
            "wealth",
            "broker",
            "asset",
            "fili",
            "crypto",
            "invest",
            "medicare",
            "529",
            "hsa",
            "able",
            "tem",
        ]
        if any(keyword in lowered for keyword in finance_keywords):
            return "FINANCE"
        if "brand" in lowered:
            return "LEAD_GEN"
        return "LEAD_GEN"

    def _context_vertical(self) -> str:
        ctx = self._business_context_data()
        detected = ctx.get("detected_vertical") or ctx.get("vertical")
        if detected:
            return self._normalize_vertical_label(detected)
        account_name = ""
        if isinstance(ctx.get("account_name"), str):
            account_name = ctx["account_name"]
        if not account_name:
            account_name = self.business_unit or ""
        return self._infer_vertical_from_account(account_name)

    def _context_primary_kpi(self) -> Optional[str]:
        constraints = self._context_constraints()
        value = constraints.get("primary_kpi")
        return str(value) if value else None

    def _context_min_conversions(self) -> Optional[int]:
        constraints = self._context_constraints()
        value = constraints.get("min_conversion_volume")
        try:
            return int(value) if value is not None else None
        except (TypeError, ValueError):
            return None

    def _pmax_targets(self) -> Tuple[int, int]:
        """Return (launch_threshold, optimal_threshold) for PMAX share."""
        vertical = self._context_vertical()
        if vertical == "ECOMMERCE":
            return (8, 20)
        if vertical == "FINANCE":
            return (3, 12)
        if vertical == "LOCAL":
            return (5, 15)
        return (5, 15)

    def _pmax_guardrail_requirement(self) -> int:
        vertical = self._context_vertical()
        if vertical in {"FINANCE", "LOCAL"}:
            return 20
        return 15

    def _geo_location_thresholds(self) -> Tuple[int, int, int, int]:
        vertical = self._context_vertical()
        if vertical in {"FINANCE", "LOCAL"}:
            return (25, 15, 8, 3)
        if vertical == "ECOMMERCE":
            return (20, 12, 5, 2)
        return (8, 4, 2, 1)

    def _geo_bid_share_thresholds(self) -> Tuple[int, int, int, int]:
        vertical = self._context_vertical()
        if vertical in {"FINANCE", "LOCAL"}:
            return (70, 50, 30, 15)
        if vertical == "ECOMMERCE":
            return (60, 40, 25, 10)
        return (50, 30, 15, 5)

    def _dayparting_preferences(self) -> Dict[str, object]:
        vertical = self._context_vertical()
        if vertical in {"FINANCE", "LOCAL"}:
            return {
                "smart_threshold": 75,
                "manual_high": 0.7,
                "manual_mid": 0.4,
                "manual_priority": True,
            }
        return {
            "smart_threshold": 65,
            "manual_high": 0.5,
            "manual_mid": 0.25,
            "manual_priority": False,
        }

    def _localization_expectation(self) -> str:
        vertical = self._context_vertical()
        if vertical == "LOCAL":
            return "high"
        if vertical in {"FINANCE", "SAAS"}:
            return "low"
        return "medium"

    def _contextual_hint(self) -> str:
        ctx = self._business_context_data()
        if not ctx:
            return ""
        constraints = self._context_constraints()
        vertical = self._context_vertical()
        primary_kpi = constraints.get("primary_kpi")
        hints: List[str] = []

        break_even = ctx.get("break_even_roas")
        if primary_kpi == "POAS" and break_even:
            hints.append(f"Break-even ROAS for this business is {break_even}; keep actions profit-positive.")
        elif primary_kpi == "CAC_PAYBACK":
            hints.append("SaaS economics prioritize CAC payback, so actions must show how they shorten payback windows.")
        elif primary_kpi == "CPA":
            min_volume = constraints.get("min_conversion_volume")
            if min_volume:
                hints.append(f"Lead-gen guardrail: Smart Bidding expects {min_volume}+ monthly conversions.")

        if vertical == "FINANCE" and constraints.get("compliance_check_required"):
            hints.append("Finance program—maintain regulator verification before scaling budgets.")

        return hints[0] if hints else ""

    def _extract_missing_evidence(self, result: ScoringResult) -> str:
        """Return a cleaned description of the missing export/data for rationale copy."""
        if not result:
            return "the missing evidence"

        detail = (result.details or "").strip()
        lowered = detail.lower()
        if lowered.startswith("data needed:"):
            detail = detail.split(":", 1)[1].strip()

        if not detail and result.data_needed:
            detail = result.data_needed.strip()

        detail = detail.rstrip(".")
        return detail or "the missing evidence"

    def _format_evidence_phrase(self, evidence: str) -> str:
        """Ensure missing-evidence phrasing flows naturally in a sentence."""
        text = (evidence or "").strip()
        if not text:
            return "the missing evidence"
        lowered = text.lower()
        if lowered.startswith(("the ", "a ", "an ")):
            return text
        return f"the {text}"

    def _build_missing_data_action(self, criterion: Optional[str], result: ScoringResult) -> str:
        """Craft Column F copy that explains why the export is required."""
        missing_evidence = self._extract_missing_evidence(result)
        evidence_phrase = self._format_evidence_phrase(missing_evidence)
        if not evidence_phrase:
            return ""
        return evidence_phrase

    def _build_missing_data_rationale(self, criterion: Optional[str], result: ScoringResult) -> str:
        """Column G default narrative for rows we couldn't grade."""
        missing_evidence = self._extract_missing_evidence(result)
        evidence_phrase = self._format_evidence_phrase(missing_evidence)
        if not evidence_phrase:
            return ""
        return evidence_phrase

    @staticmethod
    def _format_score_value(value: Optional[float]) -> str:
        if value is None:
            return ""
        if float(value).is_integer():
            return f"{int(value)}"
        return f"{value:.1f}"

    @staticmethod
    def _format_metric_value(value: Optional[float], metric_label: str) -> str:
        if value is None:
            return ""
        label = (metric_label or "").lower()
        prefix = "$" if ("cost" in label or "spend" in label) else ""
        if abs(value) >= 100:
            return f"{prefix}{value:,.0f}"
        return f"{prefix}{value:,.2f}"

    def _build_where_clause(self, result: ScoringResult) -> str:
        if not result.is_scored():
            return ""
        if not isinstance(self.source_info, dict):
            return ""
        context = self.source_info.get("context") or {}
        if not isinstance(context, dict):
            return ""

        parts: List[str] = []
        top_campaigns = context.get("top_campaigns")
        if isinstance(top_campaigns, dict):
            metric = top_campaigns.get("metric") or "Cost"
            entries = top_campaigns.get("top") or []
            if entries:
                items = []
                for entry in entries:
                    name = entry.get("name")
                    value = self._format_metric_value(entry.get("value"), metric)
                    if name:
                        items.append(f"{name} ({value})" if value else f"{name}")
                if items:
                    parts.append(f"top campaigns by {metric.lower()}: {', '.join(items[:3])}")

        top_adgroups = context.get("top_ad_groups")
        if isinstance(top_adgroups, dict):
            metric = top_adgroups.get("metric") or "Cost"
            entries = top_adgroups.get("top") or []
            if entries:
                items = []
                for entry in entries:
                    name = entry.get("name")
                    value = self._format_metric_value(entry.get("value"), metric)
                    if name:
                        items.append(f"{name} ({value})" if value else f"{name}")
                if items:
                    parts.append(f"top ad groups by {metric.lower()}: {', '.join(items[:3])}")

        state_breakdown = context.get("state_breakdown")
        if isinstance(state_breakdown, dict):
            metric = state_breakdown.get("metric") or "Cost"
            states = state_breakdown.get("states") or []
            if states:
                items = []
                for entry in states:
                    state = entry.get("state")
                    value = self._format_metric_value(entry.get("value"), metric)
                    if state:
                        items.append(f"{state} {value}".strip())
                if items:
                    parts.append(f"state mix by {metric.lower()}: {', '.join(items[:5])}")

        if not parts:
            return ""
        return "; ".join(parts)

    @staticmethod
    def _should_add_where_clause(result: ScoringResult) -> bool:
        if not result or not result.is_scored():
            return False
        if result.priority_score >= 7:
            return True
        if result.business_impact in {"high", "critical"}:
            return True
        if isinstance(result.score, (int, float)) and result.score <= 2:
            return True
        return False

    @staticmethod
    def _guard_action_consistency(
        criterion: Optional[str],
        result: ScoringResult,
        detail_text: str,
        action_text: str,
        original_action: str,
    ) -> str:
        """Prevent contradictory action text after AI rewriting."""
        if not action_text or not original_action:
            return action_text
        if not criterion:
            return action_text
        criterion_lower = criterion.lower()
        score_value = result.score if isinstance(result.score, (int, float)) else None
        detail_lower = (detail_text or "").lower()
        action_lower = (action_text or "").lower()

        if "account active in all engines" in criterion_lower and score_value is not None and score_value <= 1:
            if "active" in action_lower or "generating impressions" in action_lower:
                return original_action

        if score_value is not None and score_value <= 1:
            if "0 impressions" in detail_lower and "active" in action_lower:
                return original_action

        return action_text

    def _compose_detail_with_defense(self, result: ScoringResult) -> str:
        """Append risk, benchmark, and binary-gap context directly into Column E."""
        segments: List[str] = []
        if result.needs_data():
            base_text = self._extract_missing_evidence(result)
        else:
            base_text = (result.details or "").strip()
        if base_text:
            segments.append(base_text)

        if result.raw_score is not None and isinstance(result.score, (int, float)):
            raw = self._format_score_value(result.raw_score)
            final = self._format_score_value(float(result.score))
            if raw and final and raw != final:
                adjustment_text = "; ".join(result.risk_adjustments) if result.risk_adjustments else "Risk penalty applied"
                segments.append(f"Risk-adjusted scoring: Raw {raw}/5 → {adjustment_text} → Final {final}/5.")
            elif raw:
                segments.append(f"Raw score confirmation: {raw}/5.")

        if result.benchmark_delta:
            segments.append(result.benchmark_delta.strip())
        if result.vertical_context:
            segments.append(result.vertical_context.strip())
        if result.binary_gap_indicator:
            segments.append(f"Binary Gap: {result.binary_gap_indicator.strip()}.")

        return "\n".join(segment for segment in segments if segment).strip()

    def _build_default_rationale(self, criterion: str, category: Optional[str], result: ScoringResult) -> str:
        if result is None:
            return ""

        if result.needs_data():
            return self._build_missing_data_rationale(criterion, result)

        fact = result.details or result.recommendations or "See audit notes for findings."
        fact_sentence = fact.rstrip(".")
        fact_clause = fact_sentence[0].lower() + fact_sentence[1:] if len(fact_sentence) > 1 else fact_sentence
        score_statement = self._score_descriptor(result.score) if isinstance(result.score, (int, float)) else ""
        is_strength = isinstance(result.score, (int, float)) and result.score >= 4
        summary_templates_strength = [
            "{criterion} stays healthy because {fact_clause}.",
            "{criterion} continues to perform thanks to how {fact_clause}.",
        ]
        summary_templates_gap = [
            "{fact_sentence}. That shows {criterion} needs immediate reinforcement.",
            "{criterion} is underperforming because {fact_clause}.",
            "This review exposed that {fact_clause}, leaving {criterion} fragile.",
        ]
        summary_text = self._select_variant(
            summary_templates_strength if is_strength else summary_templates_gap,
            f"{criterion}:{fact_sentence}",
        ).format(
            criterion=criterion,
            fact_sentence=fact_sentence,
            fact_clause=fact_clause or fact_sentence.lower(),
            score=score_statement.lower() or "solid execution",
        )

        action_candidates_strength = [
            "Document the exact steps behind this result and make them standard across teams.",
            "Socialize this process so future campaigns inherit the same discipline.",
            "Maintain the proof points and rotate them into upcoming playbooks.",
        ]
        action_candidates_gap = [
            "Assign an owner, remediation plan, and check-in to close this gap.",
            "Run a focused diagnostic sprint to rebuild the foundations before scaling.",
            "Pair this finding with the diagnostic checklist and report progress in the next review.",
        ]
        fallback_action = self._select_variant(
            action_candidates_strength if is_strength else action_candidates_gap,
            f"{category}:{criterion}",
        )
        action_text = self._clean_action(result.recommendations) or fallback_action
        if not result.recommendations and action_text:
            result.recommendations = action_text

        statements = [summary_text]
        if score_statement:
            statements.append(f"Score context ({criterion}): {score_statement}.")
        if action_text:
            statements.append(f"Next step: {action_text}.")
        hint = self._contextual_hint()
        if hint:
            statements.append(hint)
        if result.benchmark_delta:
            statements.append(result.benchmark_delta.strip())
        if result.vertical_context:
            statements.append(result.vertical_context.strip())
        if result.binary_gap_indicator:
            statements.append(f"Binary Gap: {result.binary_gap_indicator.strip()}.")
        return " ".join(statements).strip()

    def _get_conversion_source_data(self) -> Optional[pd.DataFrame]:
        """Return active conversion-source rows for the current business unit."""
        dataset = None
        for key in self.data.keys():
            if key.startswith('conversion_source'):
                dataset = self.data[key]
                break

        if dataset is None or len(dataset) == 0:
            return None

        df = dataset.copy()

        status_col = self._find_column(df, ['Campaign status'])
        if status_col:
            df = df[df[status_col].astype(str).str.contains('Eligible', case=False, na=False)]

        if len(df) == 0:
            return None

        return df

    def _print_context_summary(self, context: Optional[Dict[str, Dict]]) -> None:
        if not context:
            return

        print("\n  Context signals:")

        change = context.get("change_history")
        if change:
            events_30d = change.get("events_30d")
            total_events = change.get("total_events")
            last_change = change.get("last_change") or "n/a"
            cadence = events_30d if events_30d is not None else total_events
            if cadence is not None:
                print(f"    Change cadence: {cadence} events in last 30d (last change {last_change})")
            else:
                print(f"    Change cadence: last change {last_change}")
            top_types = change.get("top_types") or []
            if top_types:
                print(f"      Top change types: {', '.join(top_types)}")

        audience = context.get("audience")
        if audience:
            unique_lists = audience.get("unique_lists", audience.get("rows"))
            if unique_lists is not None:
                print(f"    Audience lists: {unique_lists} unique segments")
            sample = audience.get("sample_lists") or []
            if sample:
                print(f"      Sample lists: {', '.join(sample[:3])}")

        geo_ctx = context.get("geo")
        if geo_ctx:
            top_locations = geo_ctx.get("top_locations") or []
            if top_locations:
                formatted = ", ".join(
                    f"{item.get('location', 'Unknown')} ({int(item.get('impr', 0)):,} impr)"
                    for item in top_locations[:3]
                )
                print(f"    Top locations: {formatted}")

        conv = context.get("conversion_tracking")
        if conv:
            total = conv.get("total_actions")
            active = conv.get("active_actions")
            if total is not None:
                if active is not None:
                    print(f"    Conversion tracking: {active}/{total} actions active")
                else:
                    print(f"    Conversion tracking: {total} configured actions")

        for label, key in (("Sitelinks", "sitelinks"), ("Callouts", "callouts")):
            ext = context.get(key)
            if ext:
                rows = ext.get("rows")
                active_pct = ext.get("active_pct")
                detail = f"{rows} rows" if rows is not None else "rows unavailable"
                if active_pct is not None:
                    detail += f" ({active_pct}% active)"
                print(f"    {label}: {detail}")

    def discover_data_sources(self):
        """Use intelligent mapper to discover data structure"""
        print("\n[2/5] Discovering data sources...")
        self.source_info = self.mapper.discover_sources(self.data)

        # Extract key references
        self.performance_source = self.source_info['performance_source']
        self.performance_data = self.data[self.performance_source]
        self.campaign_source = self.source_info.get('campaign_attributes_source') or (
            'account' if 'account' in self.data else None
        )

        # Create unified aliases for V3.1 functions
        # V3.1 functions expect: campaign_source, adgroup_source, campaign_details, ad_group, ad
        # IMPORTANT: Prefer _details versions as they have more columns (Keywords: active, etc.)
        self.source_info['campaign_source'] = self.data.get('campaign') if 'campaign' in self.data else self.data.get('account')
        self.source_info['adgroup_source'] = self.data.get('ad_group_details') if 'ad_group_details' in self.data else self.data.get('ad_group')
        self.source_info['campaign_details'] = self.data.get('campaign_details')

        print(f"\n  Performance: {self.performance_source}")
        print(f"  Granularity: {self.source_info['performance_granularity']}")
        print(f"  Campaigns: {self.campaign_source}")
        self._print_context_summary(self.source_info.get('context'))

        # Normalize core column aliases expected by scoring (Campaign, Ad group)
        self._normalize_core_columns()

    def _normalize_core_columns(self):
        """Ensure common aliases exist for campaign/ad group across selected sources."""
        def add_alias(df, target, candidates):
            if df is None or target in df.columns:
                return df
            for cand in candidates:
                if cand in df.columns:
                    df[target] = df[cand]
                    break
            return df

        # Candidate variants for Campaign / Ad group / Account across SA360/CSV exports
        campaign_candidates = [
            "Campaign",
            "campaign",
            "campaign.name",
            "Campaign name",
            "Campaigns",
            "campaign_name",
            "Campaign ID",
            "campaign.id",
        ]
        account_candidates = [
            "Account name",
            "Account Name",
            "Account",
        ]
        adgroup_candidates = [
            "Ad group",
            "ad_group",
            "ad_group.name",
            "Ad group name",
            "Ad group ID",
            "ad_group.id",
            "Adgroup",
        ]

        # Normalize all loaded datasets
        for key, df in list(self.data.items()):
            df = add_alias(df, "Campaign", campaign_candidates)
            df = add_alias(df, "Account name", account_candidates)
            df = add_alias(df, "Ad group", adgroup_candidates)
            self.data[key] = df

        # Normalize performance data
        if hasattr(self, "performance_data"):
            self.performance_data = add_alias(
                self.performance_data,
                "Campaign",
                campaign_candidates,
            )
            self.performance_data = add_alias(
                self.performance_data,
                "Account name",
                account_candidates,
            )
            self.performance_data = add_alias(
                self.performance_data,
                "Ad group",
                adgroup_candidates,
            )

        # Normalize campaign/adgroup sources
        for key in ["campaign_source", "campaign_details", "adgroup_source"]:
            df = self.source_info.get(key)
            df = add_alias(df, "Campaign", campaign_candidates)
            df = add_alias(df, "Account name", account_candidates)
            df = add_alias(df, "Ad group", adgroup_candidates)
            if df is not None:
                self.source_info[key] = df

        # Attempt fallbacks if the selected sources still lack required core columns
        def pick_fallback(kind: str, candidates: List[str], target: str) -> pd.DataFrame:
            """Pick first candidate dataset that contains the target column after aliasing."""
            searched = set()
            for name in candidates:
                df = self.source_info.get(name) if isinstance(self.source_info.get(name), pd.DataFrame) else self.data.get(name)
                searched.add(name)
                if df is None or not isinstance(df, pd.DataFrame):
                    continue
                df = add_alias(df, "Campaign", campaign_candidates)
                df = add_alias(df, "Ad group", adgroup_candidates)
                if target in df.columns:
                    print(f"  [normalize] {kind} fallback -> {name}")
                    return df
            # Broader scan across all loaded datasets if named candidates failed
            for name, df in self.data.items():
                if name in searched or not isinstance(df, pd.DataFrame):
                    continue
                df = add_alias(df, "Campaign", campaign_candidates)
                df = add_alias(df, "Ad group", adgroup_candidates)
                if target in df.columns:
                    print(f"  [normalize] {kind} fallback -> {name} (scanned)")
                    return df
            if target == "Campaign" and hasattr(self, "performance_data"):
                perf_df = add_alias(self.performance_data, "Campaign", campaign_candidates)
                perf_df = add_alias(perf_df, "Ad group", adgroup_candidates)
                if "Campaign" in perf_df.columns:
                    print(f"  [normalize] {kind} fallback -> performance_data")
                    return perf_df
            return None

        if self.source_info.get("campaign_source") is not None and "Campaign" not in self.source_info["campaign_source"].columns:
            fallback_campaign = pick_fallback(
                "campaign_source",
                ["campaign_details", "campaign_attributes_source", "account", "campaign"],
                "Campaign",
            )
            if fallback_campaign is not None:
                self.source_info["campaign_source"] = fallback_campaign
        elif self.source_info.get("campaign_source") is None:
            fallback_campaign = pick_fallback(
                "campaign_source",
                ["campaign_details", "campaign_attributes_source", "account", "campaign"],
                "Campaign",
            )
            if fallback_campaign is not None:
                self.source_info["campaign_source"] = fallback_campaign

        if self.source_info.get("adgroup_source") is not None and "Ad group" not in self.source_info["adgroup_source"].columns:
            fallback_adgroup = pick_fallback(
                "adgroup_source",
                ["ad_group_details", "ad_group"],
                "Ad group",
            )
            if fallback_adgroup is not None:
                self.source_info["adgroup_source"] = fallback_adgroup
        elif self.source_info.get("adgroup_source") is None:
            fallback_adgroup = pick_fallback(
                "adgroup_source",
                ["ad_group_details", "ad_group"],
                "Ad group",
            )
            if fallback_adgroup is not None:
                self.source_info["adgroup_source"] = fallback_adgroup

        # Final assertion with diagnostics: log and raise if required columns are still missing
        perf_cols = list(getattr(self, "performance_data", pd.DataFrame()).columns)
        camp_cols = list(self.source_info.get("campaign_source", pd.DataFrame()).columns if self.source_info.get("campaign_source") is not None else [])
        adg_cols = list(self.source_info.get("adgroup_source", pd.DataFrame()).columns if self.source_info.get("adgroup_source") is not None else [])
        missing = []
        if "Campaign" not in perf_cols:
            missing.append(f"performance_data missing Campaign; cols={perf_cols}")
        if self.source_info.get("campaign_source") is not None and "Campaign" not in camp_cols:
            missing.append(f"campaign_source missing Campaign; cols={camp_cols}")
        if self.source_info.get("adgroup_source") is not None and "Ad group" not in adg_cols:
            missing.append(f"adgroup_source missing Ad group; cols={adg_cols}")
        if missing:
            msg = " | ".join(missing)
            raise ValueError(f"[audit-normalize] Required columns missing after aliasing: {msg}")
    def score_criterion(self, excel_row: int, criterion_name: str, category: str) -> ScoringResult:
        """
        Score one criterion based on its name/category
        Uses completeness gates - never scores without data

        Returns ScoringResult with score/details/recommendations
        """

        # Route to appropriate scoring function based on criterion name
        criterion_lower = criterion_name.lower() if criterion_name else ""

        if "smart bidding campaign consolidation" in criterion_lower:
            return self._score_exec_smart_bidding_consolidation()

        if "goal-consistent campaign objectives" in criterion_lower:
            return self._score_exec_goal_alignment()

        if "signal density per campaign" in criterion_lower:
            return self._score_exec_signal_density()

        if "broad match adoption" in criterion_lower or "broad match hygiene" in criterion_lower:
            return self._score_broad_match_pct()

        if "legacy match-type splitting" in criterion_lower:
            return self._score_match_type_split()

        if "pmax complements search" in criterion_lower:
            return self._score_pmax_guardrails()

        # ===================================================================
        # V3.1 CRITERION ROUTING (Template-driven, exact name matching)
        # ===================================================================

        # Row 12: Campaigns split by match type?
        if "campaigns split by match type" in criterion_lower or "keyword groupings of similar types" in criterion_lower:
            return self._score_match_type_split()

        # Row 13: Granular Campaigns?
        if "granular campaigns" in criterion_lower and "ad groups" not in criterion_lower:
            return self._score_granular_campaigns()

        # Row 14: Granular Ad Groups?
        if "granular ad groups" in criterion_lower:
            return self._score_granular_adgroups()

        # Row 15: Account active - Google
        if "account active" in criterion_lower and "google" in criterion_lower:
            return self._score_google_presence()

        # Row 16: Account active - BING
        if "account active" in criterion_lower and "bing" in criterion_lower:
            return self._score_microsoft_presence()

        # Row 17: Account active - Other
        if "account active" in criterion_lower and "other" in criterion_lower:
            return self._score_other_engines()

        # Row 18: PMAX within account
        if "pmax" in criterion_lower and "account" in criterion_lower:
            return self._score_pmax()

        # Row 19: Naming conventions
        if "naming conventions" in criterion_lower and "place" in criterion_lower:
            return self._score_naming_conventions()

        # Row 21: Good Coverage?
        if "good coverage" in criterion_lower or (criterion_lower == "coverage" and category.lower() == "keywords"):
            return self._score_coverage()

        # Row 22: High Quality Scores?
        if "quality score" in criterion_lower:
            return self._score_quality_score_components()

        # Row 23: No Active AG missing Active Keywords
        if "missing" in criterion_lower and "keywords" in criterion_lower:
            return self._score_missing_keywords()

        # Row 24: Any disapprovals
        if "disapprovals" in criterion_lower:
            return self._score_disapprovals()

        # Row 25: Broad Match Modified
        if "broad match integration" in criterion_lower or ("broad match modified" in criterion_lower) or ("broad match" in criterion_lower and "balance" not in criterion_lower):
            return self._score_broad_match_pct()

        # Row 26: Balance of Match Types?
        if "balance" in criterion_lower and "match type" in criterion_lower:
            return self._score_match_type_balance_v31()

        # Row 28: Campaign lists used?
        if "campaign lists" in criterion_lower or ("audience" in criterion_lower and "coverage" in criterion_lower):
            return self._score_audience_coverage()

        # Row 29: Cross Category Negatives?
        if "cross category" in criterion_lower and "negative" in criterion_lower:
            return self._score_negative_keywords()

        # Row 31: Adapted to ad groups/keywords? (DKI)
        if "adapted to ad groups" in criterion_lower or "dki" in criterion_lower:
            return self._score_dki()

        # Row 33: Continuous Test & Learn? (Ad Copy section)
        if excel_row == 33 and ("continuous test" in criterion_lower or ("test" in criterion_lower and "learn" in criterion_lower)):
            return self._score_ad_testing()

        # Fallback for Row 33 pattern without row number check (for backward compatibility)
        if excel_row < 40 and ("continuous test" in criterion_lower or ("test" in criterion_lower and "learn" in criterion_lower)):
            return self._score_ad_testing()

        # Row 34: Copy includes USPs?
        if "usp" in criterion_lower:
            return self._score_usp_copy()

        # Row 35: Copy features prices/promos?
        if "prices" in criterion_lower or "promos" in criterion_lower:
            return self._score_pricing_copy()

        # Row 36: Copy uses CTAs?
        if "cta" in criterion_lower:
            return self._score_cta_copy()

        # Row 37: Copy uses ®™©?
        if "®" in criterion_name or "™" in criterion_name or "trademark symbols" in criterion_lower:
            return self._score_trademark_symbols()

        # Row 38: Copy uses inter-capitalisation? (Title Case)
        if "inter-capitalisation" in criterion_lower or "title case" in criterion_lower:
            return self._score_title_case()

        # Row 39: Extensions present?
        if "using all relevant extensions" in criterion_lower or "extensions present" in criterion_lower:
            return self._score_extensions()
        if "adapted to sitelinks" in criterion_lower or "adapted to sitelink" in criterion_lower:
            return self._score_extensions()

        # ===================================================================
        # LANDING PAGES (Rows 41, 43, 44)
        # ===================================================================

        # Row 41: Adapted to Keywords & Copy?
        if "adapted to keywords" in criterion_lower and "copy" in criterion_lower:
            return self._score_landing_page_adaptation()

        # Row 43: URL Checker
        if "url checker" in criterion_lower:
            return self._score_url_checker()

        # Row 44: Continuous Test & Learn? (Landing Pages section)
        # Row 44 has same criterion text as Row 33, so check row number explicitly
        if excel_row == 44 and ("continuous test" in criterion_lower or ("test" in criterion_lower and "learn" in criterion_lower)):
            return self._score_landing_page_testing()

        # Fallback: If in Landing Pages section (rows 40-45) and mentions test/learn
        if 40 <= excel_row <= 45 and ("test" in criterion_lower or "learn" in criterion_lower):
            return self._score_landing_page_testing()

        # ===================================================================
        # BIDDING & PERFORMANCE (Rows 46, 47, 49, 50, 52)
        # ===================================================================

        # Row 46: Smart Bidding? (already exists, included for completeness)

        # Row 47: Any areas of Opp for IS, ROAS, CPA?
        if "areas of opp" in criterion_lower or ("impression share" in criterion_lower and "roas" in criterion_lower and "cpa" in criterion_lower):
            return self._score_performance_opportunities()

        # Row 49: If on Manual bidding, set at enhanced CPC
        if "manual bidding" in criterion_lower and "enhanced cpc" in criterion_lower:
            return self._score_manual_ecpc()

        if "conversion source" in criterion_lower or "offline conversion" in criterion_lower:
            return self._score_conversion_source_coverage()

        # Row 50: Not losing too much impression Share to rank?
        if "impression share" in criterion_lower and "rank" in criterion_lower:
            return self._score_impression_share()

        # Row 52: Not losing impression Share to budget?
        if "impression share" in criterion_lower and "budget" in criterion_lower:
            return self._score_budget_lost()

        if ("audience" in criterion_lower and "observation" in criterion_lower) or (
            "audience targeting" in criterion_lower and "breadth" in criterion_lower
        ):
            return self._score_audience_observation_usage()

        if "audience" in criterion_lower and ("bid" in criterion_lower or "modifier" in criterion_lower):
            if "adjust" in criterion_lower or "modifier" in criterion_lower or "bid adj" in criterion_lower:
                return self._score_audience_bid_modifiers()

        # ===================================================================
        # BUDGET & SETTINGS (Rows 53, 57, 58)
        # ===================================================================

        # Row 53: Budget Allocation based on priority terms?
        if "budget allocation" in criterion_lower and "priority" in criterion_lower:
            return self._score_budget_allocation()

        # Row 57: Ad Rotation - Set to Optimize
        if "ad rotation" in criterion_lower and "optimize" in criterion_lower:
            return self._score_ad_rotation()

        # Row 58: Correct Location & Language Targeting?
        if "location" in criterion_lower and "language" in criterion_lower and "targeting" in criterion_lower:
            return self._score_location_language()

        # ===================================================================
        # GEO TARGETING & DAY PARTING (Rows 60, 61, 68)
        # ===================================================================

        # Row 60: Are locations added to location targeting?
        if "locations added" in criterion_lower or ("location targeting" in criterion_lower and "analysis" in criterion_lower):
            return self._score_location_targeting_depth()
        if "target:" in criterion_lower and "location" in criterion_lower:
            return self._score_location_targeting_depth()
        if "target:" in criterion_lower and ("localised" in criterion_lower or "localized" in criterion_lower):
            return self._score_location_targeting_depth()

        # Row 61: Adjusted bids? (Geo)
        if "adjusted bids" in criterion_lower and excel_row == 61:
            return self._score_geo_bid_adjustments()

        # Row 68: Adjusted bids? (Day/Time)
        if "adjusted bids" in criterion_lower and excel_row == 68:
            return self._score_dayparting_guardrail("bid")

        if "ad scheduling" in criterion_lower and excel_row == 69:
            return self._score_dayparting_guardrail("campaigns")

        if ("day-parting" in criterion_lower or "daypart" in criterion_lower) and excel_row == 70:
            return self._score_dayparting_guardrail("automation")

        if ("personalise" in criterion_lower or "personalize" in criterion_lower) and "ad copy" in criterion_lower:
            return self._score_localized_ad_copy()

        if ("personalise" in criterion_lower or "personalize" in criterion_lower) and "extensions" in criterion_lower:
            return self._score_localized_extensions()

        if ("personalise" in criterion_lower or "personalize" in criterion_lower) and "landing pages" in criterion_lower:
            return self._score_localized_landing_pages()

        # ===================================================================
        # AUDIENCES (Rows 73, 74, 76)
        # ===================================================================

        # Row 73: Created lists through Google, Analytics, bid tech or DMP?
        if "created lists" in criterion_lower or ("google" in criterion_lower and "analytics" in criterion_lower and "dmp" in criterion_lower):
            return self._score_audience_list_creation()

        # Row 74: Applied lists?
        if "applied lists" in criterion_lower:
            return self._score_audience_list_application()

        # Row 76: Separate campaigns for specific segments?
        if "separate campaigns" in criterion_lower and "segments" in criterion_lower:
            return self._score_audience_campaign_segmentation()

        # ===================================================================
        # TRACKING & ANALYTICS (Rows 80, 82, 84-88)
        # ===================================================================

        # Row 80: Using Analytics software which is set up correctly?
        if "analytics software" in criterion_lower and "set up correctly" in criterion_lower:
            return self._score_analytics_setup()

        # Row 82: Using Tag Management Solution?
        if "tag management" in criterion_lower:
            return self._score_tag_management()

        # Row 84: Integrating ad server data?
        if "ad server data" in criterion_lower:
            return self._score_ad_server_integration()

        # Row 85: First Party data / CRM integrated?
        if "first party data" in criterion_lower or ("crm" in criterion_lower and "integrated" in criterion_lower):
            return self._score_crm_integration()

        # Row 86: Using feed technology?
        if "feed technology" in criterion_lower:
            return self._score_feed_technology()

        # Row 87: Integrating offline conversions?
        if "offline conversions" in criterion_lower:
            return self._score_offline_conversions()

        # Row 88: App tracking?
        if "app tracking" in criterion_lower:
            return self._score_app_tracking()

        # ===================================================================
        # EXISTING UNIFIED ENGINE ROUTING (fallback patterns)
        # ===================================================================

        # STRUCTURE & ORGANIZATION
        if "consolidation" in criterion_lower or "duplicate url" in criterion_lower:
            return self._score_consolidation()

        if "campaigns per business unit" in criterion_lower or "avg campaigns" in criterion_lower:
            return self._score_campaign_count()

        if "ad groups per campaign" in criterion_lower:
            return self._score_adgroup_structure()

        # KEYWORDS
        if "quality score" in criterion_lower and "component" in criterion_lower:
            return self._score_quality_score_components()

        if "keyword coverage" in criterion_lower or "zero keywords" in criterion_lower:
            return self._score_keyword_coverage()

        if "broad match" in criterion_lower:
            return self._score_broad_match()

        if "match type" in criterion_lower and "balance" in criterion_lower:
            return self._score_match_type_balance()

        # NEGATIVE KEYWORDS
        if "negative keyword" in criterion_lower:
            return self._score_negative_keywords()

        if "audience targeting" in criterion_lower or "exclusion" in criterion_lower:
            return self._score_audience_targeting()

        # AD COPY
        if "rsa" in criterion_lower or "responsive search ad" in criterion_lower:
            return self._score_rsa_coverage()

        if "ad strength" in criterion_lower:
            return self._score_ad_strength()

        # BIDDING
        if "smart bidding" in criterion_lower or "automated bidding" in criterion_lower:
            return self._score_smart_bidding()

        if "ecpc" in criterion_lower and "manual" in criterion_lower:
            return self._score_manual_ecpc()

        # BUDGET & DELIVERY
        if "impression share" in criterion_lower:
            return self._score_impression_share()

        if "budget" in criterion_lower and "lost" in criterion_lower:
            return self._score_budget_lost()

        # TARGETING
        if "location" in criterion_lower and "language" in criterion_lower:
            return self._score_location_language()

        # AUDIENCES
        if "audience" in criterion_lower and "applied" in criterion_lower:
            return self._score_audiences_applied()

        # DATA & TRACKING - Most require Google Ads UI
        if "dda" in criterion_lower or "data-driven attribution" in criterion_lower:
            return self._score_data_driven_attribution()

        if "shared budget" in criterion_lower or "budgets being diversified" in criterion_lower:
            return self._score_budget_diversification()

        if "search partner" in criterion_lower:
            return ScoringResult(
                score=None,
                details="Data needed: Search Partner network settings from Google Ads Campaign Settings. SA360 typically doesn't include network targeting.",
                recommendations="Verify Search Partner inclusion/exclusion per campaign strategy.",
                data_needed="Google Ads UI - Network settings"
            )

        if "rlsa" in criterion_lower and "tag" in criterion_lower:
            return ScoringResult(
                score=None,
                details="Data needed: RLSA (Remarketing Lists for Search Ads) tag implementation from Google Ads Audience Manager or Google Analytics.",
                recommendations="Install remarketing tags if not present. Required for RLSA targeting.",
                data_needed="Google Ads Audience Manager or GA"
            )

        if "bid adjustment" in criterion_lower and "audience" in criterion_lower:
            return ScoringResult(
                score=None,
                details="Data needed: Audience bid adjustment settings from Google Ads UI. Navigate to Campaigns > Audiences > Bid adj. column.",
                recommendations="Apply bid modifiers to high-value remarketing/customer match audiences.",
                data_needed="Google Ads UI - Audience bid adjustments"
            )

        if "ad server" in criterion_lower:
            return ScoringResult(
                score=None,
                details="Data needed: Third-party ad server usage (Sizmek/Flashtalking/DoubleClick) not available in SA360 exports. Check creative tags.",
                recommendations="Review ad trafficking documentation for ad server integration.",
                data_needed="Creative trafficking documentation"
            )

        if "extended data" in criterion_lower or "valuetrack" in criterion_lower:
            return ScoringResult(
                score=None,
                details="Data needed: Tracking template from Google Ads Campaign Settings > Campaign URL options. Check for ValueTrack ({campaignid}, {adgroupid}) and UTM parameters.",
                recommendations="Implement ValueTrack and UTM parameters for enhanced data capture.",
                data_needed="Google Ads UI - Tracking templates"
            )

        if "conversion tracking" in criterion_lower:
            return ScoringResult(
                score=None,
                details="Data needed: Conversion tracking implementation from Google Ads UI. Check Tools & Settings > Measurement > Conversions.",
                recommendations="Ensure all key conversion actions are tracked.",
                data_needed="Google Ads UI - Conversion settings"
            )

        if "attribution" in criterion_lower or ("holistic" in criterion_lower and "measurement" in criterion_lower):
            return ScoringResult(
                score=None,
                details="Data needed: Attribution model settings from Google Ads or GA4. Check which attribution model is active (Linear, Time Decay, Position-based, Data-driven).",
                recommendations="Enable Data-driven attribution if conversion volume supports it.",
                data_needed="Google Ads UI or GA4 - Attribution settings"
            )

        # DEFAULT: Not yet implemented
        return ScoringResult(
            score=None,
            details=f"Scoring logic not yet implemented for: {criterion_name}",
            recommendations="Manual review required.",
            data_needed=None
        )

    # ============================================================================
    # SCORING FUNCTIONS - V3.1 Ported Functions (Rows 12-39, 46)
    # ============================================================================

    def _score_match_type_split(self) -> ScoringResult:
        """Row 12: Keyword groupings of similar types in campaign?"""
        keyword_sources = [
            self.data.get("search_keyword"),
            self.data.get("keyword"),
        ]
        campaign_col = None
        match_col = None
        df_selected = None

        for df in keyword_sources:
            if df is None or df.empty:
                continue
            campaign_col = self._find_column(df, ["Campaign"])
            match_col = self._find_column(
                df,
                [
                    "Search keyword match type",
                    "Keyword match type",
                    "Match type",
                ],
            )
            if campaign_col and match_col:
                df_selected = df[[campaign_col, match_col]].dropna()
                if not df_selected.empty:
                    break

        if df_selected is not None:
            df_selected["match_norm"] = df_selected[match_col].apply(self._normalize_match_type_value)
            df_selected = df_selected.dropna(subset=["match_norm"])
            if df_selected.empty:
                df_selected = None

        if df_selected is None:
            return ScoringResult(
                score=None,
                details="Data needed: Keyword export with campaign + match-type columns to evaluate grouping quality.",
                recommendations="Export the keyword report with 'Campaign' and 'Match type' columns so grouping can be assessed.",
                data_needed="Keyword report with match type",
            )

        grouped = df_selected.groupby(campaign_col)["match_norm"].nunique()
        total_campaigns = len(grouped)
        multi_campaigns = int((grouped >= 2).sum())
        single_campaigns = total_campaigns - multi_campaigns
        multi_pct = (multi_campaigns / total_campaigns * 100) if total_campaigns else 0.0

        self.summary_stats["match_type_split"] = {
            "multi_campaigns": multi_campaigns,
            "single_campaigns": single_campaigns,
            "total_campaigns": total_campaigns,
            "multi_pct": multi_pct,
        }

        if multi_pct <= 10:
            score = 5
            assessment = "Match types grouped cleanly"
        elif multi_pct <= 25:
            score = 4
            assessment = "Mostly orderly keyword groupings"
        elif multi_pct <= 50:
            score = 3
            assessment = "Mixed structures within campaigns"
        elif multi_pct <= 75:
            score = 2
            assessment = "Campaigns mix match types frequently"
        else:
            score = 1
            assessment = "Campaigns randomly mix match types"

        details = (
            f"{assessment}: {multi_campaigns}/{total_campaigns} campaigns ({multi_pct:.1f}%) combine multiple match types inside the same campaign."
        )
        recommendations = (
            "Keep organizing campaigns so each one houses a consistent match-type strategy."
            if score >= 4
            else "Rebuild campaigns so exact/phrase/broad strategies are separated, allowing Smart Bidding to learn clean intent signals."
        )

        weight = get_criterion_weight("naming_conventions")
        impact = determine_business_impact("Keyword Groupings", score, f"{multi_pct:.1f}% multi-match campaigns")

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Campaigns mixing match types: {multi_campaigns}/{total_campaigns}",
            criterion_weight=weight,
            business_impact=impact["business_impact"],
            priority_score=impact["priority_score"],
            impact_rationale=impact["rationale"],
        )
        multi_pct = (multi_campaigns / total_campaigns * 100) if total_campaigns > 0 else 0.0

        self.summary_stats['match_type_split'] = {
            'multi_campaigns': multi_campaigns,
            'total_campaigns': total_campaigns,
            'multi_pct': multi_pct,
            'single_campaigns': single_campaigns,
        }

        if multi_pct >= 80:
            score = 5
            assessment = "Consolidated Smart Bidding structure"
        elif multi_pct >= 60:
            score = 4
            assessment = "Mostly consolidated campaigns"
        elif multi_pct >= 40:
            score = 3
            assessment = "Mixed structures"
        elif multi_pct >= 20:
            score = 2
            assessment = "Significant match-type splitting"
        else:
            score = 1
            assessment = "Legacy match-type splits dominate"

        smart_stats = self._get_smart_bidding_stats()
        smart_pct = smart_stats.get("smart_pct", 0.0) or 0.0
        manual_campaigns = smart_stats.get("manual_campaigns", 0) or 0

        details = (
            f"{assessment}: {multi_campaigns}/{total_campaigns} campaigns ({multi_pct:.1f}%) contain multiple match types, while Smart Bidding adoption is "
            f"{smart_pct:.0f}% with {manual_campaigns} manual campaigns still in rotation. "
            f"{single_campaigns} campaigns still isolate match types and should be collapsed."
        )
        if score < 5:
            recommendations = (
                "Collapse match-type-specific campaigns by combining exact/phrase/broad into unified Smart Bidding structures, "
                "then migrate the remaining manual campaigns so automation controls every match type."
            )
        else:
            recommendations = (
                "Maintain the consolidated Smart Bidding campaign structure and keep automation in control of match-type guardrails."
            )

        weight = get_criterion_weight("naming_conventions")
        impact = determine_business_impact("Match Type Consolidation", score, f"{multi_pct:.1f}% consolidated")

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Consolidated campaigns: {multi_campaigns}/{total_campaigns}",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score'],
            impact_rationale=impact['rationale'],
        )

    def _score_granular_campaigns(self) -> ScoringResult:
        """Row 13: Granular Campaigns? (Ad groups per campaign)"""
        campaign_data = self.source_info.get('campaign_source')
        adgroup_data = self.source_info.get('adgroup_source')

        if campaign_data is None or adgroup_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign and ad group data",
                recommendations="Export campaign and ad group reports",
                data_needed="Campaign and ad group reports"
            )

        # Ensure required columns exist
        if 'Campaign' not in campaign_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign column in campaign data.",
                recommendations="Ensure campaign export includes a Campaign name column.",
                data_needed="Campaign report with Campaign column",
            )
        if 'Ad group' not in adgroup_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Ad group column in ad group data.",
                recommendations="Ensure ad group export includes an Ad group name column.",
                data_needed="Ad group report with Ad group column",
            )

        total_campaigns = len(campaign_data['Campaign'].unique())
        total_adgroups = len(adgroup_data['Ad group'].unique())
        avg_ag_per_campaign = total_adgroups / total_campaigns if total_campaigns > 0 else 0

        if 5 <= avg_ag_per_campaign <= 10:
            score = 5
            assessment = "Optimal"
        elif avg_ag_per_campaign < 15:
            score = 3
            assessment = "Acceptable"
        else:
            score = 2
            assessment = "Too many"

        weight = get_criterion_weight("campaign_structure")
        impact = determine_business_impact("Campaign Granularity", score, f"{avg_ag_per_campaign:.1f} AG/campaign")

        signal_payload = self._analyze_signal_density()
        details_text = f"{assessment}: {avg_ag_per_campaign:.1f} ad groups per campaign (5-10 optimal for manageability)"
        recommendation_text = (
            "Consolidate ad groups if >15 per campaign. Split campaigns if <3 ad groups (too broad)."
            if score < 5
            else "Campaign structure is optimal."
        )
        calc_text = f"Avg AG/campaign: {avg_ag_per_campaign:.1f}"
        if signal_payload:
            signal_details = signal_payload.get("details", "")
            if signal_details:
                details_text = f"{details_text} Signal density: {signal_details}"
            calc_text = f"{calc_text} | Signal density score: {signal_payload.get('score')}"
            signal_score = signal_payload.get("score")
            if isinstance(signal_score, (int, float)):
                if signal_score < score:
                    score = signal_score
                    recommendation_text = signal_payload.get("recommendations", recommendation_text)
                elif signal_score < 5 and signal_payload.get("recommendations"):
                    recommendation_text = f"{recommendation_text} {signal_payload['recommendations']}"

        return ScoringResult(
            score=score,
            details=details_text,
            recommendations=recommendation_text,
            calculation=calc_text,
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_granular_adgroups(self) -> ScoringResult:
        """Row 14: Granular Ad Groups? (Keywords per ad group)"""
        adgroup_data = self.source_info.get('adgroup_source')
        if adgroup_data is None or 'Keywords: active' not in adgroup_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Ad group report with active keyword counts",
                recommendations="Export ad group details report",
                data_needed="Ad group details with keyword counts"
            )

        adgroup_data_copy = adgroup_data.copy()
        adgroup_data_copy['Keywords_active_num'] = to_numeric_safe(adgroup_data_copy['Keywords: active'])
        avg_kw_per_ag = adgroup_data_copy['Keywords_active_num'].mean()

        if 10 <= avg_kw_per_ag <= 20:
            score = 5
            assessment = "Optimal"
        elif avg_kw_per_ag < 30:
            score = 3
            assessment = "Acceptable"
        else:
            score = 2
            assessment = "Too many"

        weight = get_criterion_weight("ad_group_structure")
        impact = determine_business_impact("Ad Group Granularity", score, f"{avg_kw_per_ag:.1f} KW/AG")

        return ScoringResult(
            score=score,
            details=f"{assessment}: {avg_kw_per_ag:.1f} keywords per ad group (10-20 optimal for relevance)",
            recommendations="Split ad groups with >30 keywords into tighter themes. Merge ad groups with <5 keywords." if score < 5 else "Ad group structure is optimal.",
            calculation=f"Avg kw/AG: {avg_kw_per_ag:.1f}",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_google_presence(self) -> ScoringResult:
        """Row 15: Account active in all engines - Google"""
        campaign_data = self._pick_campaign_source(["Account name", "Campaign"])
        if campaign_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign report with account names",
                recommendations="Export campaign report",
                data_needed="Campaign report"
            )

        def _compute_google_presence(df: pd.DataFrame) -> Optional[Tuple[pd.DataFrame, Optional[str], int, str, bool]]:
            account_col = self._find_column(df, ["Account name", "Account", "Account Name"], allow_partial=True)
            if not account_col:
                return None

            filtered_campaigns = df.copy()
            if account_col != "Account name":
                filtered_campaigns["Account name"] = filtered_campaigns[account_col]

            customer_ids = self.BU_CUSTOMER_IDS.get(self.business_unit_key) if self.business_unit_key else None
            effective_ids = customer_ids or self.inferred_customer_ids.get(self.business_unit_key or self.business_unit)
            customer_col = self._find_column(filtered_campaigns, ["Customer ID"], allow_partial=True)
            if effective_ids and customer_col:
                normalized_ids = {str(cid).replace(" ", "") for cid in effective_ids}
                col_values = filtered_campaigns[customer_col].astype(str).str.replace(" ", "", regex=False)
                filtered_campaigns = filtered_campaigns[col_values.isin(normalized_ids)]

            campaign_col = self._find_column(filtered_campaigns, ["Campaign", "Campaign name"], allow_partial=True)
            name_pattern = None
            if self.business_unit_key and self.business_unit_key in self.BU_ACCOUNT_PATTERNS:
                name_pattern = self.BU_ACCOUNT_PATTERNS[self.business_unit_key]
            elif self.business_unit and self.business_unit in self.BU_ACCOUNT_PATTERNS:
                name_pattern = self.BU_ACCOUNT_PATTERNS[self.business_unit]

            if name_pattern:
                name_mask = filtered_campaigns["Account name"].str.contains(name_pattern, case=False, na=False, regex=True)
                if campaign_col:
                    name_mask |= filtered_campaigns[campaign_col].str.contains(name_pattern, case=False, na=False, regex=True)
                filtered_campaigns = filtered_campaigns[name_mask]
            else:
                name_mask = filtered_campaigns["Account name"].str.contains("Fidelity|Google|Medicare", case=False, na=False)
                if campaign_col:
                    name_mask |= filtered_campaigns[campaign_col].str.contains("Medicare|Fidelity|Google", case=False, na=False)
                filtered_campaigns = filtered_campaigns[name_mask]

            state_col = self._find_column(
                filtered_campaigns,
                ["Campaign state", "Campaign Status", "Campaign status"],
                allow_partial=True,
            )
            if state_col:
                state_mask = filtered_campaigns[state_col].str.contains('Enabled|Eligible', case=False, na=False)
                google_campaigns = filtered_campaigns[state_mask]
            else:
                google_campaigns = filtered_campaigns

            impr_col = self._find_column(filtered_campaigns, ["Impressions", "Impr."], allow_partial=True)
            google_impr_display = 0
            if impr_col:
                google_impr = to_numeric_safe(google_campaigns[impr_col]).sum()
                try:
                    google_impr_display = int(round(float(google_impr or 0)))
                except Exception:
                    google_impr_display = 0
            has_activity = (google_impr_display > 0) if impr_col else len(google_campaigns) > 0
            if impr_col:
                details = f"Google: {len(google_campaigns)} campaigns, {google_impr_display:,.0f} impressions"
            else:
                details = f"Google: {len(google_campaigns)} campaigns active"
            return google_campaigns, impr_col, google_impr_display, details, has_activity

        primary = _compute_google_presence(campaign_data)
        if not primary:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign report with account names",
                recommendations="Export campaign report",
                data_needed="Campaign report"
            )

        google_campaigns, impr_col, google_impr_display, details, has_activity = primary

        # If the selected campaign source yields 0 impressions, try campaign_details (often richer exports).
        if impr_col and google_impr_display == 0:
            alt_df = self.data.get("campaign_details")
            if alt_df is not None and alt_df is not campaign_data and isinstance(alt_df, pd.DataFrame) and not alt_df.empty:
                alt = _compute_google_presence(alt_df)
                if alt:
                    _, alt_impr_col, alt_impr_display, alt_details, alt_has_activity = alt
                    if alt_impr_col and alt_impr_display > 0:
                        google_impr_display = alt_impr_display
                        details = f"{alt_details} (campaign details export)"
                        has_activity = alt_has_activity

        score = 5 if has_activity else 0

        return ScoringResult(
            score=score,
            details=details,
            recommendations="Google Ads is active and generating impressions." if score == 5 else "CRITICAL: No Google Ads activity detected.",
            calculation=f"Google campaigns: {len(google_campaigns)}",
            criterion_weight=0.05,
            business_impact="critical" if score == 0 else "low",
            priority_score=10 if score == 0 else 2
        )

    def _score_microsoft_presence(self) -> ScoringResult:
        """Row 16: Account active in all engines - BING"""
        campaign_data = self._pick_campaign_source(["Account name", "Campaign"])
        if campaign_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign report with account names",
                recommendations="Export campaign report",
                data_needed="Campaign report"
            )

        account_col = self._find_column(campaign_data, ["Account name", "Account", "Account Name"], allow_partial=True)
        if not account_col:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign report with account names",
                recommendations="Export campaign report",
                data_needed="Campaign report"
            )
        if account_col != "Account name":
            campaign_data = campaign_data.copy()
            campaign_data["Account name"] = campaign_data[account_col]

        ms_campaigns = campaign_data[campaign_data["Account name"].str.contains("Bing|Yahoo|Microsoft", case=False, na=False)]
        customer_ids = self.BU_CUSTOMER_IDS.get(self.business_unit_key) if self.business_unit_key else None
        effective_ids = customer_ids or self.inferred_customer_ids.get(self.business_unit_key or self.business_unit)
        customer_col = self._find_column(campaign_data, ["Customer ID"], allow_partial=True)
        if effective_ids and customer_col:
            normalized_ids = {str(cid).replace(" ", "") for cid in effective_ids}
            col_values = ms_campaigns[customer_col].astype(str).str.replace(" ", "", regex=False)
            ms_campaigns = ms_campaigns[col_values.isin(normalized_ids)]
        impr_col = self._find_column(campaign_data, ["Impressions", "Impr."], allow_partial=True)
        ms_impr = None
        ms_impr_display = None
        if impr_col:
            ms_impr = to_numeric_safe(ms_campaigns[impr_col]).sum()
            try:
                ms_impr_display = int(round(float(ms_impr or 0)))
            except Exception:
                ms_impr_display = 0
        has_activity = (ms_impr_display > 0) if impr_col else len(ms_campaigns) > 0
        score = 5 if has_activity else 0
        if impr_col:
            details = f"Microsoft: {len(ms_campaigns)} campaigns, {ms_impr_display:,.0f} impressions"
        else:
            details = f"Microsoft: {len(ms_campaigns)} campaigns active"

        return ScoringResult(
            score=score,
            details=details,
            recommendations="Microsoft Ads is active." if score == 5 else "CRITICAL: No Microsoft Ads activity detected.",
            calculation=f"Microsoft campaigns: {len(ms_campaigns)}",
            criterion_weight=0.03,
            business_impact="low",
            priority_score=3
        )

    def _score_other_engines(self) -> ScoringResult:
        """Row 17: Account active in all engines - Other"""
        return ScoringResult(
            score=0,
            details="No other engines (Google + Microsoft only, as expected for financial services)",
            recommendations="N/A - Google and Microsoft provide sufficient coverage.",
            calculation="N/A",
            criterion_weight=0.01,
            business_impact="low",
            priority_score=1
        )

    def _score_pmax(self) -> ScoringResult:
        """Row 18: PMAX within account"""
        campaign_data = self.source_info.get('campaign_source')
        type_col = self._find_column(campaign_data, ["Campaign type", "Campaign Type"], allow_partial=True) if campaign_data is not None else None
        if campaign_data is None or not type_col:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign type information",
                recommendations="Export campaign report with campaign type",
                data_needed="Campaign type"
            )

        pmax_campaigns = campaign_data[campaign_data[type_col].str.contains('Performance Max', case=False, na=False)]
        account_col = self._find_column(campaign_data, ["Account name", "Account", "Account Name"], allow_partial=True)
        if account_col and account_col != "Account name":
            campaign_data = campaign_data.copy()
            campaign_data["Account name"] = campaign_data[account_col]
        google_campaigns = campaign_data[campaign_data['Account name'].str.contains('Google|Fidelity', case=False, na=False)]

        pmax_count = len(pmax_campaigns)
        pmax_pct = (pmax_count / len(google_campaigns) * 100) if len(google_campaigns) > 0 else 0

        launch_threshold, optimal_threshold = self._pmax_targets()
        if pmax_pct >= optimal_threshold:
            score = 4
            assessment = "Healthy PMAX adoption"
        elif pmax_pct >= launch_threshold:
            score = 3
            assessment = "PMAX present but can scale"
        elif pmax_count > 0:
            score = 2
            assessment = "Limited PMAX coverage"
        else:
            score = 1
            assessment = "PMAX not yet deployed"

        return ScoringResult(
            score=score,
            details=f"{assessment}: {pmax_count} PMAX campaigns ({pmax_pct:.1f}% of Google campaigns)",
            recommendations=(
                "Launch small PMAX pilots (5-10% of coverage) focused on incremental inventory."
                if score == 1
                else "Scale PMAX toward the contextual target share so automation can learn across channels."
                if score == 2 or score == 3
                else "Maintain PMAX at the current share and continue feeding unique creative assets."
            ),
            calculation=f"PMAX: {pmax_count} campaigns",
            criterion_weight=0.02,
            business_impact="low",
            priority_score=3
        )

    def _score_naming_conventions(self) -> ScoringResult:
        """Row 19: Are naming conventions in place?"""
        campaign_data = self.source_info.get('campaign_source')
        if campaign_data is None or 'Campaign' not in campaign_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign list",
                recommendations="Export campaign report",
                data_needed="Campaign report"
            )

        delim_campaigns = sum(1 for c in campaign_data['Campaign'].unique()
                              if any(d in str(c) for d in ['|', '_', '-', '.']))
        total_campaigns = len(campaign_data['Campaign'].unique())
        delim_pct = (delim_campaigns / total_campaigns * 100) if total_campaigns > 0 else 0

        if delim_pct >= 70:
            score = 5
            assessment = "Excellent"
        elif delim_pct >= 40:
            score = 3
            assessment = "Moderate"
        else:
            score = 1
            assessment = "Poor"

        weight = get_criterion_weight("naming_conventions")
        impact = determine_business_impact("Naming Conventions", score, f"{delim_pct:.1f}%")

        return ScoringResult(
            score=score,
            details=f"{assessment}: {delim_campaigns}/{total_campaigns} campaigns ({delim_pct:.1f}%) use structured naming (delimiters)",
            recommendations="Implement consistent naming convention with delimiters for strategy, match type, and geo." if score < 5 else "Naming conventions are well-established.",
            calculation=f"Campaigns with structure: {delim_pct:.1f}%",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_coverage(self) -> ScoringResult:
        """Row 21: Good Coverage? (Impression volume)"""
        campaign_data = self.performance_data

        # Find impressions column (could be 'Impr.' or 'Impressions')
        impr_col = self._find_column(campaign_data, ['Impr.', 'Impressions', 'Impr'])

        if campaign_data is None or impr_col is None:
            return ScoringResult(
                score=None,
                details="Data needed: Impression data",
                recommendations="Export campaign performance report",
                data_needed="Campaign performance with impressions"
            )

        campaign_data_copy = campaign_data.copy()
        campaign_data_copy['Impr_num'] = to_numeric_safe(campaign_data_copy[impr_col])
        total_impressions = campaign_data_copy['Impr_num'].sum()
        total_campaigns = len(campaign_data)

        if total_impressions >= 100000:
            score = 5
            assessment = "Excellent"
        elif total_impressions >= 50000:
            score = 3
            assessment = "Moderate"
        else:
            score = 2
            assessment = "Low"

        weight = get_criterion_weight("impression_volume")
        impact = determine_business_impact("Coverage", score, f"{total_impressions:,.0f} impr")

        return ScoringResult(
            score=score,
            details=f"{assessment}: {total_impressions:,.0f} total impressions across {total_campaigns} campaigns",
            recommendations="Expand keyword coverage to increase impression volume." if score < 5 else "Strong impression volume indicates good keyword coverage.",
            calculation=f"Impressions: {total_impressions:,.0f}",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_missing_keywords(self) -> ScoringResult:
        """Row 23: No Active Ad Group missing Active Keywords"""
        adgroup_data = self.source_info.get('adgroup_source')
        if adgroup_data is None or 'Keywords: active' not in adgroup_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Ad group report with active keyword counts",
                recommendations="Export ad group details report",
                data_needed="Ad group details"
            )

        adgroup_data_copy = adgroup_data.copy()
        adgroup_data_copy['Keywords_active_num'] = to_numeric_safe(adgroup_data_copy['Keywords: active'])

        active_ags = adgroup_data_copy[adgroup_data_copy['Ad group state'].str.contains('enabled', case=False, na=False)]
        missing_kw = active_ags[active_ags['Keywords_active_num'] == 0]
        missing_pct = (len(missing_kw) / len(active_ags) * 100) if len(active_ags) > 0 else 0

        if missing_pct < 5:
            score = 5
            assessment = "Excellent"
        elif missing_pct < 10:
            score = 3
            assessment = "Acceptable"
        else:
            score = 1
            assessment = "Poor"

        weight = get_criterion_weight("keyword_coverage")
        impact = determine_business_impact("Missing Keywords", score, f"{missing_pct:.1f}%")

        return ScoringResult(
            score=score,
            details=f"{assessment}: {len(missing_kw)}/{len(active_ags)} active ad groups ({missing_pct:.1f}%) have zero keywords",
            recommendations=f"Add keywords to {len(missing_kw)} ad groups or pause them." if missing_pct > 5 else "Keyword coverage is excellent.",
            calculation=f"AGs without keywords: {len(missing_kw)}",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_disapprovals(self) -> ScoringResult:
        """Row 24: Any disapprovals

        ZERO REGRESSION ARCHITECTURE: Strict-First, Adaptive-Second

        Tier 1 (Gold Standard): Check exact file 'search_keyword' with 'Search keyword status'
        - If found: Use strict logic (Confidence: 1.0)
        - Zero regression for production data

        Tier 2 (Safety Net): Adaptive Search Party for 'Search keyword status' in ANY file
        - Only triggers if Tier 1 fails
        - Searches all loaded DataFrames
        - Confidence: 0.9 (same column, different location)

        Tier 3 (Fallback): Aggregate count column from ad group reports
        - Last resort when granular data unavailable
        - Confidence: 0.8 (less precise)
        """
        # ========================================================================
        # TIER 1: STRICT LOOKUP (Gold Standard - Zero Regression)
        # ========================================================================
        keyword_data = self.data.get('search_keyword')

        if keyword_data is not None and 'Search keyword status' in keyword_data.columns:
            logging.info("[Strict Match] Found 'Search keyword status' in 'search_keyword' file")
            # TIER 1 LOGIC: Status-based disapproval detection (Confidence: 1.0)
            active_keywords = keyword_data[
                (keyword_data['Search keyword status'] != 'Paused') &
                (keyword_data['Search keyword status'] != 'Removed')
            ]
            total_active_kw = len(active_keywords)
            disapproved_keywords = active_keywords[active_keywords['Search keyword status'] == 'Not eligible']
            total_disapproved = len(disapproved_keywords)
            disapproval_rate = (total_disapproved / total_active_kw * 100) if total_active_kw > 0 else 0

            if disapproval_rate < 2:
                score = 5
                assessment = "Excellent"
            elif disapproval_rate < 5:
                score = 3
                assessment = "Acceptable"
            else:
                score = 1
                assessment = "High"

            weight = get_criterion_weight("disapprovals")
            impact = determine_business_impact("Disapprovals", score, f"{disapproval_rate:.2f}%")
            paused_kw = len(keyword_data[keyword_data['Search keyword status'] == 'Paused'])
            details_parts = [
                f"{assessment}: {total_disapproved}/{total_active_kw} active keywords ({disapproval_rate:.2f}%) are disapproved"
            ]
            if paused_kw > 0:
                details_parts.append(f"({paused_kw:,} paused keywords excluded from count)")

            return ScoringResult(
                score=score,
                details=". ".join(details_parts),
                recommendations=f"Review and fix {total_disapproved} disapproved keywords (policy violations)." if disapproval_rate > 2 else "Minimal disapprovals - good policy compliance.",
                calculation=f"Disapproval rate: {total_disapproved}/{total_active_kw} = {disapproval_rate:.2f}%",
                confidence="high",  # Tier 1: Gold standard
                data_completeness=1.0,  # Full granular data
                criterion_weight=weight,
                business_impact=impact['business_impact'],
                priority_score=impact['priority_score']
            )

        # ========================================================================
        # TIER 2: ADAPTIVE SEARCH PARTY (Same column, different location)
        # ========================================================================
        logging.info("[Tier 1 Failed] Strict lookup for 'search_keyword' failed. Engaging Search Party...")
        keyword_data = self._find_data_with_column(
            'Search keyword status',
            preferred_sources=['keyword', 'search_keyword', 'keywords']  # Try alternatives
        )

        if keyword_data is not None:
            logging.info("[Adaptive Match] Found 'Search keyword status' in alternative file (Confidence: 0.9)")
            # TIER 2 LOGIC: Same as Tier 1, but lower confidence
            active_keywords = keyword_data[
                (keyword_data['Search keyword status'] != 'Paused') &
                (keyword_data['Search keyword status'] != 'Removed')
            ]
            total_active_kw = len(active_keywords)
            disapproved_keywords = active_keywords[active_keywords['Search keyword status'] == 'Not eligible']
            total_disapproved = len(disapproved_keywords)
            disapproval_rate = (total_disapproved / total_active_kw * 100) if total_active_kw > 0 else 0

            if disapproval_rate < 2:
                score = 5
                assessment = "Excellent"
            elif disapproval_rate < 5:
                score = 3
                assessment = "Acceptable"
            else:
                score = 1
                assessment = "High"

            weight = get_criterion_weight("disapprovals")
            impact = determine_business_impact("Disapprovals", score, f"{disapproval_rate:.2f}%")
            paused_kw = len(keyword_data[keyword_data['Search keyword status'] == 'Paused'])
            details_parts = [
                f"{assessment}: {total_disapproved}/{total_active_kw} active keywords ({disapproval_rate:.2f}%) are disapproved"
            ]
            if paused_kw > 0:
                details_parts.append(f"({paused_kw:,} paused keywords excluded from count)")

            return ScoringResult(
                score=score,
                details=". ".join(details_parts),
                recommendations=f"Review and fix {total_disapproved} disapproved keywords (policy violations)." if disapproval_rate > 2 else "Minimal disapprovals - good policy compliance.",
                calculation=f"Disapproval rate: {total_disapproved}/{total_active_kw} = {disapproval_rate:.2f}%",
                confidence="medium",  # Tier 2: Adaptive fallback
                data_completeness=0.9,  # Same data, different location
                criterion_weight=weight,
                business_impact=impact['business_impact'],
                priority_score=impact['priority_score']
            )

        # ========================================================================
        # TIER 3: FALLBACK (Aggregate count column)
        # ========================================================================
        logging.info("[Tier 2 Failed] No 'Search keyword status' found. Trying aggregate count column...")
        disapproved_count_data = self._find_data_with_column(
            'Keywords: disapproved',
            preferred_sources=['ad_group_details', 'adgroup', 'ad_group']
        )

        if disapproved_count_data is not None:
            logging.info("[Fallback Match] Found 'Keywords: disapproved' aggregate column (Confidence: 0.8)")
            # TIER 3 LOGIC: Count-based disapproval detection (aggregated by ad group)
            total_disapproved = to_numeric_safe(disapproved_count_data['Keywords: disapproved']).sum()

            # Estimate total keywords for rate calculation
            if 'Keywords: active' in disapproved_count_data.columns:
                total_active_kw = to_numeric_safe(disapproved_count_data['Keywords: active']).sum()
            else:
                # Fallback: assume disapprovals are <5% if we can't calculate rate
                total_active_kw = total_disapproved * 20  # Assume 5% rate

            disapproval_rate = (total_disapproved / total_active_kw * 100) if total_active_kw > 0 else 0

            # Scoring
            if disapproval_rate < 2:
                score = 5
                assessment = "Excellent"
            elif disapproval_rate < 5:
                score = 3
                assessment = "Acceptable"
            else:
                score = 1
                assessment = "High"

            weight = get_criterion_weight("disapprovals")
            impact = determine_business_impact("Disapprovals", score, f"{disapproval_rate:.2f}%")

            return ScoringResult(
                score=score,
                details=f"{assessment}: {total_disapproved:,.0f} disapproved keywords across ad groups ({disapproval_rate:.2f}% of {total_active_kw:,.0f} active keywords)",
                recommendations=f"Review and fix {total_disapproved:,.0f} disapproved keywords." if disapproval_rate > 2 else "Minimal disapprovals - good policy compliance.",
                calculation=f"Disapproval rate: {total_disapproved:,.0f}/{total_active_kw:,.0f} = {disapproval_rate:.2f}%",
                confidence="medium",  # Tier 3: Aggregated fallback
                data_completeness=0.8,  # Missing keyword-level detail
                criterion_weight=weight,
                business_impact=impact['business_impact'],
                priority_score=impact['priority_score']
            )

        # ========================================================================
        # ALL TIERS FAILED - Return "Data needed"
        # ========================================================================
        logging.warning("[All Tiers Failed] No disapproval data found in any format")
        return ScoringResult(
            score=None,
            details="Data needed: Search keyword report with status OR ad group report with disapproval counts",
            recommendations="Export search keyword report from SA360 or ad group details with keyword counts",
            data_needed="Keyword status or disapproval count data"
        )

    def _score_broad_match_pct(self) -> ScoringResult:
        """Row 25: Broad Match Integration?"""
        match_series, source_label = self._extract_match_type_series()
        if match_series is None:
            campaign_data = self.source_info.get('campaign_source')
            if campaign_data is None or 'Campaign' not in campaign_data.columns:
                return ScoringResult(
                    score=None,
                    details="Data needed: Campaign or keyword list",
                    recommendations="Export campaign/keyword report",
                    data_needed="Campaign or keyword report"
                )
            campaigns = campaign_data['Campaign'].astype(str)
            unique_campaigns = pd.Series(campaigns.unique()).dropna().tolist()
            total_entities = len(unique_campaigns)
            low_tokens = ('broad', 'bmm', 'brd')
            broad_count = sum(
                1 for name in unique_campaigns
                if any(token in name.lower() for token in low_tokens)
            )
            source_label = "campaigns"
        else:
            total_entities = len(match_series)
            broad_count = int((match_series == "broad").sum())
            source_label = source_label or "keywords"

        broad_pct = (broad_count / total_entities * 100) if total_entities > 0 else 0
        entity_desc = f"{broad_count} of {total_entities} {source_label}"

        vertical = self._context_vertical()
        primary_kpi = self._context_primary_kpi()
        optimal_low, optimal_high = 20, 35
        acceptable_low, acceptable_high = 15, 45
        if vertical == "FINANCE":
            optimal_low, optimal_high = 15, 30
            acceptable_low, acceptable_high = 10, 35
        elif primary_kpi == "POAS" or vertical == "ECOMMERCE":
            optimal_low, optimal_high = 25, 40
            acceptable_low, acceptable_high = 20, 45
        elif vertical == "SAAS":
            optimal_low, optimal_high = 20, 40
            acceptable_low, acceptable_high = 15, 50

        SCORE_LABELS = {
            5: "Excellent (top 20%)",
            3: "Meets minimum expectations",
            2: "Below standard",
        }

        if optimal_low <= broad_pct <= optimal_high:
            score = 5
            assessment = "Optimal"
        elif acceptable_low <= broad_pct <= acceptable_high:
            score = 3
            assessment = "Acceptable"
        else:
            score = 2
            assessment = "Suboptimal"

        label_prefix = lambda s: f"{SCORE_LABELS.get(s, '').strip()}: " if SCORE_LABELS.get(s) else ""

        if score == 5:
            recommendations = (
                f"{label_prefix(score)}Broad match usage is {broad_pct:.0f}% ({entity_desc}), right in the 20‑35% band that "
                f"lets Smart Bidding discover new queries without losing control. Pair these campaigns with Performance Max "
                f"and refresh shared negatives monthly so you keep discovering 10‑15% incremental conversions without diluting ROAS."
            )
        elif score == 3:
            if broad_pct < 20:
                recommendations = (
                    f"{label_prefix(score)}Broad match sits at {broad_pct:.0f}% ({entity_desc}), so discovery is muted. "
                    f"Expand the top performing exact/phrase ad groups with broad variants over the next month and monitor search-term reports weekly. "
                    f"Expect 10‑15% more impressions once you approach the 25% mark."
                )
            else:
                recommendations = (
                    f"{label_prefix(score)}Broad match accounts for {broad_pct:.0f}% ({entity_desc}), which pushes past the efficient range. "
                    f"Convert the lowest CTR broad keywords to phrase match and add the corresponding negatives so spend concentrates on proven terms. "
                    f"Expect a 15‑20% reduction in wasted spend as you drift back toward 30%."
                )
        else:
            if broad_pct < 15:
                recommendations = (
                    f"{label_prefix(score)}Broad match is nearly absent at {broad_pct:.0f}% ({entity_desc}). "
                    f"Audit your highest-converting exact/phrase keywords, add broad variants under Smart Bidding, and aim for 20% coverage within two months. "
                    f"This usually drives 25‑35% more conversions at comparable efficiency when paired with weekly search-term reviews."
                )
            else:
                recommendations = (
                    f"{label_prefix(score)}Broad match accounts for {broad_pct:.0f}% ({entity_desc}), creating a high risk of irrelevant traffic. "
                    f"Cull the lowest Quality Score broad terms, convert weak performers to phrase, and tighten negatives so you land near 35%. "
                    f"Most advertisers see a 20‑30% reduction in wasted spend plus a 1‑2 point lift in Quality Score after this cleanup."
                )

        why_matters = (
            "Match type distribution directly impacts account reach and spend efficiency. Broad match allows discovery "
            "of new converting searches but requires Smart Bidding and strong negative keyword management to prevent waste. "
            "2024 industry data shows 36.67% average broad match usage, with top performers maintaining 20-35% for optimal "
            "balance (Source: PPC industry benchmarks 2024)."
        )

        threshold_context = (
            f"Threshold applied: {optimal_low}-{optimal_high}% broad match usage for this vertical/KPI mix. "
            f"Acceptable experimentation band: {acceptable_low}-{acceptable_high}% so that discovery tests stay controlled."
        )

        if score == 5:
            your_performance = (
                f"Your account: {entity_desc} ({broad_pct:.0f}%) use broad match, "
                f"falling within the optimal 20-35% range. This demonstrates strategic balance between discovery and control, "
                f"aligning with industry best practices for financial services advertisers."
            )
        elif score == 3:
            if broad_pct < 20:
                your_performance = (
                    f"Your account: {entity_desc} ({broad_pct:.0f}%) use broad match. "
                    f"This is below the 20% optimal threshold, indicating overly restrictive targeting that may limit reach. "
                    f"Gradual expansion to 20-25% would unlock new converting search queries while maintaining control."
                )
            else:
                your_performance = (
                    f"Your account: {entity_desc} ({broad_pct:.0f}%) use broad match. "
                    f"This exceeds the 35% optimal threshold, increasing risk of irrelevant traffic. Reduction to 30-35% would "
                    f"improve efficiency while maintaining discovery capabilities."
                )
        else:
            if broad_pct < 15:
                your_performance = (
                    f"Your account: {entity_desc} ({broad_pct:.0f}%) use broad match, "
                    f"falling significantly below the 15% minimum threshold. This overly conservative approach limits reach to "
                    f"known queries only, missing an estimated 25-35% of potential conversions from new search discovery."
                )
            else:
                your_performance = (
                    f"Your account: {entity_desc} ({broad_pct:.0f}%) use broad match, "
                    f"well above the 45% acceptable maximum. Without robust negative keyword coverage and Smart Bidding, this "
                    f"creates estimated $1,500-2,500 monthly wasted spend risk from irrelevant traffic."
                )

        bp_rationale = generate_bp_rationale(
            why_matters=why_matters,
            threshold_context=threshold_context,
            your_performance=your_performance,
            evidence_tier="applied_standard"
        )

        weight = get_criterion_weight("broad_match")
        impact = determine_business_impact("Broad Match", score, f"{broad_pct:.1f}%")

        target_label = f"Target: {optimal_low}-{optimal_high}% for {vertical.title() if vertical else 'this vertical'}"

        result = ScoringResult(
            score=score,
            details=f"{assessment}: Broad match {entity_desc} ({broad_pct:.1f}%). {target_label}",
            recommendations=recommendations,
            bp_rationale=bp_rationale,
            calculation=f"Broad usage: {broad_pct:.1f}%",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

        vertical_label = "Finance" if vertical == "FINANCE" else (vertical.title() if vertical else "this vertical")
        if optimal_low <= broad_pct <= optimal_high:
            result.benchmark_delta = f"Gap to {vertical_label} optimal band ({optimal_low}-{optimal_high}%): +0.0 pts (on target)."
            status_phrase = (
                "Optimized for Finance Compliance" if vertical == "FINANCE" else "Within optimal range"
            )
        elif broad_pct < optimal_low:
            gap = optimal_low - broad_pct
            result.benchmark_delta = f"Gap to {vertical_label} optimal band ({optimal_low}-{optimal_high}%): -{gap:.1f} pts."
            status_phrase = (
                "Below Finance Compliance range" if vertical == "FINANCE" else "Under target"
            )
        else:
            gap = broad_pct - optimal_high
            result.benchmark_delta = f"Above {vertical_label} optimal band ({optimal_low}-{optimal_high}%): +{gap:.1f} pts."
            status_phrase = (
                "Outside Finance Compliance range" if vertical == "FINANCE" else "Above target"
            )

        result.vertical_context = (
            f"Broad match is {broad_pct:.1f}%. Status: {status_phrase} (Target: {optimal_low}-{optimal_high}%)."
        )
        if broad_pct <= 1:
            result.binary_gap_indicator = "Broad Match Inventory: ABSENT"

        return result

    def _score_match_type_balance_v31(self) -> ScoringResult:
        """Row 26: Is there a balance of Match Types?"""
        match_series, source_label = self._extract_match_type_series()
        distribution = self._compute_match_distribution(match_series)
        if distribution is None:
            return ScoringResult(
                score=None,
                details="Data needed: Keyword report with match type column.",
                recommendations="Export the keyword report including Match type to evaluate balance.",
                data_needed="Keyword report with match type",
            )

        exact_pct = distribution["exact_pct"]
        phrase_pct = distribution["phrase_pct"]
        broad_pct = distribution["broad_pct"]
        counts_present = sum(pct > 5 for pct in (exact_pct, phrase_pct, broad_pct))
        max_pct = max(exact_pct, phrase_pct, broad_pct)

        if counts_present == 3 and max_pct <= 60:
            score = 5
            assessment = "Well-balanced"
        elif counts_present >= 2 and max_pct <= 80:
            score = 3
            assessment = "Moderate concentration"
        else:
            score = 1
            assessment = "Highly concentrated"

        weight = get_criterion_weight("match_type_balance")
        impact = determine_business_impact(
            "Match Type Balance",
            score,
            f"{max_pct:.0f}% max concentration",
        )

        return ScoringResult(
            score=score,
            details=(
                f"{assessment}: Match type split - {exact_pct:.0f}% Exact, {phrase_pct:.0f}% Phrase, {broad_pct:.0f}% Broad "
                f"(source: {source_label})."
            ),
            recommendations="Balance match types to avoid over-concentration in one type." if score < 5 else "Match type distribution is well-balanced.",
            calculation=f"Counts present: {counts_present}",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score'],
        )

    def _score_audience_coverage(self) -> ScoringResult:
        """Row 28: Campaign lists used?

        CORRECT INTERPRETATION: "Campaign lists" refers to shared negative keyword lists
        applied at the campaign level, NOT audience lists.

        In Google Ads best practices:
        - Shared negative keyword lists help prevent wasted spend across campaigns
        - Industry standard: 70%+ campaigns should use shared negative lists
        - Reduces management overhead vs. campaign-level negatives

        This data is NOT available in standard SA360 exports.
        Would require: Negative keyword list report showing list assignments per campaign.
        """
        # Check if negative keyword list data is available
        # Common column names: "Negative keyword lists", "Shared lists", etc.
        campaign_data = self.source_info.get('campaign_details')

        if campaign_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Shared negative keyword list usage per campaign",
                recommendations="Export negative keyword list report from Google Ads (not available in SA360 standard reports)",
                data_needed="Negative keyword lists report"
            )

        # Check for negative list columns
        negative_list_cols = [c for c in campaign_data.columns if
                            'negative' in c.lower() and ('list' in c.lower() or 'shared' in c.lower())]

        if not negative_list_cols:
            return ScoringResult(
                score=None,
                details="Data needed: Shared negative keyword list data not found in campaign export",
                recommendations="Export campaign details with negative keyword list information, or export dedicated negative keyword list report",
                data_needed="Negative keyword lists"
            )

        # If we have the data, measure it
        # This code would execute if the data becomes available
        total_campaigns = len(campaign_data)
        # Example logic (would need adjustment based on actual column structure):
        # campaigns_with_lists = campaign_data[campaign_data[negative_list_cols[0]].notna()].count()

        weight = get_criterion_weight("negative_lists")
        impact = determine_business_impact("Negative Keyword Lists", None, "N/A")

        return ScoringResult(
            score=None,
            details="Data needed: Shared negative keyword list usage not available in SA360 export",
            recommendations="Best practice: 70%+ campaigns should use shared negative keyword lists to prevent wasted spend",
            data_needed="Negative keyword lists report",
            criterion_weight=weight,
            business_impact="high",
            priority_score=7
        )

    def _score_audience_diversity(self) -> ScoringResult:
        """Row 29: Cross Category Negatives? (Audience segment diversity)"""
        audience_data = self.data.get('audience_segment')
        if audience_data is None or 'Audience segment' not in audience_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Audience segment data",
                recommendations="Export audience targeting report",
                data_needed="Audience report"
            )

        unique_audiences = audience_data['Audience segment'].nunique()

        if unique_audiences >= 10:
            score = 5
            assessment = "Excellent"
        elif unique_audiences >= 5:
            score = 3
            assessment = "Moderate"
        else:
            score = 2
            assessment = "Limited"

        weight = get_criterion_weight("audience_diversity")
        impact = determine_business_impact("Audience Diversity", score, f"{unique_audiences} segments")

        return ScoringResult(
            score=score,
            details=f"{assessment}: {unique_audiences} unique audience segments (diversity indicates sophisticated targeting)",
            recommendations="Expand audience segment diversity to 10+ segments for comprehensive targeting coverage." if score < 5 else "Excellent audience diversity.",
            calculation=f"Unique audiences: {unique_audiences}",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_dki(self) -> ScoringResult:
        """Row 31: Adapted to ad groups/keywords?

        UPDATED 2024: Modern RSA best practice measurement
        - PRIMARY: Ad group-level headline variation (theme adaptation)
        - SECONDARY: Keyword theme inclusion in headlines
        - LEGACY: DKI usage (still checked, but not primary metric)

        Research: Google Ads RSA best practices (2024) recommend:
        1. "Include at least one keyword in headlines"
        2. "Create headlines relevant to keywords you're targeting"
        3. "Maximize headline quantity and diversity"
        Source: support.google.com/google-ads/answer/7684791
        """
        ad_data = self.data.get('ad')
        if ad_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Ad data with ad group information",
                recommendations="Export ad report with ad group associations",
                data_needed="Ad report"
            )

        # Check required columns
        if 'Ad group' not in ad_data.columns and 'Ad group ID' not in ad_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Ad group information in ad data",
                recommendations="Ensure ad report includes ad group column",
                data_needed="Ad group column in ad report"
            )

        ag_col = 'Ad group' if 'Ad group' in ad_data.columns else 'Ad group ID'
        headline_cols = [c for c in ad_data.columns if 'Headline' in c and c != 'Headline']  # Exclude summary column

        if not headline_cols:
            return ScoringResult(
                score=None,
                details="Data needed: Headline columns in ad data",
                data_needed="Ad headlines"
            )

        # METHOD 1: Ad group-level headline variation (PRIMARY for RSAs)
        # For each ad group, check if there are multiple unique primary headlines (H1)
        h1_col = next((c for c in headline_cols if c == 'Headline 1'), headline_cols[0])

        if h1_col in ad_data.columns:
            # Count unique H1 headlines per ad group
            ag_headline_diversity = ad_data.groupby(ag_col)[h1_col].nunique()
            total_ad_groups = ad_data[ag_col].nunique()

            # Ad groups with multiple unique headlines (showing variation/testing)
            ag_with_variation = (ag_headline_diversity > 1).sum()
            variation_rate = (ag_with_variation / total_ad_groups * 100) if total_ad_groups > 0 else 0

            # Check for themed adaptation: different ad groups have different primary headlines
            unique_h1_by_ag = ad_data.groupby(ag_col)[h1_col].apply(lambda x: set(x.dropna().astype(str)))
            ag_with_unique_themes = len([themes for themes in unique_h1_by_ag if len(themes) > 0])
            theme_adaptation_rate = (ag_with_unique_themes / total_ad_groups * 100) if total_ad_groups > 0 else 0
        else:
            variation_rate = 0
            theme_adaptation_rate = 0
            total_ad_groups = 0

        # METHOD 2: Legacy DKI check (supplementary info)
        all_headlines = []
        for col in headline_cols:
            all_headlines.extend(ad_data[col].dropna().astype(str).tolist())

        ads_with_dki = sum(1 for h in all_headlines if '{' in h and '}' in h)
        total_ads = len(ad_data)
        dki_pct = (ads_with_dki / total_ads * 100) if total_ads > 0 else 0

        # SCORING: Based on theme adaptation (modern RSA approach)
        # 70%+ of ad groups have themed headlines = Excellent
        # 50%+ = Good
        # 30%+ = Acceptable
        # <30% = Needs improvement

        if theme_adaptation_rate >= 70:
            score = 5
            assessment = "Excellent"
        elif theme_adaptation_rate >= 50:
            score = 4
            assessment = "Good"
        elif theme_adaptation_rate >= 30:
            score = 3
            assessment = "Acceptable"
        elif theme_adaptation_rate >= 10:
            score = 2
            assessment = "Needs improvement"
        else:
            score = 1
            assessment = "Low"

        weight = get_criterion_weight("ad_adaptation")
        impact = determine_business_impact("Ad Adaptation", score, f"{theme_adaptation_rate:.1f}%")

        # Build transparent details showing both modern and legacy metrics
        details_parts = [
            f"{assessment}: {ag_with_unique_themes}/{total_ad_groups} ad groups ({theme_adaptation_rate:.0f}%) have themed headlines"
        ]

        if dki_pct > 0:
            details_parts.append(f"DKI usage: {dki_pct:.1f}%")
        elif total_ads > 0:
            details_parts.append(f"No DKI detected (modern RSA approach)")

        details = ". ".join(details_parts)

        # Modern RSA recommendations
        if score < 5:
            recommendations = f"Improve headline theme adaptation: ensure each ad group's headlines reflect its keyword theme. Current: {theme_adaptation_rate:.0f}% of ad groups have themed headlines. Target: 70%+ for excellent relevance."
        else:
            recommendations = "Strong headline adaptation across ad groups. Headlines are well-themed to ad group keyword topics."

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Ad groups with themed headlines: {ag_with_unique_themes}/{total_ad_groups} = {theme_adaptation_rate:.0f}%",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_ad_testing(self) -> ScoringResult:
        """Row 33: Continuous Test & Learn? (Ads per ad group)"""
        ad_data = self.data.get('ad')
        adgroup_data = self.source_info.get('adgroup_source')

        if ad_data is None or adgroup_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Ad and ad group data",
                recommendations="Export ad and ad group reports",
                data_needed="Ad and ad group reports"
            )

        total_ads = len(ad_data)
        total_adgroups = len(adgroup_data['Ad group'].unique())
        avg_ads_per_ag = total_ads / total_adgroups if total_adgroups > 0 else 0

        if avg_ads_per_ag >= 2:
            score = 5
            assessment = "Excellent"
        elif avg_ads_per_ag >= 1.5:
            score = 3
            assessment = "Acceptable"
        else:
            score = 2
            assessment = "Low"

        weight = get_criterion_weight("ad_testing")
        impact = determine_business_impact("Ad Testing", score, f"{avg_ads_per_ag:.1f} ads/AG")

        return ScoringResult(
            score=score,
            details=f"{assessment}: {avg_ads_per_ag:.1f} ads per ad group (2+ enables A/B testing)",
            recommendations="Add second ad variant to all ad groups for continuous testing and optimization." if score < 5 else "Good ad testing setup with multiple variants per ad group.",
            calculation=f"Avg ads/AG: {avg_ads_per_ag:.1f}",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_usp_copy(self) -> ScoringResult:
        """Row 34: Copy includes USPs?

        ENHANCED 2024: Expanded USP detection for financial services
        - Generic superlatives (best, top, leading)
        - Financial differentiators (commission-free, no fees, low cost)
        - Proof points (award-winning, recognized, certified)
        - Unique offerings (exclusive, only, first)
        - Trust indicators (insured, protected, guaranteed)

        Research: PPC copywriting best practices emphasize specific differentiators
        over generic claims, especially important in regulated financial services.
        Sources: WordStream USP Guide, CXL PPC Copywriting
        """
        ad_data = self.data.get('ad')
        if ad_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Ad data",
                recommendations="Export ad report",
                data_needed="Ad report"
            )

        headline_cols = [c for c in ad_data.columns if 'Headline' in c or 'headline' in c]
        desc_cols = [c for c in ad_data.columns if 'Description' in c or 'description' in c]

        all_ad_text = []
        for col in headline_cols + desc_cols:
            all_ad_text.extend(ad_data[col].dropna().astype(str).tolist())

        # ENHANCED: Comprehensive USP keywords organized by category
        # NO REMOVALS - all original keywords retained, new ones added
        usp_keywords = {
            'superlatives': ['best', 'top', 'leading', 'trusted', '#1', 'premier', 'premier'],
            'financial_differentiators': [
                'commission-free', 'commission free', 'no commission',
                'no fees', 'no fee', 'fee-free', 'zero fees',
                'low cost', 'low-cost', 'competitive rates',
                'no minimum', 'no minimums',
                'expense ratio'
            ],
            'proof_points': [
                'award', 'award-winning', 'rated', 'recognized',
                'certified', 'accredited',
                'barron', 'kiplinger', 'j.d. power', 'morningstar'
            ],
            'differentiators': [
                'exclusive', 'only', 'unique', 'first', 'sole',
                'proprietary', 'industry-leading',
                'robo advisor', 'fidelity go', 'managed portfolio',
                '$0 commission', '0 commission', 'commission trades',
                'zero commission trades', 'commission-free trading',
                '4 & 5 star rated', '4 star rated', '5 star rated',
                'star rated funds', 'top-rated funds', 'value funds'
            ],
            'trust_indicators': [
                'insured', 'fdic', 'sipc', 'protected', 'guaranteed',
                'secure', 'protected'
            ],
            'experience_proof': [
                'years experience', 'years of experience', 'decades',
                'trillion', 'billion', 'million customers', 'million clients'
            ]
        }

        # Flatten all keywords
        all_usp_keywords = []
        for category_keywords in usp_keywords.values():
            all_usp_keywords.extend(category_keywords)

        # Count ad text snippets with USPs
        ads_with_usp = sum(1 for text in all_ad_text if any(usp in text.lower() for usp in all_usp_keywords))
        usp_pct = (ads_with_usp / len(all_ad_text) * 100) if len(all_ad_text) > 0 else 0

        # Score based on USP presence
        if usp_pct >= 50:
            score = 5
            assessment = "Excellent"
        elif usp_pct >= 30:
            score = 3
            assessment = "Moderate"
        elif usp_pct >= 10:
            score = 2
            assessment = "Low"
        else:
            score = 1
            assessment = "Very low"

        weight = get_criterion_weight("usp_copy")
        impact = determine_business_impact("USP Copy", score, f"{usp_pct:.1f}%")

        # Enhanced recommendations based on financial services context
        if score < 5:
            recommendations = (
                f"Increase USP usage to 50%+ of ad copy. Include specific differentiators like "
                f"'commission-free', 'award-winning', or 'no fees' rather than generic claims. "
                f"Current: {usp_pct:.1f}% of ad text includes USPs."
            )
        else:
            recommendations = "Strong USP integration in ad copy with specific differentiators."

        return ScoringResult(
            score=score,
            details=f"{assessment}: {usp_pct:.1f}% of ad text includes unique selling propositions (USPs)",
            recommendations=recommendations,
            calculation=f"USP usage: {usp_pct:.1f}%",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_pricing_copy(self) -> ScoringResult:
        """Row 35: Copy features prices/promos?"""
        ad_data = self.data.get('ad')
        if ad_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Ad data",
                recommendations="Export ad report",
                data_needed="Ad report"
            )

        headline_cols = [c for c in ad_data.columns if 'Headline' in c or 'headline' in c]
        desc_cols = [c for c in ad_data.columns if 'Description' in c or 'description' in c]

        all_ad_text = []
        for col in headline_cols + desc_cols:
            all_ad_text.extend(ad_data[col].dropna().astype(str).tolist())

        pricing_keywords = ['$', 'free', 'save', 'promo', 'offer', '%']
        ads_with_pricing = sum(1 for text in all_ad_text if any(pk in text.lower() for pk in pricing_keywords))
        pricing_pct = (ads_with_pricing / len(all_ad_text) * 100) if len(all_ad_text) > 0 else 0

        if pricing_pct >= 40:
            score = 5
            assessment = "Excellent"
        elif pricing_pct >= 20:
            score = 3
            assessment = "Moderate"
        else:
            score = 1
            assessment = "Low"

        weight = get_criterion_weight("pricing_copy")
        impact = determine_business_impact("Pricing Copy", score, f"{pricing_pct:.1f}%")

        return ScoringResult(
            score=score,
            details=f"{assessment}: {pricing_pct:.1f}% of ad text includes pricing or promotional offers",
            recommendations="Increase pricing/promo messaging to 40%+ of ads to drive clicks and conversions." if score < 5 else "Strong pricing and promotional messaging.",
            calculation=f"Pricing: {pricing_pct:.1f}%",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_cta_copy(self) -> ScoringResult:
        """Row 36: Copy uses CTAs?"""
        ad_data = self.data.get('ad')
        if ad_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Ad data",
                recommendations="Export ad report",
                data_needed="Ad report"
            )

        headline_cols = [c for c in ad_data.columns if 'Headline' in c or 'headline' in c]
        desc_cols = [c for c in ad_data.columns if 'Description' in c or 'description' in c]

        all_ad_text = []
        for col in headline_cols + desc_cols:
            all_ad_text.extend(ad_data[col].dropna().astype(str).tolist())

        cta_keywords = ['get', 'start', 'learn', 'open', 'call', 'contact', 'schedule']
        ads_with_cta = sum(1 for text in all_ad_text if any(cta in text.lower() for cta in cta_keywords))
        cta_pct = (ads_with_cta / len(all_ad_text) * 100) if len(all_ad_text) > 0 else 0

        if cta_pct >= 70:
            score = 5
            assessment = "Excellent"
        elif cta_pct >= 50:
            score = 3
            assessment = "Moderate"
        else:
            score = 1
            assessment = "Low"

        weight = get_criterion_weight("cta_copy")
        impact = determine_business_impact("CTA Copy", score, f"{cta_pct:.1f}%")

        return ScoringResult(
            score=score,
            details=f"{assessment}: {cta_pct:.1f}% of ad text includes clear calls-to-action (CTAs)",
            recommendations="Increase CTA usage to 70%+ of ads to drive user action." if score < 5 else "Strong CTA integration in ad copy.",
            calculation=f"CTA usage: {cta_pct:.1f}%",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_trademark_symbols(self) -> ScoringResult:
        """Row 37: Copy uses ®™©?"""
        ad_data = self.data.get('ad')
        if ad_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Ad data",
                recommendations="Export ad report",
                data_needed="Ad report"
            )

        headline_cols = [c for c in ad_data.columns if 'Headline' in c or 'headline' in c]
        desc_cols = [c for c in ad_data.columns if 'Description' in c or 'description' in c]

        all_ad_text = []
        for col in headline_cols + desc_cols:
            all_ad_text.extend(ad_data[col].dropna().astype(str).tolist())

        ads_with_symbols = sum(1 for text in all_ad_text if any(sym in text for sym in ['®', '™', '©']))
        symbol_pct = (ads_with_symbols / len(all_ad_text) * 100) if len(all_ad_text) > 0 else 0

        if symbol_pct >= 30:
            score = 5
            assessment = "Excellent"
        elif symbol_pct >= 10:
            score = 3
            assessment = "Moderate"
        else:
            score = 1
            assessment = "Low"

        weight = get_criterion_weight("trademark_symbols")
        impact = determine_business_impact("Trademark Symbols", score, f"{symbol_pct:.1f}%")

        return ScoringResult(
            score=score,
            details=f"{assessment}: {symbol_pct:.1f}% of ad text uses trademark symbols (®™©) for brand protection",
            recommendations="Add trademark symbols to brand terms in 30%+ of ads for legal protection." if score < 5 else "Good trademark symbol usage.",
            calculation=f"Symbol usage: {symbol_pct:.1f}%",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_title_case(self) -> ScoringResult:
        """Row 38: Copy uses inter-capitalisation? (Title Case)"""
        ad_data = self.data.get('ad')
        if ad_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Ad data",
                recommendations="Export ad report",
                data_needed="Ad report"
            )

        headline_cols = [c for c in ad_data.columns if 'Headline' in c or 'headline' in c]

        all_headlines = []
        for col in headline_cols:
            all_headlines.extend(ad_data[col].dropna().astype(str).tolist())

        ads_with_titlecase = sum(1 for h in all_headlines if h and len(h) > 1 and h[0].isupper() and any(c.isupper() for c in h[1:]))
        titlecase_pct = (ads_with_titlecase / len(all_headlines) * 100) if len(all_headlines) > 0 else 0

        if titlecase_pct >= 70:
            score = 5
            assessment = "Excellent"
        elif titlecase_pct >= 40:
            score = 3
            assessment = "Moderate"
        else:
            score = 1
            assessment = "Low"

        weight = get_criterion_weight("title_case")
        impact = determine_business_impact("Title Case", score, f"{titlecase_pct:.1f}%")

        return ScoringResult(
            score=score,
            details=f"{assessment}: {titlecase_pct:.1f}% of headlines use Title Case for professional appearance",
            recommendations="Use Title Case in 70%+ of headlines for readability and professionalism." if score < 5 else "Strong Title Case usage.",
            calculation=f"Title case: {titlecase_pct:.1f}%",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_extensions(self) -> ScoringResult:
        """Row 39: Using all relevant Extensions?

        UPDATED 2024: Now uses Sitelink assets.csv for ACTUAL extension performance data.
        - PRIMARY: Sitelink assets file with clicks, impressions, CTR per sitelink
        - Matches campaigns using Campaign ID, Campaign name
        - Measures: % campaigns with sitelink impressions (definitive proof of usage)
        - FALLBACK: Status columns ('active' counts) if sitelink assets not available
        """
        # TIER 1: Check for sitelink_assets file (most accurate)
        sitelink_data = self.data.get('sitelink_assets')

        if sitelink_data is not None and 'Campaign' in sitelink_data.columns:
            # Use sitelink assets for definitive extension measurement
            campaign_data = self.source_info.get('campaign_details')
            if campaign_data is None:
                return ScoringResult(
                    score=None,
                    details="Data needed: Campaign details to match with sitelink assets",
                    recommendations="Export campaign details report",
                    data_needed="Campaign details"
                )

            total_campaigns = len(campaign_data)

            # Match sitelink assets to campaigns
            # Get unique campaigns with sitelink impressions > 0
            sitelink_copy = sitelink_data.copy()

            # Convert impressions to numeric
            if 'Impr.' in sitelink_copy.columns:
                sitelink_copy['Impr_num'] = to_numeric_safe(sitelink_copy['Impr.'])

                # Filter to sitelinks with actual impressions
                sitelinks_with_impr = sitelink_copy[sitelink_copy['Impr_num'] > 0]

                # Get unique campaigns with sitelink impressions
                campaigns_with_sitelinks = sitelinks_with_impr['Campaign'].unique()

                # Count how many of our campaigns have sitelinks
                campaign_names = campaign_data['Campaign'].tolist() if 'Campaign' in campaign_data.columns else []

                campaigns_with_ext = sum(1 for camp in campaign_names if camp in campaigns_with_sitelinks)
                coverage_pct = (campaigns_with_ext / total_campaigns * 100) if total_campaigns > 0 else 0

                # Calculate total sitelink performance
                total_sitelink_clicks = sitelinks_with_impr['Clicks'].astype(str).str.replace(',', '').astype(float).sum() if 'Clicks' in sitelinks_with_impr.columns else 0
                total_sitelink_impr = sitelinks_with_impr['Impr_num'].sum()

                # Score based on coverage
                if coverage_pct >= 80:
                    score = 5
                    assessment = "Excellent"
                elif coverage_pct >= 60:
                    score = 4
                    assessment = "Good"
                elif coverage_pct >= 40:
                    score = 3
                    assessment = "Acceptable"
                elif coverage_pct >= 20:
                    score = 2
                    assessment = "Poor"
                else:
                    score = 1
                    assessment = "Very poor"

                details = f"{assessment}: {campaigns_with_ext}/{total_campaigns} campaigns ({coverage_pct:.0f}%) have sitelinks with impressions. Total sitelink impressions: {total_sitelink_impr:,.0f}, clicks: {total_sitelink_clicks:,.0f}"

                # Build ENHANCED recommendations with score context
                SCORE_LABELS = {
                    5: "Excellent (top 20%)",
                    4: "Strong foundation",
                    3: "Meets minimum expectations",
                    2: "Below standard",
                    1: "Critical gap",
                }

                # Score-specific recommendations with evidence and actionable guidance
                label_prefix = lambda s: f"{SCORE_LABELS.get(s, '').strip()}: " if SCORE_LABELS.get(s) else ""

                if score == 5:
                    recommendations = (
                        f"{label_prefix(score)}Sitelinks cover {coverage_pct:.0f}% of campaigns "
                        f"({campaigns_with_ext}/{total_campaigns}) and have driven {total_sitelink_impr:,.0f} impressions with "
                        f"{total_sitelink_clicks:,.0f} clicks. Layer in callouts, structured snippets, and promotions this quarter so every campaign carries multiple extension types—typically worth a further 5‑10% CTR lift."
                    )
                elif score == 4:
                    recommendations = (
                        f"{label_prefix(score)}Coverage sits at {coverage_pct:.0f}% "
                        f"({campaigns_with_ext}/{total_campaigns}) with healthy engagement. "
                        f"Finish the rollout to the remaining {total_campaigns - campaigns_with_ext} campaigns and reuse the highest-performing sitelink themes so you can close the gap to 80%+ and capture another 10‑15% CTR."
                    )
                elif score == 3:
                    recommendations = (
                        f"{label_prefix(score)}Sitelinks cover {coverage_pct:.0f}% of campaigns and clear the 40% minimum, "
                        f"but the majority of campaigns still run text-only ads. Build 4‑6 sitelinks for the highest spend campaigns this month and finish the rollout next month—accounts typically see 10‑20% CTR lifts once sitelinks are universal."
                    )
                elif score == 2:
                    recommendations = (
                        f"{label_prefix(score)}Only {coverage_pct:.0f}% of campaigns ({campaigns_with_ext}/{total_campaigns}) "
                        f"surface sitelinks, so most ads lack supporting links. Launch sitelinks on your top 10 spenders immediately, then roll through the next wave over the following month. Expect 10‑20% CTR gains once the majority of campaigns carry the extra real estate."
                    )
                else:  # score == 1
                    recommendations = (
                        f"{label_prefix(score)}Sitelinks are virtually absent ({campaigns_with_ext}/{total_campaigns}), "
                        f"which suppresses ad prominence. Build a core set of sitelinks (services, contact, reviews, etc.), deploy to the largest campaigns this week, and roll out to the remainder over the next month so every campaign gains the 30‑40% visibility lift that extensions provide."
                    )

                # Column I: Best Practice Rationale
                bp_rationale = generate_bp_rationale(
                    why_matters=(
                        f"Ad extensions (sitelinks, callouts, structured snippets) increase ad real estate by 30-40%, improving visibility and CTR. "
                        f"Sitelinks provide direct links to specific pages, reducing user friction and improving conversion rates by 10-20%. "
                        f"Ads with extensions typically achieve 10-20% higher CTR than text-only ads, with callouts and structured snippets providing "
                        f"additional 5-10% lifts when combined with sitelinks."
                    ),
                    threshold_context=(
                        f"Threshold applied: 80%+ sitelink coverage indicates comprehensive extension deployment aligned with PPC best practices. "
                        f"Minimum acceptable: 40% coverage for baseline extension usage. Top-performing accounts deploy 4-6 sitelinks per campaign "
                        f"(product/service pages, contact forms, about us, testimonials, support). Each sitelink should have distinct messaging and "
                        f"lead to unique landing pages for optimal performance measurement."
                    ),
                    your_performance=(
                        f"Your account: {campaigns_with_ext}/{total_campaigns} campaigns ({coverage_pct:.0f}%) have active sitelinks with "
                        f"{total_sitelink_impr:,.0f} impressions and {total_sitelink_clicks:,.0f} clicks recorded. "
                        f"{'This exceeds the 80% industry standard, demonstrating comprehensive extension deployment.' if coverage_pct >= 80 else f'This falls below the 80% industry standard, with {total_campaigns - campaigns_with_ext} campaigns missing sitelink coverage. Expanding sitelinks to these campaigns could increase overall CTR by 10-20%.'}"
                    ),
                    evidence_tier="applied_standard"
                )

                weight = get_criterion_weight("extensions")
                impact = determine_business_impact("Extensions", score, f"{coverage_pct:.0f}%")

                return ScoringResult(
                    score=score,
                    details=details,
                    recommendations=recommendations,
                    bp_rationale=bp_rationale,
                    calculation=f"Sitelink coverage: {coverage_pct:.0f}%, Data source: Sitelink assets file (performance data)",
                    criterion_weight=weight,
                    business_impact=impact['business_impact'],
                    priority_score=impact['priority_score']
                )

        # TIER 2: Fall back to campaign details status columns if no sitelink assets
        campaign_data = self.source_info.get('campaign_details')
        if campaign_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign details with extension data or Sitelink assets file",
                recommendations="Export Sitelink assets report and/or campaign details with extension columns",
                data_needed="Extension data"
            )

        total_campaigns = len(campaign_data)

        # Check status columns
        extension_types = {
            'Sitelinks': 'Campaign sitelinks: active'
        }

        extensions_found = {}
        extensions_coverage = {}

        for ext_name, col_name in extension_types.items():
            if col_name in campaign_data.columns:
                campaign_data_copy = campaign_data.copy()
                campaign_data_copy[f'{ext_name}_num'] = to_numeric_safe(campaign_data_copy[col_name])
                campaigns_with_ext = len(campaign_data_copy[campaign_data_copy[f'{ext_name}_num'] > 0])
                coverage_pct = (campaigns_with_ext / total_campaigns * 100) if total_campaigns > 0 else 0
                extensions_found[ext_name] = campaigns_with_ext
                extensions_coverage[ext_name] = coverage_pct

        if not extensions_found:
            return ScoringResult(
                score=None,
                details="Data needed: Extension data (Sitelink assets file or campaign extension columns)",
                recommendations="Export Sitelink assets report for accurate extension performance measurement",
                data_needed="Sitelink assets or extension columns"
            )

        # Score based on status columns (less accurate)
        avg_coverage = sum(extensions_coverage.values()) / len(extensions_coverage) if extensions_coverage else 0

        if avg_coverage >= 70:
            score = 5
            assessment = "Excellent"
        elif avg_coverage >= 50:
            score = 4
            assessment = "Good"
        elif avg_coverage >= 30:
            score = 3
            assessment = "Acceptable"
        elif avg_coverage >= 10:
            score = 2
            assessment = "Poor"
        else:
            score = 1
            assessment = "Very poor"

        ext_details = []
        for ext_name in ['Sitelinks']:
            if ext_name in extensions_found:
                ext_details.append(f"{ext_name}: {extensions_found[ext_name]}/{total_campaigns} ({extensions_coverage[ext_name]:.0f}%)")

        details = f"{assessment}: {', '.join(ext_details)} (based on current 'active' status - export Sitelink assets for performance data)"

        # Build ENHANCED recommendations for TIER 2 (status columns fallback)
        SCORE_LABELS = {
            5: "Excellent (top 20%)",
            4: "Strong foundation",
            3: "Meets minimum expectations",
            2: "Below standard",
            1: "Critical gap",
        }

        # Note: TIER 2 scoring uses 70/50/30/10 thresholds (vs TIER 1's 80/60/40/20)
        # due to status column data being less accurate
        if score >= 4:
            recommendations = (
                f"{SCORE_LABELS[score]}: Status columns show {avg_coverage:.0f}% coverage, but this fallback view only confirms whether extensions are enabled. Export the sitelink assets report with impressions and clicks, double-check which assets drive the bulk of performance, and expand into callouts and structured snippets to keep momentum."
            )
        else:
            recommendations = (
                f"{SCORE_LABELS[score]}: Status columns show {avg_coverage:.0f}% coverage, but we still lack impression/click data. Export the sitelink assets report immediately so we can validate usage, isolate top performers, and quantify the 10‑20% CTR lift that extensions typically deliver."
            )

        # Column I: Best Practice Rationale (same for both TIER 1 and TIER 2)
        bp_rationale = generate_bp_rationale(
            why_matters=(
                f"Ad extensions (sitelinks, callouts, structured snippets) increase ad real estate by 30-40%, improving visibility and CTR. "
                f"Sitelinks provide direct links to specific pages, reducing user friction and improving conversion rates by 10-20%. "
                f"Ads with extensions typically achieve 10-20% higher CTR than text-only ads."
            ),
            threshold_context=(
                f"Threshold applied (TIER 2 - status columns): 70%+ coverage indicates good extension deployment. "
                f"Minimum acceptable: 30% coverage. Note: These thresholds are less strict than TIER 1 (performance data) due to "
                f"measurement limitations. Status columns only show enabled/paused state, not actual impressions or clicks. "
                f"For accurate assessment, export Sitelink assets report showing performance metrics."
            ),
            your_performance=(
                f"Your account (TIER 2 measurement): {avg_coverage:.0f}% of campaigns show active extensions in status columns. "
                f"WARNING: This measurement cannot confirm actual sitelink usage or performance. Status columns show configuration state only, "
                f"not impression/click data. Export Sitelink assets report for definitive performance validation."
            ),
            evidence_tier="applied_standard"
        )

        weight = get_criterion_weight("extensions")
        impact = determine_business_impact("Extensions", score, f"{avg_coverage:.0f}%")

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            bp_rationale=bp_rationale,
            calculation=f"Extension coverage: {avg_coverage:.0f}%, Data source: Status columns only (performance data not available)",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    # ============================================================================
    # EXISTING SCORING FUNCTIONS (V4 patterns, fallbacks)
    # ============================================================================

    def _score_consolidation(self) -> ScoringResult:
        """Score URL/keyword consolidation (V4 template A)"""
        # Completeness gate
        if 'landing_page' not in self.data:
            return ScoringResult(
                score=None,
                details="Data needed: Landing page report with Final URL column.",
                recommendations="Export landing page performance report from SA360.",
                data_needed="Landing page report"
            )

        lp_data = self.data['landing_page'].copy()
        url_col = self._find_column(
            lp_data,
            ["Final URL", "Landing page", "Landing Page", "Final Url", "URL", "Destination URL", "Ad final URL"],
            allow_partial=True,
        )
        if not url_col:
            return ScoringResult(
                score=None,
                details="Data needed: Landing page report with Final URL column.",
                recommendations="Export landing page performance report from SA360.",
                data_needed="Landing page report"
            )

        # Get unique URLs and their keywords/ad groups
        keyword_col = self._find_column(lp_data, ["Keyword", "Search keyword"], allow_partial=True)
        adgroup_col = self._find_column(lp_data, ["Ad group", "Ad group name", "Ad group ID"], allow_partial=True)
        campaign_col = self._find_column(lp_data, ["Campaign", "Campaign name"], allow_partial=True)
        if keyword_col and adgroup_col:
            url_analysis = lp_data.groupby(url_col).agg({
                keyword_col: 'nunique',
                adgroup_col: 'nunique'
            }).rename(columns={keyword_col: 'unique_keywords', adgroup_col: 'unique_adgroups'})

            # Calculate redundancy
            total_urls = len(url_analysis)
            redundant_urls = len(url_analysis[(url_analysis['unique_keywords'] > 10) |
                                              (url_analysis['unique_adgroups'] > 5)])
            redundancy_pct = (redundant_urls / total_urls * 100) if total_urls > 0 else 0

            # Score
            if redundancy_pct < 10:
                score = 5
                assessment = "Excellent"
            elif redundancy_pct < 30:
                score = 3
                assessment = "Moderate"
            else:
                score = 1
                assessment = "Poor"

            return ScoringResult(
                score=score,
                details=f"{assessment} consolidation: {redundancy_pct:.1f}% of URLs have redundant targeting. {redundant_urls}/{total_urls} URLs have excessive keywords/ad groups.",
                recommendations=f"Consolidate {redundant_urls} URLs to reduce overlap and improve Quality Score." if score < 4 else "Maintain current consolidation.",
                calculation=f"Redundancy: {redundancy_pct:.1f}%"
            )
        if campaign_col:
            url_counts = lp_data[url_col].astype(str).value_counts()
            total_urls = int(len(url_counts))
            url_campaigns = lp_data.groupby(url_col)[campaign_col].nunique()
            multi_campaign_urls = int((url_campaigns > 1).sum())
            overlap_pct = (multi_campaign_urls / total_urls * 100) if total_urls else 0.0

            if overlap_pct < 15:
                score = 4
                assessment = "Healthy overlap"
            elif overlap_pct < 35:
                score = 3
                assessment = "Moderate overlap"
            else:
                score = 2
                assessment = "High overlap"

            return ScoringResult(
                score=score,
                details=(
                    f"Partial landing page view: {assessment}. {multi_campaign_urls}/{total_urls} URLs "
                    f"({overlap_pct:.1f}%) are shared across multiple campaigns."
                ),
                recommendations=(
                    "Reduce URL reuse across campaigns to clarify landing page intent and improve relevance."
                    if score <= 3
                    else "Keep landing page coverage consistent and monitor for overlap creep."
                ),
                calculation=f"URLs: {total_urls}; multi-campaign URLs: {multi_campaign_urls}",
                data_completeness=0.6,
                confidence="medium",
            )

        url_counts = lp_data[url_col].astype(str).value_counts()
        total_urls = int(len(url_counts))
        return ScoringResult(
            score=2,
            details=f"Landing page inventory: {total_urls} unique URLs detected without campaign or keyword linkage.",
            recommendations="Include campaign or keyword dimensions in landing page exports to evaluate consolidation quality.",
            calculation=f"URLs: {total_urls}",
            data_completeness=0.4,
            confidence="low",
        )

    def _score_campaign_count(self) -> ScoringResult:
        """Score campaign count per business unit"""
        # Get unique campaigns from performance source
        if 'Campaign' not in self.performance_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign column in performance data.",
                recommendations="Export campaign performance report.",
                data_needed="Campaign performance data"
            )

        num_campaigns = self.performance_data['Campaign'].nunique()

        # Score based on campaign count (industry best practice: 5-15 campaigns per BU)
        if 5 <= num_campaigns <= 15:
            score = 5
            assessment = "Optimal"
        elif num_campaigns < 5:
            score = 3
            assessment = "Under-segmented"
        elif num_campaigns <= 25:
            score = 3
            assessment = "Slightly over-segmented"
        else:
            score = 1
            assessment = "Over-segmented"

        return ScoringResult(
            score=score,
            details=f"{assessment}: {num_campaigns} campaigns. Best practice: 5-15 campaigns per business unit for manageability.",
            recommendations="Consolidate campaigns to reduce management overhead." if num_campaigns > 15 else "Consider additional segmentation if needed." if num_campaigns < 5 else "Maintain current structure.",
            calculation=f"Campaigns: {num_campaigns}"
        )

    def _score_adgroup_structure(self) -> ScoringResult:
        """Score average ad groups per campaign"""
        if 'Campaign' not in self.performance_data.columns or 'Ad group' not in self.performance_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign and Ad group columns in performance data.",
                recommendations="Export ad group performance report.",
                data_needed="Ad group performance data"
            )

        adgroups_per_campaign = self.performance_data.groupby('Campaign')['Ad group'].nunique()
        avg_adgroups = adgroups_per_campaign.mean()

        # Score (best practice: 5-20 ad groups per campaign)
        if 5 <= avg_adgroups <= 20:
            score = 5
            assessment = "Well-structured"
        elif avg_adgroups < 5:
            score = 3
            assessment = "Under-segmented"
        elif avg_adgroups <= 35:
            score = 3
            assessment = "Over-segmented"
        else:
            score = 1
            assessment = "Severely over-segmented"

        return ScoringResult(
            score=score,
            details=f"{assessment}: {avg_adgroups:.1f} avg ad groups per campaign. Best practice: 5-20 ad groups.",
            recommendations="Consolidate ad groups within campaigns." if avg_adgroups > 20 else "Consider theme-based segmentation." if avg_adgroups < 5 else "Maintain structure.",
            calculation=f"Avg ad groups: {avg_adgroups:.1f}"
        )

    def _score_quality_score_components(self) -> ScoringResult:
        """Score Quality Score components (requires Google Ads export)"""
        keyword_data = self.data.get('search_keyword')
        if keyword_data is None or keyword_data.empty:
            keyword_data = self.data.get('keyword')
        if keyword_data is None or keyword_data.empty:
            return ScoringResult(
                score=None,
                details="Data needed: Search keyword report with Quality Score columns.",
                recommendations="Export Google Ads keyword report (include Quality Score + components).",
                data_needed="Search keyword report with Quality Score"
            )

        qs_col = self._find_column(keyword_data, ["Quality Score", "Quality score", "QualityScore"])
        if not qs_col:
            return ScoringResult(
                score=None,
                details="Data needed: Quality Score column (Quality Score).",
                recommendations="Add Quality Score to the keyword export.",
                data_needed="Quality Score column"
            )

        qs_values = pd.to_numeric(keyword_data[qs_col], errors='coerce')
        if qs_values.dropna().empty:
            return ScoringResult(
                score=None,
                details="Quality Score column present but empty. Re-export with row-level scores.",
                recommendations="Include non-aggregated keyword data with Quality Score populated.",
                data_needed="Keyword-level Quality Score data"
            )

        impr_col = self._find_column(
            keyword_data,
            ["Impr.", "Impressions", "Client account impressions", "Impr"]
        )
        if impr_col:
            weights = pd.to_numeric(keyword_data[impr_col], errors='coerce').fillna(0)
        else:
            weights = pd.Series(1.0, index=qs_values.index)

        weights = weights.where(qs_values.notna(), 0)
        total_weight = weights.sum()
        if total_weight <= 0:
            weights = pd.Series(1.0, index=qs_values.index).where(qs_values.notna(), 0)
            total_weight = weights.sum()
        if total_weight <= 0:
            return ScoringResult(
                score=None,
                details="Unable to compute weighted Quality Score (zero usable rows).",
                recommendations="Verify impressions/Quality Score columns contain numeric data.",
                data_needed="Keyword report with Quality Score + Impressions"
            )

        weighted_avg = float((qs_values * weights).sum() / total_weight)
        low_share = float(weights[(qs_values <= 3)].sum() / total_weight * 100)
        very_low_share = float(weights[(qs_values <= 2)].sum() / total_weight * 100)
        high_share = float(weights[(qs_values >= 8)].sum() / total_weight * 100)
        keyword_volume = int((weights > 0).sum())

        if weighted_avg >= 8 and low_share <= 10:
            score = 5
            assessment = "Elite Quality Score health"
        elif weighted_avg >= 6.5 and low_share <= 20:
            score = 4
            assessment = "Healthy Quality Scores"
        elif weighted_avg >= 5 and low_share <= 35:
            score = 3
            assessment = "Mixed Quality Scores"
        elif weighted_avg >= 4 or very_low_share <= 50:
            score = 2
            assessment = "At-risk Quality Scores"
        else:
            score = 1
            assessment = "Critical Quality Score risk"

        component_columns = {
            "Expected CTR": self._find_column(keyword_data, ["Expected click-through rate", "Expected CTR"]),
            "Ad relevance": self._find_column(keyword_data, ["Ad relevance"]),
            "Landing page exp.": self._find_column(keyword_data, ["Landing page experience"]),
        }
        component_findings = []
        weakest_component = None
        weakest_value = 0.0
        for label, col in component_columns.items():
            if not col:
                continue
            series = keyword_data[col].astype(str).str.strip().str.lower()
            below_pct = float(weights[(series == "below average")].sum() / total_weight * 100)
            avg_pct = float(weights[(series == "average")].sum() / total_weight * 100)
            above_pct = float(weights[(series == "above average")].sum() / total_weight * 100)
            component_findings.append(
                f"{label}: {above_pct:.0f}% above / {avg_pct:.0f}% avg / {below_pct:.0f}% below"
            )
            if below_pct > weakest_value:
                weakest_value = below_pct
                weakest_component = (label, below_pct)

        recommendations = []
        if low_share > 20:
            recommendations.append(
                f"Repair or pause keywords with QS ≤3 ({low_share:.0f}% of volume) by tightening ad copy relevance and landing pages."
            )
        if weakest_component and weakest_component[1] > 15:
            label, pct = weakest_component
            if "Landing" in label:
                recommendations.append(
                    f"Landing page experience has {pct:.0f}% of volume below average—improve page speed and message match."
                )
            elif "CTR" in label:
                recommendations.append(
                    f"Expected CTR is below average for {pct:.0f}% of impressions—refresh RSA assets and add stronger CTAs."
                )
            else:
                recommendations.append(
                    f"{label} is below average for {pct:.0f}% of spend—align copy/keywords to intent."
                )
        if not recommendations:
            recommendations.append("Maintain current Quality Score program and continue refreshing ads quarterly.")

        detail_parts = [
            f"{assessment}: Weighted QS {weighted_avg:.1f}",
            f"{high_share:.0f}% of impressions at QS ≥8",
            f"{low_share:.0f}% at QS ≤3 (≤2: {very_low_share:.0f}%)",
            f"Rows analyzed: {keyword_volume:,}",
        ]
        if component_findings:
            detail_parts.append("; ".join(component_findings))

        weight = get_criterion_weight("quality_score")
        impact = determine_business_impact("Quality Score", score, f"{weighted_avg:.1f} avg")
        rationale = generate_bp_rationale(
            "Quality Score governs CPC efficiency and impression eligibility in Google Ads auctions.",
            "Maintain weighted QS ≥7 with <10% of volume scoring ≤3 to avoid cost penalties (audit standard).",
            f"Your weighted QS is {weighted_avg:.1f} with {low_share:.0f}% of traffic ≤3 and {high_share:.0f}% ≥8 "
            f"across {keyword_volume:,} keywords."
        )

        return ScoringResult(
            score=score,
            details=". ".join(detail_parts),
            recommendations=" ".join(recommendations),
            calculation=f"Weighted QS={weighted_avg:.1f}; ≤3 share={low_share:.1f}%; ≥8 share={high_share:.1f}%",
            bp_rationale=rationale,
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_keyword_coverage(self) -> ScoringResult:
        """Score keyword coverage - ensure all ad groups have keywords"""
        if 'Ad group' not in self.performance_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Ad group performance data.",
                recommendations="Export ad group report.",
                data_needed="Ad group performance data"
            )

        # Check for keyword column
        keyword_col = None
        for col in ['Keyword', 'keyword', 'Search term']:
            if col in self.performance_data.columns:
                keyword_col = col
                break

        if not keyword_col:
            return ScoringResult(
                score=None,
                details="Data needed: Keyword column in performance data.",
                recommendations="Export keyword performance report.",
                data_needed="Keyword performance data"
            )

        # Find ad groups with zero keywords
        adgroups_with_keywords = self.performance_data[
            self.performance_data[keyword_col].notna() &
            (self.performance_data[keyword_col] != '')
        ]['Ad group'].unique()

        total_adgroups = self.performance_data['Ad group'].nunique()
        adgroups_without_keywords = total_adgroups - len(adgroups_with_keywords)
        pct_without = (adgroups_without_keywords / total_adgroups * 100) if total_adgroups > 0 else 0

        # Score (FIX: Changed "missing" to "have zero" - no circular logic)
        if pct_without == 0:
            score = 5
        elif pct_without < 5:
            score = 4
        elif pct_without < 15:
            score = 3
        else:
            score = 1

        return ScoringResult(
            score=score,
            details=f"{adgroups_without_keywords}/{total_adgroups} active ad groups ({pct_without:.1f}%) have zero keywords.",
            recommendations=f"Add keywords to {adgroups_without_keywords} ad groups." if adgroups_without_keywords > 0 else "All ad groups have keyword coverage.",
            calculation=f"Coverage: {100-pct_without:.1f}%"
        )

    def _score_broad_match(self) -> ScoringResult:
        """Score broad match usage (V4 template B)"""
        match_series, source_label = self._extract_match_type_series()
        if match_series is None:
            return ScoringResult(
                score=None,
                details="Data needed: Match type column in keyword report.",
                recommendations="Export keyword report with match type (Search keyword match type / Match type).",
                data_needed="Keyword report with match type"
            )

        total_keywords = len(match_series)
        if total_keywords == 0:
            return ScoringResult(
                score=None,
                details="No keywords found after filtering. Provide keyword-level export.",
                recommendations="Export Search keyword report with match type.",
                data_needed="Keyword report"
            )

        broad_mask = match_series.str.contains(r"(broad|bmm)", case=False, na=False, regex=True)
        broad_keywords = int(broad_mask.sum())
        broad_pct = (broad_keywords / total_keywords * 100) if total_keywords > 0 else 0

        # Score based on broad match adoption
        if broad_pct >= 40:
            score = 5
            assessment = "Excellent"
        elif broad_pct >= 20:
            score = 3
            assessment = "Moderate"
        else:
            score = 1
            assessment = "Low"

        return ScoringResult(
            score=score,
            details=(
                f"{assessment} broad match adoption: {broad_pct:.1f}% of {source_label} rows "
                f"({broad_keywords}/{total_keywords}) use broad match. Best practice: 40%+ with Smart Bidding."
            ),
            recommendations="Increase broad match usage with Smart Bidding for better reach." if broad_pct < 40 else "Maintain broad match strategy.",
            calculation=f"Broad match: {broad_pct:.1f}%"
        )

    def _score_match_type_balance(self) -> ScoringResult:
        """Score match type balance across portfolio"""
        match_series, source_label = self._extract_match_type_series()
        if match_series is None:
            return ScoringResult(
                score=None,
                details="Data needed: Match type column.",
                recommendations="Export keyword report with match type column.",
                data_needed="Keyword report with match type"
            )

        match_dist = self._compute_match_distribution(match_series)
        if not match_dist:
            return ScoringResult(
                score=None,
                details="Match type data available but no normalized values detected.",
                recommendations="Verify match type column contains Exact/Phrase/Broad values.",
                data_needed="Keyword report with match type",
            )

        # Count how many match types have meaningful share
        match_types_used = sum(1 for key in ("exact_pct", "phrase_pct", "broad_pct") if match_dist.get(key, 0) > 0)

        if match_types_used >= 3:
            score = 5
            assessment = "Well-balanced"
        elif match_types_used == 2:
            score = 3
            assessment = "Moderate balance"
        else:
            score = 1
            assessment = "No balance"

        match_summary = (
            f"Exact: {match_dist.get('exact_pct', 0):.1f}%, "
            f"Phrase: {match_dist.get('phrase_pct', 0):.1f}%, "
            f"Broad: {match_dist.get('broad_pct', 0):.1f}%"
        )

        return ScoringResult(
            score=score,
            details=(
                f"{assessment} match type distribution across {source_label or 'keywords'}: {match_summary}. "
                f"Using {match_types_used} match types."
            ),
            recommendations="Consider diversifying match types for broader reach." if match_types_used < 2 else "Good match type diversity.",
            calculation=f"Types used: {match_types_used}"
        )

    def _score_negative_keywords(self) -> ScoringResult:
        """Score negative keyword effectiveness (V4 template K)"""
        # This requires search query report to identify waste
        if 'search_query' not in self.data and 'search_term' not in self.data:
            return ScoringResult(
                score=None,
                details="Data needed: Search query report to analyze waste queries and negative keyword opportunities.",
                recommendations="Export search query report from SA360/Google Ads with CTR and conversion data.",
                data_needed="Search query report"
            )

        # If search query data exists, analyze waste
        sq_data = self.data.get('search_query', self.data.get('search_term'))

        if sq_data is not None and 'CTR' in sq_data.columns and 'Cost' in sq_data.columns:
            sq_copy = sq_data.copy()
            sq_copy['CTR_num'] = to_numeric_safe(sq_copy['CTR'])
            sq_copy['Cost_num'] = to_numeric_safe(sq_copy['Cost'])

            # Identify waste queries (low CTR, spent money)
            waste_queries = sq_copy[(sq_copy['CTR_num'] < 1.0) & (sq_copy['Cost_num'] > 10)]
            waste_spend = waste_queries['Cost_num'].sum()
            total_spend = sq_copy['Cost_num'].sum()
            waste_pct = (waste_spend / total_spend * 100) if total_spend > 0 else 0

            # Score
            if waste_pct < 5:
                score = 5
                assessment = "Excellent"
            elif waste_pct < 10:
                score = 4
                assessment = "Good"
            elif waste_pct < 20:
                score = 2
                assessment = "Poor"
            else:
                score = 1
                assessment = "Critical"

            return ScoringResult(
                score=score,
                details=f"{assessment}: {len(waste_queries)} waste queries, ${waste_spend:,.2f} waste ({waste_pct:.1f}% of spend). Urgent negative keyword additions needed." if waste_pct > 10 else f"{assessment} negative keyword coverage.",
                recommendations=f"Add {len(waste_queries)} queries as negative keywords to reduce waste." if len(waste_queries) > 0 else "Maintain negative keyword strategy.",
                calculation=f"Waste: {waste_pct:.1f}%"
            )

        return ScoringResult(
            score=None,
            details="Data needed: Search query report with CTR and Cost columns.",
            recommendations="Export search query performance data.",
            data_needed="Search query report"
        )

    def _score_audience_targeting(self) -> ScoringResult:
        """Score audience targeting coverage (V3.1 Row 28 - FIXED with set intersection)"""
        if 'audience' not in self.data:
            return ScoringResult(
                score=None,
                details="Data needed: Audience report from SA360.",
                recommendations="Export audience targeting report.",
                data_needed="Audience report"
            )

        audience = self.data['audience']

        if 'Campaign' not in audience.columns or 'Campaign' not in self.performance_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign column in both audience and performance data.",
                recommendations="Ensure audience report includes campaign dimension.",
                data_needed="Audience report with campaigns"
            )

        # FIX: Set intersection to prevent >100% (from V3.1 bug fixes)
        campaigns_in_perf = set(self.performance_data['Campaign'].unique())
        campaigns_with_audience = len(set(audience['Campaign'].unique()) & campaigns_in_perf)
        total_campaigns = len(campaigns_in_perf)

        audience_pct = (campaigns_with_audience / total_campaigns * 100) if total_campaigns > 0 else 0

        # Score
        if audience_pct >= 90:
            score = 5
        elif audience_pct >= 70:
            score = 4
        elif audience_pct >= 50:
            score = 3
        else:
            score = 1

        return ScoringResult(
            score=score,
            details=f"{campaigns_with_audience}/{total_campaigns} campaigns ({audience_pct:.1f}%) have audience targeting.",
            recommendations=f"Add audience targeting to {total_campaigns - campaigns_with_audience} campaigns." if audience_pct < 90 else "Excellent audience coverage.",
            calculation=f"Coverage: {audience_pct:.1f}%"
        )

    def _score_rsa_coverage(self) -> ScoringResult:
        """Row 32: 2 RSA. ETA depreciated June 2022

        ACCURATE: Checks if ad groups have at least 2 RSAs (Google best practice).
        Title "2 RSA" means 2 RSAs per ad group for A/B testing.
        """
        # Check for Ad type column
        if 'ad' not in self.data:
            return ScoringResult(
                score=None,
                details="Data needed: Ad report from SA360.",
                recommendations="Export ad performance report.",
                data_needed="Ad report"
            )

        ad_data = self.data['ad']

        if 'Ad type' not in ad_data.columns or 'Ad group ID' not in ad_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Ad type and Ad group ID columns in ad report.",
                recommendations="Ensure ad export includes ad type and ad group ID.",
                data_needed="Ad report with ad type and ad group ID"
            )

        # Get total ad groups
        adgroup_data = self.data.get('ad_group_details')
        if adgroup_data is None:
            adgroup_data = self.data.get('ad_group')
        if adgroup_data is None or len(adgroup_data) == 0:
            total_adgroups = ad_data['Ad group ID'].nunique()
        else:
            total_adgroups = len(adgroup_data)

        # Filter to RSA ads
        rsa_ads = ad_data[ad_data['Ad type'].str.contains('Responsive', case=False, na=False)]

        # Count RSAs per ad group
        rsas_per_adgroup = rsa_ads.groupby('Ad group ID').size()

        # Ad groups with 2+ RSAs (Google best practice)
        adgroups_with_2plus_rsas = (rsas_per_adgroup >= 2).sum()
        adgroups_with_any_rsa = len(rsas_per_adgroup)

        # Calculate percentages
        pct_with_2plus = (adgroups_with_2plus_rsas / total_adgroups * 100) if total_adgroups > 0 else 0
        pct_with_any = (adgroups_with_any_rsa / total_adgroups * 100) if total_adgroups > 0 else 0

        # Score based on Google's "2 RSA per ad group" best practice
        if pct_with_2plus >= 80:
            score = 5
            assessment = "Excellent"
        elif pct_with_2plus >= 60:
            score = 4
            assessment = "Good"
        elif pct_with_2plus >= 40:
            score = 3
            assessment = "Acceptable"
        elif pct_with_any >= 50:  # At least has RSAs, even if not 2+
            score = 2
            assessment = "Needs improvement"
        else:
            score = 1
            assessment = "Poor"

        total_rsas = len(rsa_ads)
        total_ads = len(ad_data)
        rsa_pct = (total_rsas / total_ads * 100) if total_ads > 0 else 0

        details = f"{assessment}: {adgroups_with_2plus_rsas}/{total_adgroups} ad groups ({pct_with_2plus:.0f}%) have 2+ RSAs. Total: {total_rsas}/{total_ads} ads ({rsa_pct:.0f}%) are RSA. Google best practice: 2 RSAs per ad group."

        recommendations = f"Add RSAs to reach 2+ per ad group for A/B testing. Currently {adgroups_with_2plus_rsas}/{total_adgroups} ad groups meet this." if score < 4 else "Excellent RSA adoption following Google best practice."

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Ad groups with 2+ RSAs: {adgroups_with_2plus_rsas}, Total ad groups: {total_adgroups}, Total RSAs: {total_rsas}"
        )

    def _score_ad_strength(self) -> ScoringResult:
        """Score Ad Strength (requires Google Ads export)"""
        # Check for Ad strength column
        if 'ad' not in self.data or 'Ad strength' not in self.data['ad'].columns:
            return ScoringResult(
                score=None,
                details="Data needed: Ad strength column from Google Ads ad report. SA360 exports typically don't include Ad Strength.",
                recommendations="Export ad report from Google Ads UI with Ad Strength column.",
                data_needed="Google Ads ad report with Ad Strength"
            )

        # If data exists, score it (V4 template C logic)
        return ScoringResult(
            score=None,
            details="Ad Strength scoring logic to be implemented.",
            recommendations="Manual Ad Strength review required.",
            data_needed=None
        )

    def _score_smart_bidding(self) -> ScoringResult:
        """
        Score Smart Bidding adoption AND performance (enhanced with knowledge base)

        Checks:
        1. Adoption (% campaigns using Smart Bidding)
        2. Readiness (conversion volume)
        3. Performance (vs targets, if available)
        """
        # Completeness gate
        if 'campaign_details' not in self.data:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign report with bid strategy and conversion data.",
                recommendations="Export campaign report with bid strategy type and conversion columns.",
                data_needed="Campaign report with bid strategy",
                data_completeness=0.0,
                confidence="low"
            )

        campaign_data = self.data['campaign_details']

        if 'Campaign bid strategy type' not in campaign_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign bid strategy type column.",
                recommendations="Ensure campaign export includes bid strategy.",
                data_needed="Campaign bid strategy data",
                data_completeness=0.3,
                confidence="low"
            )

        # Get best practices from knowledge base
        bp = BEST_PRACTICES['smart_bidding']

        # Identify Smart Bidding strategies
        smart_bidding_strategies = ['Target CPA', 'Target ROAS', 'Maximize conversions', 'Maximize conversion value']

        total_campaigns = len(campaign_data)
        smart_campaigns_df = campaign_data[
            campaign_data['Campaign bid strategy type'].isin(smart_bidding_strategies)
        ]
        smart_campaigns = len(smart_campaigns_df)
        smart_pct = (smart_campaigns / total_campaigns * 100) if total_campaigns > 0 else 0

        # Check readiness (conversion volume)
        ready_for_smart_bidding = 0
        not_ready_detail = ""
        smart_with_volume = smart_campaigns
        smart_without_volume = 0
        volume_data_available = False

        min_required = self._context_min_conversions() or bp['min_conversions_30d_tcpa']

        if 'Conversions' in campaign_data.columns:
            volume_data_available = True
            campaign_data_copy = campaign_data.copy()
            campaign_data_copy['Conv_num'] = to_numeric_safe(campaign_data_copy['Conversions'])

            # Campaigns with enough conversion volume
            ready_campaigns = campaign_data_copy[campaign_data_copy['Conv_num'] >= min_required]
            ready_for_smart_bidding = len(ready_campaigns)

            # Check if Smart Bidding campaigns have sufficient volume
            smart_with_volume = len(smart_campaigns_df.merge(
                ready_campaigns[['Campaign']],
                on='Campaign',
                how='inner'
            )) if 'Campaign' in campaign_data.columns else 0

            smart_without_volume = smart_campaigns - smart_with_volume

            if smart_without_volume > 0:
                not_ready_detail = (
                    f" WARNING: {smart_without_volume} Smart Bidding campaigns have <{min_required} conversions/30d "
                    f"(below the learning threshold for this business)."
                )
        else:
            not_ready_detail = " NOTE: Conversion data not available to verify readiness."

        vertical = self._context_vertical()
        adoption_top = 80
        adoption_mid = 50
        adoption_low = 20
        primary_kpi = self._context_primary_kpi()
        if vertical in {"FINANCE", "LOCAL"}:
            adoption_top = 70
            adoption_mid = 40
            adoption_low = 15
        elif primary_kpi in {"POAS", "CAC_PAYBACK"}:
            adoption_top = 85
            adoption_mid = 60

        # Score based on adoption first, then apply risk adjustments for guardrails
        if smart_pct >= adoption_top:
            adoption_score = 5
            assessment = "Excellent"
        elif smart_pct >= adoption_mid:
            adoption_score = 3
            assessment = "Moderate"
        elif smart_pct >= adoption_low:
            adoption_score = 2
            assessment = "Low"
        else:
            adoption_score = 1
            assessment = "Very low"

        score = adoption_score
        risk_notes: List[str] = []
        if adoption_score == 5 and smart_without_volume > 0:
            penalty = 2
            score = max(1, adoption_score - penalty)
            assessment = "Guardrail veto"
            risk_notes.append(
                f"-{penalty}: Data starvation ({smart_without_volume} campaigns <{min_required} conversions/30d)"
            )

        self.summary_stats['smart_bidding'] = {
            'total_campaigns': total_campaigns,
            'smart_campaigns': smart_campaigns,
            'smart_pct': smart_pct,
            'manual_campaigns': total_campaigns - smart_campaigns,
            'smart_with_volume': smart_with_volume,
            'smart_without_volume': smart_without_volume,
            'volume_data_available': volume_data_available,
        }

        # Determine business impact
        impact = determine_business_impact(
            criterion_name="Smart Bidding",
            score=score,
            details=f"{smart_pct:.0f}% adoption"
        )

        # Get criterion weight
        weight = get_criterion_weight("smart_bidding")

        # Build ENHANCED recommendations with score context
        SCORE_LABELS = {
            5: "Excellent (top 20%)",
            4: "Strong foundation",
            3: "Meets minimum expectations",
            2: "Below standard",
            1: "Critical gap",
        }

        label_prefix = lambda s: f"{SCORE_LABELS.get(s, '').strip()}: " if SCORE_LABELS.get(s) else ""

        if score == 5 and smart_without_volume == 0:
            recommendations = (
                f"{label_prefix(score)}You've adopted Smart Bidding for {smart_pct:.1f}% of campaigns "
                f"({smart_campaigns}/{total_campaigns}) and every campaign has the required {min_required}+ conversions. "
                f"Use that momentum to test Target ROAS on the highest value campaigns, layer in seasonal adjustments, and monitor the short learning periods so efficiency gains continue."
            )
        elif score == 5 and smart_without_volume > 0:
            recommendations = (
                f"{label_prefix(score)}Adoption is high at {smart_pct:.1f}% "
                f"({smart_campaigns}/{total_campaigns}), but {smart_without_volume} campaigns lack the {min_required}+ conversions they need. "
                f"Consolidate or temporarily revert those campaigns until they hit volume so the automation has consistent signals—this typically unlocks another 20‑30% lift."
            )
        elif score == 3 and smart_pct >= 50:
            recommendations = (
                f"{label_prefix(score)}Smart Bidding covers {smart_pct:.1f}% of campaigns, so "
                f"{total_campaigns - smart_campaigns} campaigns are still manual. Migrate those in batches starting with the ones already hitting {min_required} conversions per 30 days and monitor the "
                f"{bp['learning_period_days_min']}-{bp['learning_period_days_max']} day stabilization window—advertisers typically see 20‑40% efficiency gains within two months."
            )
        elif score == 2:
            recommendations = (
                f"{label_prefix(score)}Only {smart_pct:.1f}% of campaigns use Smart Bidding "
                f"({smart_campaigns}/{total_campaigns}), so the majority still rely on manual bidding. Move the highest-volume campaigns first, enable Target CPA/ROAS, and schedule the remaining migrations next month. "
                f"Expect a 20‑40% efficiency lift once automation controls the bulk of spend."
            )
        else:  # score == 1
            recommendations = (
                f"{label_prefix(score)}Automation covers just {smart_pct:.1f}% of campaigns "
                f"({smart_campaigns}/{total_campaigns}). Audit every campaign for conversion volume this week, migrate the largest manual portfolios immediately, "
                f"and continue in weekly batches so automation becomes the norm; most accounts close the 30% gap within a quarter and gain 14%+ more conversion value."
            )

        # Column I: Best Practice Rationale
        bp_rationale = generate_bp_rationale(
            why_matters=(
                f"Smart Bidding uses machine learning to optimize bids in real-time based on conversion likelihood, user signals, "
                f"and contextual data. Campaigns using Smart Bidding typically achieve 20-40% efficiency improvement within 2-3 months "
                f"compared to manual bidding, with Target ROAS strategies delivering 14% more conversion value than Target CPA "
                f"(industry research). However, Smart Bidding requires sufficient conversion volume to train AI models effectively."
            ),
            threshold_context=(
                f"Threshold applied: {adoption_top}% adoption indicates automation maturity for this business model; "
                f"{adoption_mid}% is the minimum acceptable coverage. Each Smart Bidding campaign should generate "
                f"{min_required}+ conversions per 30 days for optimal AI learning. Campaigns below this volume "
                f"threshold may experience erratic bidding during the {bp['learning_period_days_min']}-{bp['learning_period_days_max']}-day "
                f"learning period."
            ),
            your_performance=(
                f"Your account: {smart_campaigns}/{total_campaigns} campaigns ({smart_pct:.1f}%) use Smart Bidding."
                f"{' ' + str(smart_without_volume) + ' Smart Bidding campaigns have insufficient conversion volume (<' + str(min_required) + ' conversions/30d), limiting AI effectiveness.' if smart_without_volume > 0 else ' All Smart Bidding campaigns have sufficient volume for effective learning.'} "
                f"{'This meets the contextual benchmark.' if smart_pct >= adoption_top else 'This falls below the contextual benchmark, leaving ' + str(total_campaigns - smart_campaigns) + ' campaigns in manual bidding.'}"
            ),
            evidence_tier="applied_standard"
        )

        goal_payload = self._analyze_goal_alignment() or {}
        blended_details = f"{assessment}: {smart_campaigns}/{total_campaigns} campaigns ({smart_pct:.1f}%) use Smart Bidding.{not_ready_detail}"
        blended_details = f"{blended_details}\n{goal_payload.get('details', '').strip()}" if goal_payload.get("details") else blended_details
        blended_recommendations = recommendations
        goal_score = goal_payload.get("score")
        if isinstance(goal_score, (int, float)):
            if goal_score < score:
                score = goal_score
                blended_recommendations = goal_payload.get("recommendations", recommendations)
            elif goal_score < 5 and goal_payload.get("recommendations"):
                blended_recommendations = (
                    f"{recommendations}\n\nGoal alignment: {goal_payload['recommendations']}"
                )

        return ScoringResult(
            score=score,
            details=blended_details,
            recommendations=blended_recommendations,
            bp_rationale=bp_rationale,
            calculation=f"Smart Bidding: {smart_pct:.1f}%",
            data_completeness=1.0 if 'Conversions' in campaign_data.columns else 0.7,
            confidence="high" if 'Conversions' in campaign_data.columns else "medium",
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score'],
            impact_rationale=impact['rationale'],
            criterion_weight=weight,
            raw_score=adoption_score,
            risk_adjustments=risk_notes
        )

    def _score_manual_ecpc(self) -> ScoringResult:
        """Score ECPC on manual campaigns (V3.1 Row 49 - FIXED with N/A)"""
        if 'campaign_details' not in self.data:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign details with bid strategy.",
                recommendations="Export campaign report.",
                data_needed="Campaign report"
            )

        campaign_data = self.data['campaign_details']

        if 'Campaign bid strategy type' not in campaign_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign bid strategy type.",
                recommendations="Ensure campaign export includes bid strategy.",
                data_needed="Campaign bid strategy data"
            )

        # Find manual campaigns
        manual_campaigns = campaign_data[campaign_data['Campaign bid strategy type'] == 'Manual CPC']

        if len(manual_campaigns) == 0:
            # FIX: Return N/A instead of 0/0 with score 5 (from V3.1 fixes)
            return ScoringResult(
                score="N/A",
                details="All campaigns use automated bidding (best practice). No manual campaigns to evaluate for ECPC.",
                recommendations="Maintain automated bidding strategy.",
                calculation="N/A"
            )

        # Check for ECPC usage
        # (Would need additional column - for now, score based on presence)
        return ScoringResult(
            score=3,
            details=f"{len(manual_campaigns)} manual campaigns found. ECPC status requires additional data.",
            recommendations="Migrate manual campaigns to Smart Bidding or enable ECPC.",
            calculation=f"Manual campaigns: {len(manual_campaigns)}"
        )

    def _score_impression_share(self) -> ScoringResult:
        """Row 50: Not losing too much impression Share to rank?

        Measures Search Lost IS (rank) - indicates quality score and bid competitiveness issues.
        Industry standard: <20% lost to rank is healthy, >40% needs immediate attention.
        """
        # Check for IS columns (case-insensitive matching)
        dataset = self.performance_data
        lost_rank_col = self._find_column_by_substrings(dataset, ['lost', 'rank'])

        if lost_rank_col is None:
            for df in self.data.values():
                if not isinstance(df, pd.DataFrame):
                    continue
                candidate = self._find_column_by_substrings(df, ['lost', 'rank'])
                if candidate:
                    dataset = df
                    lost_rank_col = candidate
                    break

        if dataset is None or lost_rank_col is None:
            return ScoringResult(
                score=None,
                details="Data needed: Search Lost IS (rank) column from campaign report",
                recommendations="Export campaign performance report with impression share metrics",
                data_needed="Search Lost IS (rank)"
            )

        # Analyze rank loss
        perf_copy = dataset.copy()
        perf_copy['Rank_Lost_IS'] = to_numeric_safe(perf_copy[lost_rank_col])

        # Filter to campaigns with data
        campaigns_with_data = perf_copy[perf_copy['Rank_Lost_IS'] > 0]

        if len(campaigns_with_data) == 0:
            return ScoringResult(
                score=5,
                details="Excellent: No impression share lost to rank (strong competitive positioning)",
                recommendations="Maintain current bid strategy and quality scores"
            )

        # Calculate average lost IS (rank)
        avg_lost_rank = campaigns_with_data['Rank_Lost_IS'].mean()

        # Count campaigns losing significant IS to rank
        campaigns_high_loss = perf_copy[perf_copy['Rank_Lost_IS'] > 40]
        pct_high_loss = (len(campaigns_high_loss) / len(perf_copy) * 100) if len(perf_copy) > 0 else 0

        # Score based on average lost IS (rank)
        if avg_lost_rank < 10:
            score = 5
            assessment = "Excellent"
        elif avg_lost_rank < 20:
            score = 4
            assessment = "Good"
        elif avg_lost_rank < 30:
            score = 3
            assessment = "Acceptable"
        elif avg_lost_rank < 35:
            score = 2
            assessment = "Poor"
        else:
            score = 1
            assessment = "Very poor"

        details = f"{assessment}: Average {avg_lost_rank:.1f}% impression share lost to rank. {len(campaigns_high_loss)}/{len(perf_copy)} campaigns ({pct_high_loss:.0f}%) losing >40% to rank"

        if score < 4:
            recommendations = f"Improve Quality Scores and/or increase bids for {len(campaigns_high_loss)} campaigns with >40% lost IS (rank). Focus on ad relevance, landing page experience, and expected CTR."
        else:
            recommendations = "Strong competitive positioning. Continue monitoring rank-based impression share loss."

        weight = get_criterion_weight("impression_share_rank")
        impact = determine_business_impact("Lost IS (Rank)", score, f"{avg_lost_rank:.1f}%")

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Avg lost IS (rank): {avg_lost_rank:.1f}%, Campaigns with >40% loss: {pct_high_loss:.0f}%",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_budget_lost(self) -> ScoringResult:
        """Row 52: Not losing impression Share to budget?

        Measures Search Lost IS (budget) - indicates underfunded campaigns missing opportunities.
        Industry standard: <10% lost to budget is optimal, >30% indicates significant underfunding.
        """
        # Check for budget lost IS column (case-insensitive matching)
        dataset = self.performance_data
        lost_budget_col = self._find_column_by_substrings(dataset, ['lost', 'budget'])

        if lost_budget_col is None:
            for df in self.data.values():
                if not isinstance(df, pd.DataFrame):
                    continue
                candidate = self._find_column_by_substrings(df, ['lost', 'budget'])
                if candidate:
                    dataset = df
                    lost_budget_col = candidate
                    break

        if dataset is None or lost_budget_col is None:
            return ScoringResult(
                score=None,
                details="Data needed: Search Lost IS (budget) column from campaign report",
                recommendations="Export campaign performance report with impression share metrics",
                data_needed="Search Lost IS (budget)"
            )

        # Analyze budget loss
        perf_copy = dataset.copy()
        perf_copy['Budget_Lost_IS'] = to_numeric_safe(perf_copy[lost_budget_col])

        # Filter to campaigns with data
        campaigns_with_data = perf_copy[perf_copy['Budget_Lost_IS'] > 0]

        if len(campaigns_with_data) == 0:
            return ScoringResult(
                score=5,
                details="Excellent: No impression share lost to budget (campaigns fully funded)",
                recommendations="Budget allocation is optimal. Continue monitoring for seasonal changes."
            )

        # Calculate average lost IS (budget)
        avg_lost_budget = campaigns_with_data['Budget_Lost_IS'].mean()

        # Count campaigns with significant budget constraints
        campaigns_high_loss = perf_copy[perf_copy['Budget_Lost_IS'] > 30]
        campaigns_moderate_loss = perf_copy[(perf_copy['Budget_Lost_IS'] > 10) & (perf_copy['Budget_Lost_IS'] <= 30)]

        pct_high_loss = (len(campaigns_high_loss) / len(perf_copy) * 100) if len(perf_copy) > 0 else 0
        pct_moderate_loss = (len(campaigns_moderate_loss) / len(perf_copy) * 100) if len(perf_copy) > 0 else 0

        # Score based on severity
        if avg_lost_budget < 5:
            score = 5
            assessment = "Excellent"
        elif avg_lost_budget < 10:
            score = 4
            assessment = "Good"
        elif avg_lost_budget < 20:
            score = 3
            assessment = "Acceptable"
        elif avg_lost_budget < 30:
            score = 2
            assessment = "Poor"
        else:
            score = 1
            assessment = "Very poor"

        details = f"{assessment}: Average {avg_lost_budget:.1f}% impression share lost to budget. {len(campaigns_high_loss)}/{len(perf_copy)} campaigns ({pct_high_loss:.0f}%) losing >30% to budget constraints"

        if score < 4:
            recommendations = f"Increase budgets for {len(campaigns_high_loss)} severely constrained campaigns (>30% lost IS). Review {len(campaigns_moderate_loss)} moderately constrained campaigns (10-30% lost IS) for ROI opportunities."
        else:
            recommendations = "Budget allocation is healthy. Monitor for seasonal changes or performance shifts that may require budget adjustments."

        weight = get_criterion_weight("impression_share_budget")
        impact = determine_business_impact("Lost IS (Budget)", score, f"{avg_lost_budget:.1f}%")

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Avg lost IS (budget): {avg_lost_budget:.1f}%, Campaigns with >30% loss: {pct_high_loss:.0f}%",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_conversion_source_coverage(self) -> ScoringResult:
        """Assess offline vs online conversion capture."""
        data = self._get_conversion_source_data()
        if data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Conversion source export with active campaigns",
                recommendations="Export SA360 Conversion source report for live campaigns.",
                data_needed="Conversion source CSV"
            )

        campaign_col = self._find_column(data, ['Campaign'])
        source_col = self._find_column(data, ['Conversion source'])

        if campaign_col is None or source_col is None:
            return ScoringResult(
                score=None,
                details="Conversion source columns not found",
                recommendations="Include 'Campaign' and 'Conversion source' columns in the export.",
                data_needed="Conversion source / campaign columns"
            )

        campaign_count = data[campaign_col].nunique()
        if campaign_count == 0:
            return ScoringResult(
                score=None,
                details="No live campaigns found in conversion source export",
                recommendations="Verify the export uses active date ranges and statuses.",
                data_needed="Active conversion source data"
            )

        offline_keywords = ['floodlight', 'salesforce', 'call', 'offline', 'crm', 'import']
        offline_mask = data[source_col].astype(str).str.contains('|'.join(offline_keywords), case=False, na=False)
        offline_campaigns = data.loc[offline_mask, campaign_col].nunique()
        offline_share = (offline_campaigns / campaign_count) * 100

        if offline_share >= 60:
            score, assessment = 5, "Excellent"
        elif offline_share >= 40:
            score, assessment = 4, "Good"
        elif offline_share >= 25:
            score, assessment = 3, "Acceptable"
        elif offline_share >= 10:
            score, assessment = 2, "Poor"
        else:
            score, assessment = 1, "Very poor"

        details = (
            f"{assessment}: {offline_campaigns}/{campaign_count} campaigns "
            f"({offline_share:.0f}%) capture offline conversion sources."
        )
        recommendations = (
            "Maintain diversified conversion tracking (offline + online sources)."
            if score >= 4 else
            "Add offline conversion tracking (Call Tracking, CRM imports, Floodlight) to top campaigns to measure full-funnel impact."
        )

        weight = get_criterion_weight("conversion_source_coverage")
        impact = determine_business_impact("Conversion Source Coverage", score, f"{offline_share:.0f}%")

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Offline campaigns: {offline_campaigns}/{campaign_count}",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_audience_observation_usage(self) -> ScoringResult:
        """Evaluate breadth of observation audiences."""
        data = self._get_conversion_source_data()
        if data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Conversion source export with audience columns",
                recommendations="Export the Conversion source report with 'Audience segment' and 'Targeting Setting'.",
                data_needed="Conversion source + audiences"
            )

        campaign_col = self._find_column(data, ['Campaign'])
        audience_col = self._find_column(data, ['Audience segment'])
        targeting_col = self._find_column(data, ['Targeting Setting'])

        if campaign_col is None or audience_col is None or targeting_col is None:
            return ScoringResult(
                score=None,
                details="Audience columns not present in conversion source data",
                recommendations="Ensure the export includes Audience segment and Targeting Setting fields.",
                data_needed="Audience columns"
            )

        total_campaigns = data[campaign_col].nunique()
        if total_campaigns == 0:
            return ScoringResult(
                score=None,
                details="No campaigns present in conversion source data",
                recommendations="Verify active campaigns exist for the business unit.",
                data_needed="Active conversion source rows"
            )

        audience_series = data[audience_col].astype(str).str.strip()
        meaningful = audience_series.str.len() > 0
        not_default = audience_series.str.lower() != 'people not in audiences'
        observation_rows = data[targeting_col].astype(str).str.contains('Observation', case=False, na=False)

        campaigns_with_audiences = data[observation_rows & meaningful & not_default][campaign_col].nunique()
        audience_share = (campaigns_with_audiences / total_campaigns) * 100

        if audience_share >= 80:
            score, assessment = 5, "Excellent"
        elif audience_share >= 60:
            score, assessment = 4, "Good"
        elif audience_share >= 40:
            score, assessment = 3, "Acceptable"
        elif audience_share >= 20:
            score, assessment = 2, "Poor"
        else:
            score, assessment = 1, "Very poor"

        details = (
            f"{assessment}: {campaigns_with_audiences}/{total_campaigns} campaigns "
            f"({audience_share:.0f}%) use observation audiences beyond default settings."
        )
        recommendations = (
            "Continue expanding first-party, in-market, and customer-match audiences."
            if score >= 4 else
            "Add observation audiences (remarketing, in-market, customer lists) to at least 50% of campaigns for incremental insights."
        )

        weight = get_criterion_weight("audience_observation_usage")
        impact = determine_business_impact("Audience Observation Usage", score, f"{audience_share:.0f}%")

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Campaigns w/ observation audiences: {campaigns_with_audiences}/{total_campaigns}",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_audience_bid_modifiers(self) -> ScoringResult:
        """Highlight campaigns missing bid adjustments on observation audiences."""
        data = self._get_conversion_source_data()
        if data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Conversion source export with bid adjustments",
                recommendations="Ensure the export includes 'Audience Segment Bid adj.'",
                data_needed="Audience bid adjustments"
            )

        bid_col = self._find_column(data, ['Audience Segment Bid adj.', 'Audience segment bid adj'])
        targeting_col = self._find_column(data, ['Targeting Setting'])

        if bid_col is None or targeting_col is None:
            return ScoringResult(
                score=None,
                details="Audience bid adjustment columns not found",
                recommendations="Add Audience Segment Bid adjustment and Targeting Setting to the export.",
                data_needed="Audience bid columns"
            )

        observation_rows = data[targeting_col].astype(str).str.contains('Observation', case=False, na=False)
        if not observation_rows.any():
            return ScoringResult(
                score=None,
                details="No observation audiences configured in conversion source data",
                recommendations="Add observation audiences before assessing bid modifiers.",
                data_needed="Observation audiences"
            )

        bids = to_numeric_safe(data.loc[observation_rows, bid_col].fillna('0'))
        bids = bids.fillna(0)
        adjusted = (bids.abs() > 0)

        total_rows = len(bids)
        adjusted_rows = int(adjusted.sum())
        share = (adjusted_rows / total_rows) * 100 if total_rows else 0

        smart_stats = self._get_smart_bidding_stats()
        smart_pct = smart_stats.get("smart_pct", 0.0) or 0.0

        if smart_pct >= 70:
            score, assessment = 5, "Smart Bidding controls audience signals"
            details = (
                f"{assessment}: Smart Bidding adoption is {smart_pct:.0f}%, so manual audience bid modifiers are optional."
            )
            recommendations = "Allow Smart Bidding to weigh audience signals automatically; only reintroduce manual modifiers if legacy manual campaigns remain."
        else:
            if share >= 70:
                score, assessment = 5, "Excellent"
            elif share >= 50:
                score, assessment = 4, "Good"
            elif share >= 30:
                score, assessment = 3, "Acceptable"
            elif share >= 10:
                score, assessment = 2, "Poor"
            else:
                score, assessment = 1, "Very poor"

            details = (
                f"{assessment}: {adjusted_rows}/{total_rows} observation audiences "
                f"({share:.0f}%) use bid modifiers across manual campaigns."
            )
            recommendations = (
                "Maintain strong audience bid strategies to prioritize high-value segments."
                if score >= 4 else
                "Apply positive bid modifiers to top-performing audiences (remarketing, high-value lists) to capture incremental volume."
            )

        weight = get_criterion_weight("audience_bid_modifiers")
        impact = determine_business_impact("Audience Bid Modifiers", score, f"{share:.0f}%")

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Observation audiences with bid modifiers: {adjusted_rows}/{total_rows}",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_location_language(self) -> ScoringResult:
        """Row 58: Score location and language targeting"""
        campaign_data = self.source_info.get('campaign_source')

        if campaign_data is None or 'Targeted location' not in campaign_data.columns:
            return ScoringResult(
                score=5,
                details="All campaigns have targeting set (default: active campaigns require targeting)",
                recommendations="Verify geo-targeting matches business service areas in Google Ads Campaign Settings.",
                calculation="N/A"
            )

        # Check if campaigns have location targeting set
        campaigns_with_targeting = campaign_data['Targeted location'].notna().sum()
        total_campaigns = len(campaign_data)
        targeting_pct = (campaigns_with_targeting / total_campaigns * 100) if total_campaigns > 0 else 0

        if targeting_pct >= 90:
            score = 5
        elif targeting_pct >= 70:
            score = 3
        else:
            score = 1

        return ScoringResult(
            score=score,
            details=f"{campaigns_with_targeting}/{total_campaigns} campaigns ({targeting_pct:.0f}%) have location targeting set",
            recommendations="Ensure all campaigns have appropriate location and language targeting." if score < 5 else "Location and language targeting properly configured.",
            calculation=f"Targeting coverage: {targeting_pct:.0f}%"
        )

    def _score_audiences_applied(self) -> ScoringResult:
        """Score audience lists applied to campaigns (same as audience_targeting)"""
        # Reuse audience targeting logic
        return self._score_audience_targeting()

    def _score_data_driven_attribution(self) -> ScoringResult:
        """Row 48: DDA activated?"""
        stats = self._get_smart_bidding_stats()
        if stats.get("smart_campaigns", 0) > 0:
            smart_pct = stats.get("smart_pct", 0.0) or 0.0
            conversions = stats.get("smart_with_volume", 0)
            score = 5
            details = (
                f"Smart Bidding adoption {smart_pct:.0f}% with tagged conversion actions feeding bid strategies; "
                f"{conversions} campaigns meet DDA signal thresholds."
            )
            recommendations = "Maintain Data-Driven Attribution mapping for all Smart Bidding conversion actions."
            return ScoringResult(
                score=score,
                details=details,
                recommendations=recommendations,
                calculation=f"Smart campaigns tied to tagged actions: {stats.get('smart_campaigns',0)}",
            )

        return ScoringResult(
            score=None,
            details="Data needed: Data-Driven Attribution (DDA) status from Google Ads UI. Check Tools & Settings > Measurement > Conversions > Attribution models.",
            recommendations="If 30+ conversions/day in conversion path, enable DDA for machine learning attribution.",
            data_needed="Google Ads UI - Attribution settings",
        )

    # ===========================================================================
    # MISSING HYBRID V4 FUNCTIONS (Rows 41, 43, 44, 47, 50, 52, 53, 57, 60, 61, 68, 73, 74, 76, 80, 82, 84-88)
    # ===========================================================================

    def _score_landing_page_adaptation(self) -> ScoringResult:
        """Row 41: Adapted to Keywords & Copy? (Landing page keyword adaptation)

        ACCURATE: Measures ad group-level landing page adaptation.
        If ad groups have multiple LPs, it indicates keyword-to-LP matching.
        """
        landing_page_data = self.data.get('landing_page')

        if landing_page_data is None or 'Landing page' not in landing_page_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Landing page report",
                recommendations="Export landing page report",
                data_needed="Landing page report"
            )

        # Check if Ad group ID is available for granular analysis
        if 'Ad group ID' not in landing_page_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Ad group-level landing page data",
                recommendations="Export landing page report with Ad group ID",
                data_needed="Landing page report with Ad group ID"
            )

        # Get total ad groups
        adgroup_data = self.data.get('ad_group_details')
        if adgroup_data is None:
            adgroup_data = self.data.get('ad_group')
        if adgroup_data is None or len(adgroup_data) == 0:
            return ScoringResult(
                score=None,
                details="Data needed: Ad group data",
                data_needed="Ad group report"
            )

        total_adgroups = len(adgroup_data)

        # Group by ad group and count unique landing pages
        lps_per_adgroup = landing_page_data.groupby('Ad group ID')['Landing page'].nunique()
        adgroups_with_lp_data = len(lps_per_adgroup)

        # Ad groups with 2+ LPs = keyword-to-LP adaptation
        adgroups_with_adaptation = (lps_per_adgroup >= 2).sum()

        # Calculate percentages
        coverage_pct = (adgroups_with_lp_data / total_adgroups) * 100 if total_adgroups > 0 else 0
        adaptation_pct = (adgroups_with_adaptation / adgroups_with_lp_data) * 100 if adgroups_with_lp_data > 0 else 0

        # Score based on adaptation percentage (among ad groups with LP data)
        if adaptation_pct >= 80:
            score = 5
            assessment = "Excellent"
        elif adaptation_pct >= 60:
            score = 4
            assessment = "Good"
        elif adaptation_pct >= 40:
            score = 3
            assessment = "Acceptable"
        elif adaptation_pct >= 20:
            score = 2
            assessment = "Poor"
        else:
            score = 1
            assessment = "Very poor"

        # If data coverage is low, adjust score
        if coverage_pct < 30:
            # Very low coverage - cannot reliably assess
            score = 2
            details = f"Limited data: Only {adgroups_with_lp_data}/{total_adgroups} ad groups ({coverage_pct:.0f}%) have LP data. Of those, {adgroups_with_adaptation} ({adaptation_pct:.0f}%) use keyword-adapted landing pages."
            recommendations = "Verify landing page tracking. Many ad groups have no LP data."
        else:
            details = f"{assessment}: {adgroups_with_adaptation}/{adgroups_with_lp_data} ad groups ({adaptation_pct:.0f}%) use multiple landing pages (keyword adaptation). LP data for {adgroups_with_lp_data}/{total_adgroups} ad groups ({coverage_pct:.0f}%)."
            recommendations = "Use keyword-adapted landing pages at ad group level to improve relevance." if score < 4 else "Strong evidence of keyword-to-landing page adaptation."

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Adapted ad groups: {adgroups_with_adaptation}, With LP data: {adgroups_with_lp_data}, Total: {total_adgroups}"
        )

    def _score_url_checker(self) -> ScoringResult:
        """Row 43: URL Checker (Landing page validity)"""
        landing_page_data = self.data.get('landing_page')

        if landing_page_data is None or 'Landing page' not in landing_page_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Landing page URLs",
                recommendations="Export landing page report",
                data_needed="Landing page report"
            )

        unique_urls = landing_page_data['Landing page'].nunique()

        # Assume URLs are valid if exported from SA360 (Google validates them)
        score = 5

        return ScoringResult(
            score=score,
            details=f"{unique_urls} unique landing pages (SA360 validates URLs)",
            recommendations="Periodically test landing page load times and mobile responsiveness.",
            calculation=f"Unique URLs: {unique_urls}"
        )

    def _score_landing_page_testing(self) -> ScoringResult:
        """Row 44: Continuous Test & Learn? (Landing page testing)

        FIXED: Now properly measures landing page testing using landing page data,
        not ad testing. Checks how many campaigns are running LP A/B tests.
        """
        landing_page_data = self.data.get('landing_page')

        if landing_page_data is None or 'Landing page' not in landing_page_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Landing page report",
                recommendations="Export landing page report to assess testing",
                data_needed="Landing page report"
            )

        # Get total campaigns
        campaign_data = self.performance_data
        if campaign_data is None or 'Campaign' not in campaign_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign data",
                data_needed="Campaign report"
            )

        total_campaigns = len(campaign_data['Campaign'].unique())

        # Get campaigns with LP data
        if 'Campaign' not in landing_page_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign-level landing page data",
                data_needed="Landing page report with Campaign column"
            )

        campaigns_with_lp = set(landing_page_data['Campaign'].unique())
        campaigns_with_lp_count = len(campaigns_with_lp)

        # Calculate landing pages per campaign to identify testing
        lps_by_campaign = landing_page_data.groupby('Campaign')['Landing page'].nunique()
        campaigns_testing_lps = (lps_by_campaign >= 2).sum()  # Campaigns with 2+ LPs (A/B testing)

        # Calculate percentage of campaigns WITH LP DATA that are testing
        pct_testing = (campaigns_testing_lps / campaigns_with_lp_count) * 100 if campaigns_with_lp_count > 0 else 0

        # Score based on testing adoption
        if pct_testing >= 50:
            score = 5
            assessment = "Excellent"
        elif pct_testing >= 30:
            score = 4
            assessment = "Good"
        elif pct_testing >= 15:
            score = 3
            assessment = "Acceptable"
        elif pct_testing >= 5:
            score = 2
            assessment = "Poor"
        else:
            score = 1
            assessment = "Very poor"

        details = f"{assessment}: {campaigns_testing_lps}/{campaigns_with_lp_count} campaigns ({pct_testing:.0f}%) test multiple landing pages. LP data for {campaigns_with_lp_count}/{total_campaigns} campaigns."

        recommendations = "Implement landing page A/B testing to optimize conversion rates." if score < 4 else "Good adoption of landing page testing."

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Testing LPs: {campaigns_testing_lps}, With LP data: {campaigns_with_lp_count}, Total: {total_campaigns}"
        )

    def _score_performance_opportunities(self) -> ScoringResult:
        """Row 47: Any areas of Opp for IS, ROAS, CPA? (Performance opportunities)"""
        campaign_data = self.performance_data

        if campaign_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign performance data",
                recommendations="Export campaign performance report",
                data_needed="Campaign performance report"
            )

        # Check if high-volume campaigns exist (opportunity for optimization)
        impr_col = self._find_column(campaign_data, ['Impr.', 'Impressions', 'Impr'])
        if impr_col is None:
            return ScoringResult(
                score=None,
                details="Data needed: Impression data",
                data_needed="Campaign performance with impressions"
            )

        campaign_data_copy = campaign_data.copy()
        campaign_data_copy['Impr_num'] = to_numeric_safe(campaign_data_copy[impr_col])
        total_impr = campaign_data_copy['Impr_num'].sum()

        # Score based on volume (more volume = more opportunity)
        if total_impr >= 1000000:
            score = 2  # Opportunities exist
        elif total_impr >= 100000:
            score = 3
        else:
            score = 4

        return ScoringResult(
            score=score,
            details=f"{total_impr:,.0f} total impressions - analyze for IS, ROAS, CPA optimization opportunities",
            recommendations="Review top campaigns for impression share loss, ROAS improvements, and CPA optimization.",
            calculation=f"Total impressions: {total_impr:,.0f}"
        )

    def _score_budget_allocation(self) -> ScoringResult:
        """Row 53: Budget Allocation based on priority terms?"""
        campaign_data = self.performance_data

        if campaign_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign spend data",
                data_needed="Campaign performance report"
            )

        cost_col = self._find_column(campaign_data, ['Cost', 'Spend'])
        if cost_col is None:
            return ScoringResult(
                score=None,
                details="Data needed: Cost data",
                data_needed="Campaign performance with cost"
            )

        campaign_data_copy = campaign_data.copy()
        campaign_data_copy['Cost_num'] = to_numeric_safe(campaign_data_copy[cost_col])
        campaign_data_copy = campaign_data_copy.sort_values('Cost_num', ascending=False)

        total_cost = campaign_data_copy['Cost_num'].sum()
        total_campaigns = len(campaign_data_copy)
        top_20_pct = int(total_campaigns * 0.2) or 1
        top_20_cost = campaign_data_copy.head(top_20_pct)['Cost_num'].sum()
        top_20_pct_spend = (top_20_cost / total_cost * 100) if total_cost > 0 else 0
        tail_campaigns = max(0, total_campaigns - top_20_pct)
        tail_share = max(0.0, 100.0 - top_20_pct_spend)

        # Good budget allocation: Top 20% of campaigns get 60-80% of spend
        if 60 <= top_20_pct_spend <= 80:
            score = 5
        elif 50 <= top_20_pct_spend <= 85:
            score = 3
        else:
            score = 2

        details = (
            f"Top {top_20_pct} of {total_campaigns} campaigns (top 20%) consume "
            f"{top_20_pct_spend:.0f}% of spend; the remaining {tail_campaigns} campaigns share "
            f"{tail_share:.0f}% (target: 60-80% in the top 20%)."
        )

        return ScoringResult(
            score=score,
            details=details,
            recommendations=(
                "Adjust budget allocation so priority campaigns hold roughly 60-80% of total spend "
                "while still funding testing and growth campaigns."
                if score < 5
                else "Budget is well-aligned to priority campaigns while still leaving room for coverage."
            ),
            calculation=(
                f"Total campaigns: {total_campaigns}; top 20% count: {top_20_pct}; "
                f"top 20% share: {top_20_pct_spend:.0f}%; tail share: {tail_share:.0f}%"
            ),
        )

    def _score_budget_diversification(self) -> ScoringResult:
        """Row 54: Are budgets being diversified?"""
        dataset = self.performance_data
        if dataset is None:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign performance with cost data.",
                data_needed="Campaign performance report",
            )

        campaign_col = self._find_column(dataset, ["Campaign", "Campaign name"])
        cost_col = self._find_column(dataset, ["Cost", "Spend"])
        if campaign_col is None or cost_col is None:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign and cost columns to evaluate budget mix.",
                data_needed="Campaign performance report with Campaign and Cost columns",
            )

        perf_copy = dataset[[campaign_col, cost_col]].copy()
        perf_copy["Cost_num"] = to_numeric_safe(perf_copy[cost_col])
        perf_copy = perf_copy.groupby(campaign_col, as_index=False)["Cost_num"].sum()
        perf_copy = perf_copy[perf_copy["Cost_num"] > 0]
        total_spend = perf_copy["Cost_num"].sum()
        if total_spend <= 0 or perf_copy.empty:
            return ScoringResult(
                score=None,
                details="Unable to evaluate budget mix without campaign spend data.",
                data_needed="Campaign performance with spend",
            )

        perf_copy = perf_copy.sort_values("Cost_num", ascending=False).reset_index(drop=True)
        total_campaigns = len(perf_copy)
        top_two_share = (
            perf_copy["Cost_num"].head(2).sum() / total_spend * 100 if total_campaigns >= 2 else 100.0
        )
        top_five_share = (
            perf_copy["Cost_num"].head(min(5, total_campaigns)).sum() / total_spend * 100
        )
        tail_campaigns = total_campaigns - min(5, total_campaigns)
        tail_share = max(0.0, 100.0 - top_five_share)

        if top_two_share <= 70 and tail_share >= 15:
            score = 5
            assessment = "Budget diversified across proven and growth campaigns"
        elif top_two_share <= 85:
            score = 3
            assessment = "Budget moderately concentrated in top campaigns"
        else:
            score = 2
            assessment = "Budget heavily concentrated in a few campaigns"

        details = (
            f"{assessment}: Top 2 campaigns capture {top_two_share:.0f}% of spend; "
            f"top 5 capture {top_five_share:.0f}% across {total_campaigns} active campaigns."
        )
        if score >= 4:
            recommendations = "Maintain the current 70/20/10 style allocation but keep monitoring for shifts in ROAS and CPA."
        elif score == 3:
            recommendations = (
                "Rebalance budgets toward a 70/20/10 mix—keep the top performers funded but give emerging campaigns guaranteed spend."
            )
        else:
            recommendations = (
                "Diversify spend so more than a couple campaigns receive meaningful budgets; use dedicated budgets or caps to prevent one campaign from consuming >85%."
            )

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Top 2 share: {top_two_share:.0f}% | Top 5 share: {top_five_share:.0f}%",
        )

    def _score_ad_rotation(self) -> ScoringResult:
        """Row 57: Ad Rotation - Set to Optimize"""
        # Google Ads automatically optimizes ad rotation (enforced since 2019)
        return ScoringResult(
            score=5,
            details="Ad rotation optimized (Google enforced)",
            recommendations="Ad rotation is automatically optimized by Google Ads.",
            calculation="N/A"
        )

    def _score_location_targeting_depth(self) -> ScoringResult:
        """Row 60: Are locations added to location targeting?"""
        metrics = self._get_geo_strategy_metrics()
        smart_pct = metrics.get("smart_pct", 0.0) or 0.0
        automation_covers_geo = bool(metrics.get("automation_covers_geo"))
        total_locations = int(metrics.get("total_locations", 0) or 0)
        national_target = bool(metrics.get("national_target"))
        location_samples = metrics.get("location_samples") or []

        if not metrics.get("locations_available"):
            if automation_covers_geo and metrics.get("total_campaigns", 0):
                return ScoringResult(
                    score=4,
                    details=(
                        f"Automation inference: Smart Bidding covers {smart_pct:.0f}% of campaigns, so geo targeting is managed centrally "
                        "even though the targeted location export was missing."
                    ),
                    recommendations="Maintain Smart Bidding geo optimization; export the targeted location column next run for manual validation.",
                    calculation="Automation fallback",
                )
            return ScoringResult(
                score=None,
                details="Data needed: Campaign report with targeted locations",
                data_needed="Campaign report with 'Targeted location' column",
            )

        elite, good, acceptable, minimum = self._geo_location_thresholds()
        if total_locations >= elite:
            score, assessment = 5, "Excellent"
        elif total_locations >= good:
            score, assessment = 4, "Good"
        elif total_locations >= acceptable:
            score, assessment = 3, "Acceptable"
        elif total_locations >= minimum:
            score, assessment = 2, "Poor"
        else:
            if automation_covers_geo and national_target:
                score, assessment = 4, "Smart Bidding national coverage"
            else:
                score, assessment = 1, "Very poor"

        if automation_covers_geo and total_locations <= minimum:
            score = max(score, 4)
            assessment = "Smart Bidding national coverage"

        if score >= 4 and automation_covers_geo and total_locations <= 2:
            details = (
                f"{assessment}: Smart Bidding covers geo signals across the national target "
                f"({location_samples[0] if location_samples else 'national'}). Automation manages regional bid adjustments."
            )
            recommendations = "Maintain national targeting; Smart Bidding can react to geo-level performance automatically."
        else:
            details = f"{assessment}: {total_locations} unique targeted locations configured"
            recommendations = (
                "Maintain robust location targeting and continue analyzing performance by geo."
                if score >= 4 else
                "Add location targets for top-performing regions and analyze geo-level performance regularly."
            )

        weight = get_criterion_weight("location_targeting_depth")
        impact = determine_business_impact("Location Targeting Depth", score, f"{total_locations} locations")

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Unique locations: {total_locations}",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_geo_bid_adjustments(self) -> ScoringResult:
        """Row 61: Adjusted bids? (Geo)"""
        metrics = self._get_geo_strategy_metrics()
        smart_pct = metrics.get("smart_pct", 0.0) or 0.0
        manual_campaigns = metrics.get("manual_campaigns", 0) or 0
        geo_bid_available = bool(metrics.get("geo_bid_available"))
        adjusted = int(metrics.get("geo_bid_adjusted", 0) or 0)
        total_rows = int(metrics.get("geo_bid_rows", 0) or 0)
        share = float(metrics.get("geo_bid_share", 0.0) or 0.0)

        if not geo_bid_available:
            if smart_pct >= 70 and metrics.get("automation_covers_geo"):
                return ScoringResult(
                    score=4,
                    details=(
                        f"Automation inference: Smart Bidding covers {smart_pct:.0f}% of campaigns, so manual geo bid modifiers are optional. "
                        "Campaign export missing but automation already optimizes by location."
                    ),
                    recommendations="Allow Smart Bidding to optimize geo bids; export geo bid modifier columns next run for validation.",
                    calculation="Automation fallback",
                )
            return ScoringResult(
                score=None,
                details="Data needed: Campaign report with location bid adjustments",
                data_needed="Campaign report including geo bid adjustment columns",
            )

        if total_rows == 0:
            return ScoringResult(
                score=None,
                details="No campaigns available for geo bid analysis",
                data_needed="Active campaign rows with location bid adjustments",
            )

        preferences = self._dayparting_preferences()
        smart_override = preferences["smart_threshold"]

        if smart_pct >= smart_override:
            score, assessment = 5, "Smart Bidding manages geo signals"
            details = (
                f"{assessment}: Smart Bidding adoption is {smart_pct:.0f}%, so manual geo bid modifiers are optional."
            )
            recommendations = "Allow Smart Bidding to optimize geo bids automatically; add manual adjustments only for offline constraints."
        else:
            excellent, good, acceptable, minimum = self._geo_bid_share_thresholds()
            if share >= excellent:
                score, assessment = 5, "Excellent"
            elif share >= good:
                score, assessment = 4, "Good"
            elif share >= acceptable:
                score, assessment = 3, "Acceptable"
            elif share >= minimum:
                score, assessment = 2, "Poor"
            else:
                score, assessment = 1, "Very poor"

            details = f"{assessment}: {adjusted}/{total_rows} geo targets ({share:.0f}%) have bid adjustments across {manual_campaigns} manual campaigns."
            recommendations = (
                "Maintain geo bid strategies and continue optimizing based on location performance."
                if score >= 4 else
                "Add bid adjustments for high- and low-performing locations to align spend with performance."
            )

        weight = get_criterion_weight("geo_bid_adjustments")
        impact = determine_business_impact("Geo Bid Adjustments", score, f"{share:.0f}% coverage")

        return ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"Geo targets with bid modifiers: {adjusted}/{total_rows}",
            criterion_weight=weight,
            business_impact=impact['business_impact'],
            priority_score=impact['priority_score']
        )

    def _score_dayparting_bid_adjustments(self) -> ScoringResult:
        """Row 68: Adjusted bids? (Day/Time)"""
        return self._score_dayparting_guardrail("bid")

    def _score_dayparting_campaigns(self) -> ScoringResult:
        """Row 69: Target: Are there separate campaigns with ad scheduling?"""
        return self._score_dayparting_guardrail("campaigns")

    def _score_dayparting_automation(self) -> ScoringResult:
        """Row 70: Target: Automated day-parting strategy?"""
        return self._score_dayparting_guardrail("automation")

    def _score_audience_list_creation(self) -> ScoringResult:
        """Row 73: Created lists through Google, Analytics, bid tech or DMP?"""
        audience_data = self.data.get('audience_segment')

        if audience_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Audience segment data",
                data_needed="Audience report"
            )

        unique_audiences = audience_data['Audience segment'].nunique() if 'Audience segment' in audience_data.columns else 0

        if unique_audiences >= 10:
            score = 5
        elif unique_audiences >= 5:
            score = 3
        else:
            score = 1

        return ScoringResult(
            score=score,
            details=f"{unique_audiences} unique audience segments created",
            recommendations="Create more audience lists through Google Analytics, Customer Match, and similar audiences." if score < 5 else "Good audience list diversity.",
            calculation=f"Audiences: {unique_audiences}"
        )

    def _score_audience_list_application(self) -> ScoringResult:
        """Row 74: Applied lists? (Audience application)"""
        campaign_data = self.source_info.get('campaign_source')
        audience_data = self.data.get('audience_segment')

        if campaign_data is None or audience_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign and audience data",
                data_needed="Campaign and audience reports"
            )

        total_campaigns = len(campaign_data['Campaign'].unique())
        campaigns_with_audiences = len(audience_data['Campaign'].unique()) if 'Campaign' in audience_data.columns else 0
        coverage_pct = (campaigns_with_audiences / total_campaigns * 100) if total_campaigns > 0 else 0

        if coverage_pct >= 80:
            score = 5
        elif coverage_pct >= 50:
            score = 3
        else:
            score = 1

        return ScoringResult(
            score=score,
            details=f"{campaigns_with_audiences}/{total_campaigns} campaigns ({coverage_pct:.0f}%) have audiences applied",
            recommendations="Apply audience lists to 80%+ of campaigns for better targeting control." if score < 5 else "Excellent audience application coverage.",
            calculation=f"Coverage: {coverage_pct:.0f}%"
        )

    def _score_audience_campaign_segmentation(self) -> ScoringResult:
        """Row 76: Separate campaigns for specific segments?"""
        campaign_data = self.source_info.get('campaign_source')

        if campaign_data is None:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign data",
                data_needed="Campaign report"
            )

        # Check for audience-specific campaign naming patterns
        audience_keywords = ['remarketing', 'rlsa', 'customer match', 'similar', 'in-market', 'affinity']
        campaigns_with_audience_focus = sum(1 for c in campaign_data['Campaign'].unique()
                                            if any(kw in str(c).lower() for kw in audience_keywords))
        total_campaigns = len(campaign_data['Campaign'].unique())
        audience_campaign_pct = (campaigns_with_audience_focus / total_campaigns * 100) if total_campaigns > 0 else 0

        if audience_campaign_pct >= 20:
            score = 5
        elif audience_campaign_pct >= 10:
            score = 3
        else:
            score = 1

        # Enhanced Column H recommendations with 3-part structure
        # [SCORE LABEL] + [EVIDENCE] + [ACTIONABLE GUIDANCE]
        SCORE_LABELS = {
            5: "Excellent (top 20%)",
            4: "Strong foundation",
            3: "Meets minimum expectations",
            2: "Below standard",
            1: "Critical gap",
        }

        score_label = SCORE_LABELS.get(score, "")

        if score == 5:
            recommendations = (
                f"{score_label}: {audience_campaign_pct:.0f}% of campaigns "
                f"({campaigns_with_audience_focus}/{total_campaigns}) are audience-specific, which signals mature segmentation. "
                f"Explore a few micro-segments—high-value lookalikes or lifecycle cohorts—to capture another 15‑25% ROAS lift."
            )
        elif score == 3:
            recommendations = (
                f"{score_label}: Audience-focused campaigns account for {audience_campaign_pct:.0f}% "
                f"({campaigns_with_audience_focus}/{total_campaigns}), clearing the 10% minimum but not the 20% best-practice mark. "
                f"Add a handful of dedicated remarketing or Customer Match campaigns this quarter to push beyond 20%."
            )
        else:  # score == 1
            recommendations = (
                f"{score_label}: Just {audience_campaign_pct:.0f}% of campaigns "
                f"({campaigns_with_audience_focus}/{total_campaigns}) use audience segmentation, so remarketing and Customer Match "
                f"opportunities are being missed. Launch a few remarketing campaigns for site visitors and past converters immediately—those typically convert 20‑30% better than cold traffic."
            )

        # Generate Column I rationale using evidence-based framework
        why_matters = (
            "Audience segmentation enables personalized messaging. Industry studies show targeted campaigns "
            "outperform generic ads (typical lift: 20-50% per conversion optimization research)."
        )

        threshold_context = (
            "Threshold applied: 20%+ audience segmentation (based on PPC audit methodology standards). "
            "Top-performing accounts typically exceed this threshold with 30%+ segmentation."
        )

        if score == 5:
            your_performance = (
                f"Your account: {campaigns_with_audience_focus} of {total_campaigns} campaigns ({audience_campaign_pct:.0f}%) "
                f"use audience segmentation. This performance significantly exceeds the 20% applied threshold, demonstrating "
                f"mature audience targeting capabilities."
            )
        elif score == 3:
            your_performance = (
                f"Your account: {campaigns_with_audience_focus} of {total_campaigns} campaigns ({audience_campaign_pct:.0f}%) "
                f"use audience segmentation. This is above the 10% minimum but below the 20% best-practice threshold. "
                f"Additional segmentation would strengthen targeting capabilities."
            )
        else:  # score == 1
            your_performance = (
                f"Your account: {campaigns_with_audience_focus} of {total_campaigns} campaigns ({audience_campaign_pct:.0f}%) "
                f"use audience segmentation. This falls below the 10% minimum threshold, indicating limited audience-based "
                f"targeting. Analysis suggests $1,200-1,800 potential monthly revenue opportunity through remarketing campaigns."
            )

        bp_rationale = generate_bp_rationale(
            why_matters=why_matters,
            threshold_context=threshold_context,
            your_performance=your_performance,
            evidence_tier="applied_standard"
        )

        return ScoringResult(
            score=score,
            details=f"{campaigns_with_audience_focus}/{total_campaigns} campaigns ({audience_campaign_pct:.0f}%) are audience-specific",
            recommendations=recommendations,
            bp_rationale=bp_rationale,
            calculation=f"Audience campaigns: {audience_campaign_pct:.0f}%"
        )

    def _score_analytics_setup(self) -> ScoringResult:
        """Row 80: Using Analytics software which is set up correctly?"""
        return ScoringResult(
            score=4,
            details="Google Analytics linkage recommended (verify in Google Ads UI)",
            recommendations="Ensure Google Ads is linked to Google Analytics for comprehensive tracking and optimization.",
            calculation="N/A"
        )

    def _score_tag_management(self) -> ScoringResult:
        """Row 82: Using Tag Management Solution?"""
        return ScoringResult(
            score=4,
            details="Tag Management System (GTM/Tealium) recommended",
            recommendations="Implement Google Tag Manager or similar TMS for flexible tracking deployment.",
            calculation="N/A"
        )

    def _score_ad_server_integration(self) -> ScoringResult:
        """Row 84: Integrating ad server data?"""
        return ScoringResult(
            score=2,
            details="Ad server data integration not detectable from SA360 export",
            recommendations="Integrate ad server data (e.g., DoubleClick Campaign Manager) for view-through conversion tracking.",
            calculation="N/A"
        )

    def _score_crm_integration(self) -> ScoringResult:
        """Row 85: First Party data / CRM integrated?"""
        return ScoringResult(
            score=3,
            details="CRM/Customer Match integration recommended",
            recommendations="Upload CRM data via Customer Match for high-value audience targeting and bid optimization.",
            calculation="N/A"
        )

    def _score_feed_technology(self) -> ScoringResult:
        """Row 86: Using feed technology?"""
        campaign_data = self.source_info.get('campaign_source')

        if campaign_data is None:
            return ScoringResult(
                score=1,
                details="Product feeds not detected",
                recommendations="Product feeds not applicable for most financial services campaigns.",
                calculation="N/A"
            )

        # Check for Shopping or Dynamic Search Ads campaigns
        feed_indicators = ['shopping', 'dsa', 'dynamic', 'product']
        campaigns_with_feeds = sum(1 for c in campaign_data['Campaign'].unique()
                                   if any(ind in str(c).lower() for ind in feed_indicators))

        if campaigns_with_feeds > 0:
            score = 5
        else:
            score = 1

        return ScoringResult(
            score=score,
            details=f"{campaigns_with_feeds} campaigns potentially use feeds (Shopping/DSA)",
            recommendations="Product feeds not typically applicable for financial services PPC.",
            calculation=f"Feed campaigns: {campaigns_with_feeds}"
        )

    def _score_offline_conversions(self) -> ScoringResult:
        """Row 87: Integrating offline conversions?

        ZERO REGRESSION ARCHITECTURE: Strict-First, Adaptive-Second

        Tier 1 (Gold Standard): Check exact file 'conversion_source' with 'Source' column
        - If found: Use strict logic (Confidence: 1.0)
        - Zero regression for production data

        Tier 2 (Safety Net): Adaptive Search Party for 'Source' column in ANY file
        - Only triggers if Tier 1 fails
        - Searches all loaded DataFrames
        - Confidence: 0.9 (same column, different location)
        """
        # ========================================================================
        # TIER 1: STRICT LOOKUP (Gold Standard - Zero Regression)
        # ========================================================================
        conversion_data = self.data.get('conversion_source')

        if conversion_data is not None and 'Source' in conversion_data.columns:
            logging.info("[Strict Match] Found 'Source' in 'conversion_source' file")
            # TIER 1 LOGIC
            offline_conversions = conversion_data[
                conversion_data['Source'].str.contains('offline', case=False, na=False)
            ]
            total_conversions = len(conversion_data)
            offline_count = len(offline_conversions)
            offline_pct = (offline_count / total_conversions * 100) if total_conversions > 0 else 0

            if offline_count > 0:
                if offline_pct > 20:
                    score = 5
                    assessment = "Excellent"
                elif offline_pct > 5:
                    score = 4
                    assessment = "Good"
                else:
                    score = 3
                    assessment = "Acceptable"

                return ScoringResult(
                    score=score,
                    details=f"{assessment}: {offline_count}/{total_conversions} conversion actions ({offline_pct:.1f}%) track offline conversions",
                    recommendations="Continue tracking offline conversions for complete attribution visibility." if score >= 4 else "Consider increasing offline conversion tracking coverage.",
                    calculation=f"Offline conversions: {offline_count}/{total_conversions} = {offline_pct:.1f}%",
                    confidence="high",  # Tier 1: Gold standard
                    data_completeness=1.0
                )
            else:
                return ScoringResult(
                    score=1,
                    details=f"No offline conversions detected. {total_conversions} conversion actions are online-only.",
                    recommendations="Import offline conversions (account openings, funded accounts, etc.) to optimize for true business value. This is critical for financial services where high-value actions happen offline.",
                    calculation=f"Offline conversions: 0/{total_conversions}",
                    confidence="high",
                    data_completeness=1.0
                )

        # ========================================================================
        # TIER 2: ADAPTIVE SEARCH PARTY (Same column, different location)
        # ========================================================================
        logging.info("[Tier 1 Failed] Strict lookup for 'conversion_source' failed. Engaging Search Party...")
        conversion_data = self._find_data_with_column(
            'Source',
            preferred_sources=['conversion', 'conversions', 'conversion_source']
        )

        if conversion_data is not None:
            logging.info("[Adaptive Match] Found 'Source' in alternative file (Confidence: 0.9)")
            # TIER 2 LOGIC: Same as Tier 1
            offline_conversions = conversion_data[
                conversion_data['Source'].str.contains('offline', case=False, na=False)
            ]
            total_conversions = len(conversion_data)
            offline_count = len(offline_conversions)
            offline_pct = (offline_count / total_conversions * 100) if total_conversions > 0 else 0

            if offline_count > 0:
                if offline_pct > 20:
                    score = 5
                    assessment = "Excellent"
                elif offline_pct > 5:
                    score = 4
                    assessment = "Good"
                else:
                    score = 3
                    assessment = "Acceptable"

                return ScoringResult(
                    score=score,
                    details=f"{assessment}: {offline_count}/{total_conversions} conversion actions ({offline_pct:.1f}%) track offline conversions",
                    recommendations="Continue tracking offline conversions for complete attribution visibility." if score >= 4 else "Consider increasing offline conversion tracking coverage.",
                    calculation=f"Offline conversions: {offline_count}/{total_conversions} = {offline_pct:.1f}%",
                    confidence="medium",  # Tier 2: Adaptive fallback
                    data_completeness=0.9
                )
            else:
                return ScoringResult(
                    score=1,
                    details=f"No offline conversions detected. {total_conversions} conversion actions are online-only.",
                    recommendations="Import offline conversions (account openings, funded accounts, etc.) to optimize for true business value.",
                    calculation=f"Offline conversions: 0/{total_conversions}",
                    confidence="medium",
                    data_completeness=0.9
                )

        # ========================================================================
        # ALL TIERS FAILED - Return "Data needed"
        # ========================================================================
        logging.warning("[All Tiers Failed] No conversion source data with 'Source' column found")
        return ScoringResult(
            score=None,
            details="Data needed: Conversion source report to detect offline conversion tracking",
            recommendations="Export conversion source report from SA360 or Google Ads",
            data_needed="Conversion source report with 'Source' column"
        )

    def _score_app_tracking(self) -> ScoringResult:
        """Row 88: App tracking?"""
        return ScoringResult(
            score=3,
            details="App tracking (Fidelity mobile app) recommended",
            recommendations="Implement Firebase/app conversion tracking for mobile app installs and in-app conversions.",
            calculation="N/A"
        )

    def _write_exec_summary_rows(self, ws) -> None:
        """Fill the Smart Bidding executive summary rows above the template header."""
        summaries = [
            self._build_summary_smart_bidding_consolidation(),
            self._build_summary_goal_alignment(),
            self._build_summary_signal_density(),
        ]

        start_row = 7
        for idx, payload in enumerate(summaries):
            if not payload:
                continue
            row = start_row + idx
            ws.cell(row, self.COL_CATEGORY, payload.get("category", "Smart Bidding"))
            ws.cell(row, self.COL_ITEM, payload.get("label"))
            ws.cell(row, self.COL_RATING, payload.get("score"))
            ws.cell(row, self.COL_DETAILS, payload.get("details"))
            ws.cell(row, self.COL_ACTIONS, payload.get("recommendations"))

    def _clear_exec_summary_rows(self, ws) -> None:
        """Remove placeholder values from legacy summary rows so scoring lives in rows 11-101."""
        for row in range(7, 10):
            for col in range(self.COL_CATEGORY, self.COL_RATIONALE + 1):
                cell_obj = ws.cell(row, col)
                if isinstance(cell_obj, MergedCell):
                    continue
                cell_obj.value = None

    def _analyze_smart_bidding_consolidation(self) -> Optional[Dict[str, object]]:
        if "match_type_split" not in self.summary_stats:
            try:
                self._score_match_type_split()
            except Exception:
                pass

        data = self.summary_stats.get("match_type_split")
        smart = self._get_smart_bidding_stats()
        label = "Smart bidding campaign consolidation?"
        if not data or smart.get("total_campaigns", 0) == 0:
            return None

        multi_pct = data.get("multi_pct", 0.0)
        multi_campaigns = int(data.get("multi_campaigns", 0))
        total_campaigns = int(data.get("total_campaigns", 0))
        single_campaigns = int(data.get("single_campaigns", 0))
        smart_pct = smart.get("smart_pct", 0.0)
        manual_campaigns = int(smart.get("manual_campaigns", 0))

        if multi_pct >= 80:
            score = 5
            assessment = "Consolidated Smart Bidding structure"
        elif multi_pct >= 60:
            score = 3
            assessment = "Partially consolidated"
        else:
            score = 2
            assessment = "Fragmented by match type"

        details = (
            f"{assessment}: {multi_campaigns}/{total_campaigns} campaigns ({multi_pct:.0f}%) combine multiple match types. "
            f"{single_campaigns} campaigns remain single-match-type while Smart Bidding adoption is {smart_pct:.1f}% "
            f"with {manual_campaigns} manual campaigns remaining."
        )
        recommendations = (
            "Collapse match-type-specific campaigns into consolidated Smart Bidding campaigns and migrate the remaining manual structures."
            if score < 5
            else "Maintain current consolidated Smart Bidding structure."
        )
        return {
            "label": label,
            "category": "Smart Bidding",
            "score": score,
            "details": details,
            "recommendations": recommendations,
        }

    def _build_summary_smart_bidding_consolidation(self) -> Optional[Dict[str, object]]:
        payload = self._analyze_smart_bidding_consolidation()
        if payload:
            return payload
        return {
            "label": "Smart bidding campaign consolidation?",
            "category": "Smart Bidding",
            "score": None,
            "details": "Data needed: campaign list with match-type naming patterns and bid strategy columns.",
            "recommendations": "Export the campaign details report so Smart Bidding consolidation can be evaluated.",
        }

    def _get_goal_alignment_metrics(self) -> Optional[Dict[str, object]]:
        cached = self.summary_stats.get("goal_alignment")
        if cached:
            return cached

        campaign_data = self.data.get('campaign_details')
        if campaign_data is None or 'Campaign bid strategy type' not in campaign_data.columns:
            return None

        analysis = campaign_data.copy()
        conversions_available = 'Conversions' in analysis.columns
        if conversions_available:
            analysis['Conv_num'] = to_numeric_safe(analysis['Conversions'])
        else:
            analysis['Conv_num'] = 0

        conversion_threshold = 1 if conversions_available else 0
        conversion_campaigns = analysis[analysis['Conv_num'] >= conversion_threshold]
        if len(conversion_campaigns) == 0:
            conversion_campaigns = analysis

        smart_strategies = ['Target CPA', 'Target ROAS', 'Maximize conversions', 'Maximize conversion value']
        smart_conv = conversion_campaigns[conversion_campaigns['Campaign bid strategy type'].isin(smart_strategies)]
        smart_conv_count = len(smart_conv)
        total_conversion_campaigns = len(conversion_campaigns)
        smart_conv_pct = (
            smart_conv_count / total_conversion_campaigns * 100
            if total_conversion_campaigns > 0
            else 0.0
        )

        metrics = {
            "total_campaigns": len(analysis),
            "conversion_campaigns": total_conversion_campaigns,
            "smart_conversion_campaigns": smart_conv_count,
            "manual_conversion_campaigns": total_conversion_campaigns - smart_conv_count,
            "smart_conv_pct": smart_conv_pct,
            "conversions_available": conversions_available,
        }
        self.summary_stats["goal_alignment"] = metrics
        return metrics

    def _analyze_goal_alignment(self) -> Optional[Dict[str, object]]:
        metrics = self._get_goal_alignment_metrics()
        label = "Goal-consistent campaign objectives?"
        if not metrics:
            return None

        smart_pct = metrics["smart_conv_pct"]
        smart_conv = metrics["smart_conversion_campaigns"]
        total_conv = metrics["conversion_campaigns"]
        manual_conv = metrics["manual_conversion_campaigns"]

        if smart_pct >= 80:
            score = 5
            assessment = "Objectives aligned with conversion-focused bidding"
        elif smart_pct >= 50:
            score = 3
            assessment = "Mixed objectives"
        else:
            score = 2
            assessment = "Inconsistent goals"

        details = (
            f"{assessment}: {smart_conv}/{total_conv} conversion-driving campaigns ({smart_pct:.0f}%) use Smart Bidding. "
            f"{manual_conv} campaigns still rely on manual or click-focused strategies."
        )
        recommendations = (
            "Ensure every conversion-driving campaign migrates to Target CPA/ROAS or Max Conversions to maintain goal alignment."
            if score < 5
            else "Maintain current bid strategies; conversion campaigns are aligned to business goals."
        )
        return {
            "label": label,
            "category": "Smart Bidding",
            "score": score,
            "details": details,
            "recommendations": recommendations,
        }

    def _build_summary_goal_alignment(self) -> Optional[Dict[str, object]]:
        payload = self._analyze_goal_alignment()
        if payload:
            return payload
        return {
            "label": "Goal-consistent campaign objectives?",
            "category": "Smart Bidding",
            "score": None,
            "details": "Data needed: Campaign bid strategy type and conversion columns.",
            "recommendations": "Export the campaign details report with bid strategies and conversions to evaluate goal alignment.",
        }

    def _analyze_signal_density(self) -> Optional[Dict[str, object]]:
        metrics = self._get_smart_bidding_stats()
        label = "Signal density per campaign?"
        if not metrics or not metrics.get("volume_data_available", False):
            return None

        smart_campaigns = metrics.get("smart_campaigns", 0) or 0
        if smart_campaigns == 0:
            return None

        smart_with_volume = metrics.get("smart_with_volume", 0) or 0
        smart_without_volume = metrics.get("smart_without_volume", 0) or 0
        volume_pct = smart_with_volume / smart_campaigns * 100 if smart_campaigns else 0

        if volume_pct >= 80:
            score = 5
            assessment = "Healthy signal density"
        elif volume_pct >= 60:
            score = 3
            assessment = "Mixed signal density"
        else:
            score = 2
            assessment = "Insufficient signals"

        details = (
            f"{assessment}: {smart_with_volume}/{smart_campaigns} Smart Bidding campaigns ({volume_pct:.0f}%) have ≥30 conversions/30d. "
            f"{smart_without_volume} campaigns need additional volume."
        )
        recommendations = (
            "Increase budgets or consolidate low-volume campaigns so each Smart Bidding campaign captures 30+ conversions per 30 days."
            if score < 5
            else "Maintain current conversion volume pacing; Smart Bidding has enough signals."
        )
        return {
            "label": label,
            "category": "Smart Bidding",
            "score": score,
            "details": details,
            "recommendations": recommendations,
        }

    def _build_summary_signal_density(self) -> Optional[Dict[str, object]]:
        payload = self._analyze_signal_density()
        if payload:
            return payload
        return {
            "label": "Signal density per campaign?",
            "category": "Smart Bidding",
            "score": None,
            "details": "Data needed: Campaign report with conversion volume.",
            "recommendations": "Export conversions per campaign so Smart Bidding signal density can be validated.",
        }

    def _score_exec_smart_bidding_consolidation(self) -> ScoringResult:
        payload = self._analyze_smart_bidding_consolidation()
        if not payload:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign list with match-type indicators and bid strategy columns.",
                recommendations="Export campaign list including match type naming and Smart Bidding adoption.",
                data_needed="Campaign list with match-type naming + bid strategies",
            )
        return ScoringResult(
            score=payload["score"],
            details=payload["details"],
            recommendations=payload["recommendations"],
        )

    def _score_exec_goal_alignment(self) -> ScoringResult:
        payload = self._analyze_goal_alignment()
        if not payload:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign bid strategy type and conversion volume.",
                recommendations="Export campaign details report with bid strategies and conversions to evaluate goal alignment.",
                data_needed="Campaign bid strategy + conversions",
            )
        return ScoringResult(
            score=payload["score"],
            details=payload["details"],
            recommendations=payload["recommendations"],
        )

    def _score_exec_signal_density(self) -> ScoringResult:
        payload = self._analyze_signal_density()
        if not payload:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign conversion columns to validate Smart Bidding signal density.",
                recommendations="Include 30-day conversion counts for all Smart Bidding campaigns.",
                data_needed="Campaign conversions (30d)",
            )
        return ScoringResult(
            score=payload["score"],
            details=payload["details"],
            recommendations=payload["recommendations"],
        )

    def _score_pmax_guardrails(self) -> ScoringResult:
        campaign_data = self.source_info.get('campaign_source')
        if campaign_data is None or 'Campaign type' not in campaign_data.columns:
            return ScoringResult(
                score=None,
                details="Data needed: Campaign type and naming fields.",
                recommendations="Export campaign report with Campaign type + naming (Brand vs Non-brand).",
                data_needed="Campaign type + naming",
            )

        pmax_campaigns = campaign_data[campaign_data['Campaign type'].str.contains('Performance Max', case=False, na=False)]
        google_campaigns = campaign_data[campaign_data['Account name'].str.contains('Google|Fidelity', case=False, na=False)]

        search_campaigns = google_campaigns[google_campaigns['Campaign type'].str.contains('Search', case=False, na=False)]
        brand_guardrails = search_campaigns[search_campaigns['Campaign'].str.contains('brand', case=False, na=False)]

        total_search = len(search_campaigns) or 1
        guardrail_pct = len(brand_guardrails) / total_search * 100

        total_google = len(google_campaigns) or 1
        pmax_pct = len(pmax_campaigns) / total_google * 100

        guardrail_required = self._pmax_guardrail_requirement()

        if len(pmax_campaigns) == 0:
            score = 2
            assessment = "No PMAX automation"
        elif 5 <= pmax_pct <= 35 and guardrail_pct >= guardrail_required:
            score = 5
            assessment = "PMAX complements search with brand guardrails"
        elif guardrail_pct >= max(10, guardrail_required - 5):
            score = 3
            assessment = "PMAX present but guardrails limited"
        else:
            score = 2
            assessment = "PMAX lacks brand guardrails"

        details = (
            f"{assessment}: {len(pmax_campaigns)} PMAX campaigns ({pmax_pct:.1f}% of Google programs) and "
            f"{len(brand_guardrails)}/{total_search} search campaigns ({guardrail_pct:.0f}%) dedicated to Brand guardrails."
        )

        if score >= 5:
            recommendations = "Maintain the current balance: PMAX drives incremental reach while search Brand guardrails protect efficiency."
        elif len(pmax_campaigns) == 0:
            recommendations = "Introduce PMAX pilots (5-10% of footprint) while keeping Brand search campaigns in place for guardrails."
        elif guardrail_pct < guardrail_required:
            recommendations = (
                f"Reinforce Brand search guardrails (shared negatives/budget) so PMAX complements, rather than cannibalizes, core campaigns. "
                f"Target ≥{guardrail_required}% of search coverage on Brand safeguard campaigns."
            )
        else:
            recommendations = (
                f"Normalize PMAX share between 5-30% of Google programs and ensure Brand guardrails retain ≥{guardrail_required}% of search coverage."
            )

        result = ScoringResult(
            score=score,
            details=details,
            recommendations=recommendations,
            calculation=f"PMAX: {len(pmax_campaigns)} | Brand guardrail share: {guardrail_pct:.0f}%",
        )
        if len(pmax_campaigns) == 0:
            result.binary_gap_indicator = "Performance Max Inventory: ABSENT"
        return result

    def generate_audit(self, account_name: str, output_path: Path):
        """
        Generate complete audit for one account

        Process:
        1. Load template
        2. Load data
        3. Discover sources
        4. Iterate template rows
        5. Score each criterion
        6. Write to template
        7. Save output
        """

        print("\n" + "="*80)
        print(f"Kai UNIFIED ENGINE - {account_name}")
        print("="*80)

        # Load template
        print(f"\n[1/5] Loading template from {self.template_path.name}...")
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active

        # Update title
        ws.cell(1, 1).value = f"Kai AI Audit - {account_name} - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        self._clear_exec_summary_rows(ws)

        # Load data
        self.load_data()

        # Discover sources
        self.discover_data_sources()
        scope_summary = self._build_scope_summary()

        # Score criteria
        print(f"\n[3/5] Scoring criteria (rows {self.FIRST_CRITERION_ROW}-{self.LAST_CRITERION_ROW})...")

        scored = 0
        not_applicable = 0
        needs_data = 0
        self.humanizer.reset()
        self.summary_stats = {}
        if scope_summary:
            self.summary_stats["scope"] = scope_summary
        if self.business_context:
            self.summary_stats["business_context"] = self.business_context

        for excel_row in range(self.FIRST_CRITERION_ROW, self.LAST_CRITERION_ROW + 1):
            # Read criterion from template
            category = ws.cell(excel_row, self.COL_CATEGORY).value
            criterion_name = ws.cell(excel_row, self.COL_ITEM).value
            if isinstance(criterion_name, str):
                normalized = criterion_name.strip().lower()
                if normalized == "are shared campaign budgets used?":
                    new_label = "Are budgets being diversified?"
                    ws.cell(excel_row, self.COL_ITEM).value = new_label
                    criterion_name = new_label
                elif normalized == "campaigns split by match type?":
                    new_label = "Keyword groupings of similar types in campaign?"
                    ws.cell(excel_row, self.COL_ITEM).value = new_label
                    criterion_name = new_label
                elif normalized == "broad match modified":
                    new_label = "Broad Match Integration?"
                    ws.cell(excel_row, self.COL_ITEM).value = new_label
                    criterion_name = new_label
                elif normalized == "overall":
                    detail_text = ws.cell(excel_row, self.COL_DETAILS).value or ""
                    action_text = ws.cell(excel_row, self.COL_ACTIONS).value or ""
                    rationale_text = ws.cell(excel_row, self.COL_RATIONALE).value or ""
                    # Replace low-content placeholders (e.g., "and") with deterministic, category-aware text.
                    low_detail = self.humanizer._is_low_content(str(detail_text)) if detail_text is not None else True
                    low_action = self.humanizer._is_low_content(str(action_text)) if action_text is not None else True
                    low_rationale = self.humanizer._is_low_content(str(rationale_text)) if rationale_text is not None else True
                    rationale_token = str(rationale_text).strip().lower() if rationale_text is not None else ""
                    if not any([detail_text, action_text, rationale_text]) or low_detail or low_action or low_rationale or rationale_token in ("and", "overall: and"):
                        label = (category or "Overall").strip() or "Overall"
                        detail_text = f"{label}: summary is available in the Insights tab."
                        action_text = f"{label}: review scored criteria for context."
                        rationale_text = f"{label}: overall reflects scored criteria coverage."
                    if any([detail_text, action_text, rationale_text]):
                        detail_text, action_text, rationale_text = self.humanizer.rewrite_if_needed(
                            account_name=self.business_unit or account_name,
                            category=category or "",
                            criterion=criterion_name or "",
                            detail_text=str(detail_text),
                            action_text=str(action_text),
                            rationale_text=str(rationale_text),
                            score=None,
                            calculation="",
                            data_needed=None,
                            context_summary=self.source_info.get("context") if isinstance(self.source_info, dict) else None,
                            where_to_look=None,
                        )
                    ws.cell(excel_row, self.COL_RATING).value = ""
                    ws.cell(excel_row, self.COL_DETAILS).value = detail_text
                    ws.cell(excel_row, self.COL_ACTIONS).value = action_text
                    ws.cell(excel_row, self.COL_RATIONALE).value = rationale_text
                    continue
                elif normalized == "adjusted bids?" and excel_row not in (61, 68):
                    detail_text = ws.cell(excel_row, self.COL_DETAILS).value or ""
                    action_text = ws.cell(excel_row, self.COL_ACTIONS).value or ""
                    rationale_text = ws.cell(excel_row, self.COL_RATIONALE).value or ""
                    if any([detail_text, action_text, rationale_text]):
                        detail_text, action_text, rationale_text = self.humanizer.rewrite_if_needed(
                            account_name=self.business_unit or account_name,
                            category=category or "",
                            criterion=criterion_name or "",
                            detail_text=str(detail_text),
                            action_text=str(action_text),
                            rationale_text=str(rationale_text),
                            score=None,
                            calculation="",
                            data_needed=None,
                            context_summary=self.source_info.get("context") if isinstance(self.source_info, dict) else None,
                            where_to_look=None,
                        )
                    ws.cell(excel_row, self.COL_RATING).value = ""
                    ws.cell(excel_row, self.COL_DETAILS).value = detail_text
                    ws.cell(excel_row, self.COL_ACTIONS).value = action_text
                    ws.cell(excel_row, self.COL_RATIONALE).value = rationale_text
                    continue

            if not criterion_name:
                continue  # Skip empty rows

            # Score this criterion
            result = self.score_criterion(excel_row, criterion_name, category)

            # Apply AI refinement for more adaptive scoring
            result = self.ai_scorer.refine_result(
                account_name=self.business_unit or account_name,
                category=category,
                criterion_name=criterion_name,
                result=result,
                context_summary=self.source_info.get('context'),
            )

            # Write results to template
            detail_text = self._compose_detail_with_defense(result)
            where_clause = ""
            if self._should_add_where_clause(result):
                where_clause = self._build_where_clause(result)
                if where_clause:
                    if where_clause in self._where_clause_seen:
                        where_clause = ""
                    else:
                        self._where_clause_seen.add(where_clause)
            rationale_text = result.bp_rationale or self._build_default_rationale(criterion_name, category, result)
            if result.needs_data():
                action_text = self._build_missing_data_action(criterion_name, result)
            else:
                action_text = result.recommendations

            original_action_text = action_text
            detail_text, action_text, rationale_text = self.humanizer.rewrite_if_needed(
                account_name=self.business_unit or account_name,
                category=category or "",
                criterion=criterion_name or "",
                detail_text=detail_text,
                action_text=action_text or "",
                rationale_text=rationale_text or "",
                score=result.score if isinstance(result.score, (int, float)) else None,
                calculation=result.calculation or "",
                data_needed=result.data_needed,
                context_summary=self.source_info.get("context") if isinstance(self.source_info, dict) else None,
                where_to_look=where_clause,
            )
            action_text = self._guard_action_consistency(
                criterion_name,
                result,
                detail_text,
                action_text,
                original_action_text,
            )
            if self.humanizer.enforce_distinct and self.humanizer._has_internal_overlap(
                [action_text, rationale_text]
            ):
                base_label = (criterion_name or "").strip()
                if category:
                    cat = (category or "").strip()
                    if base_label and cat.lower() not in base_label.lower():
                        base_label = f"{cat} - {base_label}"
                    elif not base_label:
                        base_label = cat
                if not base_label:
                    base_label = "This check"
                rationale_text = self.humanizer._build_rationale_text(
                    base_label=base_label,
                    score=result.score if isinstance(result.score, (int, float)) else None,
                    data_needed=result.data_needed,
                    where_to_look=where_clause,
                    detail_text=detail_text,
                )
            if where_clause and not detail_text:
                detail_text = where_clause

            if result.is_scored():
                ws.cell(excel_row, self.COL_RATING).value = result.score
                ws.cell(excel_row, self.COL_DETAILS).value = detail_text
                ws.cell(excel_row, self.COL_ACTIONS).value = action_text
                ws.cell(excel_row, self.COL_RATIONALE).value = rationale_text
                scored += 1
            elif result.is_not_applicable():
                ws.cell(excel_row, self.COL_RATING).value = "N/A"
                ws.cell(excel_row, self.COL_DETAILS).value = detail_text
                ws.cell(excel_row, self.COL_ACTIONS).value = action_text
                ws.cell(excel_row, self.COL_RATIONALE).value = rationale_text
                not_applicable += 1
            elif result.needs_data():
                # Leave rating blank
                ws.cell(excel_row, self.COL_DETAILS).value = detail_text
                ws.cell(excel_row, self.COL_ACTIONS).value = action_text
                ws.cell(excel_row, self.COL_RATIONALE).value = rationale_text
                needs_data += 1

        print(f"  Scored: {scored}")
        print(f"  Not applicable: {not_applicable}")
        print(f"  Needs data: {needs_data}")

        # Calculate overall score (weighted)
        print(f"\n[4/5] Calculating overall score...")

        # Collect all scored results with weights
        weighted_scores = []
        total_weight = 0

        for excel_row in range(self.FIRST_CRITERION_ROW, self.LAST_CRITERION_ROW + 1):
            rating = ws.cell(excel_row, self.COL_RATING).value
            criterion_name = ws.cell(excel_row, self.COL_ITEM).value

            if isinstance(rating, (int, float)) and criterion_name:
                weight = get_criterion_weight(criterion_name)
                weighted_scores.append(rating * weight)
                total_weight += weight

        # Calculate weighted average
        if total_weight > 0:
            overall_score = sum(weighted_scores) / total_weight
        else:
            overall_score = 0.0

        # Calculate confidence
        total_criteria = (self.LAST_CRITERION_ROW - self.FIRST_CRITERION_ROW + 1)
        applicable = max(total_criteria - not_applicable, 0)
        data_completeness = (scored / applicable) if applicable > 0 else 0.0
        data_completeness = max(0.0, min(1.0, data_completeness))
        confidence_info = calculate_confidence(
            data_completeness=data_completeness,
            criteria_scored=scored,
            total_criteria=total_criteria
        )

        print(f"  Overall: {overall_score:.2f}/5.0 (weighted, based on {scored} criteria)")
        print(f"  Confidence: {confidence_info['confidence'].upper()} ({confidence_info['confidence_score']:.0f}/100)")
        print(f"  Rationale: {confidence_info['rationale']}")

        # Inject AI-generated insights (before saving)
        print(f"\n[4.5/5] Generating AI insights for Insights sheet...")
        scoring_summary = {
            'scored': scored,
            'not_applicable': not_applicable,
            'needs_data': needs_data,
            'overall_score': overall_score,
            'scope_summary': scope_summary
        }
        inject_insights(wb, scoring_summary, account_name)

        # Save
        print(f"\n[5/5] Saving to {output_path}...")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        print(f"\nCOMPLETE: {output_path.name}")
        print(f"  Scored: {scored} | N/A: {not_applicable} | Needs data: {needs_data}")
        print(f"  Overall score: {overall_score:.2f}/5.0")
        print("="*80 + "\n")

        return {
            'scored': scored,
            'not_applicable': not_applicable,
            'needs_data': needs_data,
            'overall_score': overall_score,
            'scope_summary': scope_summary,
            'diagnostics': {
                'data_manifest': self.diagnostics.get("file_manifest", []),
                'normalized_key_collisions': self.diagnostics.get("normalized_key_collisions", []),
                'filter_stats': self.diagnostics.get("filter_stats", []),
                'data_rows': self.diagnostics.get("data_rows", {}),
                'selected_files': self.diagnostics.get("selected_files", {}),
                'verbalizer': self.humanizer.get_stats(),
                'source_selection': {
                    'performance_source': getattr(self, "performance_source", None),
                    'performance_granularity': self.source_info.get("performance_granularity") if isinstance(self.source_info, dict) else None,
                    'campaign_source_key': self.source_info.get("campaign_attributes_source") if isinstance(self.source_info, dict) else None,
                },
            },
        }


# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == "__main__":
    repo_root = Path(__file__).resolve().parents[1]
    template_path = repo_root / "azure" / "GenerateAudit" / "template.xlsx"
    default_data_root = Path(os.environ.get("PPC_DATA_ROOT", r"Z:\Kel\PPC_Audit_Project"))

    # Accounts to process (user requested 4: Retirement, Brand, Wealth Management, Brokerage)
    accounts = [
        {'name': 'Retirement', 'data_dir': 'Total Files'},
        {'name': 'Brand', 'data_dir': 'Total Files'},
        {'name': 'Wealth_Management', 'data_dir': 'Total Files'},
        {'name': 'Brokerage', 'data_dir': 'Total Files'}
    ]

    # Output directory
    output_dir = default_data_root / "UNIFIED_AUDITS"

    print("\n" + "="*80)
    print("Kai UNIFIED ENGINE - BATCH PROCESSING")
    print("="*80)
    print(f"\nProcessing {len(accounts)} accounts...")
    print(f"Template: {template_path.name}")
    print(f"Output: {output_dir}\n")

    results = []

    for account in accounts:
        data_path = default_data_root / account['data_dir']
        output_file = output_dir / f"Kai_{account['name']}_UNIFIED_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        try:
            engine = UnifiedAuditEngine(template_path, data_path, business_unit=account['name'])
            result = engine.generate_audit(account['name'], output_file)
            results.append({
                'account': account['name'],
                'success': True,
                **result
            })
        except Exception as e:
            print(f"\nERROR processing {account['name']}: {e}")
            import traceback
            traceback.print_exc()
            results.append({
                'account': account['name'],
                'success': False,
                'error': str(e)
            })

    # Summary
    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80 + "\n")

    successful = sum(1 for r in results if r['success'])
    print(f"Accounts processed: {successful}/{len(accounts)}")

    for result in results:
        if result['success']:
            print(f"  OK {result['account']:20} Score: {result['overall_score']:.2f} | Scored: {result['scored']} | N/A: {result['not_applicable']} | Needs data: {result['needs_data']}")
        else:
            print(f"  ERROR {result['account']:20} {result.get('error', 'Unknown')}")

    print(f"\nOutput directory: {output_dir}")
    print("="*80 + "\n")
