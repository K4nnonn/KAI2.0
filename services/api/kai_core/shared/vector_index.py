from __future__ import annotations

import logging
import os
import re
import time
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
from xml.etree import ElementTree as ET

from azure.core.credentials import AzureKeyCredential
from azure.search.documents import SearchClient
from openpyxl import load_workbook

from kai_core.shared.vector_search import get_embedding, load_vector_config
from kai_core.config import is_azure_embeddings_enabled

_FIRST_CRITERION_ROW = 11
_LAST_CRITERION_ROW = 92
_COL_CATEGORY = 4
_COL_ITEM = 5
_COL_DETAILS = 7
_COL_ACTIONS = 8
_COL_RATIONALE = 9
_HEADING_RE = re.compile(r"^(#{1,6})\s+(.*)$")


def _safe_id(value: str) -> str:
    # Azure Search keys cannot contain dots, so keep only alnum, underscore, dash.
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", value or "").strip("_")
    return safe or "unknown"


def _collect_insights(sheet) -> List[str]:
    texts: List[str] = []
    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.strip():
                texts.append(value.strip())
    return texts


def _build_content(details: Optional[str], actions: Optional[str], rationale: Optional[str]) -> str:
    parts = []
    if details:
        parts.append(f"Details: {details}")
    if actions:
        parts.append(f"Next steps: {actions}")
    if rationale:
        parts.append(f"Rationale: {rationale}")
    return "\n".join(parts).strip()


def _split_paragraphs(text: str, max_chars: int = 1800) -> List[str]:
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", text or "") if p.strip()]
    chunks: List[str] = []
    current: List[str] = []
    current_len = 0

    for paragraph in paragraphs:
        if len(paragraph) > max_chars:
            if current:
                chunks.append("\n\n".join(current).strip())
                current = []
                current_len = 0
            for i in range(0, len(paragraph), max_chars):
                chunk = paragraph[i : i + max_chars].strip()
                if chunk:
                    chunks.append(chunk)
            continue

        if current and (current_len + len(paragraph) + 2) > max_chars:
            chunks.append("\n\n".join(current).strip())
            current = [paragraph]
            current_len = len(paragraph)
        else:
            current.append(paragraph)
            current_len += len(paragraph) + 2

    if current:
        chunks.append("\n\n".join(current).strip())
    return chunks


def _read_docx_text(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as zf:
            xml_bytes = zf.read("word/document.xml")
    except Exception:
        return ""

    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return ""

    parts: List[str] = []
    for node in root.iter():
        tag = node.tag
        if tag.endswith("}t"):
            if node.text:
                parts.append(node.text)
        elif tag.endswith("}tab"):
            parts.append("\t")
        elif tag.endswith("}br") or tag.endswith("}cr"):
            parts.append("\n")
        elif tag.endswith("}p"):
            parts.append("\n")

    text = "".join(parts)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _chunk_markdown(text: str, max_chars: int = 1800) -> List[Dict[str, str]]:
    sections: List[Dict[str, str]] = []
    current_title = "Overview"
    buffer: List[str] = []
    for line in (text or "").splitlines():
        match = _HEADING_RE.match(line.strip())
        if match:
            if buffer:
                sections.append({"title": current_title, "body": "\n".join(buffer).strip()})
                buffer = []
            current_title = match.group(2).strip() or current_title
            continue
        buffer.append(line)
    if buffer:
        sections.append({"title": current_title, "body": "\n".join(buffer).strip()})

    if not sections and (text or "").strip():
        sections = [{"title": "Overview", "body": text.strip()}]

    chunks: List[Dict[str, str]] = []
    for section in sections:
        for part in _split_paragraphs(section["body"], max_chars=max_chars):
            if part:
                chunks.append({"title": section["title"], "content": part})
    return chunks


def extract_audit_documents(workbook_path: Path, account_name: str, audit_date: Optional[str] = None) -> List[Dict]:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb.active

    # Use second precision to avoid '.' in microseconds; colons are sanitized by _safe_id.
    audit_date = audit_date or datetime.utcnow().isoformat(timespec="seconds")
    account_safe = _safe_id(account_name)
    docs: List[Dict] = []

    for row in range(_FIRST_CRITERION_ROW, _LAST_CRITERION_ROW + 1):
        category = ws.cell(row, _COL_CATEGORY).value
        criterion = ws.cell(row, _COL_ITEM).value
        details = ws.cell(row, _COL_DETAILS).value
        actions = ws.cell(row, _COL_ACTIONS).value
        rationale = ws.cell(row, _COL_RATIONALE).value
        content = _build_content(
            details if isinstance(details, str) else None,
            actions if isinstance(actions, str) else None,
            rationale if isinstance(rationale, str) else None,
        )
        if not content:
            continue

        doc_id = _safe_id(f"{account_safe}_{audit_date}_row_{row}")
        docs.append(
            {
                "id": doc_id,
                "accountName": account_name,
                "auditDate": audit_date,
                "section": str(category or "Audit"),
                "title": str(criterion or "Audit Item"),
                "content": content,
                "source": "audit_workbook",
            }
        )

    if "Insights" in wb.sheetnames:
        insight_texts = _collect_insights(wb["Insights"])
        insight_texts = [t for t in insight_texts if t]
        if insight_texts:
            content = "\n".join(dict.fromkeys(insight_texts))
            doc_id = _safe_id(f"{account_safe}_{audit_date}_insights")
            docs.append(
                {
                    "id": doc_id,
                    "accountName": account_name,
                    "auditDate": audit_date,
                    "section": "Insights",
                    "title": "Insights Summary",
                    "content": content,
                    "source": "audit_workbook",
                }
            )
    return docs


def extract_knowledge_documents(
    root: Path,
    namespace: str = "knowledge_base",
    chunk_chars: Optional[int] = None,
) -> List[Dict]:
    if not root.exists():
        return []

    if chunk_chars is None:
        try:
            chunk_chars = int(os.environ.get("KNOWLEDGE_CHUNK_CHARS", "4000") or 4000)
        except ValueError:
            chunk_chars = 4000
    chunk_chars = max(500, chunk_chars)

    indexed_at = datetime.utcnow().isoformat(timespec="seconds")
    docs: List[Dict] = []
    md_mode = (os.environ.get("KNOWLEDGE_MD_CHUNK_MODE", "sectioned") or "sectioned").strip().lower()

    for path in sorted(root.glob("**/*")):
        suffix = path.suffix.lower()
        if suffix not in {".md", ".txt", ".docx"}:
            continue
        if suffix == ".docx":
            text = _read_docx_text(path)
        else:
            try:
                text = path.read_text(encoding="utf-8")
            except Exception:
                try:
                    text = path.read_text(encoding="utf-8", errors="ignore")
                except Exception:
                    continue

        if not text.strip():
            continue

        file_label = path.stem.replace("_", " ").strip() or path.stem
        if suffix == ".md":
            if md_mode == "flat":
                chunks = [{"title": file_label, "content": part} for part in _split_paragraphs(text, max_chars=chunk_chars)]
            else:
                chunks = _chunk_markdown(text, max_chars=chunk_chars)
        else:
            chunks = [{"title": file_label, "content": part} for part in _split_paragraphs(text, max_chars=chunk_chars)]

        for idx, chunk in enumerate(chunks, start=1):
            doc_id = _safe_id(f"{namespace}_{path.stem}_{idx}")
            docs.append(
                {
                    "id": doc_id,
                    "accountName": "KnowledgeBase",
                    "auditDate": indexed_at,
                    "section": file_label,
                    "title": chunk.get("title") or file_label,
                    "content": chunk.get("content") or "",
                    "source": "knowledge_base",
                }
            )
    return docs


def vector_indexing_enabled() -> bool:
    return os.environ.get("ENABLE_VECTOR_INDEXING", "false").strip().lower() in {"1", "true", "yes"}


def index_audit_workbook(workbook_path: Path, account_name: str, audit_date: Optional[str] = None) -> Dict[str, object]:
    cfg = load_vector_config()
    if not vector_indexing_enabled() or not cfg.enabled:
        return {"status": "skipped", "reason": "vector_indexing_disabled"}
    if not is_azure_embeddings_enabled() and not cfg.local_embedding_endpoint:
        return {"status": "skipped", "reason": "embeddings_disabled"}

    docs = extract_audit_documents(workbook_path, account_name, audit_date=audit_date)
    if not docs:
        return {"status": "skipped", "reason": "no_documents"}

    indexed = 0
    failed = 0
    for doc in docs:
        embedding = get_embedding(doc["content"], cfg)
        if not embedding:
            failed += 1
            continue
        doc[cfg.vector_field] = embedding
        indexed += 1

    if indexed == 0:
        return {"status": "error", "reason": "embeddings_failed", "docs": len(docs)}

    client = SearchClient(
        endpoint=cfg.search_endpoint,
        index_name=cfg.search_index,
        credential=AzureKeyCredential(cfg.search_key),
    )
    try:
        client.merge_or_upload_documents(docs)
        return {"status": "success", "indexed": indexed, "failed": failed, "docs": len(docs)}
    except Exception as exc:
        logging.warning("Vector index upload failed: %s", exc)
        return {"status": "error", "reason": str(exc), "indexed": indexed, "failed": failed, "docs": len(docs)}


def _purge_knowledge_documents(client: SearchClient, source: str = "knowledge_base") -> int:
    deleted = 0
    try:
        results = client.search(
            search_text="*",
            filter=f"source eq '{source}'",
            select=["id"],
            top=1000,
        )
        ids = [doc["id"] for doc in results if doc.get("id")]
        for i in range(0, len(ids), 500):
            batch = [{"id": doc_id} for doc_id in ids[i : i + 500]]
            client.delete_documents(documents=batch)
            deleted += len(batch)
    except Exception as exc:
        logging.warning("Vector index purge failed: %s", exc)
    return deleted


def index_knowledge_base(
    root: Path,
    namespace: str = "knowledge_base",
    purge_existing: bool = False,
) -> Dict[str, object]:
    cfg = load_vector_config()
    if not vector_indexing_enabled() or not cfg.enabled:
        return {"status": "skipped", "reason": "vector_indexing_disabled"}
    if not is_azure_embeddings_enabled() and not cfg.local_embedding_endpoint:
        return {"status": "skipped", "reason": "embeddings_disabled"}

    docs = extract_knowledge_documents(root, namespace=namespace)
    if not docs:
        return {"status": "skipped", "reason": "no_documents"}

    indexed = 0
    failed = 0
    try:
        delay = float(os.environ.get("KNOWLEDGE_EMBEDDING_DELAY_SECONDS", "0") or 0)
    except ValueError:
        delay = 0.0
    for doc in docs:
        if delay > 0:
            time.sleep(delay)
        embedding = get_embedding(doc["content"], cfg)
        if not embedding:
            failed += 1
            continue
        doc[cfg.vector_field] = embedding
        indexed += 1

    if indexed == 0:
        return {"status": "error", "reason": "embeddings_failed", "docs": len(docs)}

    client = SearchClient(
        endpoint=cfg.search_endpoint,
        index_name=cfg.search_index,
        credential=AzureKeyCredential(cfg.search_key),
    )
    try:
        purged = 0
        if purge_existing:
            purged = _purge_knowledge_documents(client, source="knowledge_base")
        client.merge_or_upload_documents(docs)
        result = {"status": "success", "indexed": indexed, "failed": failed, "docs": len(docs)}
        if purge_existing:
            result["purged"] = purged
        return result
    except Exception as exc:
        logging.warning("Vector index upload failed: %s", exc)
        return {"status": "error", "reason": str(exc), "indexed": indexed, "failed": failed, "docs": len(docs)}
